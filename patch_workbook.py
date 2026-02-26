#!/usr/bin/env python3
"""
Audit Patch — Adds all missing modules and fixes identified gaps.
Run AFTER generate_fnid_workbook.py or integrate into it.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# Import shared styles from main generator
from generate_fnid_workbook import (
    NAVY, DARK_BLUE, MED_BLUE, LIGHT_BLUE, GOLD, RED, GREEN, AMBER,
    WHITE, BLACK, LIGHT_GRAY, MED_GRAY,
    HEADER_FONT, HEADER_FILL, SUBHEADER_FILL, SUBHEADER_FONT,
    TITLE_FONT, SECTION_FONT, BODY_FONT, BODY_FONT_BOLD,
    CENTER, LEFT_WRAP, THIN_BORDER,
    GREEN_FILL, AMBER_FILL, RED_FILL, LIGHT_BLUE_FILL, LIGHT_GRAY_FILL, GOLD_FILL,
    TABLE_STYLE,
    write_headers, add_table, add_data_validation_list,
    _portal_header, _portal_kpi_summary, _portal_quick_stats,
    _dashboard_banner
)

# ══════════════════════════════════════════════════════════════════════
#  GAP 1 — PERSONNEL / ROSTER TABLE (Admin/HRM back-end)
# ══════════════════════════════════════════════════════════════════════

def build_personnel_table(wb):
    ws = wb.create_sheet("tbl_Personnel")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "PersonnelID","FullName","Rank","BadgeNo","Unit","Station",
        "DateOfEnlistment","DatePostedToFNID","CurrentRole","Specialization",
        "Status","Phone","Email","EmergencyContact",
        "QualificationsHeld","LastTrainingDate","NextTrainingDue",
        "PerformanceRating","DisciplinaryFlag",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,22,14,14,16,16,14,14,20,18,12,16,22,20,24,14,14,16,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","STAFF-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Personnel", ref)
    add_data_validation_list(ws, "C2:C1000", "=Lookups!$Q$2:$Q$11")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    # Training overdue
    ws.conditional_formatting.add(
        "Q2:Q1000",
        FormulaRule(formula=['=AND(Q2<>"",Q2<TODAY())'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 2 — TRAINING LOG TABLE (Training back-end)
# ══════════════════════════════════════════════════════════════════════

def build_training_table(wb):
    ws = wb.create_sheet("tbl_Training")
    ws.sheet_properties.tabColor = "548235"
    headers = [
        "TrainingID","Date","CourseTitle","TrainingType","Trainer","Location",
        "Duration_Hrs","Attendees","AttendeeNames","CertificateIssued",
        "ExpiryDate","LinkedUnit","CostJMD","FundingSource",
        "Outcome","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,24,18,18,20,12,10,30,14,14,14,12,16,24,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","TRN-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=21, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=22, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=23, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=24, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Training", ref)
    add_data_validation_list(ws, "L2:L1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "J2:J1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 3 — COURT TRACKING TABLE
# ══════════════════════════════════════════════════════════════════════

def build_court_dates_table(wb):
    ws = wb.create_sheet("tbl_CourtDates")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "CourtID","LinkedCaseID","CourtName","CourtType","HearingDate","HearingTime",
        "HearingType","JudgeMagistrate","Prosecutor","DefenseAttorney",
        "Defendant","ChargesSummary",
        "WitnessesRequired","WitnessesAttended","ExhibitsRequired",
        "Outcome","AdjournmentReason","NextHearingDate",
        "BailStatus","BailAmount","RemandStatus",
        "Verdict","Sentence","SentenceDate",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,20,14,14,12,16,20,18,18,20,28,
              20,14,18,18,24,14,14,14,14,16,20,14,
              24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","CRT-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=30, value=f'=IF(E{r}="","",WEEKNUM(E{r}))')
        ws.cell(row=r, column=31, value=f'=IF(E{r}="","",MONTH(E{r}))')
        ws.cell(row=r, column=32, value=f'=IF(E{r}="","",ROUNDUP(MONTH(E{r})/3,0))')
        ws.cell(row=r, column=33, value=f'=IF(E{r}="","",YEAR(E{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_CourtDates", ref)
    # Upcoming hearing highlight (within 7 days)
    ws.conditional_formatting.add(
        "E2:E1000",
        FormulaRule(formula=['=AND(E2<>"",E2>=TODAY(),E2<=TODAY()+7)'], fill=AMBER_FILL)
    )
    # Overdue (past hearing with no outcome)
    ws.conditional_formatting.add(
        "E2:E1000",
        FormulaRule(formula=['=AND(E2<>"",E2<TODAY(),P2="")'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 5 — ARREST REGISTER
# ══════════════════════════════════════════════════════════════════════

def build_arrests_table(wb):
    ws = wb.create_sheet("tbl_Arrests")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "ArrestID","LinkedOpsID","LinkedCaseID","ArrestDate","ArrestTime",
        "ArrestLocation","Parish","Station",
        "SuspectName","SuspectDOB","SuspectAddress","SuspectGender",
        "ChargeDescription","OffenceType","ArrestingOfficer","ArrestingOfficerBadge",
        "CautionGiven","RightsRead","BailGranted","BailAmount",
        "RemandFacility","RemandDate","CourtDate",
        "Status","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,14,12,22,16,16,22,14,24,10,28,22,20,16,12,12,12,14,18,14,14,14,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","ARR-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=30, value=f'=IF(D{r}="","",WEEKNUM(D{r}))')
        ws.cell(row=r, column=31, value=f'=IF(D{r}="","",MONTH(D{r}))')
        ws.cell(row=r, column=32, value=f'=IF(D{r}="","",ROUNDUP(MONTH(D{r})/3,0))')
        ws.cell(row=r, column=33, value=f'=IF(D{r}="","",YEAR(D{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Arrests", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "N2:N1000", "=Lookups!$C$2:$C$22")
    add_data_validation_list(ws, "Q2:Q1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "R2:R1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "S2:S1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 4 — SEIZURES REGISTER (Individual seized items)
# ══════════════════════════════════════════════════════════════════════

def build_seizures_table(wb):
    ws = wb.create_sheet("tbl_Seizures")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "SeizureID","LinkedOpsID","LinkedCaseID","LinkedArrestID",
        "SeizureDate","SeizureLocation","Parish",
        "ItemType","ItemDescription","Quantity","UnitMeasure",
        "SerialNumber","Make","Model","Calibre",
        "EstimatedValue","LinkedExhibitID","PhotoRef",
        "SeizedBy","WitnessPresent",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,14,14,22,16,18,28,10,12,16,14,14,12,14,14,14,18,18,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","SEZ-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=26, value=f'=IF(E{r}="","",WEEKNUM(E{r}))')
        ws.cell(row=r, column=27, value=f'=IF(E{r}="","",MONTH(E{r}))')
        ws.cell(row=r, column=28, value=f'=IF(E{r}="","",ROUNDUP(MONTH(E{r})/3,0))')
        ws.cell(row=r, column=29, value=f'=IF(E{r}="","",YEAR(E{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Seizures", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "H2:H1000", "=Lookups!$K$2:$K$16")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 8 — WARRANT REGISTER
# ══════════════════════════════════════════════════════════════════════

def build_warrants_table(wb):
    ws = wb.create_sheet("tbl_Warrants")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "WarrantID","WarrantType","LinkedCaseID","LinkedLeadID","LinkedOpsID",
        "ApplicationDate","IssuedDate","IssuedBy","JP_Name","Court",
        "TargetName","TargetAddress","Parish",
        "ValidUntil","ExecutionDate","ExecutedBy","ExecutionOutcome",
        "ReturnToJP_Date","ReturnToJP_By","Status",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,16,14,14,14,14,14,16,18,18,22,24,16,14,14,18,24,14,18,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","WAR-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Warrants", ref)
    add_data_validation_list(ws, "M2:M1000", "=Lookups!$D$2:$D$16")
    # Expired warrant
    ws.conditional_formatting.add(
        "N2:N1000",
        FormulaRule(formula=['=AND(N2<>"",N2<TODAY(),O2="")'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 6 — INFORMANT MANAGEMENT
# ══════════════════════════════════════════════════════════════════════

def build_informants_table(wb):
    ws = wb.create_sheet("tbl_Informants")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "InformantID","CodeName","Handler","HandlerBadge","Unit",
        "RegistrationDate","Credibility","Parish","Area_Coverage",
        "TotalTipsProvided","TipsConvertedToCase","TipsConvertedToOp",
        "ConversionRate","LastContactDate","NextContactDue",
        "PaymentsToDate","LastPaymentDate","LastPaymentAmount",
        "RiskToInformant","Status",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,16,18,14,14,14,18,16,20,14,16,16,14,14,14,14,14,14,14,12,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","INF-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=13, value=f'=IF(OR(J{r}="",J{r}=0),"",((K{r}+L{r})/J{r}))')
        ws.cell(row=r, column=13).number_format = '0%'
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Informants", ref)
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$J$2:$J$7")
    add_data_validation_list(ws, "H2:H1000", "=Lookups!$D$2:$D$16")
    # Overdue contact
    ws.conditional_formatting.add(
        "O2:O1000",
        FormulaRule(formula=['=AND(O2<>"",O2<TODAY())'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 7 — DUTY ROSTER / SHIFT TABLE
# ══════════════════════════════════════════════════════════════════════

def build_duty_roster(wb):
    ws = wb.create_sheet("tbl_DutyRoster")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "RosterID","Date","ShiftType","StartTime","EndTime",
        "OfficerName","Rank","BadgeNo","Unit","Station",
        "Assignment","PostLocation","Supervisor",
        "OvertimeHrs","LeaveType","Replacement",
        "Notes","CreatedBy","CreatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,12,12,22,14,14,14,16,20,20,18,12,14,18,24,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","DUTY-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=20, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=21, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=22, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=23, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_DutyRoster", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$Q$2:$Q$11")
    add_data_validation_list(ws, "I2:I1000", "=Lookups!$E$2:$E$11")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 10 — CHAIN OF CUSTODY TRANSFER LOG
# ══════════════════════════════════════════════════════════════════════

def build_custody_transfers(wb):
    ws = wb.create_sheet("tbl_CustodyTransfers")
    ws.sheet_properties.tabColor = "538135"
    headers = [
        "TransferID","ExhibitID","LinkedCaseID",
        "TransferDate","TransferTime",
        "ReleasedBy","ReleasedByRank","ReleasedByBadge",
        "ReceivedBy","ReceivedByRank","ReceivedByBadge",
        "Purpose","FromLocation","ToLocation",
        "ConditionAtTransfer","WitnessName","WitnessBadge",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,14,14,12,18,14,14,18,14,14,24,20,20,18,18,14,24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","COC-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_CustodyTransfers", ref)
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 9 — SCENE PROCESSING LOG
# ══════════════════════════════════════════════════════════════════════

def build_scene_log(wb):
    ws = wb.create_sheet("tbl_SceneLog")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "SceneID","LinkedCaseID","LinkedOpsID","Date","Time",
        "Location","Parish","GPS_Coordinates",
        "SceneType","SceneOfficer","Rank","BadgeNo",
        "PhotosTaken","VideoRecorded","SketchDrawn",
        "GSR_Collected","ShellCasingsFound","ShellCasingCount",
        "BloodSamplesCollected","FingerprintsLifted",
        "WeaponsFound","WeaponDetails",
        "WitnessesAtScene","WeatherConditions","Lighting",
        "SceneSealedTime","SceneReleasedTime",
        "OtherEvidenceNotes","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,14,12,24,16,18,16,18,14,14,12,12,12,12,14,12,14,14,12,24,20,16,12,14,14,28,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","SCN-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_SceneLog", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "M2:M1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "N2:N1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "O2:O1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "P2:P1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  GAP 11 — ADD MISSING DATA VALIDATIONS TO EXISTING TABLES
# ══════════════════════════════════════════════════════════════════════

def patch_missing_validations(wb):
    """Add data validations that were missed in original build."""

    # tbl_Actions — Status dropdown
    ws = wb["tbl_Actions"]
    # Add ActionStatus to Lookups first
    lk = wb["Lookups"]
    lk.cell(row=1, column=21, value="ActionStatus").font = HEADER_FONT
    lk.cell(row=1, column=21).fill = HEADER_FILL
    for i, s in enumerate(["Pending","In Progress","Completed","Cancelled","On Hold","Deferred"], 2):
        lk.cell(row=i, column=21, value=s)

    # ActionType lookup
    lk.cell(row=1, column=22, value="ActionType").font = HEADER_FONT
    lk.cell(row=1, column=22).fill = HEADER_FILL
    for i, s in enumerate(["Interview Witness","Interview Suspect","Collect Evidence","Submit Exhibit",
                            "Conduct Search","Surveillance","Arrest","Court Preparation",
                            "Lab Follow-Up","Supervisory Review","Report Writing","Other"], 2):
        lk.cell(row=i, column=22, value=s)

    # ReviewType lookup
    lk.cell(row=1, column=23, value="ReviewType").font = HEADER_FONT
    lk.cell(row=1, column=23).fill = HEADER_FILL
    for i, s in enumerate(["Supervisory Review","Peer Review","Compliance Audit","Quality Check","Command Review","Monthly Review"], 2):
        lk.cell(row=i, column=23, value=s)

    # HearingType lookup
    lk.cell(row=1, column=24, value="HearingType").font = HEADER_FONT
    lk.cell(row=1, column=24).fill = HEADER_FILL
    for i, s in enumerate(["Mention","Preliminary Inquiry","Trial","Sentencing","Bail Hearing","Appeal","Other"], 2):
        lk.cell(row=i, column=24, value=s)

    # ShiftType lookup
    lk.cell(row=1, column=25, value="ShiftType").font = HEADER_FONT
    lk.cell(row=1, column=25).fill = HEADER_FILL
    for i, s in enumerate(["Day Shift","Night Shift","Evening Shift","24-Hour Duty","On Call","Rest Day"], 2):
        lk.cell(row=i, column=25, value=s)

    # WarrantType lookup
    lk.cell(row=1, column=26, value="WarrantType").font = HEADER_FONT
    lk.cell(row=1, column=26).fill = HEADER_FILL
    for i, s in enumerate(["Search Warrant","Arrest Warrant","Bench Warrant","Seizure Order","Surveillance Order"], 2):
        lk.cell(row=i, column=26, value=s)

    # WarrantStatus lookup
    lk.cell(row=1, column=27, value="WarrantStatus").font = HEADER_FONT
    lk.cell(row=1, column=27).fill = HEADER_FILL
    for i, s in enumerate(["Applied","Issued","Executed","Returned to JP","Expired","Cancelled"], 2):
        lk.cell(row=i, column=27, value=s)

    # BailStatus lookup
    lk.cell(row=1, column=28, value="BailStatus").font = HEADER_FONT
    lk.cell(row=1, column=28).fill = HEADER_FILL
    for i, s in enumerate(["Granted","Denied","Not Applied","Revoked","Surety Pending"], 2):
        lk.cell(row=i, column=28, value=s)

    # PersonnelStatus lookup
    lk.cell(row=1, column=29, value="PersonnelStatus").font = HEADER_FONT
    lk.cell(row=1, column=29).fill = HEADER_FILL
    for i, s in enumerate(["Active","On Leave","Seconded","Suspended","Transferred","Retired","Deceased"], 2):
        lk.cell(row=i, column=29, value=s)

    # SceneType lookup
    lk.cell(row=1, column=30, value="SceneType").font = HEADER_FONT
    lk.cell(row=1, column=30).fill = HEADER_FILL
    for i, s in enumerate(["Shooting Scene","Drug Lab","Stash House","Arrest Scene","Vehicle Search","Residential Search","Commercial Search","Open Area","Other"], 2):
        lk.cell(row=i, column=30, value=s)

    lk.column_dimensions[get_column_letter(21)].width = 22
    lk.column_dimensions[get_column_letter(22)].width = 22
    lk.column_dimensions[get_column_letter(23)].width = 22
    lk.column_dimensions[get_column_letter(24)].width = 22
    lk.column_dimensions[get_column_letter(25)].width = 22
    lk.column_dimensions[get_column_letter(26)].width = 22
    lk.column_dimensions[get_column_letter(27)].width = 22
    lk.column_dimensions[get_column_letter(28)].width = 22
    lk.column_dimensions[get_column_letter(29)].width = 22
    lk.column_dimensions[get_column_letter(30)].width = 22

    # Now apply validations
    add_data_validation_list(ws, "C2:C1000", "=Lookups!$V$2:$V$13")  # ActionType
    add_data_validation_list(ws, "I2:I1000", "=Lookups!$U$2:$U$7")   # ActionStatus
    add_data_validation_list(ws, "F2:F1000", "=Lookups!$E$2:$E$11")  # Unit

    # tbl_Reviews — ReviewType and Status
    ws_rev = wb["tbl_Reviews"]
    add_data_validation_list(ws_rev, "C2:C1000", "=Lookups!$W$2:$W$7")  # ReviewType
    add_data_validation_list(ws_rev, "I2:I1000", "=Lookups!$U$2:$U$7")  # Status (reuse ActionStatus)

    # tbl_Warrants — WarrantType and Status
    ws_war = wb["tbl_Warrants"]
    add_data_validation_list(ws_war, "B2:B1000", "=Lookups!$Z$2:$Z$6")  # WarrantType
    add_data_validation_list(ws_war, "T2:T1000", "=Lookups!$AA$2:$AA$7")  # WarrantStatus

    # tbl_DutyRoster — ShiftType
    ws_duty = wb["tbl_DutyRoster"]
    add_data_validation_list(ws_duty, "C2:C1000", "=Lookups!$Y$2:$Y$7")  # ShiftType

    # tbl_Personnel — Status
    ws_pers = wb["tbl_Personnel"]
    add_data_validation_list(ws_pers, "K2:K1000", "=Lookups!$AC$2:$AC$8")  # PersonnelStatus

    # tbl_CourtDates — HearingType and BailStatus
    ws_crt = wb["tbl_CourtDates"]
    add_data_validation_list(ws_crt, "G2:G1000", "=Lookups!$X$2:$X$8")  # HearingType
    add_data_validation_list(ws_crt, "S2:S1000", "=Lookups!$AB$2:$AB$6")  # BailStatus

    # tbl_SceneLog — SceneType
    ws_scn = wb["tbl_SceneLog"]
    add_data_validation_list(ws_scn, "I2:I1000", "=Lookups!$AD$2:$AD$10")  # SceneType


# ══════════════════════════════════════════════════════════════════════
#  GAP 12 — AGING / SLA FORMULAS
# ══════════════════════════════════════════════════════════════════════

def patch_aging_formulas(wb):
    """Add aging columns to Cases, Leads, Exhibits, and Actions."""

    # tbl_Cases — add DaysOpen column after Year (col AH = 34 → we'll add at col 35)
    ws = wb["tbl_Cases"]
    # Insert aging formula in the Notes column area (we'll use an auxiliary approach)
    # Since we can't easily insert columns into existing tables, we'll add DaysOpen
    # as a manual-reference formula in each portal/dashboard instead.
    # Instead, let's add to the executive dashboard as computed metrics.
    pass  # Handled in dashboard patches below


# ══════════════════════════════════════════════════════════════════════
#  GAP 13 — MONTHLY TREND ROWS IN EXECUTIVE DASHBOARD
# ══════════════════════════════════════════════════════════════════════

def patch_executive_trends(wb):
    """Add a 12-month trend section to the Executive Dashboard."""
    ws = wb["Dash_Executive"]
    # Find the last used row
    max_row = ws.max_row + 2
    r = max_row

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    cell = ws.cell(row=r, column=1, value="12-MONTH TREND ANALYSIS (Current Year)")
    cell.font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1

    # Header row: Metric | Jan | Feb | ... | Dec
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    ws.cell(row=r, column=1, value="Metric").font = SUBHEADER_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for i, m in enumerate(months):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1

    trend_metrics = [
        ("Cases Registered", 'COUNTIFS(tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
        ("Cases Closed", 'COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
        ("Leads Received", 'COUNTIFS(tbl_Leads[Month],{m},tbl_Leads[Year],YEAR(TODAY()))'),
        ("Operations Conducted", 'COUNTIFS(tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("FCSI Operations", 'COUNTIFS(tbl_Operations[FCSI_Flag],"Yes",tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Firearms Seized", 'SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Narcotics Seized (kg)", 'SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Arrests Made", 'SUMIFS(tbl_Operations[Arrests],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Exhibits Collected", 'COUNTIFS(tbl_Exhibits[Month],{m},tbl_Exhibits[Year],YEAR(TODAY()))'),
        ("Lectures Delivered", 'COUNTIFS(tbl_DemandReduction[Month],{m},tbl_DemandReduction[Year],YEAR(TODAY()))'),
        ("Attendance (Demand Red)", 'SUMIFS(tbl_DemandReduction[Attendance],tbl_DemandReduction[Month],{m},tbl_DemandReduction[Year],YEAR(TODAY()))'),
        ("Vehicle Trips", 'COUNTIFS(tbl_VehicleUsage[Month],{m},tbl_VehicleUsage[Year],YEAR(TODAY()))'),
    ]

    for metric_name, formula_template in trend_metrics:
        ws.cell(row=r, column=1, value=metric_name).font = BODY_FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        for m_idx in range(1, 13):
            formula = "=" + formula_template.format(m=m_idx)
            cell = ws.cell(row=r, column=m_idx+1, value=formula)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        r += 1

    r += 1
    # SLA / Aging summary
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    cell = ws.cell(row=r, column=1, value="AGING & SLA ANALYSIS")
    cell.font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1

    aging_metrics = [
        ("Avg Case Age (days)", '=IFERROR(AVERAGE(IF(tbl_Cases[CaseStatus]<>"Closed*",TODAY()-tbl_Cases[DateReceived])),"—")'),
        ("Oldest Open Case (days)", '=IFERROR(MAX(IF(tbl_Cases[CaseStatus]<>"Closed*",TODAY()-tbl_Cases[DateReceived])),"—")'),
        ("Avg Lead Age (days)", '=IFERROR(AVERAGE(IF(tbl_Leads[Outcome]="Pending",TODAY()-tbl_Leads[DateReceived])),"—")'),
        ("Avg Forensic Turnaround", '=IFERROR(AVERAGE(tbl_Exhibits[TurnaroundDays]),"—")'),
        ("Overdue Actions (count)", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Forensic Returns", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Cases > 90 days Open", '=COUNTIFS(tbl_Cases[DateReceived],"<"&TODAY()-90,tbl_Cases[CaseStatus],"<>Closed*")'),
        ("Leads Pending > 7 days", '=COUNTIFS(tbl_Leads[DateReceived],"<"&TODAY()-7,tbl_Leads[TriageDecision],"")'),
    ]

    for i, (label, formula) in enumerate(aging_metrics):
        col = (i % 4) * 3 + 1
        if i > 0 and i % 4 == 0:
            r += 2
        ws.cell(row=r, column=col, value=label).font = BODY_FONT_BOLD
        ws.cell(row=r, column=col).border = THIN_BORDER
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=MED_GRAY)
        ws.cell(row=r, column=col+1, value=formula).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        ws.cell(row=r, column=col+1).border = THIN_BORDER
        ws.cell(row=r, column=col+1).alignment = CENTER


# ══════════════════════════════════════════════════════════════════════
#  UPDATE HOME NAVIGATION TO INCLUDE NEW TABLES
# ══════════════════════════════════════════════════════════════════════

def patch_home_nav(wb):
    """Add new tables to the HOME navigation sheet."""
    ws = wb["HOME"]
    # Find the last used row
    max_row = ws.max_row + 2
    r = max_row

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell = ws.cell(row=r, column=1, value="NEW TABLES (Audit Patch — Gap Remediation)")
    cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor="C00000")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 26
    r += 1

    new_tables = [
        ("tbl_Personnel", "Personnel Register — Staff details, qualifications, postings"),
        ("tbl_Training", "Training Log — Courses, attendance, certifications, costs"),
        ("tbl_CourtDates", "Court Tracking — Hearings, outcomes, bail, sentencing"),
        ("tbl_Arrests", "Arrest Register — Detailed arrest records, charges, bail/remand"),
        ("tbl_Seizures", "Seizures Register — Individual seized items linked to ops/cases"),
        ("tbl_Warrants", "Warrant Register — Applications, issuance, execution, return to JP"),
        ("tbl_Informants", "Informant Management — Code names, handlers, reliability, payments"),
        ("tbl_DutyRoster", "Duty Roster — Shifts, assignments, overtime, leave tracking"),
        ("tbl_CustodyTransfers", "Chain of Custody Transfers — Exhibit handover audit trail"),
        ("tbl_SceneLog", "Scene Processing Log — Crime scene details, evidence collection"),
    ]

    for sheet_name, description in new_tables:
        ws.cell(row=r, column=1, value=f"  >> {sheet_name}").font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        ws.cell(row=r, column=1).alignment = Alignment(indent=2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2, value=description).font = Font(name="Calibri", size=10, color="333333")
        ws.cell(row=r, column=2).alignment = LEFT_WRAP
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="hair", color="CCCCCC"))
        r += 1


# ══════════════════════════════════════════════════════════════════════
#  UPDATE STATISTICS HUB WITH NEW TABLE METRICS
# ══════════════════════════════════════════════════════════════════════

def patch_statistics_hub(wb):
    """Add metrics for new tables to the Statistics Hub."""
    ws = wb["Statistics_Hub"]
    r = ws.max_row + 2

    new_sections = [
        ("COURT & LEGAL", "1F3864", [
            ("Total Court Dates", '=COUNTA(tbl_CourtDates[CourtID])-COUNTBLANK(tbl_CourtDates[CourtID])'),
            ("Upcoming (7 days)", '=COUNTIFS(tbl_CourtDates[HearingDate],">="&TODAY(),tbl_CourtDates[HearingDate],"<="&TODAY()+7)'),
            ("Adjournments", '=COUNTIFS(tbl_CourtDates[AdjournmentReason],"<>")'),
            ("Convictions", '=COUNTIF(tbl_CourtDates[Verdict],"Guilty")'),
            ("Acquittals", '=COUNTIF(tbl_CourtDates[Verdict],"Not Guilty")'),
            ("Bail Granted", '=COUNTIF(tbl_CourtDates[BailStatus],"Granted")'),
        ]),
        ("ARRESTS", "C00000", [
            ("Total Arrests", '=COUNTA(tbl_Arrests[ArrestID])-COUNTBLANK(tbl_Arrests[ArrestID])'),
            ("Arrests This Month", '=COUNTIFS(tbl_Arrests[Month],MONTH(TODAY()),tbl_Arrests[Year],YEAR(TODAY()))'),
            ("Bail Granted", '=COUNTIF(tbl_Arrests[BailGranted],"Yes")'),
            ("Remanded", '=COUNTIFS(tbl_Arrests[RemandFacility],"<>")'),
            ("Cautions Given", '=COUNTIF(tbl_Arrests[CautionGiven],"Yes")'),
            ("Rights Read", '=COUNTIF(tbl_Arrests[RightsRead],"Yes")'),
        ]),
        ("SEIZURES (DETAILED)", "C00000", [
            ("Total Seized Items", '=COUNTA(tbl_Seizures[SeizureID])-COUNTBLANK(tbl_Seizures[SeizureID])'),
            ("Firearms Seized", '=COUNTIF(tbl_Seizures[ItemType],"Firearm")'),
            ("Ammunition Seized", '=COUNTIF(tbl_Seizures[ItemType],"Ammunition")'),
            ("Narcotics Items", '=COUNTIFS(tbl_Seizures[ItemType],"Narcotics*")'),
            ("Cash/Currency Items", '=COUNTIF(tbl_Seizures[ItemType],"Cash/Currency")'),
            ("Linked to Exhibits", '=COUNTIFS(tbl_Seizures[LinkedExhibitID],"<>")'),
        ]),
        ("WARRANTS", "C00000", [
            ("Total Warrants", '=COUNTA(tbl_Warrants[WarrantID])-COUNTBLANK(tbl_Warrants[WarrantID])'),
            ("Issued", '=COUNTIF(tbl_Warrants[Status],"Issued")'),
            ("Executed", '=COUNTIF(tbl_Warrants[Status],"Executed")'),
            ("Expired (Unexecuted)", '=COUNTIF(tbl_Warrants[Status],"Expired")'),
            ("Returned to JP", '=COUNTIFS(tbl_Warrants[ReturnToJP_Date],"<>")'),
            ("Active (Not Expired)", '=COUNTIFS(tbl_Warrants[ValidUntil],">="&TODAY(),tbl_Warrants[ExecutionDate],"")'),
        ]),
        ("PERSONNEL & ROSTER", "7030A0", [
            ("Total Personnel", '=COUNTA(tbl_Personnel[PersonnelID])-COUNTBLANK(tbl_Personnel[PersonnelID])'),
            ("Active", '=COUNTIF(tbl_Personnel[Status],"Active")'),
            ("On Leave", '=COUNTIF(tbl_Personnel[Status],"On Leave")'),
            ("Training Overdue", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
            ("Duty Entries", '=COUNTA(tbl_DutyRoster[RosterID])-COUNTBLANK(tbl_DutyRoster[RosterID])'),
            ("Training Sessions", '=COUNTA(tbl_Training[TrainingID])-COUNTBLANK(tbl_Training[TrainingID])'),
        ]),
        ("INFORMANTS", "BF8F00", [
            ("Registered Informants", '=COUNTA(tbl_Informants[InformantID])-COUNTBLANK(tbl_Informants[InformantID])'),
            ("Active", '=COUNTIF(tbl_Informants[Status],"Active")'),
            ("Total Tips Provided", '=SUM(tbl_Informants[TotalTipsProvided])'),
            ("Tips→Cases", '=SUM(tbl_Informants[TipsConvertedToCase])'),
            ("Tips→Ops", '=SUM(tbl_Informants[TipsConvertedToOp])'),
            ("Contact Overdue", '=COUNTIFS(tbl_Informants[NextContactDue],"<"&TODAY())'),
        ]),
        ("SCENE PROCESSING & CUSTODY", "538135", [
            ("Scenes Processed", '=COUNTA(tbl_SceneLog[SceneID])-COUNTBLANK(tbl_SceneLog[SceneID])'),
            ("GSR Collected", '=COUNTIF(tbl_SceneLog[GSR_Collected],"Yes")'),
            ("Shell Casings Found", '=SUM(tbl_SceneLog[ShellCasingCount])'),
            ("Fingerprints Lifted", '=COUNTIF(tbl_SceneLog[FingerprintsLifted],"Yes")'),
            ("Custody Transfers", '=COUNTA(tbl_CustodyTransfers[TransferID])-COUNTBLANK(tbl_CustodyTransfers[TransferID])'),
            ("Weapons Found at Scene", '=COUNTIF(tbl_SceneLog[WeaponsFound],"Yes")'),
        ]),
    ]

    for section_title, color, metrics in new_sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

        col = 1
        for label, formula in metrics:
            ws.cell(row=r, column=col, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=MED_GRAY)
            ws.cell(row=r, column=col+1, value=formula).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
            ws.cell(row=r, column=col+1).border = THIN_BORDER
            ws.cell(row=r, column=col+1).alignment = CENTER
            col += 2
            if col > 12:
                col = 1
                r += 1
        r += 2


# ══════════════════════════════════════════════════════════════════════
#  UPDATE EXECUTIVE DASHBOARD WITH NEW METRICS
# ══════════════════════════════════════════════════════════════════════

def patch_executive_new_sections(wb):
    """Add Court/Arrest/Warrant/Personnel sections to Executive Dashboard."""
    ws = wb["Dash_Executive"]
    r = ws.max_row + 2

    # Court & Legal
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="COURT & LEGAL STATUS").font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    court_metrics = [
        ("Court Dates Total", '=COUNTA(tbl_CourtDates[CourtID])-COUNTBLANK(tbl_CourtDates[CourtID])'),
        ("Upcoming (7d)", '=COUNTIFS(tbl_CourtDates[HearingDate],">="&TODAY(),tbl_CourtDates[HearingDate],"<="&TODAY()+7)'),
        ("Warrants Active", '=COUNTIFS(tbl_Warrants[ValidUntil],">="&TODAY(),tbl_Warrants[ExecutionDate],"")'),
        ("Warrants Expired", '=COUNTIF(tbl_Warrants[Status],"Expired")'),
        ("Total Arrests", '=COUNTA(tbl_Arrests[ArrestID])-COUNTBLANK(tbl_Arrests[ArrestID])'),
        ("Seizure Items", '=COUNTA(tbl_Seizures[SeizureID])-COUNTBLANK(tbl_Seizures[SeizureID])'),
    ]
    for i, (label, _) in enumerate(court_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(court_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # Personnel & Training
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="PERSONNEL, TRAINING & INFORMANTS").font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    pers_metrics = [
        ("Personnel", '=COUNTA(tbl_Personnel[PersonnelID])-COUNTBLANK(tbl_Personnel[PersonnelID])'),
        ("Active Staff", '=COUNTIF(tbl_Personnel[Status],"Active")'),
        ("Training Sessions", '=COUNTA(tbl_Training[TrainingID])-COUNTBLANK(tbl_Training[TrainingID])'),
        ("Training Overdue", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
        ("Active Informants", '=COUNTIF(tbl_Informants[Status],"Active")'),
        ("Scenes Processed", '=COUNTA(tbl_SceneLog[SceneID])-COUNTBLANK(tbl_SceneLog[SceneID])'),
        ("Custody Transfers", '=COUNTA(tbl_CustodyTransfers[TransferID])-COUNTBLANK(tbl_CustodyTransfers[TransferID])'),
        ("Duty Entries", '=COUNTA(tbl_DutyRoster[RosterID])-COUNTBLANK(tbl_DutyRoster[RosterID])'),
    ]
    for i, (label, _) in enumerate(pers_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(pers_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER


# ══════════════════════════════════════════════════════════════════════
#  ADDITIONAL KPIs IN UNIT PLANS
# ══════════════════════════════════════════════════════════════════════

def patch_additional_kpis(wb):
    """Add KPIs for new modules to tbl_UnitPlans."""
    ws = wb["tbl_UnitPlans"]
    last_row = ws.max_row

    new_kpis = [
        ("Investigations","Court Readiness","Court Dates Attended","Scheduled hearings attended","Monthly","tbl_CourtDates"),
        ("Investigations","Court Readiness","Conviction Rate","% of court outcomes resulting in conviction","Monthly","tbl_CourtDates"),
        ("Operations","Arrest Quality","Arrest-to-Charge Rate","% of arrests with charges filed","Monthly","tbl_Arrests"),
        ("Operations","Seizure Effectiveness","Items Seized per Op","Avg seized items per operation","Monthly","tbl_Seizures"),
        ("Intelligence","Informant Management","Active Informants","Number of active registered informants","Monthly","tbl_Informants"),
        ("Intelligence","Informant Management","Informant Tip Conversion %","% of informant tips converted","Monthly","tbl_Informants"),
        ("Admin/HRM","Workforce Readiness","Duty Roster Completion","% of shifts filled vs required","Monthly","tbl_DutyRoster"),
        ("Admin/HRM","Workforce Readiness","Training Compliance Rate","% of personnel with current training","Monthly","tbl_Personnel"),
        ("Training","Capacity Building","Training Hours Delivered","Total hours of training delivered","Monthly","tbl_Training"),
        ("Training","Capacity Building","Training Cost (JMD)","Total training expenditure","Monthly","tbl_Training"),
        ("Oversight","Evidence Integrity","Custody Transfers Logged","Total chain-of-custody transfers","Monthly","tbl_CustodyTransfers"),
        ("Oversight","Evidence Integrity","Scene Processing Compliance","Scenes processed per shooting case","Monthly","tbl_SceneLog"),
    ]

    for i, (unit, obj, kpi_name, kpi_desc, freq, src) in enumerate(new_kpis):
        r = last_row + 1 + i
        kpi_num = 32 + i + 1
        ws.cell(row=r, column=1, value=f'KPI-{kpi_num:04d}')
        ws.cell(row=r, column=2, value=unit)
        ws.cell(row=r, column=3, value=obj)
        ws.cell(row=r, column=4, value=kpi_name)
        ws.cell(row=r, column=5, value=kpi_desc)
        ws.cell(row=r, column=6, value=freq)
        ws.cell(row=r, column=23, value=src)

        # Variance, Achievement, Traffic Light formulas
        ws.cell(row=r, column=14, value=f'=IF(OR(K{r}="",H{r}=""),"",K{r}-H{r})')
        ws.cell(row=r, column=15, value=f'=IF(OR(L{r}="",I{r}=""),"",L{r}-I{r})')
        ws.cell(row=r, column=16, value=f'=IF(OR(M{r}="",J{r}=""),"",M{r}-J{r})')
        ws.cell(row=r, column=17, value=f'=IF(OR(K{r}="",H{r}="",H{r}=0),"",K{r}/H{r})')
        ws.cell(row=r, column=18, value=f'=IF(OR(L{r}="",I{r}="",I{r}=0),"",L{r}/I{r})')
        ws.cell(row=r, column=19, value=f'=IF(OR(M{r}="",J{r}="",J{r}=0),"",M{r}/J{r})')
        for col_pct in [17, 18, 19]:
            ws.cell(row=r, column=col_pct).number_format = '0%'
        for col_ach, col_tl in [(17,20),(18,21),(19,22)]:
            cl = get_column_letter(col_ach)
            ws.cell(row=r, column=col_tl,
                    value=f'=IF({cl}{r}="","",IF({cl}{r}>=0.9,"On Track",IF({cl}{r}>=0.7,"At Risk","Off Track")))')
        for c in range(1, 30):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    # Update the table reference range
    # Remove old table and recreate with expanded range
    new_last_row = last_row + len(new_kpis)
    for tbl_name in list(ws.tables.keys()):
        del ws.tables[tbl_name]
    new_ref = f"A1:AC{new_last_row}"
    add_table(ws, "tbl_UnitPlans", new_ref)


# ══════════════════════════════════════════════════════════════════════
#  MAIN PATCH FUNCTION
# ══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("  FNID Area 3 — Audit Patch: Adding Missing Modules")
    print("=" * 70)

    fname = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"Loading {fname}...")
    wb = openpyxl.load_workbook(fname)

    print("[1/14] Adding tbl_Personnel (Admin/HRM back-end)...")
    build_personnel_table(wb)

    print("[2/14] Adding tbl_Training (Training back-end)...")
    build_training_table(wb)

    print("[3/14] Adding tbl_CourtDates (Court tracking)...")
    build_court_dates_table(wb)

    print("[4/14] Adding tbl_Arrests (Arrest register)...")
    build_arrests_table(wb)

    print("[5/14] Adding tbl_Seizures (Seizure items register)...")
    build_seizures_table(wb)

    print("[6/14] Adding tbl_Warrants (Warrant register)...")
    build_warrants_table(wb)

    print("[7/14] Adding tbl_Informants (Informant management)...")
    build_informants_table(wb)

    print("[8/14] Adding tbl_DutyRoster (Shift/roster tracking)...")
    build_duty_roster(wb)

    print("[9/14] Adding tbl_CustodyTransfers (Chain of custody log)...")
    build_custody_transfers(wb)

    print("[10/14] Adding tbl_SceneLog (Scene processing)...")
    build_scene_log(wb)

    print("[11/14] Adding missing data validations + 10 new Lookup columns...")
    patch_missing_validations(wb)

    print("[12/14] Adding 12 new KPIs to tbl_UnitPlans...")
    patch_additional_kpis(wb)

    print("[13/14] Adding 12-month trend analysis + aging/SLA to Executive Dashboard...")
    patch_executive_trends(wb)
    patch_executive_new_sections(wb)

    print("[14/14] Updating Statistics Hub + HOME navigation...")
    patch_statistics_hub(wb)
    patch_home_nav(wb)

    output_file = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"\nSaving patched workbook to: {output_file}")
    wb.save(output_file)
    print(f"\n{'=' * 70}")
    print(f"  AUDIT PATCH APPLIED SUCCESSFULLY")
    print(f"  File: {output_file}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  New sheets added:")
    new_sheets = ["tbl_Personnel","tbl_Training","tbl_CourtDates","tbl_Arrests",
                  "tbl_Seizures","tbl_Warrants","tbl_Informants","tbl_DutyRoster",
                  "tbl_CustodyTransfers","tbl_SceneLog"]
    for s in new_sheets:
        print(f"    + {s}")
    print(f"\n  Patches applied:")
    print(f"    + 10 new Lookup columns (ActionStatus, ActionType, ReviewType,")
    print(f"      HearingType, ShiftType, WarrantType, WarrantStatus,")
    print(f"      BailStatus, PersonnelStatus, SceneType)")
    print(f"    + Missing data validations on tbl_Actions, tbl_Reviews,")
    print(f"      tbl_Warrants, tbl_DutyRoster, tbl_Personnel, tbl_CourtDates, tbl_SceneLog")
    print(f"    + 12 new KPIs in tbl_UnitPlans (total: 44 KPIs)")
    print(f"    + 12-month trend analysis on Executive Dashboard")
    print(f"    + Aging/SLA metrics on Executive Dashboard")
    print(f"    + Court/Arrest/Warrant/Personnel metrics on Executive Dashboard")
    print(f"    + 7 new sections in Statistics Hub (42 new metrics)")
    print(f"    + HOME navigation updated with all new tables")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
