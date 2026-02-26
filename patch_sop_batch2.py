#!/usr/bin/env python3
"""
SOP Compliance Patch — Batch 2:
- Enhance tbl_Cases with SOP fields (rebuild with expanded columns)
- Enhance tbl_Reviews with mandatory schedule fields
- Add new SOP Lookup columns
- Rebuild Portal_Registry and Portal_Investigations with full SOP compliance metrics
- Add SOP compliance sections to Executive Dashboard & Statistics Hub
- Add SOP form templates (Investigation Plan, File Checklist, Witness Summons)
- Update HOME navigation
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from generate_fnid_workbook import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BODY_FONT_BOLD,
    CENTER, LEFT_WRAP, THIN_BORDER,
    GREEN_FILL, AMBER_FILL, RED_FILL, LIGHT_GRAY_FILL,
    TABLE_STYLE, NAVY, DARK_BLUE, LIGHT_BLUE, WHITE, MED_GRAY, MED_BLUE,
    write_headers, add_table, add_data_validation_list,
    SECTION_FONT, SUBHEADER_FONT, SUBHEADER_FILL, GOLD_FILL,
    _portal_header, _portal_kpi_summary, _portal_quick_stats,
)


# ══════════════════════════════════════════════════════════════════════
#  PART A — ADD NEW SOP LOOKUP COLUMNS
# ══════════════════════════════════════════════════════════════════════

def patch_sop_lookups(wb):
    lk = wb["Lookups"]

    lookups_to_add = {
        31: ("WitnessType", ["Eye Witness","Character Witness","Expert Witness",
                              "Police Witness","Forensic Witness","Civilian Witness",
                              "Complainant","Informant Witness"]),
        32: ("StatementType", ["Witness Statement","Cautioned Statement",
                                "Voluntary Statement","Dying Declaration",
                                "Expert Report","Victim Impact Statement",
                                "Identification Statement","Alibi Statement"]),
        33: ("AssignmentType", ["Initial Assignment","Reassignment","Temporary Cover",
                                 "Joint Assignment","Handover - Transfer",
                                 "Handover - Leave","Handover - Promotion","Escalation"]),
        34: ("ContactMethod", ["Phone Call","In Person","Email","WhatsApp",
                                "Letter","Station Visit","Home Visit","Other"]),
        35: ("PlanStatus", ["Draft","Active","Under Review","Revised","Closed","Archived"]),
        36: ("FileStatus", ["Complete","Substantially Complete","Incomplete",
                             "Critically Incomplete","Under Review"]),
        37: ("ConferenceType", ["Initial Case Conference","Progress Review",
                                 "Command Conference","Pre-Trial Conference",
                                 "Cold Case Review","Multi-Agency Conference",
                                 "Escalation Conference"]),
        38: ("DirectiveStatus", ["Issued","In Progress","Completed","Overdue","Cancelled"]),
        39: ("DisclosureMaterial", ["Witness Statement","Expert Report","Exhibit",
                                     "CCTV Footage","Forensic Report","Interview Record",
                                     "Intelligence Report","Medical Report","Phone Records",
                                     "Financial Records","Other Documentary"]),
        40: ("ReviewInterval", ["7-Day","14-Day","28-Day","56-Day","90-Day","Monthly","Quarterly"]),
        41: ("MandatoryReview", ["14-Day Initial Review","28-Day Progress Review",
                                  "56-Day Supervisor Review","90-Day Command Review",
                                  "Quarterly Quality Audit","Pre-Court Review",
                                  "Cold Case Review","File Closure Review"]),
        42: ("DisposalType", ["Convicted — Sentenced","Convicted — Fine",
                               "Acquitted","Discharged","Withdrawn by DPP",
                               "No Further Action","Referred to Other Agency",
                               "Complainant Withdrew","Accused Deceased",
                               "Cold Case — Filed"]),
    }

    for col_num, (col_name, values) in lookups_to_add.items():
        lk.cell(row=1, column=col_num, value=col_name).font = HEADER_FONT
        lk.cell(row=1, column=col_num).fill = HEADER_FILL
        for i, v in enumerate(values, 2):
            lk.cell(row=i, column=col_num, value=v)
        lk.column_dimensions[get_column_letter(col_num)].width = 24

    return lk


# ══════════════════════════════════════════════════════════════════════
#  PART B — ENHANCE tbl_Cases (add SOP columns to existing sheet)
# ══════════════════════════════════════════════════════════════════════

def patch_cases_sop_fields(wb):
    """Add SOP-required columns to tbl_Cases after the existing columns."""
    ws = wb["tbl_Cases"]

    # Current table ends at col AH (34 = Year). We'll add new columns starting at col 35.
    new_cols = [
        (35, "IncidentDate", 14),
        (36, "RegistrationSLA_Hrs", 14),
        (37, "StationDiaryRef", 16),
        (38, "InvestigationPlanID", 16),
        (39, "InvestigatorChangeCount", 16),
        (40, "LastReviewDate", 14),
        (41, "NextReviewDue", 14),
        (42, "ReviewInterval", 14),
        (43, "TotalStatements", 12),
        (44, "TotalWitnesses", 12),
        (45, "TotalExhibits", 12),
        (46, "FileCompletenessScore", 16),
        (47, "VictimLastContacted", 16),
        (48, "VictimNextContact", 16),
        (49, "ChargeDate", 14),
        (50, "ChargeDeadline_48hr", 16),
        (51, "ChargeDeadlineMet", 14),
        (52, "DisposalType", 16),
        (53, "DisposalDate", 14),
        (54, "DisposalApprovedBy", 18),
        (55, "FileLocation", 16),
        (56, "ArchiveRef", 14),
        (57, "DaysOpen", 12),
        (58, "AgingBand", 14),
    ]

    # Write new headers
    for col, name, width in new_cols:
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Add formulas for sample rows (2-6)
    for r in range(2, 7):
        # RegistrationSLA_Hrs = (DateReceived - IncidentDate) * 24
        ws.cell(row=r, column=36,
                value=f'=IF(OR(E{r}="",AI{r}=""),"",(E{r}-AI{r})*24)')
        # InvestigatorChangeCount = COUNTIF in CaseAssignments
        ws.cell(row=r, column=39,
                value=f'=IFERROR(COUNTIF(tbl_CaseAssignments[LinkedCaseID],A{r}),0)')
        # TotalStatements
        ws.cell(row=r, column=43,
                value=f'=IFERROR(COUNTIF(tbl_Statements[LinkedCaseID],A{r}),0)')
        # TotalWitnesses
        ws.cell(row=r, column=44,
                value=f'=IFERROR(COUNTIF(tbl_Witnesses[LinkedCaseID],A{r}),0)')
        # TotalExhibits
        ws.cell(row=r, column=45,
                value=f'=IFERROR(COUNTIF(tbl_Exhibits[LinkedCaseID],A{r}),0)')
        # FileCompletenessScore lookup
        ws.cell(row=r, column=46,
                value=f'=IFERROR(INDEX(tbl_CaseFileChecklist[CompletenessScore],MATCH(A{r},tbl_CaseFileChecklist[LinkedCaseID],0)),"—")')
        ws.cell(row=r, column=46).number_format = '0%'
        # ChargeDeadline_48hr = IncidentDate + 2 days
        ws.cell(row=r, column=50,
                value=f'=IF(AI{r}="","",AI{r}+2)')
        # ChargeDeadlineMet
        ws.cell(row=r, column=51,
                value=f'=IF(OR(AW{r}="",AX{r}=""),"",IF(AW{r}<=AX{r},"Yes","No"))')
        # DaysOpen
        ws.cell(row=r, column=57,
                value=f'=IF(OR(E{r}="",LEFT(O{r},6)="Closed"),"",TODAY()-E{r})')
        # AgingBand
        ws.cell(row=r, column=58,
                value=f'=IF(BE{r}="","",IF(BE{r}<=30,"0-30 days",IF(BE{r}<=60,"31-60 days",IF(BE{r}<=90,"61-90 days",IF(BE{r}<=180,"91-180 days","180+ days")))))')

        for col, _, _ in new_cols:
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).font = BODY_FONT

    # Remove old table and recreate with expanded range
    for tbl_name in list(ws.tables.keys()):
        del ws.tables[tbl_name]
    new_ref = f"A1:{get_column_letter(58)}6"
    add_table(ws, "tbl_Cases", new_ref)

    # Data validations for new columns
    add_data_validation_list(ws, f"{get_column_letter(42)}2:{get_column_letter(42)}1000",
                              "=Lookups!$AN$2:$AN$8")  # ReviewInterval
    add_data_validation_list(ws, f"{get_column_letter(52)}2:{get_column_letter(52)}1000",
                              "=Lookups!$AP$2:$AP$11")  # DisposalType

    # Conditional formatting — next review overdue
    ws.conditional_formatting.add(
        f"{get_column_letter(41)}2:{get_column_letter(41)}1000",
        FormulaRule(formula=[f'=AND({get_column_letter(41)}2<>"",{get_column_letter(41)}2<TODAY())'], fill=RED_FILL)
    )
    # Victim contact overdue
    ws.conditional_formatting.add(
        f"{get_column_letter(48)}2:{get_column_letter(48)}1000",
        FormulaRule(formula=[f'=AND({get_column_letter(48)}2<>"",{get_column_letter(48)}2<TODAY())'], fill=RED_FILL)
    )
    # 48-hour charge deadline missed
    ws.conditional_formatting.add(
        f"{get_column_letter(51)}2:{get_column_letter(51)}1000",
        FormulaRule(formula=[f'={get_column_letter(51)}2="No"'], fill=RED_FILL)
    )
    # File completeness < 50%
    ws.conditional_formatting.add(
        f"{get_column_letter(46)}2:{get_column_letter(46)}1000",
        FormulaRule(formula=[f'=AND({get_column_letter(46)}2<>"",{get_column_letter(46)}2<0.5)'], fill=RED_FILL)
    )
    # Aging band 180+
    ws.conditional_formatting.add(
        f"{get_column_letter(58)}2:{get_column_letter(58)}1000",
        FormulaRule(formula=[f'={get_column_letter(58)}2="180+ days"'], fill=RED_FILL)
    )


# ══════════════════════════════════════════════════════════════════════
#  PART C — ENHANCE tbl_Reviews (add mandatory schedule fields)
# ══════════════════════════════════════════════════════════════════════

def patch_reviews_sop(wb):
    ws = wb["tbl_Reviews"]

    new_cols = [
        (15, "MandatoryReviewType", 20),
        (16, "ReviewInterval_Days", 16),
        (17, "DirectivesIssued", 30),
        (18, "DirectiveAssignedTo", 18),
        (19, "DirectiveDueDate", 14),
        (20, "DirectiveStatus", 14),
        (21, "DirectiveCompletedDate", 16),
        (22, "EscalationRequired", 14),
        (23, "EscalatedTo", 18),
        (24, "NextReviewDue", 14),
        (25, "CaseDaysOpenAtReview", 16),
    ]

    for col, name, width in new_cols:
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for r in range(2, 7):
        # NextReviewDue = ReviewDate + ReviewInterval
        ws.cell(row=r, column=24,
                value=f'=IF(OR(E{r}="",P{r}=""),"",E{r}+P{r})')
        # CaseDaysOpenAtReview — needs case DateReceived via lookup
        ws.cell(row=r, column=25,
                value=f'=IFERROR(E{r}-INDEX(tbl_Cases[DateReceived],MATCH(B{r},tbl_Cases[CaseID],0)),"")')
        for col, _, _ in new_cols:
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).font = BODY_FONT

    # Rebuild table
    for tbl_name in list(ws.tables.keys()):
        del ws.tables[tbl_name]
    add_table(ws, "tbl_Reviews", f"A1:{get_column_letter(25)}6")

    add_data_validation_list(ws, "O2:O1000", "=Lookups!$AO$2:$AO$9")  # MandatoryReviewType
    add_data_validation_list(ws, "T2:T1000", "=Lookups!$AL$2:$AL$6")  # DirectiveStatus
    add_data_validation_list(ws, "V2:V1000", "=Lookups!$I$2:$I$3")    # EscalationRequired

    # Overdue directive
    ws.conditional_formatting.add(
        "S2:S1000",
        FormulaRule(formula=['=AND(S2<>"",S2<TODAY(),T2<>"Completed")'], fill=RED_FILL)
    )
    # Next review overdue
    ws.conditional_formatting.add(
        "X2:X1000",
        FormulaRule(formula=['=AND(X2<>"",X2<TODAY())'], fill=RED_FILL)
    )


# ══════════════════════════════════════════════════════════════════════
#  PART D — REBUILD PORTAL_REGISTRY WITH FULL SOP COMPLIANCE
# ══════════════════════════════════════════════════════════════════════

def rebuild_portal_registry(wb):
    # Delete old sheet and recreate
    if "Portal_Registry" in wb.sheetnames:
        del wb["Portal_Registry"]
    ws = wb.create_sheet("Portal_Registry")
    r = _portal_header(ws, "REGISTRY PORTAL",
                       "Case Registration, CR Compliance, SOP Adherence & File Management",
                       "Registry", "1F3864")

    # KPIs
    kpis = [
        (1, "Cases Registered"),
        (2, "CR Compliance Rate"),
        (3, "Avg Days to First Action"),
    ]
    r = _portal_kpi_summary(ws, r, "Registry", kpis)

    # Quick Stats — Original + Enhanced
    stats = [
        ("Total Cases", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
        ("Open Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Open")'),
        ("Active Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Active")'),
        ("Pending Review", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Review")'),
        ("Closed Cases", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*")'),
        ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
        ("Non-Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    # ── SOP COMPLIANCE SECTION ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="SOP COMPLIANCE METRICS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1

    sop_metrics = [
        ("CR1 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR1_Compliant],"Yes")/COUNTA(tbl_Cases[CR1_Compliant]),"—")'),
        ("CR2 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR2_Compliant],"Yes")/COUNTA(tbl_Cases[CR2_Compliant]),"—")'),
        ("CR5 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR5_Compliant],"Yes")/COUNTA(tbl_Cases[CR5_Compliant]),"—")'),
        ("All 3 CRs Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Compliant")'),
        ("Avg Registration SLA (hrs)", '=IFERROR(AVERAGE(tbl_Cases[RegistrationSLA_Hrs]),"—")'),
        ("Registration > 24hrs", '=COUNTIFS(tbl_Cases[RegistrationSLA_Hrs],">"&24)'),
        ("Files Complete (>95%)", '=COUNTIFS(tbl_Cases[FileCompletenessScore],">="&0.95)'),
        ("Files Incomplete (<50%)", '=COUNTIFS(tbl_Cases[FileCompletenessScore],"<"&0.5)'),
    ]
    for i in range(0, len(sop_metrics), 4):
        row_metrics = sop_metrics[i:i+4]
        for j, (label, _) in enumerate(row_metrics):
            cell = ws.cell(row=r, column=j*2+1, value=label)
            cell.font = BODY_FONT_BOLD
            cell.border = THIN_BORDER
            cell.fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_metrics):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── CASE AGING SECTION ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="CASE AGING ANALYSIS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    aging_bands = [
        ("0-30 days", '=COUNTIF(tbl_Cases[AgingBand],"0-30 days")'),
        ("31-60 days", '=COUNTIF(tbl_Cases[AgingBand],"31-60 days")'),
        ("61-90 days", '=COUNTIF(tbl_Cases[AgingBand],"61-90 days")'),
        ("91-180 days", '=COUNTIF(tbl_Cases[AgingBand],"91-180 days")'),
        ("180+ days", '=COUNTIF(tbl_Cases[AgingBand],"180+ days")'),
        ("Avg Days Open", '=IFERROR(AVERAGE(tbl_Cases[DaysOpen]),"—")'),
        ("Max Days Open", '=IFERROR(MAX(tbl_Cases[DaysOpen]),"—")'),
        ("Cases > 90 days", '=COUNTIFS(tbl_Cases[DaysOpen],">"&90)'),
    ]
    for i in range(0, len(aging_bands), 4):
        row_metrics = aging_bands[i:i+4]
        for j, (label, _) in enumerate(row_metrics):
            cell = ws.cell(row=r, column=j*2+1, value=label)
            cell.font = BODY_FONT_BOLD
            cell.border = THIN_BORDER
            cell.fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_metrics):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── DISPOSAL TRACKING ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="DISPOSAL & ARCHIVE TRACKING").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    disposal_stats = [
        ("Pending Disposal", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[DisposalType],"")'),
        ("Disposal Complete", '=COUNTIFS(tbl_Cases[DisposalType],"<>")'),
        ("Archived", '=COUNTIFS(tbl_Cases[ArchiveRef],"<>")'),
        ("Awaiting Archive", '=COUNTIFS(tbl_Cases[DisposalType],"<>",tbl_Cases[ArchiveRef],"")'),
    ]
    for j, (label, _) in enumerate(disposal_stats):
        ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
        ws.cell(row=r, column=j*2+1).border = THIN_BORDER
        ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
    r += 1
    for j, (_, formula) in enumerate(disposal_stats):
        cell = ws.cell(row=r, column=j*2+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    r += 2

    # Instruction text
    ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: 'tbl_Cases' for registration | 'tbl_Actions' for action items | "
                  "'tbl_Reviews' for supervisory reviews | 'tbl_CaseFileChecklist' for file compliance | "
                  "'tbl_CaseAssignments' for handovers | 'tbl_Statements' for statement register | "
                  "'tbl_Disclosure' for prosecution disclosure").font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.cell(row=r, column=1).alignment = LEFT_WRAP
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART E — REBUILD PORTAL_INVESTIGATIONS WITH FULL SOP COMPLIANCE
# ══════════════════════════════════════════════════════════════════════

def rebuild_portal_investigations(wb):
    if "Portal_Investigations" in wb.sheetnames:
        del wb["Portal_Investigations"]
    ws = wb.create_sheet("Portal_Investigations")
    r = _portal_header(ws, "INVESTIGATIONS PORTAL",
                       "Case Investigation, Witness Management, Evidence Tracking, Court Readiness & SOP Compliance",
                       "Investigations", "1F3864")

    kpis = [
        (4, "Active Cases"),
        (5, "Cases Closed"),
        (6, "Case Clearance Rate"),
        (7, "Supervisory Reviews Done"),
    ]
    r = _portal_kpi_summary(ws, r, "Investigations", kpis)

    # Quick Stats
    stats = [
        ("Under Investigation", '=COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
        ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
        ("Linked Leads", '=COUNTIFS(tbl_Cases[LinkedLeadID],"<>")'),
        ("Persons of Interest", '=COUNTA(tbl_Persons[PersonID])-COUNTBLANK(tbl_Persons[PersonID])'),
        ("Diary Entries", '=COUNTA(tbl_DiaryEntries[DiaryID])-COUNTBLANK(tbl_DiaryEntries[DiaryID])'),
        ("Total Arrests", '=COUNTA(tbl_Arrests[ArrestID])-COUNTBLANK(tbl_Arrests[ArrestID])'),
        ("Scenes Processed", '=COUNTA(tbl_SceneLog[SceneID])-COUNTBLANK(tbl_SceneLog[SceneID])'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    # ── WITNESS & STATEMENT MANAGEMENT ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="WITNESS & STATEMENT MANAGEMENT").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    wit_metrics = [
        ("Total Witnesses", '=COUNTA(tbl_Witnesses[WitnessID])-COUNTBLANK(tbl_Witnesses[WitnessID])'),
        ("Statements Taken", '=COUNTA(tbl_Statements[StatementID])-COUNTBLANK(tbl_Statements[StatementID])'),
        ("Cautioned Statements", '=COUNTIF(tbl_Statements[CautionGiven],"Yes")'),
        ("Willing to Testify", '=COUNTIF(tbl_Witnesses[WillingToTestify],"Yes")'),
        ("Summons Issued", '=COUNTIF(tbl_Witnesses[SummonsIssued],"Yes")'),
        ("Summons Not Served", '=COUNTIFS(tbl_Witnesses[SummonsIssued],"Yes",tbl_Witnesses[SummonsServed],"No")'),
        ("Protection Required", '=COUNTIF(tbl_Witnesses[ProtectionRequired],"Yes")'),
        ("Vulnerable Witnesses", '=COUNTIF(tbl_Witnesses[VulnerableWitness],"Yes")'),
    ]
    for i in range(0, len(wit_metrics), 4):
        row_m = wit_metrics[i:i+4]
        for j, (label, _) in enumerate(row_m):
            ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=j*2+1).border = THIN_BORDER
            ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_m):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── SUPERVISORY REVIEW & DIRECTIVES ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="SUPERVISORY REVIEWS & DIRECTIVES").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    rev_metrics = [
        ("Reviews Completed", '=COUNTA(tbl_Reviews[ReviewID])-COUNTBLANK(tbl_Reviews[ReviewID])'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[NextReviewDue],"<"&TODAY())'),
        ("Directives Issued", '=COUNTIFS(tbl_Reviews[DirectivesIssued],"<>")'),
        ("Directives Overdue", '=COUNTIFS(tbl_Reviews[DirectiveDueDate],"<"&TODAY(),tbl_Reviews[DirectiveStatus],"<>Completed")'),
        ("Escalations", '=COUNTIF(tbl_Reviews[EscalationRequired],"Yes")'),
        ("Conferences Held", '=COUNTA(tbl_CommandConference[ConferenceID])-COUNTBLANK(tbl_CommandConference[ConferenceID])'),
        ("Conf Directives Overdue", '=COUNTIFS(tbl_CommandConference[DirectiveDueDate],"<"&TODAY(),tbl_CommandConference[DirectiveStatus],"<>Completed")'),
        ("Next Conference Due", '=IFERROR(MIN(IF(tbl_CommandConference[NextConferenceDate]>=TODAY(),tbl_CommandConference[NextConferenceDate])),"—")'),
    ]
    for i in range(0, len(rev_metrics), 4):
        row_m = rev_metrics[i:i+4]
        for j, (label, _) in enumerate(row_m):
            ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=j*2+1).border = THIN_BORDER
            ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_m):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── INVESTIGATION PLAN & ASSIGNMENT ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="INVESTIGATION PLANS & CASE ASSIGNMENTS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    plan_metrics = [
        ("Investigation Plans", '=COUNTA(tbl_InvestigationPlans[PlanID])-COUNTBLANK(tbl_InvestigationPlans[PlanID])'),
        ("Active Plans", '=COUNTIF(tbl_InvestigationPlans[PlanStatus],"Active")'),
        ("Plans Under Review", '=COUNTIF(tbl_InvestigationPlans[PlanStatus],"Under Review")'),
        ("Plan Reviews Overdue", '=COUNTIFS(tbl_InvestigationPlans[NextReviewDate],"<"&TODAY())'),
        ("Case Assignments", '=COUNTA(tbl_CaseAssignments[AssignmentID])-COUNTBLANK(tbl_CaseAssignments[AssignmentID])'),
        ("Reassignments", '=COUNTIF(tbl_CaseAssignments[AssignmentType],"Reassignment")'),
        ("Handover Briefings Done", '=COUNTIF(tbl_CaseAssignments[HandoverBriefingDone],"Yes")'),
        ("Pending Acknowledgement", '=COUNTIFS(tbl_CaseAssignments[AcknowledgedByReceiver],"No")'),
    ]
    for i in range(0, len(plan_metrics), 4):
        row_m = plan_metrics[i:i+4]
        for j, (label, _) in enumerate(row_m):
            ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=j*2+1).border = THIN_BORDER
            ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_m):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── VICTIM MANAGEMENT ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="VICTIM / COMPLAINANT MANAGEMENT").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    victim_metrics = [
        ("Victim Contacts Made", '=COUNTA(tbl_VictimContact[ContactID])-COUNTBLANK(tbl_VictimContact[ContactID])'),
        ("Victim Satisfied", '=COUNTIF(tbl_VictimContact[VictimSatisfied],"Yes")'),
        ("Concerns Raised", '=COUNTIFS(tbl_VictimContact[ConcernsRaised],"<>")'),
        ("Referrals Made", '=COUNTIF(tbl_VictimContact[ReferralMade],"Yes")'),
        ("Safety Plans in Place", '=COUNTIF(tbl_VictimContact[SafetyPlanInPlace],"Yes")'),
        ("Contact Overdue", '=COUNTIFS(tbl_VictimContact[NextScheduledContact],"<"&TODAY())'),
    ]
    for i in range(0, len(victim_metrics), 3):
        row_m = victim_metrics[i:i+3]
        for j, (label, _) in enumerate(row_m):
            ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=j*2+1).border = THIN_BORDER
            ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_m):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # ── COURT READINESS & DISCLOSURE ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="COURT READINESS & DISCLOSURE").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    court_metrics = [
        ("Upcoming Court (7d)", '=COUNTIFS(tbl_CourtDates[HearingDate],">="&TODAY(),tbl_CourtDates[HearingDate],"<="&TODAY()+7)'),
        ("Total Hearings", '=COUNTA(tbl_CourtDates[CourtID])-COUNTBLANK(tbl_CourtDates[CourtID])'),
        ("Disclosures Made", '=COUNTA(tbl_Disclosure[DisclosureID])-COUNTBLANK(tbl_Disclosure[DisclosureID])'),
        ("Redactions Required", '=COUNTIF(tbl_Disclosure[RedactionRequired],"Yes")'),
        ("48hr Charge Met", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"Yes")'),
        ("48hr Charge Missed", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"No")'),
    ]
    for i in range(0, len(court_metrics), 3):
        row_m = court_metrics[i:i+3]
        for j, (label, _) in enumerate(row_m):
            ws.cell(row=r, column=j*2+1, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=j*2+1).border = THIN_BORDER
            ws.cell(row=r, column=j*2+1).fill = PatternFill("solid", fgColor=MED_GRAY)
        r += 1
        for j, (_, formula) in enumerate(row_m):
            cell = ws.cell(row=r, column=j*2+1, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        r += 2

    # Instruction text
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY TABLES:\n"
                  "'tbl_Cases' — Case updates | 'tbl_DiaryEntries' — Investigation diary | 'tbl_Reviews' — Supervisory reviews\n"
                  "'tbl_Witnesses' — Witness management | 'tbl_Statements' — Statement register | 'tbl_VictimContact' — Victim updates\n"
                  "'tbl_InvestigationPlans' — Investigation plans | 'tbl_CaseAssignments' — Case handovers\n"
                  "'tbl_CommandConference' — Case conferences | 'tbl_Disclosure' — Prosecution disclosure\n"
                  "'tbl_CaseFileChecklist' — File completeness | 'tbl_Persons' — Persons of interest\n"
                  "'tbl_CourtDates' — Court tracking | 'tbl_SceneLog' — Scene processing | 'tbl_Arrests' — Arrest register"
            ).font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws.cell(row=r, column=1).alignment = LEFT_WRAP
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART F — ADD SOP FORM TEMPLATES
# ══════════════════════════════════════════════════════════════════════

def build_sop_form_templates(wb):
    # ── Investigation Plan Template ──
    ws1 = wb.create_sheet("Form_InvestigationPlan")
    ws1.sheet_properties.tabColor = "1F3864"
    ws1.merge_cells("A1:F1")
    ws1.cell(row=1, column=1, value="FNID AREA 3 — INVESTIGATION PLAN").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws1.cell(row=1, column=1).alignment = CENTER
    ws1.merge_cells("A2:F2")
    ws1.cell(row=2, column=1, value="(Confidential — For Official Use Only)").font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws1.cell(row=2, column=1).alignment = CENTER

    ip_fields = [
        ("Case ID:", ""), ("CR1 Number:", ""), ("Offence:", ""),
        ("Date of Incident:", ""), ("Parish/Station:", ""),
        ("Investigator:", ""), ("Rank/Badge:", ""), ("Supervisor:", ""),
        ("Plan Version:", ""), ("Plan Date:", ""),
        ("", ""),
        ("HYPOTHESIS / THEORY OF THE CASE", ""),
        ("", ""), ("", ""),
        ("OBJECTIVES OF INVESTIGATION", ""),
        ("", ""), ("", ""),
        ("LINES OF ENQUIRY", ""),
        ("1.", ""), ("2.", ""), ("3.", ""), ("4.", ""), ("5.", ""),
        ("KEY EVIDENCE REQUIRED", ""),
        ("", ""), ("", ""),
        ("WITNESSES TO INTERVIEW", ""),
        ("Name:", ""), ("Name:", ""), ("Name:", ""), ("Name:", ""),
        ("SUSPECTS / PERSONS OF INTEREST", ""),
        ("Name:", ""), ("Name:", ""),
        ("FORENSIC REQUIREMENTS", ""),
        ("", ""),
        ("SURVEILLANCE REQUIRED", ""),
        ("", ""),
        ("SEARCHES / WARRANTS PLANNED", ""),
        ("", ""),
        ("INTER-AGENCY SUPPORT REQUIRED", ""),
        ("", ""),
        ("RESOURCES REQUIRED", ""),
        ("", ""),
        ("TIMELINE / KEY MILESTONES", ""),
        ("Milestone:", "Due Date:"), ("Milestone:", "Due Date:"),
        ("Milestone:", "Due Date:"), ("Milestone:", "Due Date:"),
        ("RISK ASSESSMENT", ""),
        ("", ""),
        ("OFFICER SAFETY CONSIDERATIONS", ""),
        ("", ""),
        ("LEGAL CONSIDERATIONS", ""),
        ("", ""),
        ("DISCLOSURE OBLIGATIONS", ""),
        ("", ""),
        ("REVIEW SCHEDULE", ""),
        ("Next Review Date:", ""), ("Review Interval:", ""),
        ("", ""),
        ("APPROVALS", ""),
        ("Investigating Officer:", ""), ("Date:", ""), ("Signature:", ""),
        ("Supervisor:", ""), ("Date:", ""), ("Signature:", ""),
        ("Divisional Commander:", ""), ("Date:", ""), ("Signature:", ""),
    ]
    r = 4
    for label, val in ip_fields:
        if label and label == label.upper() and not label.endswith(":"):
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws1.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif label:
            ws1.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws1.cell(row=r, column=2, value=val).border = Border(bottom=Side(style="thin"))
        r += 1
    for c in range(1, 7):
        ws1.column_dimensions[get_column_letter(c)].width = 18

    # ── Case File Checklist Template ──
    ws2 = wb.create_sheet("Form_FileChecklist")
    ws2.sheet_properties.tabColor = "1F3864"
    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1, value="FNID AREA 3 — CASE FILE COMPLETENESS CHECKLIST").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws2.cell(row=1, column=1).alignment = CENTER

    ws2.cell(row=3, column=1, value="Case ID:").font = BODY_FONT_BOLD
    ws2.cell(row=3, column=2).border = Border(bottom=Side(style="thin"))
    ws2.cell(row=4, column=1, value="Reviewed By:").font = BODY_FONT_BOLD
    ws2.cell(row=4, column=2).border = Border(bottom=Side(style="thin"))
    ws2.cell(row=5, column=1, value="Review Date:").font = BODY_FONT_BOLD
    ws2.cell(row=5, column=2).border = Border(bottom=Side(style="thin"))

    checklist_items = [
        ("CORE DOCUMENTS", None),
        ("CR1 (Crime Report Initial)", None),
        ("CR2 (Crime Report Update)", None),
        ("CR5 (Case Disposal Form)", None),
        ("Station Diary Extract", None),
        ("Investigation Plan (current version)", None),
        ("STATEMENTS & INTERVIEWS", None),
        ("All witness statements obtained", None),
        ("Cautioned statement(s) of suspect(s)", None),
        ("Identification parade record", None),
        ("Alibi check documentation", None),
        ("EVIDENCE & FORENSICS", None),
        ("Exhibit register (complete & signed)", None),
        ("Chain of custody log for all exhibits", None),
        ("Forensic reports received", None),
        ("Scene photographs", None),
        ("Scene sketch/diagram", None),
        ("Scene processing report", None),
        ("COURT READINESS", None),
        ("Witness list (complete)", None),
        ("Witness summonses (served)", None),
        ("Witness availability confirmed", None),
        ("Prosecution brief (prepared)", None),
        ("Disclosure schedule (complete)", None),
        ("SUSPECT DOCUMENTATION", None),
        ("Suspect caution record", None),
        ("Suspect interview record", None),
        ("Identification evidence", None),
        ("Alibi checked and documented", None),
        ("ADMINISTRATIVE", None),
        ("Supervisory endorsements present", None),
        ("Command conference notes filed", None),
        ("Correspondence file complete", None),
        ("Victim contact log included", None),
    ]
    r = 7
    headers = ["Document / Item", "Present (Y/N)", "Date Checked", "Notes"]
    for i, h in enumerate(headers):
        cell = ws2.cell(row=r, column=i+1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for item, _ in checklist_items:
        if item == item.upper():
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws2.cell(row=r, column=1, value=item).font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
            ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        else:
            ws2.cell(row=r, column=1, value=item).font = BODY_FONT
            for c in range(1, 5):
                ws2.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="Total Items Checked:").font = BODY_FONT_BOLD
    ws2.cell(row=r+1, column=1, value="Items Present:").font = BODY_FONT_BOLD
    ws2.cell(row=r+2, column=1, value="Completeness Score:").font = BODY_FONT_BOLD
    ws2.cell(row=r+3, column=1, value="File Status:").font = BODY_FONT_BOLD
    r += 5
    ws2.cell(row=r, column=1, value="Reviewed By:").font = BODY_FONT_BOLD
    ws2.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
    ws2.cell(row=r+1, column=1, value="Signature:").font = BODY_FONT_BOLD
    ws2.cell(row=r+1, column=2).border = Border(bottom=Side(style="thin"))
    ws2.cell(row=r+2, column=1, value="Date:").font = BODY_FONT_BOLD
    ws2.cell(row=r+2, column=2).border = Border(bottom=Side(style="thin"))
    for c in range(1, 5):
        ws2.column_dimensions[get_column_letter(c)].width = 26

    # ── Witness Summons Template ──
    ws3 = wb.create_sheet("Form_WitnessSummons")
    ws3.sheet_properties.tabColor = "1F3864"
    ws3.merge_cells("A1:F1")
    ws3.cell(row=1, column=1, value="FNID AREA 3 — WITNESS SUMMONS TRACKING FORM").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws3.cell(row=1, column=1).alignment = CENTER
    fields = [
        ("Case ID:", ""), ("Court Name:", ""), ("Court Date:", ""),
        ("Hearing Type:", ""), ("", ""),
    ]
    r = 3
    for label, val in fields:
        if label:
            ws3.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws3.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 1

    summons_headers = ["Witness Name","Summons Date","Served By","Served Date","Served (Y/N)","Notes"]
    for i, h in enumerate(summons_headers):
        cell = ws3.cell(row=r, column=i+1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for _ in range(10):
        for c in range(1, 7):
            ws3.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    for c in range(1, 7):
        ws3.column_dimensions[get_column_letter(c)].width = 20


# ══════════════════════════════════════════════════════════════════════
#  PART G — UPDATE HOME NAVIGATION
# ══════════════════════════════════════════════════════════════════════

def patch_home_sop(wb):
    ws = wb["HOME"]
    r = ws.max_row + 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell = ws.cell(row=r, column=1, value="SOP COMPLIANCE TABLES (Case Management Policy)")
    cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 26
    r += 1

    sop_tables = [
        ("tbl_Witnesses", "Witness Register — Statements, summons, protection, availability"),
        ("tbl_Statements", "Statement Register — Formal statements, caution, legal rep, disclosure"),
        ("tbl_CaseAssignments", "Case Assignment Log — Handovers, briefings, acknowledgements"),
        ("tbl_VictimContact", "Victim Contact Log — Updates, satisfaction, referrals, safety plans"),
        ("tbl_InvestigationPlans", "Investigation Plans — Hypothesis, lines of enquiry, milestones"),
        ("tbl_CaseFileChecklist", "File Completeness Checklist — 28-point compliance scoring"),
        ("tbl_Disclosure", "Prosecution Disclosure Log — Materials, redaction, continuing duty"),
        ("tbl_CommandConference", "Command Conference Notes — Directives, escalations, minutes"),
        ("Form_InvestigationPlan", "Investigation Plan Template (Printable)"),
        ("Form_FileChecklist", "Case File Checklist Template (Printable)"),
        ("Form_WitnessSummons", "Witness Summons Tracking Form (Printable)"),
    ]

    for sheet_name, description in sop_tables:
        ws.cell(row=r, column=1, value=f"  >> {sheet_name}").font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        ws.cell(row=r, column=1).alignment = Alignment(indent=2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2, value=description).font = Font(name="Calibri", size=10, color="333333")
        ws.cell(row=r, column=2).alignment = LEFT_WRAP
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="hair", color="CCCCCC"))
        r += 1


# ══════════════════════════════════════════════════════════════════════
#  PART H — ADD SOP KPIs TO UNIT PLANS
# ══════════════════════════════════════════════════════════════════════

def patch_sop_kpis(wb):
    ws = wb["tbl_UnitPlans"]
    last_row = ws.max_row

    new_kpis = [
        ("Registry","SOP Compliance","Registration SLA Compliance","% of cases registered within 24hrs of incident","Monthly","tbl_Cases"),
        ("Registry","SOP Compliance","File Completeness Rate","% of case files scoring >95%","Monthly","tbl_CaseFileChecklist"),
        ("Registry","Case Disposal","Cases Pending Disposal","Cases closed but not yet disposed","Monthly","tbl_Cases"),
        ("Investigations","Witness Management","Witness Statement Rate","% of identified witnesses with statements","Monthly","tbl_Witnesses"),
        ("Investigations","Witness Management","Summons Service Rate","% of summonses successfully served","Monthly","tbl_Witnesses"),
        ("Investigations","SOP Compliance","Investigation Plan Coverage","% of active cases with investigation plans","Monthly","tbl_InvestigationPlans"),
        ("Investigations","SOP Compliance","Mandatory Review Compliance","% of reviews completed within schedule","Monthly","tbl_Reviews"),
        ("Investigations","Victim Rights","Victim Contact Compliance","% of cases with victim contact within 7 days","Monthly","tbl_VictimContact"),
        ("Investigations","Court Readiness","48hr Charge Compliance","% of arrests charged within 48 hours","Monthly","tbl_Cases"),
        ("Investigations","Court Readiness","Disclosure Completion Rate","% of court cases with complete disclosure","Monthly","tbl_Disclosure"),
        ("Investigations","Directive Compliance","Directive Completion Rate","% of supervisory directives completed on time","Monthly","tbl_Reviews"),
        ("Investigations","Case Quality","Handover Compliance Rate","% of reassignments with formal handover","Monthly","tbl_CaseAssignments"),
    ]

    for i, (unit, obj, kpi_name, kpi_desc, freq, src) in enumerate(new_kpis):
        r = last_row + 1 + i
        kpi_num = 44 + i + 1
        ws.cell(row=r, column=1, value=f'KPI-{kpi_num:04d}')
        ws.cell(row=r, column=2, value=unit)
        ws.cell(row=r, column=3, value=obj)
        ws.cell(row=r, column=4, value=kpi_name)
        ws.cell(row=r, column=5, value=kpi_desc)
        ws.cell(row=r, column=6, value=freq)
        ws.cell(row=r, column=23, value=src)

        ws.cell(row=r, column=14, value=f'=IF(OR(K{r}="",H{r}=""),"",K{r}-H{r})')
        ws.cell(row=r, column=15, value=f'=IF(OR(L{r}="",I{r}=""),"",L{r}-I{r})')
        ws.cell(row=r, column=16, value=f'=IF(OR(M{r}="",J{r}=""),"",M{r}-J{r})')
        ws.cell(row=r, column=17, value=f'=IF(OR(K{r}="",H{r}="",H{r}=0),"",K{r}/H{r})')
        ws.cell(row=r, column=18, value=f'=IF(OR(L{r}="",I{r}="",I{r}=0),"",L{r}/I{r})')
        ws.cell(row=r, column=19, value=f'=IF(OR(M{r}="",J{r}="",J{r}=0),"",M{r}/J{r})')
        for col_pct in [17, 18, 19]:
            ws.cell(row=r, column=col_pct).number_format = '0%'
        for col_ach, col_tl in [(17,20),(18,21),(19,22)]:
            cl = get_column_letter(col_ach)
            ws.cell(row=r, column=col_tl,
                    value=f'=IF({cl}{r}="","",IF({cl}{r}>=0.9,"On Track",IF({cl}{r}>=0.7,"At Risk","Off Track")))')
        for c in range(1, 30):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    new_last_row = last_row + len(new_kpis)
    for tbl_name in list(ws.tables.keys()):
        del ws.tables[tbl_name]
    add_table(ws, "tbl_UnitPlans", f"A1:AC{new_last_row}")


# ══════════════════════════════════════════════════════════════════════
#  PART I — ADD SOP METRICS TO STATISTICS HUB
# ══════════════════════════════════════════════════════════════════════

def patch_stats_hub_sop(wb):
    ws = wb["Statistics_Hub"]
    r = ws.max_row + 2

    sections = [
        ("SOP COMPLIANCE — REGISTRY", "1F3864", [
            ("Registration SLA Avg (hrs)", '=IFERROR(AVERAGE(tbl_Cases[RegistrationSLA_Hrs]),"—")'),
            ("Registered > 24hrs", '=COUNTIFS(tbl_Cases[RegistrationSLA_Hrs],">"&24)'),
            ("Files Complete (>95%)", '=COUNTIFS(tbl_Cases[FileCompletenessScore],">="&0.95)'),
            ("Files Critically Incomplete", '=COUNTIFS(tbl_Cases[FileCompletenessScore],"<"&0.5)'),
            ("Avg Completeness Score", '=IFERROR(AVERAGE(tbl_Cases[FileCompletenessScore]),"—")'),
            ("Pending Disposal", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[DisposalType],"")'),
        ]),
        ("SOP COMPLIANCE — INVESTIGATIONS", "1F3864", [
            ("Witnesses Registered", '=COUNTA(tbl_Witnesses[WitnessID])-COUNTBLANK(tbl_Witnesses[WitnessID])'),
            ("Statements Taken", '=COUNTA(tbl_Statements[StatementID])-COUNTBLANK(tbl_Statements[StatementID])'),
            ("Case Assignments", '=COUNTA(tbl_CaseAssignments[AssignmentID])-COUNTBLANK(tbl_CaseAssignments[AssignmentID])'),
            ("Victim Contacts", '=COUNTA(tbl_VictimContact[ContactID])-COUNTBLANK(tbl_VictimContact[ContactID])'),
            ("Investigation Plans", '=COUNTA(tbl_InvestigationPlans[PlanID])-COUNTBLANK(tbl_InvestigationPlans[PlanID])'),
            ("File Checklists Done", '=COUNTA(tbl_CaseFileChecklist[ChecklistID])-COUNTBLANK(tbl_CaseFileChecklist[ChecklistID])'),
            ("Disclosures Logged", '=COUNTA(tbl_Disclosure[DisclosureID])-COUNTBLANK(tbl_Disclosure[DisclosureID])'),
            ("Command Conferences", '=COUNTA(tbl_CommandConference[ConferenceID])-COUNTBLANK(tbl_CommandConference[ConferenceID])'),
            ("Reviews Overdue", '=COUNTIFS(tbl_Reviews[NextReviewDue],"<"&TODAY())'),
            ("Directives Overdue", '=COUNTIFS(tbl_Reviews[DirectiveDueDate],"<"&TODAY(),tbl_Reviews[DirectiveStatus],"<>Completed")'),
            ("48hr Charge Met", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"Yes")'),
            ("48hr Charge Missed", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"No")'),
        ]),
    ]

    for section_title, color, metrics in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
        col = 1
        for label, formula in metrics:
            ws.cell(row=r, column=col, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=MED_GRAY)
            ws.cell(row=r, column=col+1, value=formula).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
            ws.cell(row=r, column=col+1).border = THIN_BORDER
            ws.cell(row=r, column=col+1).alignment = CENTER
            col += 2
            if col > 12:
                col = 1
                r += 1
        r += 2


# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    fname = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"SOP Batch 2: Loading {fname}...")
    wb = openpyxl.load_workbook(fname)

    print("[1/9] Adding SOP Lookup columns (12 new lists)...")
    patch_sop_lookups(wb)

    print("[2/9] Enhancing tbl_Cases with 24 SOP fields...")
    patch_cases_sop_fields(wb)

    print("[3/9] Enhancing tbl_Reviews with directive & schedule tracking...")
    patch_reviews_sop(wb)

    print("[4/9] Rebuilding Portal_Registry with full SOP compliance...")
    rebuild_portal_registry(wb)

    print("[5/9] Rebuilding Portal_Investigations with full SOP compliance...")
    rebuild_portal_investigations(wb)

    print("[6/9] Adding SOP form templates (Investigation Plan, File Checklist, Witness Summons)...")
    build_sop_form_templates(wb)

    print("[7/9] Adding 12 new SOP KPIs to tbl_UnitPlans...")
    patch_sop_kpis(wb)

    print("[8/9] Adding SOP metrics to Statistics Hub...")
    patch_stats_hub_sop(wb)

    print("[9/9] Updating HOME navigation...")
    patch_home_sop(wb)

    wb.save(fname)
    print(f"\nSOP Batch 2 complete: {fname} — {len(wb.sheetnames)} sheets")
    print(f"\nSummary of SOP enhancements:")
    print(f"  + 24 new SOP columns added to tbl_Cases")
    print(f"  + 11 new columns added to tbl_Reviews")
    print(f"  + 12 new SOP Lookup lists")
    print(f"  + 12 new SOP KPIs (total now 56)")
    print(f"  + Portal_Registry rebuilt with: SOP compliance, case aging, disposal tracking")
    print(f"  + Portal_Investigations rebuilt with: witness management, review/directives,")
    print(f"    investigation plans, victim management, court readiness, disclosure")
    print(f"  + 3 printable SOP form templates added")
    print(f"  + Statistics Hub expanded with SOP compliance sections")
    print(f"  + HOME navigation updated")


if __name__ == "__main__":
    main()
