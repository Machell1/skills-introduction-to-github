#!/usr/bin/env python3
"""
FNID Area 3 — Integrated Operational Workbook Generator
========================================================
Generates a comprehensive Excel workbook for the Firearm and Narcotics
Investigation Division (FNID) Area 3, covering:
  - Case Management & Registry
  - Morning Crime Report (MCR) Intake & FNID Filtering
  - Intelligence Unit Funnel & Lead Management
  - Operational Reports & FCSI Tracking
  - Evidence & Exhibits Data Pool + Forensic Routing
  - Transport Unit (Fleet & Logistics)
  - Demand Reduction Lecture Tracking
  - FNID Unit Plans, Targets & Performance
  - Statistics Hub & Dashboards (Executive + per-unit)
  - Governance, Siloed Portals & Audit
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
from datetime import datetime

# ── Global style constants ──────────────────────────────────────────────
NAVY      = "1F3864"
DARK_BLUE = "2E75B6"
MED_BLUE  = "5B9BD5"
LIGHT_BLUE= "D6E4F0"
GOLD      = "FFC000"
RED       = "FF0000"
GREEN     = "00B050"
AMBER     = "FFC000"
WHITE     = "FFFFFF"
BLACK     = "000000"
LIGHT_GRAY= "F2F2F2"
MED_GRAY  = "D9D9D9"

HEADER_FONT     = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL     = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL  = PatternFill("solid", fgColor=DARK_BLUE)
SUBHEADER_FONT  = Font(name="Calibri", bold=True, size=11, color=WHITE)
TITLE_FONT      = Font(name="Calibri", bold=True, size=16, color=NAVY)
SECTION_FONT    = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
BODY_FONT       = Font(name="Calibri", size=10)
BODY_FONT_BOLD  = Font(name="Calibri", size=10, bold=True)
CENTER          = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP       = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
GREEN_FILL  = PatternFill("solid", fgColor=GREEN)
AMBER_FILL  = PatternFill("solid", fgColor=AMBER)
RED_FILL    = PatternFill("solid", fgColor=RED)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
LIGHT_GRAY_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
GOLD_FILL       = PatternFill("solid", fgColor=GOLD)

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)

wb = None  # will be set in main()

# ── Helpers ─────────────────────────────────────────────────────────────

def style_header_row(ws, row, ncols, fill=None, font=None):
    """Apply header styling to a row."""
    f = fill or HEADER_FILL
    fn = font or HEADER_FONT
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fn
        cell.fill = f
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def add_table(ws, name, ref, style=None):
    """Add an Excel Table object to the worksheet."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = style or TABLE_STYLE
    ws.add_table(tbl)
    return tbl


def write_headers(ws, headers, row=1, start_col=1, widths=None):
    """Write header row and optionally set column widths."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w


def write_title(ws, title, row=1, col=1, merge_to=6):
    """Write a large title across merged cells."""
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=merge_to
    )
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36


def write_section(ws, title, row, col=1, merge_to=6):
    """Write a section header."""
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=merge_to
    )
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = SECTION_FONT
    cell.alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 24


def add_data_validation_list(ws, sqref, formula1):
    """Add a dropdown list validation."""
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.error = "Please select from the list."
    dv.errorTitle = "Invalid Entry"
    dv.sqref = sqref
    ws.add_data_validation(dv)
    return dv


def fill_sample_rows(ws, start_row, num_rows, num_cols):
    """Fill sample placeholder rows with light formatting."""
    for r in range(start_row, start_row + num_rows):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP
            if r % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL


def auto_id_formula(prefix, row):
    """Return a formula that generates a unique ID like CASE-0001."""
    return f'="{prefix}-"&TEXT(ROW()-1,"0000")'


def time_helper_formulas(date_col_letter, row):
    """Return dict of formulas for Week, Month, Quarter, Year from a date column."""
    ref = f"{date_col_letter}{row}"
    return {
        "Week":    f'=IF({ref}="","",WEEKNUM({ref}))',
        "Month":   f'=IF({ref}="","",MONTH({ref}))',
        "Quarter": f'=IF({ref}="","",ROUNDUP(MONTH({ref})/3,0))',
        "Year":    f'=IF({ref}="","",YEAR({ref}))',
    }


# ══════════════════════════════════════════════════════════════════════
#  MODULE 0 — LOOKUPS & REFERENCE DATA
# ══════════════════════════════════════════════════════════════════════

def build_lookups(wb):
    ws = wb.create_sheet("Lookups")
    ws.sheet_properties.tabColor = "808080"

    # --- Case Status ---
    ws.cell(row=1, column=1, value="CaseStatus").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    statuses = ["Open","Active","Under Investigation","Pending Review",
                "Pending Court","Closed - Convicted","Closed - Acquitted",
                "Closed - Withdrawn","Closed - No Further Action","Cold Case"]
    for i, s in enumerate(statuses, 2):
        ws.cell(row=i, column=1, value=s)

    # --- Priority ---
    ws.cell(row=1, column=2, value="Priority").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    for i, p in enumerate(["Critical","High","Medium","Low"], 2):
        ws.cell(row=i, column=2, value=p)

    # --- OffenceType ---
    ws.cell(row=1, column=3, value="OffenceType").font = HEADER_FONT
    ws.cell(row=1, column=3).fill = HEADER_FILL
    offences = [
        "Illegal Possession of Firearm","Illegal Possession of Ammunition",
        "Shooting with Intent","Murder (Firearm)","Robbery (Firearm)",
        "Possession of Narcotics","Trafficking of Narcotics",
        "Cultivation of Narcotics","Distribution of Narcotics",
        "Possession of Cocaine","Possession of Cannabis",
        "Possession of Other Drug","Larceny of Firearm",
        "Wounding (Firearm)","Conspiracy - Firearms",
        "Conspiracy - Narcotics","Importation of Firearm",
        "Importation of Narcotics","Other Firearms Offence",
        "Other Narcotics Offence"
    ]
    for i, o in enumerate(offences, 2):
        ws.cell(row=i, column=3, value=o)

    # --- Parish ---
    ws.cell(row=1, column=4, value="Parish").font = HEADER_FONT
    ws.cell(row=1, column=4).fill = HEADER_FILL
    parishes = [
        "Kingston","St. Andrew","St. Catherine","Clarendon","Manchester",
        "St. Elizabeth","Westmoreland","Hanover","St. James","Trelawny",
        "St. Ann","St. Mary","Portland","St. Thomas"
    ]
    for i, p in enumerate(parishes, 2):
        ws.cell(row=i, column=4, value=p)

    # --- Unit ---
    ws.cell(row=1, column=5, value="Unit").font = HEADER_FONT
    ws.cell(row=1, column=5).fill = HEADER_FILL
    units = [
        "Registry","Investigations","Operations","Intelligence",
        "Transport","Demand Reduction","Admin/HRM","Training","Oversight"
    ]
    for i, u in enumerate(units, 2):
        ws.cell(row=i, column=5, value=u)

    # --- LeadSource ---
    ws.cell(row=1, column=6, value="LeadSource").font = HEADER_FONT
    ws.cell(row=1, column=6).fill = HEADER_FILL
    for i, s in enumerate(["MCR","Operations","Tip-off","Informant","Surveillance","Inter-Agency","Walk-in","Other"], 2):
        ws.cell(row=i, column=6, value=s)

    # --- TriageDecision ---
    ws.cell(row=1, column=7, value="TriageDecision").font = HEADER_FONT
    ws.cell(row=1, column=7).fill = HEADER_FILL
    for i, t in enumerate(["Action Immediately","Investigate","Monitor","Archive","Forward to External Agency"], 2):
        ws.cell(row=i, column=7, value=t)

    # --- OpsType ---
    ws.cell(row=1, column=8, value="OpsType").font = HEADER_FONT
    ws.cell(row=1, column=8).fill = HEADER_FILL
    ops = ["Search Warrant","Patrol","Checkpoint","Surveillance","Raid",
           "Arrest Operation","Seizure Operation","Joint Operation","FCSI Operation","Other"]
    for i, o in enumerate(ops, 2):
        ws.cell(row=i, column=8, value=o)

    # --- YesNo ---
    ws.cell(row=1, column=9, value="YesNo").font = HEADER_FONT
    ws.cell(row=1, column=9).fill = HEADER_FILL
    ws.cell(row=2, column=9, value="Yes")
    ws.cell(row=3, column=9, value="No")

    # --- Credibility ---
    ws.cell(row=1, column=10, value="Credibility").font = HEADER_FONT
    ws.cell(row=1, column=10).fill = HEADER_FILL
    for i, c in enumerate(["A - Reliable","B - Usually Reliable","C - Fairly Reliable","D - Not Usually Reliable","E - Unreliable","F - Unknown"], 2):
        ws.cell(row=i, column=10, value=c)

    # --- EvidenceType ---
    ws.cell(row=1, column=11, value="EvidenceType").font = HEADER_FONT
    ws.cell(row=1, column=11).fill = HEADER_FILL
    ev_types = ["Firearm","Ammunition","Narcotics - Cannabis","Narcotics - Cocaine",
                "Narcotics - Other","Cash/Currency","Electronic Device","Document",
                "Vehicle","Clothing","Biological Sample","Fingerprint","CCTV Footage",
                "Ballistic Evidence","Other Physical Evidence"]
    for i, e in enumerate(ev_types, 2):
        ws.cell(row=i, column=11, value=e)

    # --- ForensicAgency ---
    ws.cell(row=1, column=12, value="ForensicAgency").font = HEADER_FONT
    ws.cell(row=1, column=12).fill = HEADER_FILL
    agencies = ["Government Forensic Lab","Ballistics Unit","Fingerprint Bureau",
                "Cyber Forensics Unit","Chemistry Division","DNA Lab","External Lab"]
    for i, a in enumerate(agencies, 2):
        ws.cell(row=i, column=12, value=a)

    # --- AnalysisStatus ---
    ws.cell(row=1, column=13, value="AnalysisStatus").font = HEADER_FONT
    ws.cell(row=1, column=13).fill = HEADER_FILL
    for i, s in enumerate(["Pending","Submitted","In Progress","Completed","Inconclusive","N/A"], 2):
        ws.cell(row=i, column=13, value=s)

    # --- CourtReadiness ---
    ws.cell(row=1, column=14, value="CourtReadiness").font = HEADER_FONT
    ws.cell(row=1, column=14).fill = HEADER_FILL
    for i, s in enumerate(["Ready","Not Ready","Partial","Under Review"], 2):
        ws.cell(row=i, column=14, value=s)

    # --- VehicleStatus ---
    ws.cell(row=1, column=15, value="VehicleStatus").font = HEADER_FONT
    ws.cell(row=1, column=15).fill = HEADER_FILL
    for i, s in enumerate(["Available","In Use","Under Maintenance","Out of Service","Reserved"], 2):
        ws.cell(row=i, column=15, value=s)

    # --- TrafficLight ---
    ws.cell(row=1, column=16, value="TrafficLight").font = HEADER_FONT
    ws.cell(row=1, column=16).fill = HEADER_FILL
    for i, s in enumerate(["On Track","At Risk","Off Track"], 2):
        ws.cell(row=i, column=16, value=s)

    # --- Rank ---
    ws.cell(row=1, column=17, value="Rank").font = HEADER_FONT
    ws.cell(row=1, column=17).fill = HEADER_FILL
    ranks = ["Constable","Corporal","Sergeant","Inspector","ASP","DSP","SSP","ACP","DCP","Commissioner"]
    for i, r in enumerate(ranks, 2):
        ws.cell(row=i, column=17, value=r)

    # --- Frequency ---
    ws.cell(row=1, column=18, value="Frequency").font = HEADER_FONT
    ws.cell(row=1, column=18).fill = HEADER_FILL
    for i, f in enumerate(["Daily","Weekly","Monthly","Quarterly","Annually"], 2):
        ws.cell(row=i, column=18, value=f)

    # --- LeadOutcome ---
    ws.cell(row=1, column=19, value="LeadOutcome").font = HEADER_FONT
    ws.cell(row=1, column=19).fill = HEADER_FILL
    for i, o in enumerate(["Converted to Case","Converted to Op","Intel Filed","Closed - No Action","Pending"], 2):
        ws.cell(row=i, column=19, value=o)

    # --- ComplianceStatus ---
    ws.cell(row=1, column=20, value="ComplianceStatus").font = HEADER_FONT
    ws.cell(row=1, column=20).fill = HEADER_FILL
    for i, s in enumerate(["Compliant","Non-Compliant","Under Review","Overdue","N/A"], 2):
        ws.cell(row=i, column=20, value=s)

    # Set column widths
    for c in range(1, 21):
        ws.column_dimensions[get_column_letter(c)].width = 22

    return ws



# ══════════════════════════════════════════════════════════════════════
#  MODULE 0B — OFFENCE CLASSIFICATION MAP & FORENSIC ROUTING
# ══════════════════════════════════════════════════════════════════════

def build_offence_class_map(wb):
    ws = wb.create_sheet("OffenceClassMap")
    ws.sheet_properties.tabColor = "808080"
    headers = ["OffenceCode","OffenceDescription","Category","FNID_Relevant","AutoRouteToFNID"]
    widths  = [14, 40, 20, 16, 18]
    write_headers(ws, headers, widths=widths)

    data = [
        ["FC-001","Illegal Possession of Firearm","Firearms","Yes","Yes"],
        ["FC-002","Illegal Possession of Ammunition","Firearms","Yes","Yes"],
        ["FC-003","Shooting with Intent","Firearms","Yes","Yes"],
        ["FC-004","Murder (Firearm-related)","Firearms","Yes","Yes"],
        ["FC-005","Robbery with Firearm","Firearms","Yes","Yes"],
        ["FC-006","Wounding with Firearm","Firearms","Yes","Yes"],
        ["FC-007","Larceny of Firearm","Firearms","Yes","Yes"],
        ["FC-008","Conspiracy - Firearms","Firearms","Yes","Yes"],
        ["FC-009","Importation of Firearm","Firearms","Yes","Yes"],
        ["NC-001","Possession of Cannabis","Narcotics","Yes","Yes"],
        ["NC-002","Possession of Cocaine","Narcotics","Yes","Yes"],
        ["NC-003","Trafficking of Narcotics","Narcotics","Yes","Yes"],
        ["NC-004","Cultivation of Narcotics","Narcotics","Yes","Yes"],
        ["NC-005","Distribution of Narcotics","Narcotics","Yes","Yes"],
        ["NC-006","Importation of Narcotics","Narcotics","Yes","Yes"],
        ["NC-007","Conspiracy - Narcotics","Narcotics","Yes","Yes"],
        ["NC-008","Possession of Other Drug","Narcotics","Yes","Yes"],
        ["OT-001","Robbery (No Firearm)","Other","No","No"],
        ["OT-002","Burglary","Other","No","No"],
        ["OT-003","Assault (Non-Firearm)","Other","No","No"],
        ["OT-004","Fraud","Other","No","No"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
    ref = f"A1:{get_column_letter(len(headers))}{1+len(data)}"
    add_table(ws, "tbl_OffenceClassMap", ref)
    return ws


def build_forensic_routing(wb):
    ws = wb.create_sheet("ForensicRouting")
    ws.sheet_properties.tabColor = "808080"
    headers = ["EvidenceType","RecommendedForensicAgency","TestType","ExpectedTurnaround_Days","Notes"]
    widths  = [24, 28, 24, 22, 30]
    write_headers(ws, headers, widths=widths)

    data = [
        ["Firearm","Ballistics Unit","Ballistic Comparison",30,"Serial number trace included"],
        ["Ammunition","Ballistics Unit","Cartridge/Projectile Analysis",21,"Match to firearm"],
        ["Narcotics - Cannabis","Chemistry Division","Substance Identification",14,"Weight & purity"],
        ["Narcotics - Cocaine","Chemistry Division","Substance Identification",14,"Weight & purity"],
        ["Narcotics - Other","Chemistry Division","Substance Identification",21,"Compound analysis"],
        ["Fingerprint","Fingerprint Bureau","AFIS Comparison",14,"Latent print processing"],
        ["Biological Sample","DNA Lab","DNA Profiling",45,"CODIS check"],
        ["Electronic Device","Cyber Forensics Unit","Data Extraction",30,"Mobile/computer forensics"],
        ["CCTV Footage","Cyber Forensics Unit","Video Enhancement",14,"Frame analysis"],
        ["Ballistic Evidence","Ballistics Unit","Trajectory/Residue Analysis",21,"GSR included"],
        ["Document","Government Forensic Lab","Document Examination",21,"Handwriting/forgery"],
        ["Cash/Currency","Government Forensic Lab","Trace Analysis",14,"Drug trace detection"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
    ref = f"A1:{get_column_letter(len(headers))}{1+len(data)}"
    add_table(ws, "tbl_ForensicRouting", ref)
    return ws


def build_evidence_use_map(wb):
    ws = wb.create_sheet("EvidenceUseMap")
    ws.sheet_properties.tabColor = "808080"
    headers = ["EvidenceType","EvidentialUse","CaseElement_Supported","LegalRequirement","Notes"]
    widths  = [24, 30, 28, 24, 30]
    write_headers(ws, headers, widths=widths)

    data = [
        ["Firearm","Proves possession / links to crime","Actus Reus - Weapon","Firearms Act","Chain of custody critical"],
        ["Ammunition","Corroborates firearm possession","Actus Reus - Ammunition","Firearms Act","Match to firearm if possible"],
        ["Narcotics - Cannabis","Proves possession / quantity","Actus Reus - Substance","Dangerous Drugs Act","Certified weight required"],
        ["Narcotics - Cocaine","Proves possession / quantity","Actus Reus - Substance","Dangerous Drugs Act","Purity analysis for trafficking"],
        ["Fingerprint","Links suspect to scene/item","Identity","Evidence Act","AFIS hit + expert testimony"],
        ["Biological Sample","Links suspect via DNA","Identity","Evidence Act","DNA profile match"],
        ["Electronic Device","Communications / planning evidence","Mens Rea / Conspiracy","Cybercrimes Act","Extraction report as exhibit"],
        ["CCTV Footage","Visual identification / timeline","Identity / Timeline","Evidence Act","Authentication certificate"],
        ["Ballistic Evidence","Links firearm to shooting","Actus Reus - Weapon Use","Firearms Act","Expert ballistic report"],
        ["Cash/Currency","Proceeds of crime","Proceeds","Proceeds of Crime Act","Amount + trace evidence"],
        ["Document","Forged / planning documents","Mens Rea / Identity","Forgery Act","Expert document report"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP
    ref = f"A1:{get_column_letter(len(headers))}{1+len(data)}"
    add_table(ws, "tbl_EvidenceUseMap", ref)
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 1 — CASE MANAGEMENT & REGISTRY (Back-end table)
# ══════════════════════════════════════════════════════════════════════

def build_cases_table(wb):
    ws = wb.create_sheet("tbl_Cases")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "CaseID","CR1_Number","CR2_Number","CR5_Number",
        "DateReceived","OffenceType","OffenceDescription","Parish","Station",
        "Complainant","Suspect","InvestigatorAssigned","Unit",
        "Priority","CaseStatus","SupervisorReview","ReviewDate",
        "NextActionDue","ComplianceFlag","CR1_Compliant","CR2_Compliant","CR5_Compliant",
        "FCSI_Flag","LinkedLeadID","LinkedOpsID",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,14, 14,24,30,16,16, 20,20,20,16,
              12,18,20,14, 14,16,14,14,14, 12,14,14,
              30,14,14,14,14, 8,8,8,8]
    write_headers(ws, headers, widths=widths)

    # Sample formula rows (5 rows)
    NUM_SAMPLE = 5
    for r in range(2, 2 + NUM_SAMPLE):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","CASE-"&TEXT(ROW()-1,"0000"))')
        # Time helpers from column E (DateReceived)
        ws.cell(row=r, column=31, value=f'=IF(E{r}="","",WEEKNUM(E{r}))')
        ws.cell(row=r, column=32, value=f'=IF(E{r}="","",MONTH(E{r}))')
        ws.cell(row=r, column=33, value=f'=IF(E{r}="","",ROUNDUP(MONTH(E{r})/3,0))')
        ws.cell(row=r, column=34, value=f'=IF(E{r}="","",YEAR(E{r}))')
        # Compliance flag formula
        ws.cell(row=r, column=19, value=f'=IF(AND(T{r}="Yes",U{r}="Yes",V{r}="Yes"),"Compliant","Non-Compliant")')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}{1+NUM_SAMPLE}"
    add_table(ws, "tbl_Cases", ref)

    # Data validations
    add_data_validation_list(ws, f"F2:F1000", "=Lookups!$C$2:$C$22")
    add_data_validation_list(ws, f"H2:H1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, f"M2:M1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, f"N2:N1000", "=Lookups!$B$2:$B$6")
    add_data_validation_list(ws, f"O2:O1000", "=Lookups!$A$2:$A$12")
    add_data_validation_list(ws, f"T2:T1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, f"U2:U1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, f"V2:V1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, f"W2:W1000", "=Lookups!$I$2:$I$3")

    # Conditional formatting - overdue actions
    ws.conditional_formatting.add(
        f"R2:R1000",
        FormulaRule(formula=[f'=AND(R2<>"",R2<TODAY())'], fill=RED_FILL)
    )
    return ws


def build_actions_table(wb):
    ws = wb.create_sheet("tbl_Actions")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "ActionID","CaseID","ActionType","Description","AssignedTo","Unit",
        "DueDate","CompletedDate","Status","Priority","Outcome","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,18,30,18,14,14,14,14,12,24,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","ACT-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=17, value=f'=IF(G{r}="","",WEEKNUM(G{r}))')
        ws.cell(row=r, column=18, value=f'=IF(G{r}="","",MONTH(G{r}))')
        ws.cell(row=r, column=19, value=f'=IF(G{r}="","",ROUNDUP(MONTH(G{r})/3,0))')
        ws.cell(row=r, column=20, value=f'=IF(G{r}="","",YEAR(G{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Actions", ref)
    # Overdue highlight
    ws.conditional_formatting.add(
        "G2:G1000",
        FormulaRule(formula=['=AND(G2<>"",G2<TODAY(),H2="")'], fill=RED_FILL)
    )
    return ws


def build_reviews_table(wb):
    ws = wb.create_sheet("tbl_Reviews")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "ReviewID","CaseID","ReviewType","Reviewer","ReviewDate","DueDate",
        "Findings","Recommendations","Status","ComplianceStatus",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,18,18,14,14,30,30,14,16,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","REV-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Reviews", ref)
    return ws


def build_diary_table(wb):
    ws = wb.create_sheet("tbl_DiaryEntries")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "DiaryID","CaseID","EntryDate","EntryTime","EnteredBy","Unit",
        "DiaryNarrative","ActionTaken","NextSteps",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,12,18,14,40,30,30,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","DRY-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_DiaryEntries", ref)
    return ws


def build_radio_messages_table(wb):
    ws = wb.create_sheet("tbl_RadioMessages")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "MessageID","Date","Time","From","To","Priority","Subject",
        "MessageBody","ActionRequired","LinkedCaseID","LinkedLeadID",
        "CreatedBy","CreatedDate"
    ]
    widths = [14,14,12,18,18,12,24,40,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","MSG-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_RadioMessages", ref)
    return ws


def build_persons_table(wb):
    ws = wb.create_sheet("tbl_Persons")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "PersonID","FullName","Alias","DOB","Gender","Address","Parish",
        "Phone","ID_Type","ID_Number","Role","LinkedCaseID","LinkedLeadID",
        "RiskLevel","Notes","PhotoRef",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,22,16,14,10,28,16,16,14,16,14,14,14,12,24,16,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","PER-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Persons", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$D$2:$D$16")
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 2 — MCR INTAKE & FNID FILTERING
# ══════════════════════════════════════════════════════════════════════

def build_mcr_raw(wb):
    ws = wb.create_sheet("tbl_MCR_Raw")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "MCR_ID","MCR_Date","ReportingStation","Parish","OffenceCode",
        "OffenceDescription","Complainant","Location","Suspect","Summary",
        "FNID_Relevant","AutoRouted","RoutedDate","TriageStatus",
        "LinkedLeadID","Notes",
        "CreatedBy","CreatedDate","Week","Month","Quarter","Year"
    ]
    widths = [14,14,18,16,14,28,20,24,20,36,14,12,14,14,14,24,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","MCR-"&TEXT(ROW()-1,"0000"))')
        # Auto FNID relevance via VLOOKUP to OffenceClassMap
        ws.cell(row=r, column=11, value=f'=IFERROR(VLOOKUP(E{r},tbl_OffenceClassMap[#All],4,FALSE),"No")')
        ws.cell(row=r, column=12, value=f'=IFERROR(VLOOKUP(E{r},tbl_OffenceClassMap[#All],5,FALSE),"No")')
        # Time helpers
        ws.cell(row=r, column=19, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=20, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=21, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=22, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_MCR_Raw", ref)
    return ws


# ══════════════════════════════════════════════════════════════════════
#  MODULE 3 — INTELLIGENCE UNIT FUNNEL & LEAD MANAGEMENT
# ══════════════════════════════════════════════════════════════════════

def build_leads_table(wb):
    ws = wb.create_sheet("tbl_Leads")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "LeadID","DateReceived","Source","SourceDetail","Parish","Station",
        "Subject","Summary","Credibility","Priority","RiskScore",
        "TriageDecision","TriageDate","AssignedTo","Unit",
        "Outcome","OutcomeDate","ConvertedToCaseID","ConvertedToOpsID",
        "FollowUpRequired","FollowUpDueDate","FCSI_Flag",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,20,16,16,24,36,18,12,12,20,14,18,14,18,14,16,16,14,14,12,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","LEAD-"&TEXT(ROW()-1,"0000"))')
        # Risk score formula (simple: Credibility letter * Priority weight)
        ws.cell(row=r, column=11, value=f'=IFERROR(IF(OR(I{r}="",J{r}=""),"",IF(J{r}="Critical",4,IF(J{r}="High",3,IF(J{r}="Medium",2,1)))*IF(LEFT(I{r},1)="A",5,IF(LEFT(I{r},1)="B",4,IF(LEFT(I{r},1)="C",3,IF(LEFT(I{r},1)="D",2,1))))),"")' )
        ws.cell(row=r, column=28, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=29, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=30, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=31, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Leads", ref)

    add_data_validation_list(ws, "C2:C1000", "=Lookups!$F$2:$F$9")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "I2:I1000", "=Lookups!$J$2:$J$7")
    add_data_validation_list(ws, "J2:J1000", "=Lookups!$B$2:$B$6")
    add_data_validation_list(ws, "L2:L1000", "=Lookups!$G$2:$G$6")
    add_data_validation_list(ws, "O2:O1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "P2:P1000", "=Lookups!$S$2:$S$6")
    add_data_validation_list(ws, "V2:V1000", "=Lookups!$I$2:$I$3")

    # Highlight high risk scores
    ws.conditional_formatting.add(
        "K2:K1000",
        CellIsRule(operator='greaterThanOrEqual', formula=['15'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        "K2:K1000",
        CellIsRule(operator='between', formula=['8','14'], fill=AMBER_FILL)
    )
    # Overdue follow-ups
    ws.conditional_formatting.add(
        "U2:U1000",
        FormulaRule(formula=['=AND(U2<>"",U2<TODAY(),T2="Yes")'], fill=RED_FILL)
    )
    return ws


def build_followup_recs(wb):
    ws = wb.create_sheet("tbl_FollowUpRecs")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "RecID","LeadID","RecommendedBy","RecommendationDate","Recommendation",
        "Priority","AssignedTo","DueDate","Status","Outcome","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,18,16,36,12,18,14,14,24,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","REC-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_FollowUpRecs", ref)
    return ws


def build_intel_briefings(wb):
    ws = wb.create_sheet("tbl_IntelBriefings")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "BriefingID","BriefingDate","BriefingTime","Presenter","Audience",
        "Topic","KeyFindings","ActionItems","LinkedLeadIDs","LinkedCaseIDs",
        "AttendeeCount","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,12,18,20,24,36,30,18,18,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","BRIEF-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_IntelBriefings", ref)
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 4 — OPERATIONS REGISTER & FCSI TRACKING
# ══════════════════════════════════════════════════════════════════════

def build_operations_table(wb):
    ws = wb.create_sheet("tbl_Operations")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "OPS_ID","OpsDate","OpsType","OpsName","Parish","Station","Location",
        "TeamLead","TeamSize","Unit","FCSI_Flag","FCSI_Hotspot",
        "WarrantNumber","TargetDescription",
        "Searches","Arrests","FirearmsSeized","AmmunitionSeized",
        "NarcoticsSeized_kg","CashSeized","OtherSeizures",
        "IntelGained","LinkedLeadID","LinkedCaseID",
        "Outcome","OutcomeNotes","StartTime","EndTime",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,18,22,16,16,22,18,10,14,12,18,16,24,
              10,10,14,16,16,14,18,14,14,14,18,28,12,12,
              24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","OPS-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=34, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=35, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=36, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=37, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Operations", ref)

    add_data_validation_list(ws, "C2:C1000", "=Lookups!$H$2:$H$11")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "J2:J1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "K2:K1000", "=Lookups!$I$2:$I$3")

    # FCSI flag highlighting
    ws.conditional_formatting.add(
        "K2:K1000",
        FormulaRule(formula=['=K2="Yes"'], fill=GOLD_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  MODULE 5 — EVIDENCE & EXHIBITS DATA POOL
# ══════════════════════════════════════════════════════════════════════

def build_exhibits_table(wb):
    ws = wb.create_sheet("tbl_Exhibits")
    ws.sheet_properties.tabColor = "538135"
    headers = [
        "ExhibitID","LinkedCaseID","LinkedOpsID","DateCollected","CollectedBy",
        "EvidenceType","Description","Quantity","Unit_Measure",
        "StorageLocation","ChainOfCustody_Log","CurrentCustodian",
        "ForensicAgency","ForensicTestType","SubmissionDate","ExpectedReturn",
        "ActualReturnDate","AnalysisStatus","ForensicResult","TurnaroundDays",
        "EvidentialUse","CaseElementSupported","CourtReadiness",
        "CourtDate","DisposalDate","DisposalMethod",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,14,18,20,28,10,12,18,24,18,22,20,14,14,14,16,24,14,
              24,22,14,14,14,16,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)

    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","EXH-"&TEXT(ROW()-1,"0000"))')
        # Auto forensic routing lookup
        ws.cell(row=r, column=13, value=f'=IFERROR(VLOOKUP(F{r},tbl_ForensicRouting[#All],2,FALSE),"")')
        ws.cell(row=r, column=14, value=f'=IFERROR(VLOOKUP(F{r},tbl_ForensicRouting[#All],3,FALSE),"")')
        # Turnaround days calc
        ws.cell(row=r, column=20, value=f'=IF(OR(O{r}="",Q{r}=""),"",Q{r}-O{r})')
        # Evidential use lookup
        ws.cell(row=r, column=21, value=f'=IFERROR(VLOOKUP(F{r},tbl_EvidenceUseMap[#All],2,FALSE),"")')
        ws.cell(row=r, column=22, value=f'=IFERROR(VLOOKUP(F{r},tbl_EvidenceUseMap[#All],3,FALSE),"")')
        # Time helpers
        ws.cell(row=r, column=32, value=f'=IF(D{r}="","",WEEKNUM(D{r}))')
        ws.cell(row=r, column=33, value=f'=IF(D{r}="","",MONTH(D{r}))')
        ws.cell(row=r, column=34, value=f'=IF(D{r}="","",ROUNDUP(MONTH(D{r})/3,0))')
        ws.cell(row=r, column=35, value=f'=IF(D{r}="","",YEAR(D{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Exhibits", ref)

    add_data_validation_list(ws, "F2:F1000", "=Lookups!$K$2:$K$16")
    add_data_validation_list(ws, "R2:R1000", "=Lookups!$M$2:$M$7")
    add_data_validation_list(ws, "W2:W1000", "=Lookups!$N$2:$N$5")

    # Overdue forensic returns
    ws.conditional_formatting.add(
        "P2:P1000",
        FormulaRule(formula=['=AND(P2<>"",P2<TODAY(),Q2="")'], fill=RED_FILL)
    )
    # Court readiness
    ws.conditional_formatting.add(
        "W2:W1000",
        FormulaRule(formula=['=W2="Not Ready"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        "W2:W1000",
        FormulaRule(formula=['=W2="Ready"'], fill=GREEN_FILL)
    )
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 6 — TRANSPORT UNIT (FLEET & LOGISTICS)
# ══════════════════════════════════════════════════════════════════════

def build_vehicles_table(wb):
    ws = wb.create_sheet("tbl_Vehicles")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = [
        "VehicleID","RegistrationNo","Make","Model","Year","VehicleType",
        "AssignedUnit","AssignedStation","CurrentStatus","Mileage",
        "LastServiceDate","NextServiceDue","InsuranceExpiry","FitnessExpiry",
        "FuelType","TankCapacity_L","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,16,14,14,8,14,16,16,16,12,14,14,14,14,12,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","VEH-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Vehicles", ref)
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "I2:I1000", "=Lookups!$O$2:$O$6")
    # Service overdue
    ws.conditional_formatting.add(
        "L2:L1000",
        FormulaRule(formula=['=AND(L2<>"",L2<TODAY())'], fill=RED_FILL)
    )
    # Insurance/fitness expiry
    ws.conditional_formatting.add(
        "M2:M1000",
        FormulaRule(formula=['=AND(M2<>"",M2<TODAY())'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        "N2:N1000",
        FormulaRule(formula=['=AND(N2<>"",N2<TODAY())'], fill=RED_FILL)
    )
    return ws


def build_vehicle_usage_table(wb):
    ws = wb.create_sheet("tbl_VehicleUsage")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = [
        "UsageID","VehicleID","Date","Driver","Purpose","LinkedOpsID","LinkedCaseID",
        "DepartureTime","ReturnTime","StartMileage","EndMileage","DistanceKm",
        "Unit","Parish","Notes",
        "CreatedBy","CreatedDate","Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,18,22,14,14,12,12,12,12,12,14,16,24,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","USE-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=12, value=f'=IF(OR(J{r}="",K{r}=""),"",K{r}-J{r})')
        ws.cell(row=r, column=18, value=f'=IF(C{r}="","",WEEKNUM(C{r}))')
        ws.cell(row=r, column=19, value=f'=IF(C{r}="","",MONTH(C{r}))')
        ws.cell(row=r, column=20, value=f'=IF(C{r}="","",ROUNDUP(MONTH(C{r})/3,0))')
        ws.cell(row=r, column=21, value=f'=IF(C{r}="","",YEAR(C{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_VehicleUsage", ref)
    return ws


def build_fuel_log(wb):
    ws = wb.create_sheet("tbl_FuelLog")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = [
        "FuelID","VehicleID","Date","Driver","FuelType","Litres","Cost",
        "Mileage_AtFill","Station","ReceiptNo","Notes",
        "CreatedBy","CreatedDate","Week","Month","Quarter","Year"
    ]
    widths = [14,14,14,18,12,10,12,14,16,14,24,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","FUEL-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=14, value=f'=IF(C{r}="","",WEEKNUM(C{r}))')
        ws.cell(row=r, column=15, value=f'=IF(C{r}="","",MONTH(C{r}))')
        ws.cell(row=r, column=16, value=f'=IF(C{r}="","",ROUNDUP(MONTH(C{r})/3,0))')
        ws.cell(row=r, column=17, value=f'=IF(C{r}="","",YEAR(C{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_FuelLog", ref)
    return ws


def build_maintenance_table(wb):
    ws = wb.create_sheet("tbl_Maintenance")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = [
        "MaintID","VehicleID","MaintenanceType","Description","ServiceProvider",
        "DateIn","DateOut","Cost","Mileage","PartsReplaced",
        "NextServiceDue","DowntimeDays","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,18,28,20,14,14,12,12,24,14,12,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","MAINT-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=12, value=f'=IF(OR(F{r}="",G{r}=""),"",G{r}-F{r})')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Maintenance", ref)
    return ws


def build_drivers_table(wb):
    ws = wb.create_sheet("tbl_Drivers")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = [
        "DriverID","FullName","Rank","BadgeNo","Unit","Station",
        "LicenseNo","LicenseExpiry","LicenseClass","DefensiveDrivingCert",
        "CertExpiry","Status","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,22,14,14,14,16,16,14,14,20,14,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","DRV-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Drivers", ref)
    add_data_validation_list(ws, "C2:C1000", "=Lookups!$Q$2:$Q$11")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    # License expiry
    ws.conditional_formatting.add(
        "H2:H1000",
        FormulaRule(formula=['=AND(H2<>"",H2<TODAY())'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  MODULE 7 — DEMAND REDUCTION LECTURE TRACKING
# ══════════════════════════════════════════════════════════════════════

def build_demand_reduction(wb):
    ws = wb.create_sheet("tbl_DemandReduction")
    ws.sheet_properties.tabColor = "00B050"
    headers = [
        "LectureID","Date","Location","Parish","Community","Venue",
        "Topic","Presenter","TargetAudience","Attendance","Duration_Hrs",
        "MaterialsUsed","PartnerOrganizations","Outcome","Feedback",
        "FollowUpPlanned","FollowUpDate","Photos_YN",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","Month","Quarter","Year"
    ]
    widths = [14,14,22,16,18,20,24,18,18,12,12,20,22,24,24,14,14,10,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","LEC-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=24, value=f'=IF(B{r}="","",WEEKNUM(B{r}))')
        ws.cell(row=r, column=25, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=26, value=f'=IF(B{r}="","",ROUNDUP(MONTH(B{r})/3,0))')
        ws.cell(row=r, column=27, value=f'=IF(B{r}="","",YEAR(B{r}))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_DemandReduction", ref)
    add_data_validation_list(ws, "D2:D1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "R2:R1000", "=Lookups!$I$2:$I$3")
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 8 — UNIT PLANS, TARGETS & PERFORMANCE
# ══════════════════════════════════════════════════════════════════════

def build_unit_plans(wb):
    ws = wb.create_sheet("tbl_UnitPlans")
    ws.sheet_properties.tabColor = "002060"
    headers = [
        "PlanID","Unit","StrategicObjective","KPI_Name","KPI_Description",
        "Frequency","Target_Daily","Target_Monthly","Target_Quarterly","Target_Annual",
        "Actual_MTD","Actual_QTD","Actual_YTD",
        "Variance_Monthly","Variance_Quarterly","Variance_Annual",
        "Achievement_Pct_Monthly","Achievement_Pct_Quarterly","Achievement_Pct_Annual",
        "TrafficLight_Monthly","TrafficLight_Quarterly","TrafficLight_Annual",
        "DataSource","Formula_Reference","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [12,16,28,22,30,12,12,14,14,14,12,12,12,14,14,14,18,18,18,18,18,18,18,20,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)

    # Pre-populate KPI rows for all units
    kpis = [
        # (Unit, Objective, KPI Name, KPI Description, Frequency, DataSource)
        ("Registry","Case Processing Efficiency","Cases Registered","Total new cases registered","Monthly","tbl_Cases"),
        ("Registry","Case Processing Efficiency","CR Compliance Rate","% of cases with compliant CR1/CR2/CR5","Monthly","tbl_Cases"),
        ("Registry","Case Processing Efficiency","Avg Days to First Action","Average days from registration to first action","Monthly","tbl_Actions"),
        ("Investigations","Case Resolution","Active Cases","Number of cases under active investigation","Monthly","tbl_Cases"),
        ("Investigations","Case Resolution","Cases Closed","Cases resolved/closed in period","Monthly","tbl_Cases"),
        ("Investigations","Case Resolution","Case Clearance Rate","% of cases cleared vs opened","Monthly","tbl_Cases"),
        ("Investigations","Case Resolution","Supervisory Reviews Done","Reviews completed on time","Monthly","tbl_Reviews"),
        ("Intelligence","Intelligence Processing","Leads Received","Total new leads received","Monthly","tbl_Leads"),
        ("Intelligence","Intelligence Processing","Leads Triaged < 24hrs","% of leads triaged within 24 hours","Monthly","tbl_Leads"),
        ("Intelligence","Intelligence Processing","Lead-to-Case Conversion Rate","% of leads converted to cases","Monthly","tbl_Leads"),
        ("Intelligence","Intelligence Processing","Intel Briefings Delivered","Number of briefings to operations","Monthly","tbl_IntelBriefings"),
        ("Operations","Operational Output","Total Operations Conducted","All ops including FCSI","Monthly","tbl_Operations"),
        ("Operations","Operational Output","FCSI Operations Conducted","Operations under FCSI flag","Monthly","tbl_Operations"),
        ("Operations","Operational Output","Firearms Seized","Total firearms seized from ops","Monthly","tbl_Operations"),
        ("Operations","Operational Output","Narcotics Seized (kg)","Total narcotics seized in kg","Monthly","tbl_Operations"),
        ("Operations","Operational Output","Arrests Made","Total arrests from operations","Monthly","tbl_Operations"),
        ("Operations","FCSI Effectiveness","FCSI Positive Outcome Rate","% of FCSI ops with seizure/arrest","Monthly","tbl_Operations"),
        ("Operations","FCSI Effectiveness","FCSI Hotspot Coverage","% of designated hotspots covered","Monthly","tbl_Operations"),
        ("Transport","Fleet Readiness","Vehicle Availability Rate","% of fleet available for duty","Monthly","tbl_Vehicles"),
        ("Transport","Fleet Readiness","Ops Support Coverage","% of operations with vehicle support","Monthly","tbl_VehicleUsage"),
        ("Transport","Fleet Readiness","Maintenance Compliance","% of vehicles serviced on schedule","Monthly","tbl_Maintenance"),
        ("Transport","Fleet Readiness","Avg Downtime Days","Average days per vehicle out of service","Monthly","tbl_Maintenance"),
        ("Demand Reduction","Community Outreach","Lectures Delivered","Total demand-reduction sessions","Monthly","tbl_DemandReduction"),
        ("Demand Reduction","Community Outreach","Total Attendance","Cumulative attendance at sessions","Monthly","tbl_DemandReduction"),
        ("Demand Reduction","Community Outreach","Communities Reached","Unique communities with sessions","Monthly","tbl_DemandReduction"),
        ("Admin/HRM","Human Resources","Personnel Strength","Current staffing level","Monthly","Manual"),
        ("Admin/HRM","Human Resources","Vacancy Rate","% of positions unfilled","Monthly","Manual"),
        ("Training","Capacity Building","Training Sessions Conducted","In-service training sessions","Monthly","Manual"),
        ("Training","Capacity Building","Personnel Trained","Number of officers trained","Monthly","Manual"),
        ("Oversight","Compliance & Governance","Non-Conformities Identified","Total compliance issues found","Monthly","Manual"),
        ("Oversight","Compliance & Governance","Overdue Actions","Count of overdue action items","Monthly","tbl_Actions"),
        ("Oversight","Compliance & Governance","KPIs Off Track","Number of KPIs rated Off Track","Monthly","tbl_UnitPlans"),
    ]

    for i, (unit, obj, kpi_name, kpi_desc, freq, src) in enumerate(kpis, 2):
        r = i
        ws.cell(row=r, column=1, value=f'KPI-{i-1:04d}')
        ws.cell(row=r, column=2, value=unit)
        ws.cell(row=r, column=3, value=obj)
        ws.cell(row=r, column=4, value=kpi_name)
        ws.cell(row=r, column=5, value=kpi_desc)
        ws.cell(row=r, column=6, value=freq)
        ws.cell(row=r, column=23, value=src)

        # Variance formulas
        ws.cell(row=r, column=14, value=f'=IF(OR(K{r}="",H{r}=""),"",K{r}-H{r})')
        ws.cell(row=r, column=15, value=f'=IF(OR(L{r}="",I{r}=""),"",L{r}-I{r})')
        ws.cell(row=r, column=16, value=f'=IF(OR(M{r}="",J{r}=""),"",M{r}-J{r})')

        # Achievement % formulas
        ws.cell(row=r, column=17, value=f'=IF(OR(K{r}="",H{r}="",H{r}=0),"",K{r}/H{r})')
        ws.cell(row=r, column=18, value=f'=IF(OR(L{r}="",I{r}="",I{r}=0),"",L{r}/I{r})')
        ws.cell(row=r, column=19, value=f'=IF(OR(M{r}="",J{r}="",J{r}=0),"",M{r}/J{r})')

        # Traffic light formulas
        for col_ach, col_tl in [(17,20),(18,21),(19,22)]:
            cl = get_column_letter(col_ach)
            ws.cell(row=r, column=col_tl,
                    value=f'=IF({cl}{r}="","",IF({cl}{r}>=0.9,"On Track",IF({cl}{r}>=0.7,"At Risk","Off Track")))')

        # Format achievement as %
        for col_pct in [17, 18, 19]:
            ws.cell(row=r, column=col_pct).number_format = '0%'

        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT

    last_row = 1 + len(kpis)
    ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    add_table(ws, "tbl_UnitPlans", ref)

    add_data_validation_list(ws, f"B2:B1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, f"F2:F1000", "=Lookups!$R$2:$R$6")
    add_data_validation_list(ws, f"T2:T1000", "=Lookups!$P$2:$P$4")
    add_data_validation_list(ws, f"U2:U1000", "=Lookups!$P$2:$P$4")
    add_data_validation_list(ws, f"V2:V1000", "=Lookups!$P$2:$P$4")

    # Traffic light conditional formatting
    for col_letter in ['T','U','V']:
        rng = f"{col_letter}2:{col_letter}1000"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={col_letter}2="On Track"'], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={col_letter}2="At Risk"'], fill=AMBER_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={col_letter}2="Off Track"'], fill=RED_FILL))

    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 9 — UNIT-SPECIFIC PORTAL SHEETS (Siloed Views)
# ══════════════════════════════════════════════════════════════════════

def _portal_header(ws, title, subtitle, unit_name, color):
    """Standard portal header block."""
    ws.sheet_properties.tabColor = color
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value=f"FNID AREA 3 — {title}")
    cell.font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:H2")
    cell2 = ws.cell(row=2, column=1, value=subtitle)
    cell2.font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A3:H3")
    cell3 = ws.cell(row=3, column=1, value=f"Unit: {unit_name}  |  Generated: {{=TODAY()}}  |  FNID Area 3 Operational Workbook")
    cell3.font = Font(name="Calibri", italic=True, size=9, color="666666")
    cell3.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18
    return 5  # next available row


def _portal_kpi_summary(ws, start_row, unit_name, kpis_list):
    """Add a KPI summary section referencing tbl_UnitPlans."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    ws.cell(row=start_row, column=1, value="KEY PERFORMANCE INDICATORS").font = SECTION_FONT
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r = start_row + 1
    kpi_headers = ["KPI","Target (Monthly)","Actual (MTD)","Variance","Achievement %","Status"]
    for i, h in enumerate(kpi_headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    r += 1
    for kpi_row_num, kpi_name in kpis_list:
        ws.cell(row=r, column=1, value=kpi_name).font = BODY_FONT
        ws.cell(row=r, column=2, value=f"=tbl_UnitPlans[[#Data],[Target_Monthly]]").font = BODY_FONT
        # Use INDEX/MATCH for specific KPI lookup
        ws.cell(row=r, column=2, value=f'=IFERROR(INDEX(tbl_UnitPlans[Target_Monthly],MATCH("{kpi_name}",tbl_UnitPlans[KPI_Name],0)),"—")')
        ws.cell(row=r, column=3, value=f'=IFERROR(INDEX(tbl_UnitPlans[Actual_MTD],MATCH("{kpi_name}",tbl_UnitPlans[KPI_Name],0)),"—")')
        ws.cell(row=r, column=4, value=f'=IFERROR(INDEX(tbl_UnitPlans[Variance_Monthly],MATCH("{kpi_name}",tbl_UnitPlans[KPI_Name],0)),"—")')
        ws.cell(row=r, column=5, value=f'=IFERROR(INDEX(tbl_UnitPlans[Achievement_Pct_Monthly],MATCH("{kpi_name}",tbl_UnitPlans[KPI_Name],0)),"—")')
        ws.cell(row=r, column=5).number_format = '0%'
        ws.cell(row=r, column=6, value=f'=IFERROR(INDEX(tbl_UnitPlans[TrafficLight_Monthly],MATCH("{kpi_name}",tbl_UnitPlans[KPI_Name],0)),"—")')
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1

    # Traffic light conditional formatting on status column
    col_f = f"F{start_row+2}:F{r}"
    ws.conditional_formatting.add(col_f, FormulaRule(formula=[f'=F{start_row+2}="On Track"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(col_f, FormulaRule(formula=[f'=F{start_row+2}="At Risk"'], fill=AMBER_FILL))
    ws.conditional_formatting.add(col_f, FormulaRule(formula=[f'=F{start_row+2}="Off Track"'], fill=RED_FILL))
    return r + 1


def _portal_quick_stats(ws, start_row, stats_formulas):
    """Add quick stats boxes."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    ws.cell(row=start_row, column=1, value="QUICK STATISTICS").font = SECTION_FONT
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r = start_row + 1
    col = 1
    for label, formula in stats_formulas:
        ws.cell(row=r, column=col, value=label).font = BODY_FONT_BOLD
        ws.cell(row=r, column=col).border = THIN_BORDER
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=MED_GRAY)
        ws.cell(row=r+1, column=col, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r+1, column=col).border = THIN_BORDER
        ws.cell(row=r+1, column=col).alignment = CENTER
        col += 1
        if col > 8:
            col = 1
            r += 3
    return r + 3


def build_portal_registry(wb):
    ws = wb.create_sheet("Portal_Registry")
    r = _portal_header(ws, "REGISTRY PORTAL", "Case Registration, CR Tracking & Compliance", "Registry", "1F3864")

    kpis = [
        (1, "Cases Registered"),
        (2, "CR Compliance Rate"),
        (3, "Avg Days to First Action"),
    ]
    r = _portal_kpi_summary(ws, r, "Registry", kpis)

    stats = [
        ("Total Cases", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
        ("Open Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Open")'),
        ("Active Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Active")'),
        ("Pending Review", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Review")'),
        ("Closed Cases", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*")'),
        ("CR1 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR1_Compliant],"Yes")/COUNTA(tbl_Cases[CR1_Compliant]),"—")'),
        ("Non-Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    # Instruction text
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Navigate to sheet 'tbl_Cases' to register new cases. Use 'tbl_Actions' for action items. Use 'tbl_Reviews' for supervisory reviews.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_investigations(wb):
    ws = wb.create_sheet("Portal_Investigations")
    r = _portal_header(ws, "INVESTIGATIONS PORTAL", "Case Investigation Tracking & Supervisory Review", "Investigations", "1F3864")

    kpis = [
        (4, "Active Cases"),
        (5, "Cases Closed"),
        (6, "Case Clearance Rate"),
        (7, "Supervisory Reviews Done"),
    ]
    r = _portal_kpi_summary(ws, r, "Investigations", kpis)

    stats = [
        ("Under Investigation", '=COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
        ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
        ("Reviews Completed", '=COUNTA(tbl_Reviews[ReviewID])-COUNTBLANK(tbl_Reviews[ReviewID])'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[DueDate],"<"&TODAY(),tbl_Reviews[Status],"<>Completed")'),
        ("Diary Entries", '=COUNTA(tbl_DiaryEntries[DiaryID])-COUNTBLANK(tbl_DiaryEntries[DiaryID])'),
        ("Persons of Interest", '=COUNTA(tbl_Persons[PersonID])-COUNTBLANK(tbl_Persons[PersonID])'),
        ("Linked Leads", '=COUNTIFS(tbl_Cases[LinkedLeadID],"<>")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Navigate to 'tbl_Cases' for case updates, 'tbl_DiaryEntries' for investigation diary, 'tbl_Reviews' for supervisory reviews, 'tbl_Persons' for suspect/witness data.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_operations(wb):
    ws = wb.create_sheet("Portal_Operations")
    r = _portal_header(ws, "OPERATIONS PORTAL", "Operational Activity, FCSI Tracking & Outcomes", "Operations", "C00000")

    kpis = [
        (12, "Total Operations Conducted"),
        (13, "FCSI Operations Conducted"),
        (14, "Firearms Seized"),
        (15, "Narcotics Seized (kg)"),
        (16, "Arrests Made"),
        (17, "FCSI Positive Outcome Rate"),
        (18, "FCSI Hotspot Coverage"),
    ]
    r = _portal_kpi_summary(ws, r, "Operations", kpis)

    stats = [
        ("Total Ops", '=COUNTA(tbl_Operations[OPS_ID])-COUNTBLANK(tbl_Operations[OPS_ID])'),
        ("FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("Non-FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"No")'),
        ("Total Arrests", '=SUM(tbl_Operations[Arrests])'),
        ("Firearms Seized", '=SUM(tbl_Operations[FirearmsSeized])'),
        ("Ammo Seized", '=SUM(tbl_Operations[AmmunitionSeized])'),
        ("Narcotics (kg)", '=SUM(tbl_Operations[NarcoticsSeized_kg])'),
        ("Search Warrants", '=COUNTIF(tbl_Operations[OpsType],"Search Warrant")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Navigate to 'tbl_Operations' to log operational activity. Mark FCSI_Flag='Yes' for all Focused Crime Suppression Initiative operations.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_intelligence(wb):
    ws = wb.create_sheet("Portal_Intelligence")
    r = _portal_header(ws, "INTELLIGENCE PORTAL", "Lead Management, Triage & Intelligence Briefings", "Intelligence", "BF8F00")

    kpis = [
        (8, "Leads Received"),
        (9, "Leads Triaged < 24hrs"),
        (10, "Lead-to-Case Conversion Rate"),
        (11, "Intel Briefings Delivered"),
    ]
    r = _portal_kpi_summary(ws, r, "Intelligence", kpis)

    stats = [
        ("Total Leads", '=COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])'),
        ("From MCR", '=COUNTIF(tbl_Leads[Source],"MCR")'),
        ("From Ops", '=COUNTIF(tbl_Leads[Source],"Operations")'),
        ("From Tips", '=COUNTIF(tbl_Leads[Source],"Tip-off")'),
        ("High Priority", '=COUNTIFS(tbl_Leads[Priority],"High")+COUNTIFS(tbl_Leads[Priority],"Critical")'),
        ("Pending Triage", '=COUNTBLANK(tbl_Leads[TriageDecision])'),
        ("Converted to Case", '=COUNTIF(tbl_Leads[Outcome],"Converted to Case")'),
        ("Briefings Given", '=COUNTA(tbl_IntelBriefings[BriefingID])-COUNTBLANK(tbl_IntelBriefings[BriefingID])'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Use 'tbl_MCR_Raw' for MCR intake, 'tbl_Leads' for lead management, 'tbl_FollowUpRecs' for recommendations, 'tbl_IntelBriefings' for briefing logs.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_transport(wb):
    ws = wb.create_sheet("Portal_Transport")
    r = _portal_header(ws, "TRANSPORT PORTAL", "Fleet Management, Vehicle Usage & Maintenance", "Transport", "2E75B6")

    kpis = [
        (19, "Vehicle Availability Rate"),
        (20, "Ops Support Coverage"),
        (21, "Maintenance Compliance"),
        (22, "Avg Downtime Days"),
    ]
    r = _portal_kpi_summary(ws, r, "Transport", kpis)

    stats = [
        ("Total Vehicles", '=COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])'),
        ("Available", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Available")'),
        ("In Use", '=COUNTIF(tbl_Vehicles[CurrentStatus],"In Use")'),
        ("Under Maintenance", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Under Maintenance")'),
        ("Out of Service", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Out of Service")'),
        ("Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("Total Fuel (L)", '=SUM(tbl_FuelLog[Litres])'),
        ("Active Drivers", '=COUNTIF(tbl_Drivers[Status],"Active")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Use 'tbl_Vehicles' for fleet register, 'tbl_VehicleUsage' for trip logs, 'tbl_FuelLog' for fuel records, 'tbl_Maintenance' for service records, 'tbl_Drivers' for driver registry.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_demand_reduction(wb):
    ws = wb.create_sheet("Portal_DemandReduction")
    r = _portal_header(ws, "DEMAND REDUCTION PORTAL", "Community Outreach Lectures & Activity Tracking", "Demand Reduction", "00B050")

    kpis = [
        (23, "Lectures Delivered"),
        (24, "Total Attendance"),
        (25, "Communities Reached"),
    ]
    r = _portal_kpi_summary(ws, r, "Demand Reduction", kpis)

    stats = [
        ("Sessions Held", '=COUNTA(tbl_DemandReduction[LectureID])-COUNTBLANK(tbl_DemandReduction[LectureID])'),
        ("Total Attendance", '=SUM(tbl_DemandReduction[Attendance])'),
        ("Avg Attendance", '=IFERROR(AVERAGE(tbl_DemandReduction[Attendance]),"—")'),
        ("Total Hours", '=SUM(tbl_DemandReduction[Duration_Hrs])'),
        ("With Follow-Up", '=COUNTIF(tbl_DemandReduction[FollowUpPlanned],"Yes")'),
        ("Parishes Covered", '=IFERROR(SUMPRODUCT(1/COUNTIF(tbl_DemandReduction[Parish],tbl_DemandReduction[Parish])),"—")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Navigate to 'tbl_DemandReduction' to log outreach sessions, attendance, topics, and outcomes.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_admin_hrm(wb):
    ws = wb.create_sheet("Portal_AdminHRM")
    r = _portal_header(ws, "ADMIN / HRM PORTAL", "Personnel, Administration & Human Resource Management", "Admin/HRM", "7030A0")

    kpis = [
        (26, "Personnel Strength"),
        (27, "Vacancy Rate"),
    ]
    r = _portal_kpi_summary(ws, r, "Admin/HRM", kpis)

    # Admin-specific quick stats (manual-entry oriented)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="ADMINISTRATIVE SUMMARY").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    admin_items = [
        "Current Establishment Strength:", "Actual Strength:", "Vacancies:",
        "On Leave:", "Seconded Out:", "Seconded In:",
        "Pending Transfers:", "Disciplinary Matters:"
    ]
    for item in admin_items:
        ws.cell(row=r, column=1, value=item).font = BODY_FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: Admin/HRM data is primarily manual entry. Update targets and actuals in 'tbl_UnitPlans' for Admin/HRM unit.").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_training(wb):
    ws = wb.create_sheet("Portal_Training")
    r = _portal_header(ws, "TRAINING PORTAL", "Capacity Building, In-Service Training & Certifications", "Training", "548235")

    kpis = [
        (28, "Training Sessions Conducted"),
        (29, "Personnel Trained"),
    ]
    r = _portal_kpi_summary(ws, r, "Training", kpis)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="TRAINING LOG").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    train_headers = ["Date","Course/Topic","Trainer","Attendees","Duration (hrs)","Location","Certificate Issued","Notes"]
    for i, h in enumerate(train_headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    r += 1
    for row_offset in range(20):
        for c in range(1, len(train_headers)+1):
            ws.cell(row=r+row_offset, column=c).border = THIN_BORDER
            ws.cell(row=r+row_offset, column=c).font = BODY_FONT

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


def build_portal_oversight(wb):
    ws = wb.create_sheet("Portal_Oversight")
    r = _portal_header(ws, "OVERSIGHT PORTAL", "Compliance, Non-Conformities, Governance & KPI Review", "Oversight", "FF0000")

    kpis = [
        (30, "Non-Conformities Identified"),
        (31, "Overdue Actions"),
        (32, "KPIs Off Track"),
    ]
    r = _portal_kpi_summary(ws, r, "Oversight", kpis)

    stats = [
        ("Total Cases", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
        ("Non-Compliant Cases", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[DueDate],"<"&TODAY(),tbl_Reviews[Status],"<>Completed")'),
        ("Overdue Forensics", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Service Overdue Vehicles", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("KPIs Off Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"Off Track")'),
        ("KPIs At Risk", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"At Risk")'),
    ]
    r = _portal_quick_stats(ws, r, stats)

    # Non-conformity log section
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="NON-CONFORMITY LOG").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    nc_headers = ["NC_ID","Date","Unit","Description","Severity","Corrective Action","Due Date","Status"]
    for i, h in enumerate(nc_headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    r += 1
    for row_offset in range(15):
        ws.cell(row=r+row_offset, column=1, value=f'=IF(B{r+row_offset}="","","NC-"&TEXT(ROW()-{r-1},"0000"))')
        for c in range(1, len(nc_headers)+1):
            ws.cell(row=r+row_offset, column=c).border = THIN_BORDER
            ws.cell(row=r+row_offset, column=c).font = BODY_FONT

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws




# ══════════════════════════════════════════════════════════════════════
#  MODULE 10 — DASHBOARDS & STATISTICS HUB (ALL UNITS)
# ══════════════════════════════════════════════════════════════════════

def _dashboard_banner(ws, title, subtitle):
    ws.merge_cells("A1:L1")
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Calibri", bold=True, size=20, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:L2")
    cell2 = ws.cell(row=2, column=1, value=subtitle)
    cell2.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    cell2.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    return 4


def build_executive_dashboard(wb):
    ws = wb.create_sheet("Dash_Executive")
    ws.sheet_properties.tabColor = "1F3864"
    r = _dashboard_banner(ws,
        "FNID AREA 3 — EXECUTIVE DASHBOARD",
        "Integrated Operational Picture  |  Intelligence → Cases → Operations → Evidence → Performance")

    # ── SECTION 1: Operations Overview ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="OPERATIONAL OVERVIEW").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1

    overview_metrics = [
        ("Total Cases", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
        ("Active Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Active")+COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Cases Closed", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*")'),
        ("Total Leads", '=COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])'),
        ("Leads Converted", '=COUNTIF(tbl_Leads[Outcome],"Converted to Case")+COUNTIF(tbl_Leads[Outcome],"Converted to Op")'),
        ("Total Ops", '=COUNTA(tbl_Operations[OPS_ID])-COUNTBLANK(tbl_Operations[OPS_ID])'),
        ("FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("Firearms Seized", '=SUM(tbl_Operations[FirearmsSeized])'),
        ("Ammo Seized", '=SUM(tbl_Operations[AmmunitionSeized])'),
        ("Narcotics (kg)", '=SUM(tbl_Operations[NarcoticsSeized_kg])'),
        ("Arrests", '=SUM(tbl_Operations[Arrests])'),
        ("Total Exhibits", '=COUNTA(tbl_Exhibits[ExhibitID])-COUNTBLANK(tbl_Exhibits[ExhibitID])'),
    ]

    # Row of metric labels
    for i, (label, _) in enumerate(overview_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    # Row of metric values
    for i, (_, formula) in enumerate(overview_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=16, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 2

    # ── SECTION 2: Intelligence Pipeline ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="INTELLIGENCE PIPELINE").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    intel_metrics = [
        ("MCR Entries", '=COUNTA(tbl_MCR_Raw[MCR_ID])-COUNTBLANK(tbl_MCR_Raw[MCR_ID])'),
        ("FNID-Relevant MCRs", '=COUNTIF(tbl_MCR_Raw[FNID_Relevant],"Yes")'),
        ("MCR→Lead Rate", '=IFERROR(COUNTIFS(tbl_MCR_Raw[LinkedLeadID],"<>")/COUNTIF(tbl_MCR_Raw[FNID_Relevant],"Yes"),"—")'),
        ("Leads Pending Triage", '=COUNTBLANK(tbl_Leads[TriageDecision])'),
        ("High/Crit Leads", '=COUNTIFS(tbl_Leads[Priority],"High")+COUNTIFS(tbl_Leads[Priority],"Critical")'),
        ("Lead→Case Conv %", '=IFERROR(COUNTIF(tbl_Leads[Outcome],"Converted to Case")/(COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])),"—")'),
        ("Intel Briefings", '=COUNTA(tbl_IntelBriefings[BriefingID])-COUNTBLANK(tbl_IntelBriefings[BriefingID])'),
        ("Follow-Up Overdue", '=COUNTIFS(tbl_Leads[FollowUpDueDate],"<"&TODAY(),tbl_Leads[FollowUpRequired],"Yes")'),
    ]
    for i, (label, _) in enumerate(intel_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="BF8F00")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(intel_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # ── SECTION 3: FCSI Performance ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="FOCUSED CRIME SUPPRESSION INITIATIVE (FCSI)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GOLD)
    r += 1
    fcsi_metrics = [
        ("FCSI Ops Conducted", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("Non-FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"No")'),
        ("FCSI Arrests", '=SUMIFS(tbl_Operations[Arrests],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Firearms Seized", '=SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Narcotics (kg)", '=SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Positive Outcome %", '=IFERROR(COUNTIFS(tbl_Operations[FCSI_Flag],"Yes",tbl_Operations[Arrests],">"&0)/COUNTIF(tbl_Operations[FCSI_Flag],"Yes"),"—")'),
    ]
    for i, (label, _) in enumerate(fcsi_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=BLACK)
        cell.fill = GOLD_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(fcsi_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # ── SECTION 4: Evidence & Forensics ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="EVIDENCE & FORENSIC STATUS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    ev_metrics = [
        ("Total Exhibits", '=COUNTA(tbl_Exhibits[ExhibitID])-COUNTBLANK(tbl_Exhibits[ExhibitID])'),
        ("Pending Analysis", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Pending")'),
        ("Submitted", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Submitted")'),
        ("In Progress", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"In Progress")'),
        ("Completed", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Completed")'),
        ("Court Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Ready")'),
        ("Not Court Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Not Ready")'),
        ("Overdue Returns", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
    ]
    for i, (label, _) in enumerate(ev_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="538135")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(ev_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # ── SECTION 5: Transport & Logistics ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="TRANSPORT & FLEET STATUS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    transport_metrics = [
        ("Total Fleet", '=COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])'),
        ("Available", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Available")'),
        ("In Use", '=COUNTIF(tbl_Vehicles[CurrentStatus],"In Use")'),
        ("Under Maint.", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Under Maintenance")'),
        ("Out of Service", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Out of Service")'),
        ("Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("Total Fuel (L)", '=SUM(tbl_FuelLog[Litres])'),
        ("Active Drivers", '=COUNTIF(tbl_Drivers[Status],"Active")'),
    ]
    for i, (label, _) in enumerate(transport_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(transport_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # ── SECTION 6: Demand Reduction ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="DEMAND REDUCTION & COMMUNITY OUTREACH").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1
    dr_metrics = [
        ("Sessions Held", '=COUNTA(tbl_DemandReduction[LectureID])-COUNTBLANK(tbl_DemandReduction[LectureID])'),
        ("Total Attendance", '=SUM(tbl_DemandReduction[Attendance])'),
        ("Avg Attendance", '=IFERROR(AVERAGE(tbl_DemandReduction[Attendance]),"—")'),
        ("Total Hours", '=SUM(tbl_DemandReduction[Duration_Hrs])'),
        ("With Follow-Up", '=COUNTIF(tbl_DemandReduction[FollowUpPlanned],"Yes")'),
    ]
    for i, (label, _) in enumerate(dr_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="00B050")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(dr_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # ── SECTION 7: Compliance & Governance ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="COMPLIANCE & GOVERNANCE ALERTS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FF0000")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    r += 1
    alert_metrics = [
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[DueDate],"<"&TODAY(),tbl_Reviews[Status],"<>Completed")'),
        ("Non-Compliant Cases", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Forensics", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Overdue Follow-Ups", '=COUNTIFS(tbl_Leads[FollowUpDueDate],"<"&TODAY(),tbl_Leads[FollowUpRequired],"Yes")'),
        ("KPIs Off Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"Off Track")'),
        ("KPIs At Risk", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"At Risk")'),
        ("Vehicle Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
    ]
    for i, (label, _) in enumerate(alert_metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(alert_metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=16, color="C00000")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 2

    # ── SECTION 8: Unit Plan Performance Summary (ALL UNITS) ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="UNIT PERFORMANCE — PLAN vs ACTUAL (ALL UNITS)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    r += 1

    plan_headers = ["Unit","KPI","Target (Monthly)","Actual (MTD)","Variance","Achievement %","Status (Monthly)"]
    for i, h in enumerate(plan_headers):
        cell = ws.cell(row=r, column=i+1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1

    # Reference all 32 KPIs from tbl_UnitPlans
    for kpi_idx in range(1, 33):
        ws.cell(row=r, column=1, value=f'=INDEX(tbl_UnitPlans[Unit],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=2, value=f'=INDEX(tbl_UnitPlans[KPI_Name],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=INDEX(tbl_UnitPlans[Target_Monthly],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=4, value=f'=INDEX(tbl_UnitPlans[Actual_MTD],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=INDEX(tbl_UnitPlans[Variance_Monthly],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=6, value=f'=INDEX(tbl_UnitPlans[Achievement_Pct_Monthly],{kpi_idx})').border = THIN_BORDER
        ws.cell(row=r, column=6).number_format = '0%'
        ws.cell(row=r, column=7, value=f'=INDEX(tbl_UnitPlans[TrafficLight_Monthly],{kpi_idx})').border = THIN_BORDER
        for c in range(1, 8):
            ws.cell(row=r, column=c).font = BODY_FONT
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1

    # Traffic light formatting on status column
    status_range = f"G{r-32}:G{r-1}"
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'=G{r-32}="On Track"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'=G{r-32}="At Risk"'], fill=AMBER_FILL))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'=G{r-32}="Off Track"'], fill=RED_FILL))

    # Set column widths
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 20

    return ws




def build_statistics_hub(wb):
    """Central Statistics Hub aggregating all modules."""
    ws = wb.create_sheet("Statistics_Hub")
    ws.sheet_properties.tabColor = "002060"
    r = _dashboard_banner(ws,
        "FNID AREA 3 — CENTRAL STATISTICS HUB",
        "Aggregated Statistics Across All Modules & Units")

    sections = [
        ("CASES & REGISTRY", "1F3864", [
            ("Total Cases Registered", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
            ("Open / Active", '=COUNTIF(tbl_Cases[CaseStatus],"Open")+COUNTIF(tbl_Cases[CaseStatus],"Active")'),
            ("Under Investigation", '=COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
            ("Pending Review", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Review")'),
            ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
            ("Closed (All)", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*")'),
            ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
            ("CR1 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR1_Compliant],"Yes")/COUNTA(tbl_Cases[CR1_Compliant]),"—")'),
            ("CR2 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR2_Compliant],"Yes")/COUNTA(tbl_Cases[CR2_Compliant]),"—")'),
            ("CR5 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR5_Compliant],"Yes")/COUNTA(tbl_Cases[CR5_Compliant]),"—")'),
            ("Firearms Cases", '=COUNTIFS(tbl_Cases[OffenceType],"*Firearm*")+COUNTIFS(tbl_Cases[OffenceType],"*Ammunition*")'),
            ("Narcotics Cases", '=COUNTIFS(tbl_Cases[OffenceType],"*Narcotic*")+COUNTIFS(tbl_Cases[OffenceType],"*Cannabis*")+COUNTIFS(tbl_Cases[OffenceType],"*Cocaine*")'),
        ]),
        ("INTELLIGENCE", "BF8F00", [
            ("Total Leads", '=COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])'),
            ("MCR-Sourced Leads", '=COUNTIF(tbl_Leads[Source],"MCR")'),
            ("Operations-Sourced", '=COUNTIF(tbl_Leads[Source],"Operations")'),
            ("Tip-off Sourced", '=COUNTIF(tbl_Leads[Source],"Tip-off")'),
            ("Leads Triaged", '=COUNTA(tbl_Leads[TriageDecision])-COUNTBLANK(tbl_Leads[TriageDecision])'),
            ("Pending Triage", '=COUNTBLANK(tbl_Leads[TriageDecision])'),
            ("Converted to Case", '=COUNTIF(tbl_Leads[Outcome],"Converted to Case")'),
            ("Converted to Op", '=COUNTIF(tbl_Leads[Outcome],"Converted to Op")'),
            ("Conversion Rate", '=IFERROR((COUNTIF(tbl_Leads[Outcome],"Converted to Case")+COUNTIF(tbl_Leads[Outcome],"Converted to Op"))/(COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])),"—")'),
            ("Briefings Delivered", '=COUNTA(tbl_IntelBriefings[BriefingID])-COUNTBLANK(tbl_IntelBriefings[BriefingID])'),
            ("FCSI-Flagged Leads", '=COUNTIF(tbl_Leads[FCSI_Flag],"Yes")'),
            ("Follow-Ups Overdue", '=COUNTIFS(tbl_Leads[FollowUpDueDate],"<"&TODAY(),tbl_Leads[FollowUpRequired],"Yes")'),
        ]),
        ("OPERATIONS", "C00000", [
            ("Total Operations", '=COUNTA(tbl_Operations[OPS_ID])-COUNTBLANK(tbl_Operations[OPS_ID])'),
            ("Search Warrants", '=COUNTIF(tbl_Operations[OpsType],"Search Warrant")'),
            ("Patrols", '=COUNTIF(tbl_Operations[OpsType],"Patrol")'),
            ("Checkpoints", '=COUNTIF(tbl_Operations[OpsType],"Checkpoint")'),
            ("Raids", '=COUNTIF(tbl_Operations[OpsType],"Raid")'),
            ("Surveillance", '=COUNTIF(tbl_Operations[OpsType],"Surveillance")'),
            ("FCSI Operations", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
            ("Total Arrests", '=SUM(tbl_Operations[Arrests])'),
            ("Firearms Seized", '=SUM(tbl_Operations[FirearmsSeized])'),
            ("Ammunition Seized", '=SUM(tbl_Operations[AmmunitionSeized])'),
            ("Narcotics Seized (kg)", '=SUM(tbl_Operations[NarcoticsSeized_kg])'),
            ("Cash Seized", '=SUM(tbl_Operations[CashSeized])'),
        ]),
        ("EVIDENCE & FORENSICS", "538135", [
            ("Total Exhibits", '=COUNTA(tbl_Exhibits[ExhibitID])-COUNTBLANK(tbl_Exhibits[ExhibitID])'),
            ("Pending Analysis", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Pending")'),
            ("Submitted to Lab", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Submitted")'),
            ("Analysis In Progress", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"In Progress")'),
            ("Analysis Completed", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Completed")'),
            ("Court Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Ready")'),
            ("Not Court Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Not Ready")'),
            ("Avg Turnaround (days)", '=IFERROR(AVERAGE(tbl_Exhibits[TurnaroundDays]),"—")'),
            ("Overdue Returns", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
            ("Firearms Exhibits", '=COUNTIF(tbl_Exhibits[EvidenceType],"Firearm")'),
            ("Narcotics Exhibits", '=COUNTIFS(tbl_Exhibits[EvidenceType],"Narcotics*")'),
            ("Electronic Devices", '=COUNTIF(tbl_Exhibits[EvidenceType],"Electronic Device")'),
        ]),
        ("TRANSPORT & FLEET", "2E75B6", [
            ("Total Vehicles", '=COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])'),
            ("Available", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Available")'),
            ("In Use", '=COUNTIF(tbl_Vehicles[CurrentStatus],"In Use")'),
            ("Under Maintenance", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Under Maintenance")'),
            ("Out of Service", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Out of Service")'),
            ("Availability Rate %", '=IFERROR(COUNTIF(tbl_Vehicles[CurrentStatus],"Available")/(COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])),"—")'),
            ("Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
            ("Total Fuel Used (L)", '=SUM(tbl_FuelLog[Litres])'),
            ("Total Fuel Cost", '=SUM(tbl_FuelLog[Cost])'),
            ("Total Trips", '=COUNTA(tbl_VehicleUsage[UsageID])-COUNTBLANK(tbl_VehicleUsage[UsageID])'),
            ("Total Distance (km)", '=SUM(tbl_VehicleUsage[DistanceKm])'),
            ("Active Drivers", '=COUNTIF(tbl_Drivers[Status],"Active")'),
        ]),
        ("DEMAND REDUCTION", "00B050", [
            ("Lectures Delivered", '=COUNTA(tbl_DemandReduction[LectureID])-COUNTBLANK(tbl_DemandReduction[LectureID])'),
            ("Total Attendance", '=SUM(tbl_DemandReduction[Attendance])'),
            ("Avg Attendance/Session", '=IFERROR(AVERAGE(tbl_DemandReduction[Attendance]),"—")'),
            ("Total Duration (hrs)", '=SUM(tbl_DemandReduction[Duration_Hrs])'),
            ("Sessions w/ Follow-Up", '=COUNTIF(tbl_DemandReduction[FollowUpPlanned],"Yes")'),
            ("Unique Parishes", '=IFERROR(SUMPRODUCT(1/COUNTIF(tbl_DemandReduction[Parish],tbl_DemandReduction[Parish])),"—")'),
        ]),
    ]

    for section_title, color, metrics in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

        # Lay metrics in 2 columns (Metric | Value | Metric | Value ... )
        col = 1
        for label, formula in metrics:
            ws.cell(row=r, column=col, value=label).font = BODY_FONT_BOLD
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=MED_GRAY)
            ws.cell(row=r, column=col+1, value=formula).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
            ws.cell(row=r, column=col+1).border = THIN_BORDER
            ws.cell(row=r, column=col+1).alignment = CENTER
            col += 2
            if col > 12:
                col = 1
                r += 1
        r += 2

    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18
    return ws


def build_fcsi_dashboard(wb):
    """Dedicated FCSI performance dashboard."""
    ws = wb.create_sheet("Dash_FCSI")
    ws.sheet_properties.tabColor = "FFC000"
    r = _dashboard_banner(ws,
        "FOCUSED CRIME SUPPRESSION INITIATIVE (FCSI) DASHBOARD",
        "FNID Area 3 — FCSI Operations, Outcomes & Trend Analysis")

    # Totals row
    fcsi_totals = [
        ("FCSI Ops Total", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Arrests", '=SUMIFS(tbl_Operations[Arrests],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Searches", '=SUMIFS(tbl_Operations[Searches],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Firearms", '=SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Ammo", '=SUMIFS(tbl_Operations[AmmunitionSeized],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Narcotics (kg)", '=SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[FCSI_Flag],"Yes")'),
        ("FCSI Cash Seized", '=SUMIFS(tbl_Operations[CashSeized],tbl_Operations[FCSI_Flag],"Yes")'),
        ("Positive Outcome %", '=IFERROR(COUNTIFS(tbl_Operations[FCSI_Flag],"Yes",tbl_Operations[Arrests],">"&0)/COUNTIF(tbl_Operations[FCSI_Flag],"Yes"),"—")'),
    ]
    for i, (label, _) in enumerate(fcsi_totals):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=BLACK)
        cell.fill = GOLD_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(fcsi_totals):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=16, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 2

    # FCSI by Operation Type breakdown
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="FCSI BREAKDOWN BY OPERATION TYPE").font = SECTION_FONT
    r += 1
    ops_types = ["Search Warrant","Patrol","Checkpoint","Surveillance","Raid","Arrest Operation","Joint Operation","FCSI Operation"]
    bk_headers = ["Op Type","Count","Arrests","Firearms","Ammo","Narcotics(kg)"]
    for i, h in enumerate(bk_headers):
        cell = ws.cell(row=r, column=i+1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for ot in ops_types:
        ws.cell(row=r, column=1, value=ot).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIFS(tbl_Operations[OpsType],"{ot}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=SUMIFS(tbl_Operations[Arrests],tbl_Operations[OpsType],"{ot}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=4, value=f'=SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[OpsType],"{ot}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=5, value=f'=SUMIFS(tbl_Operations[AmmunitionSeized],tbl_Operations[OpsType],"{ot}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=6, value=f'=SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[OpsType],"{ot}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1

    r += 1
    # FCSI by Parish
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="FCSI BREAKDOWN BY PARISH").font = SECTION_FONT
    r += 1
    parish_headers = ["Parish","FCSI Ops","Arrests","Firearms","Narcotics(kg)"]
    for i, h in enumerate(parish_headers):
        cell = ws.cell(row=r, column=i+1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    parishes = ["Kingston","St. Andrew","St. Catherine","Clarendon","Manchester",
                "St. Elizabeth","Westmoreland","Hanover","St. James","Trelawny",
                "St. Ann","St. Mary","Portland","St. Thomas"]
    for p in parishes:
        ws.cell(row=r, column=1, value=p).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIFS(tbl_Operations[Parish],"{p}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=SUMIFS(tbl_Operations[Arrests],tbl_Operations[Parish],"{p}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=4, value=f'=SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[Parish],"{p}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        ws.cell(row=r, column=5, value=f'=SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[Parish],"{p}",tbl_Operations[FCSI_Flag],"Yes")').font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1

    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18
    return ws




def build_unit_dashboards(wb):
    """Build individual unit dashboards — one sheet per unit."""

    # ── Registry Dashboard ──
    ws = wb.create_sheet("Dash_Registry")
    ws.sheet_properties.tabColor = "1F3864"
    r = _dashboard_banner(ws, "REGISTRY UNIT DASHBOARD", "Case Intake, CR Compliance & Processing Metrics")
    metrics = [
        ("Cases This Month", '=COUNTIFS(tbl_Cases[Month],MONTH(TODAY()),tbl_Cases[Year],YEAR(TODAY()))'),
        ("Open Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Open")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("CR1 Compliance %", '=IFERROR(COUNTIF(tbl_Cases[CR1_Compliant],"Yes")/COUNTA(tbl_Cases[CR1_Compliant]),"—")'),
        ("CR2 Compliance %", '=IFERROR(COUNTIF(tbl_Cases[CR2_Compliant],"Yes")/COUNTA(tbl_Cases[CR2_Compliant]),"—")'),
        ("CR5 Compliance %", '=IFERROR(COUNTIF(tbl_Cases[CR5_Compliant],"Yes")/COUNTA(tbl_Cases[CR5_Compliant]),"—")'),
        ("Non-Compliant Cases", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Radio Messages", '=COUNTA(tbl_RadioMessages[MessageID])-COUNTBLANK(tbl_RadioMessages[MessageID])'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Investigations Dashboard ──
    ws = wb.create_sheet("Dash_Investigations")
    ws.sheet_properties.tabColor = "1F3864"
    r = _dashboard_banner(ws, "INVESTIGATIONS UNIT DASHBOARD", "Active Cases, Resolution Rates & Review Status")
    metrics = [
        ("Active Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Active")+COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Cases Closed (YTD)", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[Year],YEAR(TODAY()))'),
        ("Clearance Rate", '=IFERROR(COUNTIFS(tbl_Cases[CaseStatus],"Closed*")/(COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])),"—")'),
        ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
        ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
        ("Reviews Done", '=COUNTA(tbl_Reviews[ReviewID])-COUNTBLANK(tbl_Reviews[ReviewID])'),
        ("Diary Entries", '=COUNTA(tbl_DiaryEntries[DiaryID])-COUNTBLANK(tbl_DiaryEntries[DiaryID])'),
        ("Persons Tracked", '=COUNTA(tbl_Persons[PersonID])-COUNTBLANK(tbl_Persons[PersonID])'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Operations Dashboard ──
    ws = wb.create_sheet("Dash_Operations")
    ws.sheet_properties.tabColor = "C00000"
    r = _dashboard_banner(ws, "OPERATIONS UNIT DASHBOARD", "Operational Output, Seizures, FCSI & Arrest Metrics")
    metrics = [
        ("Ops This Month", '=COUNTIFS(tbl_Operations[Month],MONTH(TODAY()),tbl_Operations[Year],YEAR(TODAY()))'),
        ("FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("Total Arrests", '=SUM(tbl_Operations[Arrests])'),
        ("Firearms Seized", '=SUM(tbl_Operations[FirearmsSeized])'),
        ("Ammo Seized", '=SUM(tbl_Operations[AmmunitionSeized])'),
        ("Narcotics (kg)", '=SUM(tbl_Operations[NarcoticsSeized_kg])'),
        ("Cash Seized", '=SUM(tbl_Operations[CashSeized])'),
        ("Warrants Executed", '=COUNTIF(tbl_Operations[OpsType],"Search Warrant")'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="C00000")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Intelligence Dashboard ──
    ws = wb.create_sheet("Dash_Intelligence")
    ws.sheet_properties.tabColor = "BF8F00"
    r = _dashboard_banner(ws, "INTELLIGENCE UNIT DASHBOARD", "Lead Pipeline, Triage Rates & Briefing Activity")
    metrics = [
        ("Leads This Month", '=COUNTIFS(tbl_Leads[Month],MONTH(TODAY()),tbl_Leads[Year],YEAR(TODAY()))'),
        ("Pending Triage", '=COUNTBLANK(tbl_Leads[TriageDecision])'),
        ("High/Crit Leads", '=COUNTIFS(tbl_Leads[Priority],"High")+COUNTIFS(tbl_Leads[Priority],"Critical")'),
        ("Converted to Case", '=COUNTIF(tbl_Leads[Outcome],"Converted to Case")'),
        ("Converted to Op", '=COUNTIF(tbl_Leads[Outcome],"Converted to Op")'),
        ("Conversion Rate", '=IFERROR((COUNTIF(tbl_Leads[Outcome],"Converted to Case")+COUNTIF(tbl_Leads[Outcome],"Converted to Op"))/(COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])),"—")'),
        ("Briefings Given", '=COUNTA(tbl_IntelBriefings[BriefingID])-COUNTBLANK(tbl_IntelBriefings[BriefingID])'),
        ("Follow-Ups Overdue", '=COUNTIFS(tbl_Leads[FollowUpDueDate],"<"&TODAY(),tbl_Leads[FollowUpRequired],"Yes")'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="BF8F00")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Transport Dashboard ──
    ws = wb.create_sheet("Dash_Transport")
    ws.sheet_properties.tabColor = "2E75B6"
    r = _dashboard_banner(ws, "TRANSPORT UNIT DASHBOARD", "Fleet Status, Utilization, Fuel & Maintenance")
    metrics = [
        ("Total Fleet", '=COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])'),
        ("Available", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Available")'),
        ("In Use", '=COUNTIF(tbl_Vehicles[CurrentStatus],"In Use")'),
        ("Under Maintenance", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Under Maintenance")'),
        ("Availability %", '=IFERROR(COUNTIF(tbl_Vehicles[CurrentStatus],"Available")/(COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])),"—")'),
        ("Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("Total Fuel (L)", '=SUM(tbl_FuelLog[Litres])'),
        ("Total Fuel Cost", '=SUM(tbl_FuelLog[Cost])'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="2E75B6")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Evidence Dashboard ──
    ws = wb.create_sheet("Dash_Evidence")
    ws.sheet_properties.tabColor = "538135"
    r = _dashboard_banner(ws, "EVIDENCE & FORENSICS DASHBOARD", "Exhibit Status, Forensic Turnaround & Court Readiness")
    metrics = [
        ("Total Exhibits", '=COUNTA(tbl_Exhibits[ExhibitID])-COUNTBLANK(tbl_Exhibits[ExhibitID])'),
        ("Pending", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Pending")'),
        ("Submitted", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Submitted")'),
        ("In Progress", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"In Progress")'),
        ("Completed", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Completed")'),
        ("Court Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Ready")'),
        ("Not Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Not Ready")'),
        ("Avg Turnaround", '=IFERROR(AVERAGE(tbl_Exhibits[TurnaroundDays]),"—")'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="538135")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Demand Reduction Dashboard ──
    ws = wb.create_sheet("Dash_DemandReduction")
    ws.sheet_properties.tabColor = "00B050"
    r = _dashboard_banner(ws, "DEMAND REDUCTION DASHBOARD", "Community Outreach Sessions, Attendance & Coverage")
    metrics = [
        ("Sessions This Month", '=COUNTIFS(tbl_DemandReduction[Month],MONTH(TODAY()),tbl_DemandReduction[Year],YEAR(TODAY()))'),
        ("Total Sessions", '=COUNTA(tbl_DemandReduction[LectureID])-COUNTBLANK(tbl_DemandReduction[LectureID])'),
        ("Total Attendance", '=SUM(tbl_DemandReduction[Attendance])'),
        ("Avg Attendance", '=IFERROR(AVERAGE(tbl_DemandReduction[Attendance]),"—")'),
        ("Total Hours", '=SUM(tbl_DemandReduction[Duration_Hrs])'),
        ("Follow-Up Planned", '=COUNTIF(tbl_DemandReduction[FollowUpPlanned],"Yes")'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="00B050")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Oversight Dashboard ──
    ws = wb.create_sheet("Dash_Oversight")
    ws.sheet_properties.tabColor = "FF0000"
    r = _dashboard_banner(ws, "OVERSIGHT & COMPLIANCE DASHBOARD", "Governance Alerts, Non-Conformities & KPI Health")
    metrics = [
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[DueDate],"<"&TODAY(),tbl_Reviews[Status],"<>Completed")'),
        ("Non-Compliant Cases", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Forensics", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("KPIs On Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"On Track")'),
        ("KPIs At Risk", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"At Risk")'),
        ("KPIs Off Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"Off Track")'),
        ("Service Overdue Vehicles", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
    ]
    for i, (label, _) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=label).font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="C00000")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        ws.cell(row=r, column=i+1, value=formula).font = Font(name="Calibri", bold=True, size=16, color="C00000")
        ws.cell(row=r, column=i+1).alignment = CENTER
        ws.cell(row=r, column=i+1).border = THIN_BORDER
        ws.cell(row=r, column=i+1).fill = PatternFill("solid", fgColor="FFF2CC")
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18




# ══════════════════════════════════════════════════════════════════════
#  MODULE 11 — PRINTABLE FORM TEMPLATES (CR1, CR2, CR5, Case Summary)
# ══════════════════════════════════════════════════════════════════════

def build_form_templates(wb):
    """Create printable/PDF-ready form template sheets."""

    # ── CR1 Form Template ──
    ws = wb.create_sheet("Form_CR1")
    ws.sheet_properties.tabColor = "808080"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="JAMAICA CONSTABULARY FORCE — CR1 FORM").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1, value="CRIME REPORT — INITIAL COMPLAINT").font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=2, column=1).alignment = CENTER

    fields = [
        ("Case Reference No.:", ""),("Date of Report:", ""),("Time of Report:", ""),
        ("Station:", ""),("Division:", "FNID Area 3"),("Reporting Officer:", ""),
        ("Rank:", ""),("Badge No.:", ""),
        ("",""),
        ("COMPLAINANT DETAILS",""),
        ("Full Name:", ""),("Address:", ""),("Phone:", ""),("ID Type & No.:", ""),
        ("",""),
        ("OFFENCE DETAILS",""),
        ("Offence Type:", ""),("Date of Offence:", ""),("Time of Offence:", ""),
        ("Location of Offence:", ""),("Parish:", ""),
        ("",""),
        ("NARRATIVE / PARTICULARS",""),
        ("",""),("",""),("",""),("",""),("",""),
        ("",""),
        ("SUSPECT INFORMATION",""),
        ("Suspect Name:", ""),("Alias:", ""),("Description:", ""),("Address:", ""),
        ("",""),
        ("WITNESS INFORMATION",""),
        ("Witness 1:", ""),("Witness 2:", ""),
        ("",""),
        ("OFFICER SIGNATURE:", ""),("SUPERVISOR SIGNATURE:", ""),("DATE:", ""),
    ]
    r = 4
    for label, val in fields:
        if label and label == label.upper() and not label.endswith(":"):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif label:
            ws.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=2, value=val).font = BODY_FONT
            ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 1

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── CR2 Form Template ──
    ws2 = wb.create_sheet("Form_CR2")
    ws2.sheet_properties.tabColor = "808080"
    ws2.merge_cells("A1:F1")
    ws2.cell(row=1, column=1, value="JAMAICA CONSTABULARY FORCE — CR2 FORM").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws2.cell(row=1, column=1).alignment = CENTER
    ws2.merge_cells("A2:F2")
    ws2.cell(row=2, column=1, value="CRIME REPORT — INVESTIGATION PROGRESS").font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws2.cell(row=2, column=1).alignment = CENTER

    cr2_fields = [
        ("Case Reference No.:", ""),("Linked CR1 No.:", ""),("Investigation Officer:", ""),
        ("Rank:", ""),("Badge No.:", ""),("Date:", ""),
        ("",""),
        ("INVESTIGATION PROGRESS",""),
        ("Actions Taken:", ""),("",""),("",""),
        ("Witnesses Interviewed:", ""),("",""),
        ("Evidence Collected:", ""),("",""),
        ("Suspect Status:", ""),("",""),
        ("FORENSIC SUBMISSIONS",""),
        ("Exhibits Submitted:", ""),("Lab/Agency:", ""),("Date Submitted:", ""),
        ("",""),
        ("RECOMMENDATIONS",""),
        ("",""),("",""),
        ("",""),
        ("SUPERVISOR REVIEW",""),
        ("Reviewed By:", ""),("Date:", ""),("Comments:", ""),("",""),
        ("",""),
        ("SIGNATURES",""),
        ("Investigating Officer:", ""),("Supervisor:", ""),("Date:", ""),
    ]
    r = 4
    for label, val in cr2_fields:
        if label and label == label.upper() and not label.endswith(":"):
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws2.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif label:
            ws2.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws2.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 1
    for c in range(1, 7):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # ── CR5 Form Template ──
    ws5 = wb.create_sheet("Form_CR5")
    ws5.sheet_properties.tabColor = "808080"
    ws5.merge_cells("A1:F1")
    ws5.cell(row=1, column=1, value="JAMAICA CONSTABULARY FORCE — CR5 FORM").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws5.cell(row=1, column=1).alignment = CENTER
    ws5.merge_cells("A2:F2")
    ws5.cell(row=2, column=1, value="CRIME REPORT — CASE CLOSURE / DISPOSAL").font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws5.cell(row=2, column=1).alignment = CENTER

    cr5_fields = [
        ("Case Reference No.:", ""),("Linked CR1 No.:", ""),("Linked CR2 No.:", ""),
        ("Offence:", ""),("Date of Offence:", ""),
        ("Investigating Officer:", ""),("Rank:", ""),
        ("",""),
        ("DISPOSAL DETAILS",""),
        ("Disposal Type:", ""),("Court:", ""),("Court Date:", ""),
        ("Verdict:", ""),("Sentence:", ""),
        ("",""),
        ("EVIDENCE DISPOSITION",""),
        ("Exhibits Returned:", ""),("Exhibits Destroyed:", ""),("Exhibits Retained:", ""),
        ("",""),
        ("FINAL REMARKS",""),
        ("",""),("",""),
        ("",""),
        ("APPROVALS",""),
        ("Investigating Officer:", ""),("Supervisor:", ""),
        ("Divisional Commander:", ""),("Date:", ""),
    ]
    r = 4
    for label, val in cr5_fields:
        if label and label == label.upper() and not label.endswith(":"):
            ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws5.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws5.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif label:
            ws5.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws5.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 1
    for c in range(1, 7):
        ws5.column_dimensions[get_column_letter(c)].width = 18

    # ── Case Summary Template ──
    wcs = wb.create_sheet("Form_CaseSummary")
    wcs.sheet_properties.tabColor = "808080"
    wcs.merge_cells("A1:F1")
    wcs.cell(row=1, column=1, value="FNID AREA 3 — CASE SUMMARY REPORT").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    wcs.cell(row=1, column=1).alignment = CENTER
    summary_fields = [
        ("Case ID:", ""),("CR1 Number:", ""),("CR2 Number:", ""),("CR5 Number:", ""),
        ("Offence:", ""),("Date:", ""),("Parish/Station:", ""),
        ("Investigator:", ""),("Status:", ""),("Priority:", ""),
        ("",""),
        ("CASE SYNOPSIS",""),
        ("",""),("",""),("",""),
        ("KEY EVIDENCE",""),
        ("",""),("",""),
        ("SUSPECTS",""),
        ("",""),("",""),
        ("CURRENT STATUS & NEXT STEPS",""),
        ("",""),("",""),
        ("SUPERVISORY NOTES",""),
        ("",""),
        ("Prepared By:", ""),("Date:", ""),
    ]
    r = 3
    for label, val in summary_fields:
        if label and label == label.upper() and not label.endswith(":"):
            wcs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            wcs.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            wcs.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif label:
            wcs.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
            wcs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            wcs.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 1
    for c in range(1, 7):
        wcs.column_dimensions[get_column_letter(c)].width = 18


# ══════════════════════════════════════════════════════════════════════
#  MODULE 12 — HOME / NAVIGATION SHEET
# ══════════════════════════════════════════════════════════════════════

def build_home_sheet(wb):
    ws = wb.active
    ws.title = "HOME"
    ws.sheet_properties.tabColor = "002060"

    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value="FNID AREA 3 — INTEGRATED OPERATIONAL WORKBOOK")
    cell.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 56

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1,
            value="Firearm and Narcotics Investigation Division | Area 3 | Jamaica Constabulary Force").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws.cell(row=2, column=1).alignment = CENTER

    ws.merge_cells("A3:H3")
    ws.cell(row=3, column=1,
            value="Operational Management System — Intelligence | Cases | Operations | Evidence | Transport | Demand Reduction | Performance").font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.cell(row=3, column=1).alignment = CENTER

    r = 5
    # Navigation matrix
    nav_sections = [
        ("UNIT PORTALS", "002060", [
            ("Portal_Registry", "Registry — Case Registration & CR Tracking"),
            ("Portal_Investigations", "Investigations — Case Progress & Reviews"),
            ("Portal_Operations", "Operations — Ops Activity & FCSI Tracking"),
            ("Portal_Intelligence", "Intelligence — Lead Management & Briefings"),
            ("Portal_Transport", "Transport — Fleet, Usage, Fuel & Maintenance"),
            ("Portal_DemandReduction", "Demand Reduction — Community Outreach"),
            ("Portal_AdminHRM", "Admin/HRM — Personnel & Administration"),
            ("Portal_Training", "Training — Capacity Building & Certifications"),
            ("Portal_Oversight", "Oversight — Compliance & Governance"),
        ]),
        ("DASHBOARDS", "C00000", [
            ("Dash_Executive", "Executive Dashboard — Full Operational Picture"),
            ("Dash_FCSI", "FCSI Dashboard — Focused Crime Suppression Initiative"),
            ("Statistics_Hub", "Central Statistics Hub — All Modules Aggregated"),
            ("Dash_Registry", "Registry Dashboard"),
            ("Dash_Investigations", "Investigations Dashboard"),
            ("Dash_Operations", "Operations Dashboard"),
            ("Dash_Intelligence", "Intelligence Dashboard"),
            ("Dash_Transport", "Transport Dashboard"),
            ("Dash_Evidence", "Evidence & Forensics Dashboard"),
            ("Dash_DemandReduction", "Demand Reduction Dashboard"),
            ("Dash_Oversight", "Oversight & Compliance Dashboard"),
        ]),
        ("DATA TABLES (Back-End)", "538135", [
            ("tbl_Cases", "Cases Register"),
            ("tbl_Actions", "Action Items"),
            ("tbl_Reviews", "Supervisory Reviews"),
            ("tbl_DiaryEntries", "Investigation Diary"),
            ("tbl_RadioMessages", "Radio Messages"),
            ("tbl_Persons", "Persons of Interest"),
            ("tbl_MCR_Raw", "Morning Crime Report Intake"),
            ("tbl_Leads", "Lead Register"),
            ("tbl_FollowUpRecs", "Follow-Up Recommendations"),
            ("tbl_IntelBriefings", "Intelligence Briefings"),
            ("tbl_Operations", "Operations Register"),
            ("tbl_Exhibits", "Evidence/Exhibits Pool"),
            ("tbl_Vehicles", "Vehicle Register"),
            ("tbl_VehicleUsage", "Vehicle Usage Log"),
            ("tbl_FuelLog", "Fuel Log"),
            ("tbl_Maintenance", "Maintenance Register"),
            ("tbl_Drivers", "Driver Registry"),
            ("tbl_DemandReduction", "Demand Reduction Lectures"),
            ("tbl_UnitPlans", "Unit Plans, Targets & Performance"),
        ]),
        ("REFERENCE & FORMS", "808080", [
            ("Lookups", "Dropdown Lists & Reference Data"),
            ("OffenceClassMap", "Offence Classification & FNID Routing Map"),
            ("ForensicRouting", "Evidence → Forensic Agency Routing"),
            ("EvidenceUseMap", "Evidence → Evidential Use Mapping"),
            ("Form_CR1", "CR1 Form Template (Printable)"),
            ("Form_CR2", "CR2 Form Template (Printable)"),
            ("Form_CR5", "CR5 Form Template (Printable)"),
            ("Form_CaseSummary", "Case Summary Report Template"),
        ]),
    ]

    for section_title, color, items in nav_sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

        for sheet_name, description in items:
            ws.cell(row=r, column=1, value=f"  >> {sheet_name}").font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
            ws.cell(row=r, column=1).alignment = Alignment(indent=2)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.cell(row=r, column=2, value=description).font = Font(name="Calibri", size=10, color="333333")
            ws.cell(row=r, column=2).alignment = LEFT_WRAP
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = Border(bottom=Side(style="hair", color="CCCCCC"))
            r += 1
        r += 1

    # Footer
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="FNID Area 3 Operational Workbook | Confidential — For Official Use Only | Generated " + datetime.now().strftime("%Y-%m-%d")).font = Font(name="Calibri", italic=True, size=9, color="999999")
    ws.cell(row=r, column=1).alignment = CENTER

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    ws.column_dimensions['A'].width = 28




# ══════════════════════════════════════════════════════════════════════
#  MAIN — ASSEMBLE THE WORKBOOK
# ══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("  FNID Area 3 — Integrated Operational Workbook Generator")
    print("=" * 70)

    wb = openpyxl.Workbook()

    # ── Step 1: HOME / Navigation ──
    print("[1/15] Building HOME navigation sheet...")
    build_home_sheet(wb)

    # ── Step 2: Lookups & Reference ──
    print("[2/15] Building Lookups & Reference Data...")
    build_lookups(wb)

    # ── Step 3: Offence Classification Map ──
    print("[3/15] Building Offence Classification Map...")
    build_offence_class_map(wb)

    # ── Step 4: Forensic Routing ──
    print("[4/15] Building Forensic Routing Table...")
    build_forensic_routing(wb)

    # ── Step 5: Evidence Use Map ──
    print("[5/15] Building Evidence → Evidential Use Map...")
    build_evidence_use_map(wb)

    # ── Step 6: Case Management Tables ──
    print("[6/15] Building Case Management tables (Cases, Actions, Reviews, Diary, Radio, Persons)...")
    build_cases_table(wb)
    build_actions_table(wb)
    build_reviews_table(wb)
    build_diary_table(wb)
    build_radio_messages_table(wb)
    build_persons_table(wb)

    # ── Step 7: MCR Intake ──
    print("[7/15] Building MCR Raw Intake table with auto-FNID filtering...")
    build_mcr_raw(wb)

    # ── Step 8: Intelligence Tables ──
    print("[8/15] Building Intelligence tables (Leads, Follow-Ups, Briefings)...")
    build_leads_table(wb)
    build_followup_recs(wb)
    build_intel_briefings(wb)

    # ── Step 9: Operations Register ──
    print("[9/15] Building Operations Register with FCSI tracking...")
    build_operations_table(wb)

    # ── Step 10: Evidence/Exhibits Pool ──
    print("[10/15] Building Exhibits Pool with forensic routing...")
    build_exhibits_table(wb)

    # ── Step 11: Transport Tables ──
    print("[11/15] Building Transport tables (Vehicles, Usage, Fuel, Maintenance, Drivers)...")
    build_vehicles_table(wb)
    build_vehicle_usage_table(wb)
    build_fuel_log(wb)
    build_maintenance_table(wb)
    build_drivers_table(wb)

    # ── Step 12: Demand Reduction ──
    print("[12/15] Building Demand Reduction Lecture register...")
    build_demand_reduction(wb)

    # ── Step 13: Unit Plans & Performance ──
    print("[13/15] Building Unit Plans with 32 KPIs and traffic-light tracking...")
    build_unit_plans(wb)

    # ── Step 14: Unit Portals ──
    print("[14/15] Building unit-specific Portal sheets (9 portals)...")
    build_portal_registry(wb)
    build_portal_investigations(wb)
    build_portal_operations(wb)
    build_portal_intelligence(wb)
    build_portal_transport(wb)
    build_portal_demand_reduction(wb)
    build_portal_admin_hrm(wb)
    build_portal_training(wb)
    build_portal_oversight(wb)

    # ── Step 15: Dashboards ──
    print("[15/15] Building Dashboards (Executive + FCSI + Statistics Hub + ALL unit dashboards)...")
    build_executive_dashboard(wb)
    build_statistics_hub(wb)
    build_fcsi_dashboard(wb)
    build_unit_dashboards(wb)

    # ── Step 16: Form Templates ──
    print("  [+] Building printable form templates (CR1, CR2, CR5, Case Summary)...")
    build_form_templates(wb)

    # ── Final: Set HOME as active sheet and save ──
    wb.active = wb.sheetnames.index("HOME")

    output_file = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"\nSaving workbook to: {output_file}")
    wb.save(output_file)
    print(f"\n{'=' * 70}")
    print(f"  WORKBOOK GENERATED SUCCESSFULLY")
    print(f"  File: {output_file}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Sheet list:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"    {i:2d}. {name}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
