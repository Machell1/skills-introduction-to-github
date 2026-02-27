#!/usr/bin/env python3
"""
Patch — Build all 8 empty unit dashboards with full metrics, formulas,
and monthly trend rows.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from generate_fnid_workbook import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BODY_FONT_BOLD,
    CENTER, LEFT_WRAP, THIN_BORDER,
    GREEN_FILL, AMBER_FILL, RED_FILL,
    NAVY, DARK_BLUE, LIGHT_BLUE, WHITE, MED_GRAY,
    SECTION_FONT, SUBHEADER_FONT, SUBHEADER_FILL,
)

def _dash_banner(ws, title, subtitle, color):
    ws.sheet_properties.tabColor = color
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value=f"FNID AREA 3 — {title}")
    c.font = Font(name="Calibri", bold=True, size=18, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42
    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    c2.alignment = CENTER
    ws.merge_cells("A3:L3")
    ws.cell(row=3, column=1, value="Auto-generated from operational data tables | Refresh: re-open workbook").font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws.cell(row=3, column=1).alignment = CENTER
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 16
    return 5

def _section_header(ws, r, title):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    return r + 1

def _metric_row(ws, r, metrics, label_color="1F3864"):
    for i, (label, _) in enumerate(metrics):
        cell = ws.cell(row=r, column=i+1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=label_color)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for i, (_, formula) in enumerate(metrics):
        cell = ws.cell(row=r, column=i+1, value=formula)
        cell.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return r + 2

def _trend_section(ws, r, metrics_templates):
    r = _section_header(ws, r, "MONTHLY TREND (Current Year)")
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    ws.cell(row=r, column=1, value="Metric").font = SUBHEADER_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for i, m in enumerate(months):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1
    for name, tmpl in metrics_templates:
        ws.cell(row=r, column=1, value=name).font = BODY_FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        for m in range(1, 13):
            cell = ws.cell(row=r, column=m+1, value="=" + tmpl.format(m=m))
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        r += 1
    return r + 1


# ═══════════════════════════════════════════
#  DASH_REGISTRY
# ═══════════════════════════════════════════
def build_dash_registry(wb):
    if "Dash_Registry" in wb.sheetnames:
        del wb["Dash_Registry"]
    ws = wb.create_sheet("Dash_Registry")
    r = _dash_banner(ws, "REGISTRY DASHBOARD", "Case Registration, Compliance & File Management Analytics", "1F3864")

    r = _section_header(ws, r, "CASE VOLUME")
    r = _metric_row(ws, r, [
        ("Total Cases", '=COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])'),
        ("Open", '=COUNTIF(tbl_Cases[CaseStatus],"Open")'),
        ("Active", '=COUNTIF(tbl_Cases[CaseStatus],"Active")'),
        ("Under Investigation", '=COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Pending Review", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Review")'),
        ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
        ("Closed", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*")'),
        ("Cold Case", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
    ])

    r = _section_header(ws, r, "CR COMPLIANCE")
    r = _metric_row(ws, r, [
        ("CR1 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR1_Compliant],"Yes")/COUNTA(tbl_Cases[CR1_Compliant]),"—")'),
        ("CR2 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR2_Compliant],"Yes")/COUNTA(tbl_Cases[CR2_Compliant]),"—")'),
        ("CR5 Compliant %", '=IFERROR(COUNTIF(tbl_Cases[CR5_Compliant],"Yes")/COUNTA(tbl_Cases[CR5_Compliant]),"—")'),
        ("All CRs Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Compliant")'),
        ("Non-Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Compliance Rate", '=IFERROR(COUNTIF(tbl_Cases[ComplianceFlag],"Compliant")/COUNTA(tbl_Cases[ComplianceFlag]),"—")'),
    ])

    r = _section_header(ws, r, "REGISTRATION SLA & FILE QUALITY")
    r = _metric_row(ws, r, [
        ("Avg SLA (hrs)", '=IFERROR(AVERAGE(tbl_Cases[RegistrationSLA_Hrs]),"—")'),
        ("Registered <24hrs", '=COUNTIFS(tbl_Cases[RegistrationSLA_Hrs],"<="&24)'),
        ("Registered >24hrs", '=COUNTIFS(tbl_Cases[RegistrationSLA_Hrs],">"&24)'),
        ("Avg File Score", '=IFERROR(AVERAGE(tbl_Cases[FileCompletenessScore]),"—")'),
        ("Files Complete", '=COUNTIFS(tbl_Cases[FileCompletenessScore],">="&0.95)'),
        ("Files Incomplete", '=COUNTIFS(tbl_Cases[FileCompletenessScore],"<"&0.5)'),
    ])

    r = _section_header(ws, r, "CASE AGING")
    r = _metric_row(ws, r, [
        ("0-30 days", '=COUNTIF(tbl_Cases[AgingBand],"0-30 days")'),
        ("31-60 days", '=COUNTIF(tbl_Cases[AgingBand],"31-60 days")'),
        ("61-90 days", '=COUNTIF(tbl_Cases[AgingBand],"61-90 days")'),
        ("91-180 days", '=COUNTIF(tbl_Cases[AgingBand],"91-180 days")'),
        ("180+ days", '=COUNTIF(tbl_Cases[AgingBand],"180+ days")'),
        ("Avg Days Open", '=IFERROR(AVERAGE(tbl_Cases[DaysOpen]),"—")'),
        ("Max Days Open", '=IFERROR(MAX(tbl_Cases[DaysOpen]),"—")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
    ])

    r = _section_header(ws, r, "DISPOSAL & ARCHIVE")
    r = _metric_row(ws, r, [
        ("Pending Disposal", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[DisposalType],"")'),
        ("Disposed", '=COUNTIFS(tbl_Cases[DisposalType],"<>")'),
        ("Archived", '=COUNTIFS(tbl_Cases[ArchiveRef],"<>")'),
        ("Awaiting Archive", '=COUNTIFS(tbl_Cases[DisposalType],"<>",tbl_Cases[ArchiveRef],"")'),
    ])

    r = _section_header(ws, r, "OFFENCE BREAKDOWN")
    r = _metric_row(ws, r, [
        ("Firearms Cases", '=COUNTIF(tbl_Cases[OffenceType],"*Firearm*")'),
        ("Narcotics Cases", '=COUNTIF(tbl_Cases[OffenceType],"*Narcotics*")+COUNTIF(tbl_Cases[OffenceType],"*Cannabis*")+COUNTIF(tbl_Cases[OffenceType],"*Cocaine*")+COUNTIF(tbl_Cases[OffenceType],"*Drug*")'),
        ("Shooting Cases", '=COUNTIF(tbl_Cases[OffenceType],"*Shooting*")'),
        ("Murder Cases", '=COUNTIF(tbl_Cases[OffenceType],"*Murder*")'),
        ("Conspiracy", '=COUNTIF(tbl_Cases[OffenceType],"*Conspiracy*")'),
        ("Other", '=COUNTIF(tbl_Cases[OffenceType],"Other*")'),
    ])

    r = _trend_section(ws, r, [
        ("Cases Registered", 'COUNTIFS(tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
        ("Cases Closed", 'COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
        ("Non-Compliant", 'COUNTIFS(tbl_Cases[ComplianceFlag],"Non-Compliant",tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_INVESTIGATIONS
# ═══════════════════════════════════════════
def build_dash_investigations(wb):
    if "Dash_Investigations" in wb.sheetnames:
        del wb["Dash_Investigations"]
    ws = wb.create_sheet("Dash_Investigations")
    r = _dash_banner(ws, "INVESTIGATIONS DASHBOARD", "Case Progress, Witness Management, Reviews, Court Readiness & SOP Compliance", "1F3864")

    r = _section_header(ws, r, "CASE STATUS")
    r = _metric_row(ws, r, [
        ("Under Investigation", '=COUNTIF(tbl_Cases[CaseStatus],"Under Investigation")'),
        ("Pending Court", '=COUNTIF(tbl_Cases[CaseStatus],"Pending Court")'),
        ("Cold Cases", '=COUNTIF(tbl_Cases[CaseStatus],"Cold Case")'),
        ("Clearance Rate", '=IFERROR(COUNTIFS(tbl_Cases[CaseStatus],"Closed*")/(COUNTA(tbl_Cases[CaseID])-COUNTBLANK(tbl_Cases[CaseID])),"—")'),
        ("Avg Days Open", '=IFERROR(AVERAGE(tbl_Cases[DaysOpen]),"—")'),
        ("Cases >90d", '=COUNTIFS(tbl_Cases[DaysOpen],">"&90)'),
    ])

    r = _section_header(ws, r, "WITNESS & STATEMENT MANAGEMENT")
    r = _metric_row(ws, r, [
        ("Total Witnesses", '=COUNTA(tbl_Witnesses[WitnessID])-COUNTBLANK(tbl_Witnesses[WitnessID])'),
        ("Statements Taken", '=COUNTA(tbl_Statements[StatementID])-COUNTBLANK(tbl_Statements[StatementID])'),
        ("Cautioned Stmts", '=COUNTIF(tbl_Statements[CautionGiven],"Yes")'),
        ("Willing to Testify", '=COUNTIF(tbl_Witnesses[WillingToTestify],"Yes")'),
        ("Summons Issued", '=COUNTIF(tbl_Witnesses[SummonsIssued],"Yes")'),
        ("Summons Unserved", '=COUNTIFS(tbl_Witnesses[SummonsIssued],"Yes",tbl_Witnesses[SummonsServed],"No")'),
        ("Protection Req", '=COUNTIF(tbl_Witnesses[ProtectionRequired],"Yes")'),
        ("Vulnerable", '=COUNTIF(tbl_Witnesses[VulnerableWitness],"Yes")'),
    ])

    r = _section_header(ws, r, "SUPERVISORY REVIEWS & DIRECTIVES")
    r = _metric_row(ws, r, [
        ("Reviews Done", '=COUNTA(tbl_Reviews[ReviewID])-COUNTBLANK(tbl_Reviews[ReviewID])'),
        ("Reviews Overdue", '=COUNTIFS(tbl_Reviews[NextReviewDue],"<"&TODAY())'),
        ("Directives Issued", '=COUNTIFS(tbl_Reviews[DirectivesIssued],"<>")'),
        ("Directives Overdue", '=COUNTIFS(tbl_Reviews[DirectiveDueDate],"<"&TODAY(),tbl_Reviews[DirectiveStatus],"<>Completed")'),
        ("Escalations", '=COUNTIF(tbl_Reviews[EscalationRequired],"Yes")'),
        ("Conferences Held", '=COUNTA(tbl_CommandConference[ConferenceID])-COUNTBLANK(tbl_CommandConference[ConferenceID])'),
    ])

    r = _section_header(ws, r, "INVESTIGATION PLANS & ASSIGNMENTS")
    r = _metric_row(ws, r, [
        ("Active Plans", '=COUNTIF(tbl_InvestigationPlans[PlanStatus],"Active")'),
        ("Plan Reviews Due", '=COUNTIFS(tbl_InvestigationPlans[NextReviewDate],"<"&TODAY())'),
        ("Assignments", '=COUNTA(tbl_CaseAssignments[AssignmentID])-COUNTBLANK(tbl_CaseAssignments[AssignmentID])'),
        ("Reassignments", '=COUNTIF(tbl_CaseAssignments[AssignmentType],"Reassignment")'),
        ("Pending Ack", '=COUNTIFS(tbl_CaseAssignments[AcknowledgedByReceiver],"No")'),
        ("Diary Entries", '=COUNTA(tbl_DiaryEntries[DiaryID])-COUNTBLANK(tbl_DiaryEntries[DiaryID])'),
    ])

    r = _section_header(ws, r, "VICTIM MANAGEMENT")
    r = _metric_row(ws, r, [
        ("Victim Contacts", '=COUNTA(tbl_VictimContact[ContactID])-COUNTBLANK(tbl_VictimContact[ContactID])'),
        ("Satisfied", '=COUNTIF(tbl_VictimContact[VictimSatisfied],"Yes")'),
        ("Concerns Raised", '=COUNTIFS(tbl_VictimContact[ConcernsRaised],"<>")'),
        ("Referrals", '=COUNTIF(tbl_VictimContact[ReferralMade],"Yes")'),
        ("Safety Plans", '=COUNTIF(tbl_VictimContact[SafetyPlanInPlace],"Yes")'),
        ("Contact Overdue", '=COUNTIFS(tbl_VictimContact[NextScheduledContact],"<"&TODAY())'),
    ])

    r = _section_header(ws, r, "COURT & DISCLOSURE")
    r = _metric_row(ws, r, [
        ("Upcoming Ct (7d)", '=COUNTIFS(tbl_CourtDates[HearingDate],">="&TODAY(),tbl_CourtDates[HearingDate],"<="&TODAY()+7)'),
        ("Total Hearings", '=COUNTA(tbl_CourtDates[CourtID])-COUNTBLANK(tbl_CourtDates[CourtID])'),
        ("48hr Met", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"Yes")'),
        ("48hr Missed", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"No")'),
        ("Disclosures", '=COUNTA(tbl_Disclosure[DisclosureID])-COUNTBLANK(tbl_Disclosure[DisclosureID])'),
        ("Redactions Req", '=COUNTIF(tbl_Disclosure[RedactionRequired],"Yes")'),
    ])

    r = _trend_section(ws, r, [
        ("Cases Closed", 'COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[Month],{m},tbl_Cases[Year],YEAR(TODAY()))'),
        ("Statements Taken", 'COUNTIFS(tbl_Statements[StatementDate],">="&DATE(YEAR(TODAY()),{m},1),tbl_Statements[StatementDate],"<"&DATE(YEAR(TODAY()),{m}+1,1))'),
        ("Reviews Completed", 'COUNTIFS(tbl_Reviews[ReviewDate],">="&DATE(YEAR(TODAY()),{m},1),tbl_Reviews[ReviewDate],"<"&DATE(YEAR(TODAY()),{m}+1,1))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_OPERATIONS
# ═══════════════════════════════════════════
def build_dash_operations(wb):
    if "Dash_Operations" in wb.sheetnames:
        del wb["Dash_Operations"]
    ws = wb.create_sheet("Dash_Operations")
    r = _dash_banner(ws, "OPERATIONS DASHBOARD", "Operational Activity, Seizures, Arrests, FCSI & Warrant Tracking", "C00000")

    r = _section_header(ws, r, "OPERATIONAL VOLUME")
    r = _metric_row(ws, r, [
        ("Total Ops", '=COUNTA(tbl_Operations[OPS_ID])-COUNTBLANK(tbl_Operations[OPS_ID])'),
        ("FCSI Ops", '=COUNTIF(tbl_Operations[FCSI_Flag],"Yes")'),
        ("Non-FCSI", '=COUNTIF(tbl_Operations[FCSI_Flag],"No")'),
        ("Search Warrants", '=COUNTIF(tbl_Operations[OpsType],"Search Warrant")'),
        ("Raids", '=COUNTIF(tbl_Operations[OpsType],"Raid")'),
        ("Checkpoints", '=COUNTIF(tbl_Operations[OpsType],"Checkpoint")'),
        ("Surveillance", '=COUNTIF(tbl_Operations[OpsType],"Surveillance")'),
        ("Joint Ops", '=COUNTIF(tbl_Operations[OpsType],"Joint Operation")'),
    ], label_color="C00000")

    r = _section_header(ws, r, "SEIZURES & ARRESTS (FROM OPS TABLE)")
    r = _metric_row(ws, r, [
        ("Total Arrests", '=SUM(tbl_Operations[Arrests])'),
        ("Firearms Seized", '=SUM(tbl_Operations[FirearmsSeized])'),
        ("Ammo Seized", '=SUM(tbl_Operations[AmmunitionSeized])'),
        ("Narcotics (kg)", '=SUM(tbl_Operations[NarcoticsSeized_kg])'),
        ("Cash Seized", '=SUM(tbl_Operations[CashSeized])'),
        ("Searches Done", '=SUM(tbl_Operations[Searches])'),
    ], label_color="C00000")

    r = _section_header(ws, r, "DETAILED SEIZURES & ARRESTS")
    r = _metric_row(ws, r, [
        ("Arrest Records", '=COUNTA(tbl_Arrests[ArrestID])-COUNTBLANK(tbl_Arrests[ArrestID])'),
        ("Bail Granted", '=COUNTIF(tbl_Arrests[BailGranted],"Yes")'),
        ("Remanded", '=COUNTIFS(tbl_Arrests[RemandFacility],"<>")'),
        ("Seized Items", '=COUNTA(tbl_Seizures[SeizureID])-COUNTBLANK(tbl_Seizures[SeizureID])'),
        ("Firearms (Detail)", '=COUNTIF(tbl_Seizures[ItemType],"Firearm")'),
        ("Narcotics Items", '=COUNTIFS(tbl_Seizures[ItemType],"Narcotics*")'),
    ], label_color="C00000")

    r = _section_header(ws, r, "WARRANTS")
    r = _metric_row(ws, r, [
        ("Warrants Issued", '=COUNTIF(tbl_Warrants[Status],"Issued")'),
        ("Executed", '=COUNTIF(tbl_Warrants[Status],"Executed")'),
        ("Expired", '=COUNTIF(tbl_Warrants[Status],"Expired")'),
        ("Active (Valid)", '=COUNTIFS(tbl_Warrants[ValidUntil],">="&TODAY(),tbl_Warrants[ExecutionDate],"")'),
        ("Returned to JP", '=COUNTIFS(tbl_Warrants[ReturnToJP_Date],"<>")'),
        ("Total Warrants", '=COUNTA(tbl_Warrants[WarrantID])-COUNTBLANK(tbl_Warrants[WarrantID])'),
    ], label_color="C00000")

    r = _trend_section(ws, r, [
        ("Operations", 'COUNTIFS(tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("FCSI Ops", 'COUNTIFS(tbl_Operations[FCSI_Flag],"Yes",tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Firearms Seized", 'SUMIFS(tbl_Operations[FirearmsSeized],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Arrests", 'SUMIFS(tbl_Operations[Arrests],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
        ("Narcotics (kg)", 'SUMIFS(tbl_Operations[NarcoticsSeized_kg],tbl_Operations[Month],{m},tbl_Operations[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_INTELLIGENCE
# ═══════════════════════════════════════════
def build_dash_intelligence(wb):
    if "Dash_Intelligence" in wb.sheetnames:
        del wb["Dash_Intelligence"]
    ws = wb.create_sheet("Dash_Intelligence")
    r = _dash_banner(ws, "INTELLIGENCE DASHBOARD", "Lead Management, Triage, Informant Metrics & Briefings", "BF8F00")

    r = _section_header(ws, r, "LEAD VOLUME & TRIAGE")
    r = _metric_row(ws, r, [
        ("Total Leads", '=COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])'),
        ("From MCR", '=COUNTIF(tbl_Leads[Source],"MCR")'),
        ("From Ops", '=COUNTIF(tbl_Leads[Source],"Operations")'),
        ("From Tips", '=COUNTIF(tbl_Leads[Source],"Tip-off")'),
        ("From Informants", '=COUNTIF(tbl_Leads[Source],"Informant")'),
        ("Pending Triage", '=COUNTBLANK(tbl_Leads[TriageDecision])'),
        ("High/Critical", '=COUNTIFS(tbl_Leads[Priority],"High")+COUNTIFS(tbl_Leads[Priority],"Critical")'),
        ("Overdue Follow-Up", '=COUNTIFS(tbl_Leads[FollowUpDueDate],"<"&TODAY(),tbl_Leads[FollowUpRequired],"Yes")'),
    ], label_color="BF8F00")

    r = _section_header(ws, r, "LEAD OUTCOMES")
    r = _metric_row(ws, r, [
        ("Converted to Case", '=COUNTIF(tbl_Leads[Outcome],"Converted to Case")'),
        ("Converted to Op", '=COUNTIF(tbl_Leads[Outcome],"Converted to Op")'),
        ("Intel Filed", '=COUNTIF(tbl_Leads[Outcome],"Intel Filed")'),
        ("Closed - No Action", '=COUNTIF(tbl_Leads[Outcome],"Closed - No Action")'),
        ("Pending", '=COUNTIF(tbl_Leads[Outcome],"Pending")'),
        ("Conversion Rate", '=IFERROR((COUNTIF(tbl_Leads[Outcome],"Converted to Case")+COUNTIF(tbl_Leads[Outcome],"Converted to Op"))/(COUNTA(tbl_Leads[LeadID])-COUNTBLANK(tbl_Leads[LeadID])),"—")'),
    ], label_color="BF8F00")

    r = _section_header(ws, r, "INFORMANT MANAGEMENT")
    r = _metric_row(ws, r, [
        ("Registered", '=COUNTA(tbl_Informants[InformantID])-COUNTBLANK(tbl_Informants[InformantID])'),
        ("Active", '=COUNTIF(tbl_Informants[Status],"Active")'),
        ("Total Tips", '=SUM(tbl_Informants[TotalTipsProvided])'),
        ("Tips to Cases", '=SUM(tbl_Informants[TipsConvertedToCase])'),
        ("Tips to Ops", '=SUM(tbl_Informants[TipsConvertedToOp])'),
        ("Contact Overdue", '=COUNTIFS(tbl_Informants[NextContactDue],"<"&TODAY())'),
    ], label_color="BF8F00")

    r = _section_header(ws, r, "INTELLIGENCE BRIEFINGS & MCR")
    r = _metric_row(ws, r, [
        ("Briefings Given", '=COUNTA(tbl_IntelBriefings[BriefingID])-COUNTBLANK(tbl_IntelBriefings[BriefingID])'),
        ("MCR Reports", '=COUNTA(tbl_MCR_Raw[MCR_ID])-COUNTBLANK(tbl_MCR_Raw[MCR_ID])'),
        ("MCR FNID-Relevant", '=COUNTIF(tbl_MCR_Raw[FNID_Relevant],"Yes")'),
        ("MCR Auto-Routed", '=COUNTIF(tbl_MCR_Raw[AutoRouted],"Yes")'),
        ("Follow-Up Recs", '=COUNTA(tbl_FollowUpRecs[RecID])-COUNTBLANK(tbl_FollowUpRecs[RecID])'),
        ("Avg Risk Score", '=IFERROR(AVERAGE(tbl_Leads[RiskScore]),"—")'),
    ], label_color="BF8F00")

    r = _trend_section(ws, r, [
        ("Leads Received", 'COUNTIFS(tbl_Leads[Month],{m},tbl_Leads[Year],YEAR(TODAY()))'),
        ("Leads Converted", 'COUNTIFS(tbl_Leads[Outcome],"Converted*",tbl_Leads[Month],{m},tbl_Leads[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_TRANSPORT
# ═══════════════════════════════════════════
def build_dash_transport(wb):
    if "Dash_Transport" in wb.sheetnames:
        del wb["Dash_Transport"]
    ws = wb.create_sheet("Dash_Transport")
    r = _dash_banner(ws, "TRANSPORT DASHBOARD", "Fleet Status, Usage, Fuel, Maintenance & Driver Management", "2E75B6")

    r = _section_header(ws, r, "FLEET STATUS")
    r = _metric_row(ws, r, [
        ("Total Vehicles", '=COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])'),
        ("Available", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Available")'),
        ("In Use", '=COUNTIF(tbl_Vehicles[CurrentStatus],"In Use")'),
        ("Maintenance", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Under Maintenance")'),
        ("Out of Service", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Out of Service")'),
        ("Reserved", '=COUNTIF(tbl_Vehicles[CurrentStatus],"Reserved")'),
        ("Service Overdue", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("Availability %", '=IFERROR(COUNTIF(tbl_Vehicles[CurrentStatus],"Available")/(COUNTA(tbl_Vehicles[VehicleID])-COUNTBLANK(tbl_Vehicles[VehicleID])),"—")'),
    ], label_color="2E75B6")

    r = _section_header(ws, r, "USAGE & FUEL")
    r = _metric_row(ws, r, [
        ("Total Trips", '=COUNTA(tbl_VehicleUsage[UsageID])-COUNTBLANK(tbl_VehicleUsage[UsageID])'),
        ("Total Distance (km)", '=SUM(tbl_VehicleUsage[DistanceKm])'),
        ("Avg Distance/Trip", '=IFERROR(AVERAGE(tbl_VehicleUsage[DistanceKm]),"—")'),
        ("Total Fuel (L)", '=SUM(tbl_FuelLog[Litres])'),
        ("Total Fuel Cost", '=SUM(tbl_FuelLog[Cost])'),
        ("Avg L/Fill", '=IFERROR(AVERAGE(tbl_FuelLog[Litres]),"—")'),
    ], label_color="2E75B6")

    r = _section_header(ws, r, "MAINTENANCE & DRIVERS")
    r = _metric_row(ws, r, [
        ("Maint Records", '=COUNTA(tbl_Maintenance[MaintID])-COUNTBLANK(tbl_Maintenance[MaintID])'),
        ("Total Maint Cost", '=SUM(tbl_Maintenance[Cost])'),
        ("Avg Downtime (d)", '=IFERROR(AVERAGE(tbl_Maintenance[DowntimeDays]),"—")'),
        ("Active Drivers", '=COUNTIF(tbl_Drivers[Status],"Active")'),
        ("License Expired", '=COUNTIFS(tbl_Drivers[LicenseExpiry],"<"&TODAY())'),
        ("Insurance Expired", '=COUNTIFS(tbl_Vehicles[InsuranceExpiry],"<"&TODAY())'),
    ], label_color="2E75B6")

    r = _trend_section(ws, r, [
        ("Trips", 'COUNTIFS(tbl_VehicleUsage[Month],{m},tbl_VehicleUsage[Year],YEAR(TODAY()))'),
        ("Fuel (L)", 'SUMIFS(tbl_FuelLog[Litres],tbl_FuelLog[Month],{m},tbl_FuelLog[Year],YEAR(TODAY()))'),
        ("Fuel Cost", 'SUMIFS(tbl_FuelLog[Cost],tbl_FuelLog[Month],{m},tbl_FuelLog[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_EVIDENCE
# ═══════════════════════════════════════════
def build_dash_evidence(wb):
    if "Dash_Evidence" in wb.sheetnames:
        del wb["Dash_Evidence"]
    ws = wb.create_sheet("Dash_Evidence")
    r = _dash_banner(ws, "EVIDENCE & FORENSICS DASHBOARD", "Exhibits, Forensic Routing, Chain of Custody & Court Readiness", "538135")

    r = _section_header(ws, r, "EXHIBIT VOLUME")
    r = _metric_row(ws, r, [
        ("Total Exhibits", '=COUNTA(tbl_Exhibits[ExhibitID])-COUNTBLANK(tbl_Exhibits[ExhibitID])'),
        ("Firearms", '=COUNTIF(tbl_Exhibits[EvidenceType],"Firearm")'),
        ("Ammunition", '=COUNTIF(tbl_Exhibits[EvidenceType],"Ammunition")'),
        ("Cannabis", '=COUNTIF(tbl_Exhibits[EvidenceType],"Narcotics - Cannabis")'),
        ("Cocaine", '=COUNTIF(tbl_Exhibits[EvidenceType],"Narcotics - Cocaine")'),
        ("Electronic Dev", '=COUNTIF(tbl_Exhibits[EvidenceType],"Electronic Device")'),
        ("Cash/Currency", '=COUNTIF(tbl_Exhibits[EvidenceType],"Cash/Currency")'),
        ("Other", '=COUNTIF(tbl_Exhibits[EvidenceType],"Other*")'),
    ], label_color="538135")

    r = _section_header(ws, r, "FORENSIC STATUS")
    r = _metric_row(ws, r, [
        ("Pending Submission", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Pending")'),
        ("Submitted", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Submitted")'),
        ("In Progress", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"In Progress")'),
        ("Completed", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Completed")'),
        ("Inconclusive", '=COUNTIF(tbl_Exhibits[AnalysisStatus],"Inconclusive")'),
        ("Avg Turnaround", '=IFERROR(AVERAGE(tbl_Exhibits[TurnaroundDays]),"—")'),
        ("Overdue Returns", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Max Turnaround", '=IFERROR(MAX(tbl_Exhibits[TurnaroundDays]),"—")'),
    ], label_color="538135")

    r = _section_header(ws, r, "COURT READINESS & CHAIN OF CUSTODY")
    r = _metric_row(ws, r, [
        ("Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Ready")'),
        ("Not Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Not Ready")'),
        ("Partial", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Partial")'),
        ("Custody Transfers", '=COUNTA(tbl_CustodyTransfers[TransferID])-COUNTBLANK(tbl_CustodyTransfers[TransferID])'),
        ("Scenes Processed", '=COUNTA(tbl_SceneLog[SceneID])-COUNTBLANK(tbl_SceneLog[SceneID])'),
        ("GSR Collected", '=COUNTIF(tbl_SceneLog[GSR_Collected],"Yes")'),
    ], label_color="538135")

    r = _trend_section(ws, r, [
        ("Exhibits Collected", 'COUNTIFS(tbl_Exhibits[Month],{m},tbl_Exhibits[Year],YEAR(TODAY()))'),
        ("Forensic Completed", 'COUNTIFS(tbl_Exhibits[AnalysisStatus],"Completed",tbl_Exhibits[Month],{m},tbl_Exhibits[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_DEMAND_REDUCTION
# ═══════════════════════════════════════════
def build_dash_demand_reduction(wb):
    if "Dash_DemandReduction" in wb.sheetnames:
        del wb["Dash_DemandReduction"]
    ws = wb.create_sheet("Dash_DemandReduction")
    r = _dash_banner(ws, "DEMAND REDUCTION DASHBOARD", "Community Outreach Lectures, Attendance & Coverage", "00B050")

    r = _section_header(ws, r, "SESSION VOLUME & ATTENDANCE")
    r = _metric_row(ws, r, [
        ("Sessions Held", '=COUNTA(tbl_DemandReduction[LectureID])-COUNTBLANK(tbl_DemandReduction[LectureID])'),
        ("Total Attendance", '=SUM(tbl_DemandReduction[Attendance])'),
        ("Avg Attendance", '=IFERROR(AVERAGE(tbl_DemandReduction[Attendance]),"—")'),
        ("Max Attendance", '=IFERROR(MAX(tbl_DemandReduction[Attendance]),"—")'),
        ("Total Hours", '=SUM(tbl_DemandReduction[Duration_Hrs])'),
        ("Avg Duration (hrs)", '=IFERROR(AVERAGE(tbl_DemandReduction[Duration_Hrs]),"—")'),
        ("With Follow-Up", '=COUNTIF(tbl_DemandReduction[FollowUpPlanned],"Yes")'),
        ("Parishes Covered", '=IFERROR(SUMPRODUCT(1/COUNTIF(tbl_DemandReduction[Parish],tbl_DemandReduction[Parish])),"—")'),
    ], label_color="00B050")

    r = _trend_section(ws, r, [
        ("Sessions", 'COUNTIFS(tbl_DemandReduction[Month],{m},tbl_DemandReduction[Year],YEAR(TODAY()))'),
        ("Attendance", 'SUMIFS(tbl_DemandReduction[Attendance],tbl_DemandReduction[Month],{m},tbl_DemandReduction[Year],YEAR(TODAY()))'),
        ("Hours", 'SUMIFS(tbl_DemandReduction[Duration_Hrs],tbl_DemandReduction[Month],{m},tbl_DemandReduction[Year],YEAR(TODAY()))'),
    ])
    return ws


# ═══════════════════════════════════════════
#  DASH_OVERSIGHT
# ═══════════════════════════════════════════
def build_dash_oversight(wb):
    if "Dash_Oversight" in wb.sheetnames:
        del wb["Dash_Oversight"]
    ws = wb.create_sheet("Dash_Oversight")
    r = _dash_banner(ws, "OVERSIGHT & COMPLIANCE DASHBOARD", "Governance, SOP Compliance, Audit Metrics & Risk Monitoring", "002060")

    r = _section_header(ws, r, "OVERALL COMPLIANCE STATUS")
    r = _metric_row(ws, r, [
        ("KPIs On Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"On Track")'),
        ("KPIs At Risk", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"At Risk")'),
        ("KPIs Off Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"Off Track")'),
        ("CR Compliance %", '=IFERROR(COUNTIF(tbl_Cases[ComplianceFlag],"Compliant")/COUNTA(tbl_Cases[ComplianceFlag]),"—")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[NextReviewDue],"<"&TODAY())'),
    ], label_color="002060")

    r = _section_header(ws, r, "SOP COMPLIANCE — CASE MANAGEMENT")
    r = _metric_row(ws, r, [
        ("Files Complete (>95%)", '=COUNTIFS(tbl_Cases[FileCompletenessScore],">="&0.95)'),
        ("Files Incomplete (<50%)", '=COUNTIFS(tbl_Cases[FileCompletenessScore],"<"&0.5)'),
        ("Avg File Score", '=IFERROR(AVERAGE(tbl_Cases[FileCompletenessScore]),"—")'),
        ("48hr Charge Met", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"Yes")'),
        ("48hr Missed", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"No")'),
        ("Pending Disposal", '=COUNTIFS(tbl_Cases[CaseStatus],"Closed*",tbl_Cases[DisposalType],"")'),
    ], label_color="002060")

    r = _section_header(ws, r, "DIRECTIVE & ESCALATION TRACKING")
    r = _metric_row(ws, r, [
        ("Review Directives", '=COUNTIFS(tbl_Reviews[DirectivesIssued],"<>")'),
        ("Dir. Completed", '=COUNTIF(tbl_Reviews[DirectiveStatus],"Completed")'),
        ("Dir. Overdue", '=COUNTIFS(tbl_Reviews[DirectiveDueDate],"<"&TODAY(),tbl_Reviews[DirectiveStatus],"<>Completed")'),
        ("Conf. Directives", '=COUNTIFS(tbl_CommandConference[DirectivesIssued],"<>")'),
        ("Conf. Dir. Overdue", '=COUNTIFS(tbl_CommandConference[DirectiveDueDate],"<"&TODAY(),tbl_CommandConference[DirectiveStatus],"<>Completed")'),
        ("Total Escalations", '=COUNTIF(tbl_Reviews[EscalationRequired],"Yes")'),
    ], label_color="002060")

    r = _section_header(ws, r, "EVIDENCE & FORENSIC INTEGRITY")
    r = _metric_row(ws, r, [
        ("Exhibits Not Ready", '=COUNTIF(tbl_Exhibits[CourtReadiness],"Not Ready")'),
        ("Forensic Overdue", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Custody Transfers", '=COUNTA(tbl_CustodyTransfers[TransferID])-COUNTBLANK(tbl_CustodyTransfers[TransferID])'),
        ("Warrants Expired", '=COUNTIF(tbl_Warrants[Status],"Expired")'),
        ("Victim Contact Overdue", '=COUNTIFS(tbl_VictimContact[NextScheduledContact],"<"&TODAY())'),
        ("Summons Unserved", '=COUNTIFS(tbl_Witnesses[SummonsIssued],"Yes",tbl_Witnesses[SummonsServed],"No")'),
    ], label_color="002060")

    r = _section_header(ws, r, "FLEET & PERSONNEL COMPLIANCE")
    r = _metric_row(ws, r, [
        ("Vehicle Service Due", '=COUNTIFS(tbl_Vehicles[NextServiceDue],"<"&TODAY())'),
        ("Insurance Expired", '=COUNTIFS(tbl_Vehicles[InsuranceExpiry],"<"&TODAY())'),
        ("License Expired", '=COUNTIFS(tbl_Drivers[LicenseExpiry],"<"&TODAY())'),
        ("Training Overdue", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
        ("Informant Contact Due", '=COUNTIFS(tbl_Informants[NextContactDue],"<"&TODAY())'),
        ("Plan Reviews Due", '=COUNTIFS(tbl_InvestigationPlans[NextReviewDate],"<"&TODAY())'),
    ], label_color="002060")
    return ws


def main():
    fname = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"Loading {fname}...")
    wb = openpyxl.load_workbook(fname)

    print("[1/8] Building Dash_Registry...")
    build_dash_registry(wb)
    print("[2/8] Building Dash_Investigations...")
    build_dash_investigations(wb)
    print("[3/8] Building Dash_Operations...")
    build_dash_operations(wb)
    print("[4/8] Building Dash_Intelligence...")
    build_dash_intelligence(wb)
    print("[5/8] Building Dash_Transport...")
    build_dash_transport(wb)
    print("[6/8] Building Dash_Evidence...")
    build_dash_evidence(wb)
    print("[7/8] Building Dash_DemandReduction...")
    build_dash_demand_reduction(wb)
    print("[8/8] Building Dash_Oversight...")
    build_dash_oversight(wb)

    wb.save(fname)
    print(f"All 8 dashboards built: {fname} — {len(wb.sheetnames)} sheets")

if __name__ == "__main__":
    main()
