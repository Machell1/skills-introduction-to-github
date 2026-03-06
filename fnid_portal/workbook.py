"""
JCF Case Management Operational Workbook Generator
Aligned with JCF FO 4032 (2024-Sept-12) Case Management Policy & SOPs,
DPP Prosecution Protocol 2012, and DPP Disclosure Protocol 2013.

Generates a multi-sheet Excel workbook covering:
  1. Case Register & Investigation Tracking
  2. SOP Compliance Checklist
  3. DPP File Pipeline
  4. Evidence & Chain of Custody
  5. Witness Statements Log
  6. Disclosure Schedule
  7. 48-Hour Compliance Tracker
  8. Forensic Lab Tracker
  9. Court Date Tracker
 10. Command Summary Dashboard
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import constants as cfg

# ---------------------------------------------------------------------------
# Style definitions
# ---------------------------------------------------------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")

_SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGN = Alignment(vertical="top", wrap_text=True)

_SOP_YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_SOP_NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_SOP_NA_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

_OVERDUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP = Alignment(vertical="top", wrap_text=True)


def _apply_header_row(ws, row_num, headers, widths=None):
    """Apply formatted headers to a row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    if widths:
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_title_row(ws, title, col_count, row_num=1):
    """Add a merged title row with JCF branding."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 30


def _add_subtitle_row(ws, text, col_count, row_num):
    """Add a subtitle / metadata row."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(name="Calibri", size=9, italic=True)
    cell.alignment = Alignment(horizontal="center")


def _add_dropdown(ws, col_letter, min_row, max_row, options):
    """Add data-validation dropdown to a column range."""
    if not options or max_row < min_row:
        return
    formula = '"' + ",".join(str(o)[:255] for o in options[:25]) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select a value"
    dv.promptTitle = "Controlled Input"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


def _write_data_rows(ws, rows, start_row, columns):
    """Write database rows into the worksheet."""
    for r_idx, row in enumerate(rows):
        for c_idx, col_name in enumerate(columns):
            cell = ws.cell(row=start_row + r_idx, column=c_idx + 1,
                           value=row[col_name] if row[col_name] else "")
            cell.font = _DATA_FONT
            cell.alignment = _DATA_ALIGN
            cell.border = _BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_case_register(wb, db_rows):
    """Sheet 1: Case Register & Investigation Tracker."""
    ws = wb.active
    ws.title = "Case Register"

    headers = [
        "Case ID", "Reg. Date", "Classification", "OIC Name", "OIC Rank",
        "OIC Badge", "Parish", "Division", "Offence Description",
        "Law & Section", "Suspect Name", "Suspect DOB", "Suspect Address",
        "Victim Name", "Case Status", "File Completeness",
        "SOP Compliance", "DPP Submission Date", "DPP Status",
        "Court Type", "Next Court Date", "Verdict", "Sentence",
        "POCA Referred", "POCA Status", "Record Status", "Notes",
    ]
    widths = [
        14, 12, 22, 18, 20, 12, 14, 16, 30,
        30, 18, 12, 22,
        18, 28, 28,
        28, 14, 28,
        32, 14, 14, 14,
        12, 28, 12, 30,
    ]
    col_map = [
        "case_id", "registration_date", "classification", "oic_name", "oic_rank",
        "oic_badge", "parish", "division", "offence_description",
        "law_and_section", "suspect_name", "suspect_dob", "suspect_address",
        "victim_name", "case_status", "file_completeness",
        "sop_compliance", "dpp_submission_date", "dpp_status",
        "court_type", "next_court_date", "verdict", "sentence",
        "poca_referred", "poca_status", "record_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "JAMAICA CONSTABULARY FORCE — FNID AREA 3 — CASE REGISTER", ncols)
    _add_subtitle_row(ws, f"Per JCF FO 4032 Case Management Policy & SOPs | Generated: {datetime.now():%Y-%m-%d %H:%M}", ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    # Dropdowns
    _add_dropdown(ws, "C", data_start, blank_end, cfg.CASE_CLASSIFICATIONS)
    _add_dropdown(ws, "E", data_start, blank_end, cfg.JCF_RANKS)
    _add_dropdown(ws, "G", data_start, blank_end, cfg.ALL_PARISHES)
    _add_dropdown(ws, "J", data_start, blank_end, cfg.ALL_OFFENCES[:25])
    _add_dropdown(ws, "O", data_start, blank_end, cfg.CASE_STATUS)
    _add_dropdown(ws, "P", data_start, blank_end, cfg.FILE_COMPLETENESS)
    _add_dropdown(ws, "Q", data_start, blank_end, cfg.SOP_COMPLIANCE_OVERALL)
    _add_dropdown(ws, "S", data_start, blank_end, cfg.DPP_FILE_STATUS)
    _add_dropdown(ws, "T", data_start, blank_end, cfg.COURT_TYPES)
    _add_dropdown(ws, "X", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "Y", data_start, blank_end, cfg.POCA_STATUS)
    _add_dropdown(ws, "Z", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_sop_checklist(wb, db_rows):
    """Sheet 2: SOP Compliance Checklist (per FO 4032)."""
    ws = wb.create_sheet("SOP Checklist")

    # Build flat header list from categories
    static_headers = ["Case ID", "OIC Name", "Checklist Date"]
    sop_headers = []
    sop_fields = []
    for category, items in cfg.SOP_CHECKLIST_CATEGORIES.items():
        for field_name, label in items:
            sop_headers.append(label)
            sop_fields.append(field_name)
    trailing = ["Overall Compliance", "Compliance Notes", "Record Status"]

    all_headers = static_headers + sop_headers + trailing
    ncols = len(all_headers)

    _add_title_row(ws, "JCF FO 4032 — SOP COMPLIANCE CHECKLIST", ncols)
    _add_subtitle_row(ws,
        "48-item checklist covering 7 SOP categories | Yes/No/N/A per item",
        ncols, 2)

    # Category sub-headers (row 3)
    col = 4  # start after static columns
    for category, items in cfg.SOP_CHECKLIST_CATEGORIES.items():
        span = len(items)
        ws.merge_cells(start_row=3, start_column=col,
                       end_row=3, end_column=col + span - 1)
        cell = ws.cell(row=3, column=col, value=category)
        cell.font = _SUBHEADER_FONT
        cell.fill = _SUBHEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        col += span

    _apply_header_row(ws, 4, all_headers)

    # Set widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    for i in range(4, 4 + len(sop_headers)):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(ncols - 2)].width = 22
    ws.column_dimensions[get_column_letter(ncols - 1)].width = 30
    ws.column_dimensions[get_column_letter(ncols)].width = 12

    data_start = 5
    col_map = (
        ["linked_case_id", "oic_name", "checklist_date"]
        + sop_fields
        + ["overall_compliance", "compliance_notes", "record_status"]
    )
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    # Dropdowns for Yes/No/N/A columns
    for i in range(len(sop_fields)):
        col_letter = get_column_letter(4 + i)
        _add_dropdown(ws, col_letter, data_start, blank_end, cfg.YES_NO_NA)

    # Overall compliance dropdown
    comp_col = get_column_letter(ncols - 2)
    _add_dropdown(ws, comp_col, data_start, blank_end, cfg.SOP_COMPLIANCE_OVERALL)

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{blank_end}"


def _build_dpp_pipeline(wb, db_rows):
    """Sheet 3: DPP File Pipeline (per Prosecution Protocol 2012)."""
    ws = wb.create_sheet("DPP Pipeline")

    headers = [
        "Case ID", "Classification", "OIC Name", "Suspect Name",
        "Offence Summary", "DPP File Date", "Crown Counsel",
        "DPP Status", "Evidential Sufficiency", "Public Interest Met",
        "Ruling Date", "Ruling Outcome", "Ruling Notes",
        "Voluntary Bill", "Prelim Enquiry", "Returned for Investigation",
        "Return Reason", "Resubmission Date", "Record Status",
    ]
    widths = [
        14, 22, 18, 18,
        30, 12, 18,
        28, 12, 12,
        12, 18, 30,
        10, 10, 10,
        30, 12, 12,
    ]
    col_map = [
        "linked_case_id", "classification", "oic_name", "suspect_name",
        "offence_summary", "dpp_file_date", "crown_counsel",
        "dpp_status", "evidential_sufficiency", "public_interest_met",
        "ruling_date", "ruling_outcome", "ruling_notes",
        "voluntary_bill", "prelim_exam", "returned_for_investigation",
        "return_reason", "resubmission_date", "record_status",
    ]

    ncols = len(headers)
    _add_title_row(ws, "DPP FILE PIPELINE — PROSECUTION PROTOCOL 2012", ncols)
    _add_subtitle_row(ws,
        "Two-stage test: Evidential Sufficiency + Public Interest | 6-week turnaround",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    _add_dropdown(ws, "B", data_start, blank_end, cfg.CASE_CLASSIFICATIONS)
    _add_dropdown(ws, "H", data_start, blank_end, cfg.DPP_FILE_STATUS)
    _add_dropdown(ws, "I", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "J", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "N", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "O", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "P", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "S", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_evidence_custody(wb, db_rows):
    """Sheet 4: Evidence & Chain of Custody."""
    ws = wb.create_sheet("Evidence & Custody")

    headers = [
        "Exhibit Tag", "Exhibit Type", "Description", "Linked Case ID",
        "Linked Seizure ID", "Seized Date", "Seized By", "Seized Location",
        "Current Custodian", "Storage Location",
        "Transfer Date", "Transfer From", "Transfer To", "Transfer Reason",
        "Condition", "Photos Taken", "Seal Intact",
        "Disposal Status", "Record Status", "Notes",
    ]
    widths = [
        14, 22, 25, 14,
        14, 12, 18, 20,
        18, 18,
        12, 18, 18, 20,
        14, 10, 10,
        28, 12, 25,
    ]
    col_map = [
        "exhibit_tag", "exhibit_type", "description", "linked_case_id",
        "linked_seizure_id", "seized_date", "seized_by", "seized_location",
        "current_custodian", "storage_location",
        "transfer_date", "transfer_from", "transfer_to", "transfer_reason",
        "condition", "photos_taken", "seal_intact",
        "disposal_status", "record_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "EVIDENCE & CHAIN OF CUSTODY LOG", ncols)
    _add_subtitle_row(ws,
        "Per JCF FO 4032 Exhibit Management SOP | Maintain unbroken chain of custody",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    _add_dropdown(ws, "B", data_start, blank_end, cfg.EXHIBIT_TYPES)
    _add_dropdown(ws, "P", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "Q", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "R", data_start, blank_end, cfg.EXHIBIT_DISPOSAL)
    _add_dropdown(ws, "S", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_witness_log(wb, db_rows):
    """Sheet 5: Witness Statements Log."""
    ws = wb.create_sheet("Witness Statements")

    headers = [
        "Statement ID", "Case ID", "Witness Name", "Witness Type",
        "Address", "Phone", "Relation to Case",
        "Statement Date", "Taken By", "Pages", "Signed",
        "Willing to Testify", "Special Measures", "Measures Type",
        "Available for Court", "Record Status", "Notes",
    ]
    widths = [
        14, 14, 20, 16,
        22, 14, 18,
        12, 18, 8, 8,
        10, 10, 20,
        10, 12, 25,
    ]
    col_map = [
        "statement_id", "linked_case_id", "witness_name", "witness_type",
        "witness_address", "witness_phone", "relation_to_case",
        "statement_date", "statement_taken_by", "statement_pages", "statement_signed",
        "witness_willing", "special_measures_needed", "special_measures_type",
        "available_for_court", "record_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "WITNESS STATEMENTS LOG", ncols)
    _add_subtitle_row(ws,
        "Per JCF FO 4032 Statement-Taking SOP | Track all prosecution & defence witnesses",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    witness_types = [
        "Victim / Complainant", "Civilian Eyewitness",
        "Expert / Forensic", "Investigating Officer",
        "Arresting Officer", "Scene Officer",
        "Character Witness", "Defence Witness",
    ]
    _add_dropdown(ws, "D", data_start, blank_end, witness_types)
    _add_dropdown(ws, "K", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "L", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "M", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "O", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "P", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_disclosure_schedule(wb, db_rows):
    """Sheet 6: Disclosure Schedule (per DPP Disclosure Protocol 2013)."""
    ws = wb.create_sheet("Disclosure Schedule")

    headers = [
        "Disclosure ID", "Case ID", "Disclosure Date", "Type",
        "Material Disclosed", "Served on Defence", "Defence Solicitor",
        "Service Method", "Service Date", "Acknowledgement Received",
        "PII Application", "PII Outcome",
        "Supplementary Needed", "Supplementary Date",
        "Disclosure Status", "Prepared By", "Record Status", "Notes",
    ]
    widths = [
        14, 14, 12, 20,
        30, 10, 20,
        14, 12, 10,
        10, 20,
        10, 12,
        28, 18, 12, 25,
    ]
    col_map = [
        "disclosure_id", "linked_case_id", "disclosure_date", "disclosure_type",
        "material_disclosed", "served_on_defence", "defence_solicitor",
        "service_method", "service_date", "acknowledgement_received",
        "pii_application", "pii_outcome",
        "supplementary_needed", "supplementary_date",
        "disclosure_status", "prepared_by", "record_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "DISCLOSURE SCHEDULE — DPP DISCLOSURE PROTOCOL 2013", ncols)
    _add_subtitle_row(ws,
        "Primary disclosure, PII applications, supplementary disclosure, unused material",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    disclosure_types = [
        "Primary Disclosure", "Initial Disclosure",
        "Supplementary Disclosure", "Unused Material Schedule",
        "Defence Statement Response", "Third Party Material",
    ]
    service_methods = [
        "Hand Delivered (signed receipt)", "Registered Mail",
        "Email (acknowledged)", "Court Filing", "DPP Office",
    ]
    _add_dropdown(ws, "D", data_start, blank_end, disclosure_types)
    _add_dropdown(ws, "F", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "H", data_start, blank_end, service_methods)
    _add_dropdown(ws, "J", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "K", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "M", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "O", data_start, blank_end, cfg.DISCLOSURE_STATUS)
    _add_dropdown(ws, "Q", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_48hr_tracker(wb, db_rows):
    """Sheet 7: 48-Hour Compliance Tracker (Constabulary Force Act s.15)."""
    ws = wb.create_sheet("48-Hour Tracker")

    headers = [
        "Arrest ID", "Arrest Date", "Arrest Time", "Suspect Name",
        "Parish", "Arrest Location", "Arresting Officer", "Rank",
        "Offence 1", "Law & Section",
        "48-Hour Deadline", "Charge Date", "Charged Within 48hr",
        "Bail Status", "Court Type", "First Court Date",
        "Remand Location", "Legal Rep", "Record Status",
    ]
    widths = [
        14, 12, 10, 20,
        14, 20, 18, 20,
        28, 28,
        16, 12, 12,
        28, 32, 12,
        28, 14, 12,
    ]
    col_map = [
        "arrest_id", "arrest_date", "arrest_time", "suspect_name",
        "parish", "arrest_location", "arresting_officer", "arresting_officer_rank",
        "offence_1", "law_section_1",
        "deadline_48hr", "charge_date", "charge_within_48hr",
        "bail_status", "court_type", "first_court_date",
        "remand_location", "legal_representation", "record_status",
    ]

    ncols = len(headers)
    _add_title_row(ws, "48-HOUR COMPLIANCE TRACKER — CONSTABULARY FORCE ACT s.15", ncols)
    _add_subtitle_row(ws,
        "Charge or release within 48 hours | Automatic deadline calculation",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    _add_dropdown(ws, "E", data_start, blank_end, cfg.ALL_PARISHES)
    _add_dropdown(ws, "I", data_start, blank_end, cfg.ALL_OFFENCES[:25])
    _add_dropdown(ws, "M", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "N", data_start, blank_end, cfg.BAIL_STATUS)
    _add_dropdown(ws, "O", data_start, blank_end, cfg.COURT_TYPES)
    _add_dropdown(ws, "Q", data_start, blank_end, cfg.REMAND_LOCATIONS)
    _add_dropdown(ws, "S", data_start, blank_end, cfg.RECORD_STATUS)

    # Conditional formatting: highlight non-compliant rows
    for r_idx, row in enumerate(db_rows):
        compliance = row["charge_within_48hr"] if row["charge_within_48hr"] else ""
        if compliance == "No":
            for c in range(1, ncols + 1):
                cell = ws.cell(row=data_start + r_idx, column=c)
                cell.fill = _OVERDUE_FILL

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_lab_tracker(wb, db_rows):
    """Sheet 8: Forensic Lab Submission Tracker."""
    ws = wb.create_sheet("Forensic Lab Tracker")

    headers = [
        "Lab Ref", "Exhibit Tag", "Linked Case ID",
        "Submission Date", "Lab Type", "Exam Type", "Analyst",
        "Expected Date", "Completion Date",
        "Certificate No.", "Certificate Status", "Result",
        "Collected By", "Collection Date",
        "IBIS Status", "eTrace Status", "Record Status", "Notes",
    ]
    widths = [
        14, 14, 14,
        12, 14, 18, 18,
        12, 12,
        14, 22, 18,
        18, 12,
        22, 22, 12, 25,
    ]
    col_map = [
        "lab_ref", "exhibit_tag", "linked_case_id",
        "submission_date", "lab_type", "exam_type", "analyst",
        "expected_date", "completion_date",
        "certificate_number", "certificate_status", "result",
        "collected_by", "collection_date",
        "ibis_status", "etrace_status", "record_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "FORENSIC LAB SUBMISSION TRACKER — IFSLM", ncols)
    _add_subtitle_row(ws,
        "Track ballistic certificates, drug analysis, DNA | 8-week overdue threshold",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    data_start = 4
    _write_data_rows(ws, db_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(db_rows) - 1)
    blank_end = max_data + 50

    lab_types = [
        "IFSLM - Ballistics", "IFSLM - Chemistry (Drug Analysis)",
        "IFSLM - DNA / Biology", "IFSLM - Toxicology",
        "IFSLM - Questioned Documents", "IFSLM - Digital Forensics",
        "FMO - Post Mortem / Autopsy", "External Lab",
    ]
    _add_dropdown(ws, "E", data_start, blank_end, lab_types)
    _add_dropdown(ws, "K", data_start, blank_end, cfg.FORENSIC_CERT_STATUS)
    _add_dropdown(ws, "O", data_start, blank_end, cfg.IBIS_STATUS)
    _add_dropdown(ws, "P", data_start, blank_end, cfg.ETRACE_STATUS)
    _add_dropdown(ws, "Q", data_start, blank_end, cfg.RECORD_STATUS)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_court_tracker(wb, cases_rows):
    """Sheet 9: Court Date Tracker."""
    ws = wb.create_sheet("Court Tracker")

    headers = [
        "Case ID", "Classification", "OIC Name", "Suspect Name",
        "Offence", "Court Type", "Next Court Date", "Trial Date",
        "DPP Status", "Verdict", "Sentence",
        "POCA Referred", "POCA Status", "Notes",
    ]
    widths = [
        14, 22, 18, 18,
        30, 32, 14, 14,
        28, 14, 14,
        10, 28, 25,
    ]
    col_map = [
        "case_id", "classification", "oic_name", "suspect_name",
        "law_and_section", "court_type", "next_court_date", "trial_date",
        "dpp_status", "verdict", "sentence",
        "poca_referred", "poca_status", "notes",
    ]

    ncols = len(headers)
    _add_title_row(ws, "COURT DATE TRACKER", ncols)
    _add_subtitle_row(ws,
        "Gun Court, Parish Court, Circuit Court, Court of Appeal tracking",
        ncols, 2)
    _apply_header_row(ws, 3, headers, widths)

    # Filter only cases that have court involvement
    court_rows = [r for r in cases_rows
                  if r["court_type"] or r["next_court_date"] or r["trial_date"]
                  or (r["case_status"] and "Court" in str(r["case_status"]))]

    data_start = 4
    _write_data_rows(ws, court_rows, data_start, col_map)

    max_data = max(data_start, data_start + len(court_rows) - 1)
    blank_end = max_data + 50

    _add_dropdown(ws, "F", data_start, blank_end, cfg.COURT_TYPES)
    _add_dropdown(ws, "I", data_start, blank_end, cfg.DPP_FILE_STATUS)
    _add_dropdown(ws, "L", data_start, blank_end, cfg.YES_NO)
    _add_dropdown(ws, "M", data_start, blank_end, cfg.POCA_STATUS)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{blank_end}"


def _build_command_summary(wb, conn):
    """Sheet 10: Command Summary Dashboard with KPIs."""
    ws = wb.create_sheet("Command Summary")

    ncols = 8
    _add_title_row(ws, "FNID AREA 3 — COMMAND SUMMARY DASHBOARD", ncols)
    _add_subtitle_row(ws,
        f"JCF FO 4032 Compliance Overview | Generated: {datetime.now():%Y-%m-%d %H:%M}",
        ncols, 2)

    # KPI section
    row = 4
    kpi_header = Font(name="Calibri", size=11, bold=True, color="1F3864")
    kpi_value = Font(name="Calibri", size=16, bold=True, color="C00000")

    kpis = [
        ("Total Cases", "SELECT COUNT(*) FROM cases"),
        ("Open Cases", "SELECT COUNT(*) FROM cases WHERE case_status LIKE 'Open%'"),
        ("At DPP", "SELECT COUNT(*) FROM cases WHERE case_status LIKE 'Referred to DPP%'"),
        ("Before Court", "SELECT COUNT(*) FROM cases WHERE case_status LIKE 'Before Court%'"),
        ("Closed - Convicted", "SELECT COUNT(*) FROM cases WHERE case_status LIKE 'Closed - Convicted%'"),
        ("SOP Fully Compliant", "SELECT COUNT(*) FROM cases WHERE sop_compliance LIKE 'Fully%'"),
        ("48-hr Compliant Arrests", "SELECT COUNT(*) FROM arrests WHERE charge_within_48hr='Yes'"),
        ("Total Firearms Seized", "SELECT COUNT(*) FROM firearm_seizures"),
    ]

    for col_idx, (label, query) in enumerate(kpis):
        c = col_idx + 1
        ws.cell(row=row, column=c, value=label).font = kpi_header
        ws.cell(row=row, column=c).alignment = _CENTER
        ws.cell(row=row, column=c).border = _BORDER
        val = conn.execute(query).fetchone()[0]
        ws.cell(row=row + 1, column=c, value=val).font = kpi_value
        ws.cell(row=row + 1, column=c).alignment = _CENTER
        ws.cell(row=row + 1, column=c).border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Case status breakdown
    row = 7
    ws.cell(row=row, column=1, value="CASE STATUS BREAKDOWN").font = kpi_header
    _apply_header_row(ws, row + 1, ["Status", "Count"], [35, 12])
    statuses = conn.execute(
        "SELECT case_status, COUNT(*) as cnt FROM cases GROUP BY case_status ORDER BY cnt DESC"
    ).fetchall()
    for i, s in enumerate(statuses):
        ws.cell(row=row + 2 + i, column=1, value=s["case_status"]).font = _DATA_FONT
        ws.cell(row=row + 2 + i, column=1).border = _BORDER
        ws.cell(row=row + 2 + i, column=2, value=s["cnt"]).font = _DATA_FONT
        ws.cell(row=row + 2 + i, column=2).border = _BORDER

    # DPP pipeline breakdown
    dpp_start = row
    ws.cell(row=dpp_start, column=4, value="DPP PIPELINE STATUS").font = kpi_header
    _apply_header_row(ws, dpp_start + 1, ["", "", "", "DPP Status", "Count"])
    dpp_stats = conn.execute(
        "SELECT dpp_status, COUNT(*) as cnt FROM dpp_pipeline GROUP BY dpp_status ORDER BY cnt DESC"
    ).fetchall()
    for i, d in enumerate(dpp_stats):
        ws.cell(row=dpp_start + 2 + i, column=4, value=d["dpp_status"]).font = _DATA_FONT
        ws.cell(row=dpp_start + 2 + i, column=4).border = _BORDER
        ws.cell(row=dpp_start + 2 + i, column=5, value=d["cnt"]).font = _DATA_FONT
        ws.cell(row=dpp_start + 2 + i, column=5).border = _BORDER
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12

    # Parish breakdown
    parish_start = row
    ws.cell(row=parish_start, column=7, value="CASES BY PARISH").font = kpi_header
    _apply_header_row(ws, parish_start + 1,
                      ["", "", "", "", "", "", "Parish", "Count"])
    parish_stats = conn.execute(
        "SELECT parish, COUNT(*) as cnt FROM cases GROUP BY parish ORDER BY cnt DESC"
    ).fetchall()
    for i, p in enumerate(parish_stats):
        ws.cell(row=parish_start + 2 + i, column=7, value=p["parish"]).font = _DATA_FONT
        ws.cell(row=parish_start + 2 + i, column=7).border = _BORDER
        ws.cell(row=parish_start + 2 + i, column=8, value=p["cnt"]).font = _DATA_FONT
        ws.cell(row=parish_start + 2 + i, column=8).border = _BORDER
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 12


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_operational_workbook(conn):
    """Generate a comprehensive operational workbook from current database state.

    Returns an openpyxl Workbook ready to be saved.
    """
    wb = Workbook()

    # Fetch all data
    cases = conn.execute("SELECT * FROM cases ORDER BY id DESC").fetchall()
    sop = conn.execute("SELECT * FROM sop_checklists ORDER BY id DESC").fetchall()
    dpp = conn.execute("SELECT * FROM dpp_pipeline ORDER BY id DESC").fetchall()
    custody = conn.execute("SELECT * FROM chain_of_custody ORDER BY id DESC").fetchall()
    witnesses = conn.execute("SELECT * FROM witness_statements ORDER BY id DESC").fetchall()
    disclosure = conn.execute("SELECT * FROM disclosure_log ORDER BY id DESC").fetchall()
    arrests = conn.execute("SELECT * FROM arrests ORDER BY id DESC").fetchall()
    lab = conn.execute("SELECT * FROM lab_tracking ORDER BY id DESC").fetchall()

    _build_case_register(wb, cases)
    _build_sop_checklist(wb, sop)
    _build_dpp_pipeline(wb, dpp)
    _build_evidence_custody(wb, custody)
    _build_witness_log(wb, witnesses)
    _build_disclosure_schedule(wb, disclosure)
    _build_48hr_tracker(wb, arrests)
    _build_lab_tracker(wb, lab)
    _build_court_tracker(wb, cases)
    _build_command_summary(wb, conn)

    return wb
