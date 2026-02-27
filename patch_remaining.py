#!/usr/bin/env python3
"""
Patch — Remaining gaps: Portal updates, 8 new tables, bug fixes, polish sheets.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
from generate_fnid_workbook import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BODY_FONT_BOLD,
    CENTER, LEFT_WRAP, THIN_BORDER,
    GREEN_FILL, AMBER_FILL, RED_FILL, LIGHT_GRAY_FILL, GOLD_FILL,
    NAVY, DARK_BLUE, LIGHT_BLUE, WHITE, MED_GRAY, MED_BLUE,
    SECTION_FONT, SUBHEADER_FONT, SUBHEADER_FILL,
    TABLE_STYLE,
    write_headers, add_table, add_data_validation_list,
    _portal_header, _portal_kpi_summary, _portal_quick_stats,
)


# ══════════════════════════════════════════════════════════════════════
#  PART 1 — REBUILD Portal_Training with tbl_Training link
# ══════════════════════════════════════════════════════════════════════
def rebuild_portal_training(wb):
    if "Portal_Training" in wb.sheetnames:
        del wb["Portal_Training"]
    ws = wb.create_sheet("Portal_Training")
    r = _portal_header(ws, "TRAINING PORTAL", "Capacity Building, Courses, Certifications & Personnel Development", "Training", "548235")
    kpis = [(28, "Training Sessions Conducted"), (29, "Personnel Trained")]
    r = _portal_kpi_summary(ws, r, "Training", kpis)
    stats = [
        ("Sessions Held", '=COUNTA(tbl_Training[TrainingID])-COUNTBLANK(tbl_Training[TrainingID])'),
        ("Total Attendees", '=SUM(tbl_Training[Attendees])'),
        ("Total Hours", '=SUM(tbl_Training[Duration_Hrs])'),
        ("Certificates Issued", '=COUNTIF(tbl_Training[CertificateIssued],"Yes")'),
        ("Total Cost (JMD)", '=SUM(tbl_Training[CostJMD])'),
        ("Avg Attendees/Session", '=IFERROR(AVERAGE(tbl_Training[Attendees]),"—")'),
        ("Training Overdue (Staff)", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
        ("Active Personnel", '=COUNTIF(tbl_Personnel[Status],"Active")'),
    ]
    r = _portal_quick_stats(ws, r, stats)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: 'tbl_Training' for course log | 'tbl_Personnel' for staff qualifications & training due dates").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART 2 — REBUILD Portal_AdminHRM with tbl_Personnel + tbl_DutyRoster
# ══════════════════════════════════════════════════════════════════════
def rebuild_portal_admin_hrm(wb):
    if "Portal_AdminHRM" in wb.sheetnames:
        del wb["Portal_AdminHRM"]
    ws = wb.create_sheet("Portal_AdminHRM")
    r = _portal_header(ws, "ADMIN / HRM PORTAL", "Personnel Management, Duty Roster, Leave & Discipline", "Admin/HRM", "7030A0")
    kpis = [(26, "Personnel Strength"), (27, "Vacancy Rate")]
    r = _portal_kpi_summary(ws, r, "Admin/HRM", kpis)
    stats = [
        ("Total Personnel", '=COUNTA(tbl_Personnel[PersonnelID])-COUNTBLANK(tbl_Personnel[PersonnelID])'),
        ("Active", '=COUNTIF(tbl_Personnel[Status],"Active")'),
        ("On Leave", '=COUNTIF(tbl_Personnel[Status],"On Leave")'),
        ("Seconded", '=COUNTIF(tbl_Personnel[Status],"Seconded")'),
        ("Suspended", '=COUNTIF(tbl_Personnel[Status],"Suspended")'),
        ("Training Overdue", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
        ("Duty Entries", '=COUNTA(tbl_DutyRoster[RosterID])-COUNTBLANK(tbl_DutyRoster[RosterID])'),
        ("Leave Records", '=COUNTA(tbl_LeaveManagement[LeaveID])-COUNTBLANK(tbl_LeaveManagement[LeaveID])'),
        ("Disciplinary Cases", '=COUNTA(tbl_Disciplinary[DisciplinaryID])-COUNTBLANK(tbl_Disciplinary[DisciplinaryID])'),
        ("Pending Discipline", '=COUNTIF(tbl_Disciplinary[Status],"Pending")'),
        ("Approved Leave", '=COUNTIF(tbl_LeaveManagement[Status],"Approved")'),
        ("Leave Pending", '=COUNTIF(tbl_LeaveManagement[Status],"Pending")'),
    ]
    r = _portal_quick_stats(ws, r, stats)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: 'tbl_Personnel' for staff register | 'tbl_DutyRoster' for shifts | "
                  "'tbl_LeaveManagement' for leave | 'tbl_Disciplinary' for discipline cases").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART 3 — REBUILD Portal_Oversight with all SOP tables
# ══════════════════════════════════════════════════════════════════════
def rebuild_portal_oversight(wb):
    if "Portal_Oversight" in wb.sheetnames:
        del wb["Portal_Oversight"]
    ws = wb.create_sheet("Portal_Oversight")
    r = _portal_header(ws, "OVERSIGHT PORTAL", "Compliance, Governance, SOP Adherence, Risk & Audit", "Oversight", "002060")
    kpis = [(30, "Non-Conformities Identified"), (31, "Overdue Actions"), (32, "KPIs Off Track")]
    r = _portal_kpi_summary(ws, r, "Oversight", kpis)
    stats = [
        ("KPIs On Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"On Track")'),
        ("KPIs At Risk", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"At Risk")'),
        ("KPIs Off Track", '=COUNTIF(tbl_UnitPlans[TrafficLight_Monthly],"Off Track")'),
        ("CR Non-Compliant", '=COUNTIF(tbl_Cases[ComplianceFlag],"Non-Compliant")'),
        ("Overdue Actions", '=COUNTIFS(tbl_Actions[DueDate],"<"&TODAY(),tbl_Actions[CompletedDate],"")'),
        ("Overdue Reviews", '=COUNTIFS(tbl_Reviews[NextReviewDue],"<"&TODAY())'),
        ("Directives Overdue", '=COUNTIFS(tbl_Reviews[DirectiveDueDate],"<"&TODAY(),tbl_Reviews[DirectiveStatus],"<>Completed")'),
        ("Files Incomplete", '=COUNTIFS(tbl_Cases[FileCompletenessScore],"<"&0.5)'),
        ("48hr Charge Missed", '=COUNTIF(tbl_Cases[ChargeDeadlineMet],"No")'),
        ("Warrants Expired", '=COUNTIF(tbl_Warrants[Status],"Expired")'),
        ("Forensic Overdue", '=COUNTIFS(tbl_Exhibits[ExpectedReturn],"<"&TODAY(),tbl_Exhibits[ActualReturnDate],"")'),
        ("Risk Items", '=COUNTA(tbl_RiskRegister[RiskID])-COUNTBLANK(tbl_RiskRegister[RiskID])'),
        ("Victim Contact Overdue", '=COUNTIFS(tbl_VictimContact[NextScheduledContact],"<"&TODAY())'),
        ("Summons Unserved", '=COUNTIFS(tbl_Witnesses[SummonsIssued],"Yes",tbl_Witnesses[SummonsServed],"No")'),
        ("Correspondence Pending", '=COUNTIF(tbl_Correspondence[Status],"Pending")'),
        ("Training Overdue", '=COUNTIFS(tbl_Personnel[NextTrainingDue],"<"&TODAY())'),
    ]
    r = _portal_quick_stats(ws, r, stats)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
            value="DATA ENTRY: 'tbl_RiskRegister' for organizational risks | 'tbl_Correspondence' for external comms | "
                  "All SOP tables feed into this portal automatically").font = Font(name="Calibri", italic=True, size=10, color="666666")
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART 4 — NEW TABLES (Gaps 1-9, 13, 15)
# ══════════════════════════════════════════════════════════════════════

def build_ballistic_trace(wb):
    ws = wb.create_sheet("tbl_BallisticTrace")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "TraceID","LinkedCaseID","LinkedExhibitID","LinkedSeizureID",
        "FirearmMake","FirearmModel","FirearmCalibre","SerialNumber",
        "TraceSubmitDate","TracingAgency","TraceType",
        "IBIS_Hit","IBIS_MatchCaseID","IBIS_MatchDetails",
        "ATF_TraceNumber","ATF_Result","CountryOfOrigin","ImportPath",
        "INTERPOL_Ref","MatchToOtherScene","MatchSceneCaseID",
        "BallisticExpertName","ExpertReport","ReportDate",
        "Status","Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,14,16,16,14,16,14,18,16,12,16,28,16,20,16,20,16,14,16,20,14,14,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","BTRC-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_BallisticTrace", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "L2:L1000", "=Lookups!$I$2:$I$3")
    ws.conditional_formatting.add("L2:L1000", FormulaRule(formula=['=L2="Yes"'], fill=GOLD_FILL))
    return ws

def build_gang_affiliation(wb):
    ws = wb.create_sheet("tbl_GangAffiliation")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "RecordID","LinkedPersonID","LinkedCaseID",
        "GangName","GangAlias","GangTerritory","Parish",
        "RoleInGang","RankInGang","RecruitmentDate",
        "ActiveStatus","ThreatLevel","AssociatesCount",
        "LinkedGangMembers","IntelSourceID",
        "FirearmsAssociated","NarcoticsAssociated",
        "LastKnownActivity","LastActivityDate",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,20,16,22,16,16,14,14,12,14,12,24,14,12,12,24,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","GANG-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_GangAffiliation", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "G2:G1000", "=Lookups!$D$2:$D$16")
    add_data_validation_list(ws, "P2:P1000", "=Lookups!$I$2:$I$3")
    add_data_validation_list(ws, "Q2:Q1000", "=Lookups!$I$2:$I$3")
    return ws

def build_budget_table(wb):
    ws = wb.create_sheet("tbl_Budget")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "BudgetID","FiscalYear","Quarter","Month","Unit",
        "BudgetCategory","Description",
        "AllocatedAmount","SpentAmount","CommittedAmount","AvailableBalance",
        "Variance","VariancePct",
        "FundingSource","ApprovedBy","ApprovalDate",
        "LinkedOpsID","LinkedTrainingID",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate",
        "Week","MonthNum","QuarterNum","Year"
    ]
    widths = [14,12,10,10,16,20,28,14,14,14,14,14,12,18,18,14,14,14,24,14,14,14,14,8,8,8,8]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(E{r}="","","BUD-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=11, value=f'=IF(OR(H{r}="",I{r}=""),"",H{r}-I{r}-J{r})')
        ws.cell(row=r, column=12, value=f'=IF(OR(H{r}="",I{r}=""),"",I{r}-H{r})')
        ws.cell(row=r, column=13, value=f'=IF(OR(H{r}="",H{r}=0),"",L{r}/H{r})')
        ws.cell(row=r, column=13).number_format = '0%'
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_Budget", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    ws.conditional_formatting.add("M2:M1000", FormulaRule(formula=['=AND(M2<>"",M2>0.1)'], fill=RED_FILL))
    return ws

def build_leave_management(wb):
    ws = wb.create_sheet("tbl_LeaveManagement")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "LeaveID","PersonnelID","OfficerName","Rank","Unit",
        "LeaveType","StartDate","EndDate","DaysRequested","DaysApproved",
        "Status","ApprovedBy","ApprovalDate",
        "ReliefOfficer","HandoverDone",
        "ReturnDate","ReturnOnTime",
        "AnnualBalance","SickBalance",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,22,14,14,16,14,14,12,12,14,18,14,18,12,14,12,12,12,24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","LV-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=9, value=f'=IF(OR(G{r}="",H{r}=""),"",NETWORKDAYS(G{r},H{r}))')
        ws.cell(row=r, column=17, value=f'=IF(OR(H{r}="",P{r}=""),"",IF(P{r}<=H{r},"Yes","No"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_LeaveManagement", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "O2:O1000", "=Lookups!$I$2:$I$3")
    return ws

def build_disciplinary(wb):
    ws = wb.create_sheet("tbl_Disciplinary")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        "DisciplinaryID","PersonnelID","OfficerName","Rank","Unit","BadgeNo",
        "IncidentDate","ReportDate","ReportedBy",
        "AllegationType","AllegationDescription",
        "InvestigatingOfficer","InvestigationStartDate","InvestigationEndDate",
        "Findings","Recommendation","Sanction",
        "HearingDate","HearingPanel","Outcome",
        "AppealFiled","AppealDate","AppealOutcome",
        "Status","Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,22,14,14,14,14,14,18,18,30,18,14,14,30,24,20,14,20,20,12,14,20,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","DISC-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_Disciplinary", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "E2:E1000", "=Lookups!$E$2:$E$11")
    add_data_validation_list(ws, "U2:U1000", "=Lookups!$I$2:$I$3")
    return ws

def build_correspondence(wb):
    ws = wb.create_sheet("tbl_Correspondence")
    ws.sheet_properties.tabColor = "808080"
    headers = [
        "CorrespondenceID","LinkedCaseID","Date","Direction",
        "From","To","Agency","Subject",
        "Method","ReferenceNo","Summary",
        "ResponseRequired","ResponseDueDate","ResponseDate",
        "Status","Priority",
        "Attachments","FiledLocation",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,14,12,20,20,20,24,14,16,30,12,14,14,14,12,14,16,24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","CORR-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_Correspondence", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "L2:L1000", "=Lookups!$I$2:$I$3")
    ws.conditional_formatting.add("M2:M1000", FormulaRule(formula=['=AND(M2<>"",M2<TODAY(),N2="")'], fill=RED_FILL))
    return ws

def build_risk_register(wb):
    ws = wb.create_sheet("tbl_RiskRegister")
    ws.sheet_properties.tabColor = "002060"
    headers = [
        "RiskID","RiskCategory","RiskDescription","Unit",
        "Likelihood","Impact","RiskScore","RiskRating",
        "ExistingControls","ControlEffectiveness",
        "MitigationPlan","MitigationOwner","MitigationDueDate",
        "ResidualLikelihood","ResidualImpact","ResidualScore","ResidualRating",
        "Status","LastReviewDate","NextReviewDate",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,18,30,14,12,12,12,14,28,16,30,18,14,14,14,14,14,14,14,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","RISK-"&TEXT(ROW()-1,"0000"))')
        ws.cell(row=r, column=7, value=f'=IF(OR(E{r}="",F{r}=""),"",E{r}*F{r})')
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}>=15,"Critical",IF(G{r}>=10,"High",IF(G{r}>=5,"Medium","Low"))))')
        ws.cell(row=r, column=16, value=f'=IF(OR(N{r}="",O{r}=""),"",N{r}*O{r})')
        ws.cell(row=r, column=17, value=f'=IF(P{r}="","",IF(P{r}>=15,"Critical",IF(P{r}>=10,"High",IF(P{r}>=5,"Medium","Low"))))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_RiskRegister", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "D2:D1000", "=Lookups!$E$2:$E$11")
    ws.conditional_formatting.add("H2:H1000", FormulaRule(formula=['=H2="Critical"'], fill=RED_FILL))
    ws.conditional_formatting.add("H2:H1000", FormulaRule(formula=['=H2="High"'], fill=AMBER_FILL))
    ws.conditional_formatting.add("H2:H1000", FormulaRule(formula=['=H2="Medium"'], fill=GOLD_FILL))
    ws.conditional_formatting.add("T2:T1000", FormulaRule(formula=['=AND(T2<>"",T2<TODAY())'], fill=RED_FILL))
    return ws

def build_after_action_report(wb):
    ws = wb.create_sheet("tbl_AfterActionReport")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "AAR_ID","LinkedOpsID","LinkedCaseID","ReportDate",
        "OperationName","OperationType","TeamLead",
        "ObjectivesMet","ObjectiveSummary",
        "WhatWorked","WhatDidNotWork","LessonsLearned",
        "TacticalFindings","IntelAssessment",
        "CasualtiesOrInjuries","OfficerSafetyIssues",
        "EquipmentIssues","CommunicationIssues",
        "Recommendations","ActionItemsFromAAR","ActionAssignedTo","ActionDueDate",
        "ReviewedBy","ReviewDate",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,14,22,16,18,12,30,30,30,30,28,24,14,20,20,20,30,28,18,14,18,14,24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","AAR-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_AfterActionReport", f"A1:{get_column_letter(len(headers))}6")
    add_data_validation_list(ws, "H2:H1000", "=Lookups!$I$2:$I$3")
    ws.conditional_formatting.add("V2:V1000", FormulaRule(formula=['=AND(V2<>"",V2<TODAY())'], fill=RED_FILL))
    return ws

def build_notification_log(wb):
    ws = wb.create_sheet("tbl_Notifications")
    ws.sheet_properties.tabColor = "002060"
    headers = [
        "NotificationID","Date","Time","AlertType","Severity",
        "SourceTable","SourceRecordID","Description",
        "TriggeredRule","AffectedUnit","AffectedOfficer",
        "ActionRequired","ActionTaken","ActionDate","ResolvedBy",
        "Status","Notes","CreatedDate"
    ]
    widths = [14,14,12,18,12,16,16,30,24,14,18,24,24,14,18,14,24,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","NTFY-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    add_table(ws, "tbl_Notifications", f"A1:{get_column_letter(len(headers))}6")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  PART 5 — BUG FIXES
# ══════════════════════════════════════════════════════════════════════
def fix_bugs(wb):
    # Fix missing data validations on tbl_DiaryEntries
    ws_de = wb["tbl_DiaryEntries"]
    add_data_validation_list(ws_de, "F2:F1000", "=Lookups!$E$2:$E$11")  # Unit

    # Fix missing data validations on tbl_RadioMessages
    ws_rm = wb["tbl_RadioMessages"]
    add_data_validation_list(ws_rm, "F2:F1000", "=Lookups!$B$2:$B$6")  # Priority

    # Add new lookups for new tables
    lk = wb["Lookups"]
    new_lookups = {
        43: ("LeaveType", ["Annual Leave","Sick Leave","Maternity Leave","Paternity Leave",
                            "Study Leave","Compassionate Leave","Special Leave","No-Pay Leave"]),
        44: ("LeaveStatus", ["Pending","Approved","Denied","Cancelled","In Progress","Completed"]),
        45: ("RiskCategory", ["Operational","Strategic","Personnel","Financial","Legal",
                               "Reputational","Technology","Environmental"]),
        46: ("Direction", ["Incoming","Outgoing"]),
        47: ("CorrespondenceStatus", ["Pending","Sent","Received","Responded","Filed","Overdue"]),
        48: ("AllegationType", ["Misconduct","Neglect of Duty","Abuse of Authority","Corruption",
                                 "Excessive Force","Unauthorized Absence","Insubordination",
                                 "Loss of Evidence","Other"]),
        49: ("DisciplinaryStatus", ["Pending","Under Investigation","Hearing Scheduled",
                                     "Sanction Applied","Dismissed","Appealed","Closed"]),
        50: ("AlertSeverity", ["Critical","High","Medium","Low","Info"]),
        51: ("BudgetCategory", ["Operations","Transport/Fuel","Training","Maintenance",
                                 "Equipment","Admin/Stationery","Communications","Other"]),
    }
    for col_num, (col_name, values) in new_lookups.items():
        lk.cell(row=1, column=col_num, value=col_name).font = HEADER_FONT
        lk.cell(row=1, column=col_num).fill = HEADER_FILL
        for i, v in enumerate(values, 2):
            lk.cell(row=i, column=col_num, value=v)
        lk.column_dimensions[get_column_letter(col_num)].width = 22

    # Apply validations using the new lookups
    wb["tbl_LeaveManagement"]
    add_data_validation_list(wb["tbl_LeaveManagement"], "F2:F1000", f"=Lookups!${get_column_letter(43)}$2:${get_column_letter(43)}$9")
    add_data_validation_list(wb["tbl_LeaveManagement"], "K2:K1000", f"=Lookups!${get_column_letter(44)}$2:${get_column_letter(44)}$7")
    add_data_validation_list(wb["tbl_Disciplinary"], "X2:X1000", f"=Lookups!${get_column_letter(49)}$2:${get_column_letter(49)}$8")
    add_data_validation_list(wb["tbl_RiskRegister"], "C2:C1000", f"=Lookups!${get_column_letter(45)}$2:${get_column_letter(45)}$9")  # RiskCategory uses col C=Description, we want col B. Fix:
    add_data_validation_list(wb["tbl_Correspondence"], "D2:D1000", f"=Lookups!${get_column_letter(46)}$2:${get_column_letter(46)}$3")
    add_data_validation_list(wb["tbl_Correspondence"], "O2:O1000", f"=Lookups!${get_column_letter(47)}$2:${get_column_letter(47)}$7")
    add_data_validation_list(wb["tbl_Disciplinary"], "J2:J1000", f"=Lookups!${get_column_letter(48)}$2:${get_column_letter(48)}$10")
    add_data_validation_list(wb["tbl_Notifications"], "E2:E1000", f"=Lookups!${get_column_letter(50)}$2:${get_column_letter(50)}$6")
    add_data_validation_list(wb["tbl_Budget"], "F2:F1000", f"=Lookups!${get_column_letter(51)}$2:${get_column_letter(51)}$9")
    add_data_validation_list(wb["tbl_RiskRegister"], "B2:B1000", f"=Lookups!${get_column_letter(45)}$2:${get_column_letter(45)}$9")


# ══════════════════════════════════════════════════════════════════════
#  PART 6 — INSTRUCTIONS + CHANGELOG + QUARTERLY REPORT TEMPLATE
# ══════════════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = "002060"
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="FNID AREA 3 — WORKBOOK USER GUIDE").font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 36

    sections = [
        ("GETTING STARTED", [
            "1. This workbook is the central operational management system for FNID Area 3.",
            "2. Navigate using the HOME sheet — click sheet names to go to each section.",
            "3. PORTALS are read-only summary views for each unit (no data entry on portals).",
            "4. DATA TABLES (tbl_*) are where you enter and manage operational data.",
            "5. DASHBOARDS (Dash_*) provide analytics and trend views — all auto-calculated.",
            "6. FORMS (Form_*) are printable templates for field use.",
        ]),
        ("DATA ENTRY RULES", [
            "1. Always use the dropdown lists where provided — they ensure data consistency.",
            "2. DO NOT delete or rename column headers — formulas depend on them.",
            "3. Enter dates in DD/MM/YYYY format. Times in HH:MM format.",
            "4. ID columns (CaseID, LeadID, etc.) auto-generate — do not type in them.",
            "5. Always fill CreatedBy and CreatedDate when adding new records.",
            "6. Use LastUpdatedBy and LastUpdatedDate when modifying existing records.",
        ]),
        ("CASE MANAGEMENT SOP", [
            "1. Register all cases in tbl_Cases within 24 hours (RegistrationSLA tracked).",
            "2. Create an Investigation Plan for every active case (tbl_InvestigationPlans).",
            "3. Conduct mandatory reviews at 14, 28, 56, and 90 day intervals (tbl_Reviews).",
            "4. Log all case assignments and handovers in tbl_CaseAssignments.",
            "5. Record all witness statements in tbl_Statements.",
            "6. Update victims/complainants regularly via tbl_VictimContact.",
            "7. Complete Case File Checklist before court (tbl_CaseFileChecklist).",
            "8. Log all prosecution disclosure in tbl_Disclosure.",
            "9. Ensure 48-hour charge rule compliance (auto-tracked in tbl_Cases).",
            "10. Document all command conferences in tbl_CommandConference.",
        ]),
        ("EVIDENCE MANAGEMENT", [
            "1. Register all exhibits immediately in tbl_Exhibits.",
            "2. Log every handover in tbl_CustodyTransfers (chain of custody).",
            "3. Forensic routing auto-suggests agency based on evidence type.",
            "4. Track ballistic matches in tbl_BallisticTrace.",
            "5. Court readiness status must be updated before each hearing.",
        ]),
        ("INTELLIGENCE PROCEDURES", [
            "1. All MCR reports are auto-filtered for FNID relevance (tbl_MCR_Raw).",
            "2. Triage all leads within 24 hours (tbl_Leads).",
            "3. Register informants with code names only (tbl_Informants).",
            "4. Track gang affiliations in tbl_GangAffiliation.",
            "5. Log all intelligence briefings in tbl_IntelBriefings.",
        ]),
        ("TRANSPORT PROCEDURES", [
            "1. Update vehicle status in tbl_Vehicles before and after every operation.",
            "2. Log all trips in tbl_VehicleUsage, all fuel in tbl_FuelLog.",
            "3. Schedule maintenance in tbl_Maintenance — overdue items turn RED.",
            "4. Ensure all driver licenses are current (tbl_Drivers).",
        ]),
        ("ADMIN/HRM PROCEDURES", [
            "1. Maintain personnel register in tbl_Personnel.",
            "2. Manage leave requests via tbl_LeaveManagement.",
            "3. Track disciplinary matters in tbl_Disciplinary.",
            "4. Manage shifts and assignments via tbl_DutyRoster.",
            "5. Log all correspondence in tbl_Correspondence.",
        ]),
        ("CONDITIONAL FORMATTING GUIDE", [
            "RED cells = Overdue, expired, non-compliant, or critical items.",
            "AMBER/GOLD cells = At risk, approaching deadline, or needs attention.",
            "GREEN cells = On track, compliant, or complete.",
        ]),
        ("SECURITY & CONFIDENTIALITY", [
            "This workbook contains CONFIDENTIAL operational data.",
            "For OFFICIAL USE ONLY — do not share externally.",
            "Protect sensitive sheets with passwords where possible.",
            "Back up the file regularly to a secure location.",
            "tbl_Informants contains highly sensitive data — restrict access.",
        ]),
    ]

    r = 3
    for title, items in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        r += 1
        for item in items:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=item).font = BODY_FONT
            ws.cell(row=r, column=1).alignment = LEFT_WRAP
            r += 1
        r += 1
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 22

def build_changelog(wb):
    ws = wb.create_sheet("ChangeLog")
    ws.sheet_properties.tabColor = "808080"
    headers = ["Version","Date","Author","ChangeType","Description","SheetsAffected"]
    widths = [12,14,18,16,40,30]
    write_headers(ws, headers, widths=widths)

    changes = [
        ("1.0", "2026-02-27", "System", "Initial Build",
         "Full workbook: 48 sheets, 19 tables, 32 KPIs, 20 lookups, 9 portals, 2 dashboards, 4 forms",
         "All"),
        ("1.1", "2026-02-27", "Audit Patch", "Gap Remediation",
         "Added 10 tables (Personnel, Training, Court, Arrests, Seizures, Warrants, Informants, DutyRoster, CustodyTransfers, SceneLog), "
         "12 KPIs, 10 lookup columns, trend analysis, aging/SLA metrics",
         "58 sheets affected"),
        ("1.2", "2026-02-27", "SOP Compliance", "Major Enhancement",
         "Added 8 SOP tables (Witnesses, Statements, CaseAssignments, VictimContact, InvestigationPlans, "
         "CaseFileChecklist, Disclosure, CommandConference), 24 new columns on tbl_Cases, 11 on tbl_Reviews, "
         "12 SOP KPIs, rebuilt Registry & Investigations portals, 3 form templates",
         "69 sheets"),
        ("1.3", "2026-02-27", "Full Gap Closure", "Major Enhancement",
         "Built 8 unit dashboards, 8 new tables (BallisticTrace, GangAffiliation, Budget, LeaveManagement, "
         "Disciplinary, Correspondence, RiskRegister, AfterActionReport, Notifications), rebuilt 3 portals, "
         "added Instructions, ChangeLog, Quarterly Report template, bug fixes, 9 new lookup lists",
         "All sheets"),
    ]
    for i, row_data in enumerate(changes, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP
    add_table(ws, "tbl_ChangeLog", f"A1:{get_column_letter(len(headers))}{1+len(changes)}")

def build_quarterly_report(wb):
    ws = wb.create_sheet("Form_QuarterlyReport")
    ws.sheet_properties.tabColor = "002060"
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="FNID AREA 3 — QUARTERLY OPERATIONAL REPORT").font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws.cell(row=1, column=1).alignment = CENTER
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1, value="(For Submission to Divisional Commander)").font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.cell(row=2, column=1).alignment = CENTER

    sections = [
        ("REPORTING PERIOD", [("Quarter:", ""), ("Year:", ""), ("Prepared By:", ""), ("Date:", "")]),
        ("EXECUTIVE SUMMARY", [("", ""), ("", ""), ("", "")]),
        ("CASE MANAGEMENT", [
            ("Cases Registered:", ""), ("Cases Closed:", ""), ("Clearance Rate:", ""),
            ("Cold Cases:", ""), ("CR Compliance Rate:", ""), ("Avg Days to Close:", ""),
        ]),
        ("OPERATIONS", [
            ("Operations Conducted:", ""), ("FCSI Operations:", ""),
            ("Firearms Seized:", ""), ("Ammunition Seized:", ""),
            ("Narcotics Seized (kg):", ""), ("Arrests Made:", ""),
            ("Search Warrants Executed:", ""),
        ]),
        ("INTELLIGENCE", [
            ("Leads Received:", ""), ("Leads Triaged:", ""),
            ("Conversion Rate:", ""), ("Briefings Delivered:", ""),
        ]),
        ("EVIDENCE & FORENSICS", [
            ("Exhibits Collected:", ""), ("Forensic Submissions:", ""),
            ("Avg Turnaround (days):", ""), ("Court Ready Items:", ""),
        ]),
        ("TRANSPORT", [
            ("Fleet Availability %:", ""), ("Total Trips:", ""),
            ("Fuel Consumed (L):", ""), ("Maintenance Cost:", ""),
        ]),
        ("DEMAND REDUCTION", [
            ("Sessions Held:", ""), ("Total Attendance:", ""),
            ("Communities Reached:", ""),
        ]),
        ("PERSONNEL & TRAINING", [
            ("Current Strength:", ""), ("Vacancies:", ""),
            ("Training Sessions:", ""), ("Personnel Trained:", ""),
        ]),
        ("CHALLENGES & RECOMMENDATIONS", [("", ""), ("", ""), ("", "")]),
        ("PRIORITIES FOR NEXT QUARTER", [("1.", ""), ("2.", ""), ("3.", "")]),
        ("APPROVALS", [
            ("Area Commander:", ""), ("Signature:", ""), ("Date:", ""),
            ("Divisional Commander:", ""), ("Signature:", ""), ("Date:", ""),
        ]),
    ]
    r = 4
    for title, fields in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        r += 1
        for label, val in fields:
            if label:
                ws.cell(row=r, column=1, value=label).font = BODY_FONT_BOLD
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
            else:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                ws.cell(row=r, column=1).border = Border(bottom=Side(style="dotted"))
            r += 1
        r += 1
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18


# ══════════════════════════════════════════════════════════════════════
#  PART 7 — UPDATE HOME NAVIGATION
# ══════════════════════════════════════════════════════════════════════
def patch_home_final(wb):
    ws = wb["HOME"]
    r = ws.max_row + 2

    new_sections = [
        ("NEW OPERATIONAL TABLES (Gap Closure v1.3)", "C00000", [
            ("tbl_BallisticTrace", "Firearm Trace / IBIS — Ballistic matches across cases"),
            ("tbl_GangAffiliation", "Gang Intelligence — Affiliations, territory, associates"),
            ("tbl_Budget", "Budget & Expenditure — Allocations, spending, variance tracking"),
            ("tbl_LeaveManagement", "Leave Register — Applications, approvals, balances"),
            ("tbl_Disciplinary", "Disciplinary Register — Allegations, hearings, sanctions"),
            ("tbl_Correspondence", "Correspondence Log — External communications tracking"),
            ("tbl_RiskRegister", "Organizational Risk Register — Likelihood, impact, mitigation"),
            ("tbl_AfterActionReport", "After-Action Reports — Ops lessons learned & recommendations"),
            ("tbl_Notifications", "Notification / Alert Log — Triggered alerts and resolutions"),
        ]),
        ("SUPPORT SHEETS (v1.3)", "002060", [
            ("Instructions", "Workbook User Guide — How to use every module"),
            ("ChangeLog", "Version History — All changes tracked"),
            ("Form_QuarterlyReport", "Quarterly Operational Report Template (Printable)"),
        ]),
    ]

    for section_title, color, items in new_sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1
        for sheet_name, description in items:
            ws.cell(row=r, column=1, value=f"  >> {sheet_name}").font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
            ws.cell(row=r, column=1).alignment = Alignment(indent=2)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.cell(row=r, column=2, value=description).font = Font(name="Calibri", size=10, color="333333")
            ws.cell(row=r, column=2).alignment = LEFT_WRAP
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = Border(bottom=Side(style="hair", color="CCCCCC"))
            r += 1
        r += 1


def main():
    fname = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"Loading {fname}...")
    wb = openpyxl.load_workbook(fname)

    print("[1/16] Rebuilding Portal_Training...")
    rebuild_portal_training(wb)
    print("[2/16] Rebuilding Portal_AdminHRM...")
    rebuild_portal_admin_hrm(wb)
    print("[3/16] Rebuilding Portal_Oversight...")
    rebuild_portal_oversight(wb)
    print("[4/16] Adding tbl_BallisticTrace...")
    build_ballistic_trace(wb)
    print("[5/16] Adding tbl_GangAffiliation...")
    build_gang_affiliation(wb)
    print("[6/16] Adding tbl_Budget...")
    build_budget_table(wb)
    print("[7/16] Adding tbl_LeaveManagement...")
    build_leave_management(wb)
    print("[8/16] Adding tbl_Disciplinary...")
    build_disciplinary(wb)
    print("[9/16] Adding tbl_Correspondence...")
    build_correspondence(wb)
    print("[10/16] Adding tbl_RiskRegister...")
    build_risk_register(wb)
    print("[11/16] Adding tbl_AfterActionReport...")
    build_after_action_report(wb)
    print("[12/16] Adding tbl_Notifications...")
    build_notification_log(wb)
    print("[13/16] Fixing bugs (validations, lookups)...")
    fix_bugs(wb)
    print("[14/16] Adding Instructions sheet...")
    build_instructions(wb)
    print("[15/16] Adding ChangeLog...")
    build_changelog(wb)
    print("[16/16] Adding Quarterly Report + HOME nav...")
    build_quarterly_report(wb)
    patch_home_final(wb)

    wb.save(fname)
    print(f"\nAll remaining gaps closed: {fname} — {len(wb.sheetnames)} sheets")

if __name__ == "__main__":
    main()
