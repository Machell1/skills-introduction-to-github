#!/usr/bin/env python3
"""
FNID Area 3 Operational Workbook — Version 2.0 (Purpose-Driven Redesign)

Built from research on FNID's actual day-to-day operations:
  - Intelligence-led interdiction of illegal firearms AND narcotics
  - Port/cargo interdiction with Jamaica Customs Agency
  - Roadway interceptions (Special Operations Team)
  - Drugs-for-guns pipeline disruption (Jamaica→Haiti)
  - Case file preparation for DPP submission to Gun Court

Lifecycle:  INTELLIGENCE → OPERATION → SEIZURE → ARREST → FORENSICS → DPP → COURT

Legal framework:
  - Firearms (Prohibition, Restriction and Regulation) Act, 2022 (s.91 warrants)
  - Dangerous Drugs Act (as amended 2015)
  - Gun Court Act, 1974
  - Proceeds of Crime Act (POCA), 2007
  - Bail Act, 2023
  - 48-Hour Rule (Constabulary Force Act s.15)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ═══════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════
NAVY       = "1F3864"
DARK_BLUE  = "002060"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
MED_GRAY   = "808080"
OPS_RED    = "C00000"
INTEL_GOLD = "BF8F00"
EVIDENCE_GREEN = "538135"
COURT_PURPLE   = "7030A0"
NARCO_TEAL     = "00B0F0"

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT   = Font(name="Calibri", size=10)
BODY_BOLD   = Font(name="Calibri", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL  = PatternFill("solid", fgColor="FFE699")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
GOLD_FILL   = PatternFill("solid", fgColor="FFD700")

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False)


def write_headers(ws, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if widths and c <= len(widths):
            ws.column_dimensions[get_column_letter(c)].width = widths[c-1]
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

def add_table(ws, name, ref):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TABLE_STYLE
    ws.add_table(t)

def dv_list(ws, cell_range, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Select from dropdown"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(cell_range)

def data_rows(ws, num_cols, count=5):
    for r in range(2, 2+count):
        for c in range(1, num_cols+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT


# ═══════════════════════════════════════════════════════════════
#  LOOKUPS — Jamaica-specific reference data
# ═══════════════════════════════════════════════════════════════
def build_lookups(wb):
    ws = wb.create_sheet("Lookups")
    ws.sheet_properties.tabColor = MED_GRAY

    lists = {
        1: ("IntelSource", [
            "Crime Stop (311)", "NIB (811)", "Informant", "DEA", "ATF",
            "HSI", "MOCA", "JCA Detection", "Divisional CIB", "Patrol Intercept",
            "C-TOC/SIB", "INTERPOL", "Walk-In Report", "FNID Direct (923-6184)",
            "UK NCA", "US Marshals", "Anonymous"]),
        2: ("IntelPriority", ["Critical", "High", "Medium", "Low"]),
        3: ("TriageDecision", [
            "Action — Mount Operation", "Action — Surveillance", "Action — Port Alert",
            "Refer to Divisional CIB", "Refer to MOCA", "Refer to SIB",
            "Intel Filed — Monitor", "Closed — Insufficient Info"]),
        4: ("OperationType", [
            "Search Warrant Execution", "Snap Raid", "Port/Cargo Interdiction",
            "Airport Interdiction", "Vehicle Interception", "Coastal/Beach Operation",
            "Checkpoint", "Surveillance", "Joint Op — JCA", "Joint Op — JDF",
            "Joint Op — MOCA", "Joint Op — DEA", "Fugitive Apprehension",
            "Follow-Up Secondary Op", "ZOSO Operation", "SOE Operation",
            "Courier/Parcel Investigation"]),
        5: ("WarrantBasis", [
            "Section 91 — Firearms Act 2022", "Section 21 — Dangerous Drugs Act",
            "Section 14 — POCA 2007", "ZOSO — Warrantless (s.5 ZOSO Act)",
            "SOE — Warrantless", "Sergeant Written Directive (DDA s.21)"]),
        6: ("FirearmType", [
            "Pistol — Semi-Auto", "Pistol — Revolver", "Rifle — Semi-Auto",
            "Rifle — Bolt Action", "Rifle — Assault", "Shotgun", "Submachine Gun",
            "Machine Pistol", "Improvised/Homemade", "3D-Printed", "Imitation/Replica",
            "Component — Frame/Receiver", "Component — Barrel", "Component — Slide",
            "Magazine", "Silencer/Suppressor"]),
        7: ("FirearmCalibre", [
            "9mm", ".40 S&W", ".45 ACP", ".380 ACP", ".22 LR", ".25 ACP",
            ".38 Special", ".357 Magnum", "5.56mm/.223", "7.62x39mm",
            "7.62x51mm/.308", "12 Gauge", "20 Gauge", ".44 Magnum", "Other"]),
        8: ("DrugType", [
            "Cannabis/Ganja — Compressed", "Cannabis/Ganja — Loose",
            "Cannabis/Ganja — Oil/Edibles", "Cocaine — Powder",
            "Cocaine — Crack", "Heroin", "MDMA/Ecstasy", "Methamphetamine",
            "Synthetic Cannabinoids", "Prescription Opioids", "Other Controlled"]),
        9: ("DrugUnit", ["kg", "g", "lbs", "oz", "plants", "tablets", "ml", "litres"]),
        10: ("FirearmsActOffence", [
            "s.5 — Possession of Prohibited Weapon (15-25yr)",
            "s.6 — Stockpiling (3+ weapons / 50+ rounds)",
            "s.7 — Trafficking in Prohibited Weapon (20yr min)",
            "s.8 — Possession with Intent to Traffic",
            "s.9 — Manufacture of Prohibited Weapon",
            "s.10 — Dealing in Prohibited Weapon",
            "s.12 — Diversion of Lawful Firearms",
            "s.29 — Unmarked Firearm",
            "Shooting with Intent", "Illegal Possession of Ammunition",
            "Wounding with Intent (firearm)"]),
        11: ("DDAOffence", [
            "Import/Export — Cannabis (DDA Part IIIA)",
            "Import/Export — Cocaine (DDA Part IV)",
            "Import/Export — Heroin (DDA Part IV)",
            "Dealing/Trafficking — Cannabis",
            "Dealing/Trafficking — Cocaine",
            "Dealing/Trafficking — Heroin",
            "Possession — Cannabis (>2oz)",
            "Possession — Cocaine",
            "Cultivation — Cannabis (>5 plants)",
            "Cultivation — Coca/Opium",
            "Deemed Dealing — School Premises",
            "Deemed Dealing — Heroin >1/10oz"]),
        12: ("CourtType", [
            "Gun Court — High Court Division (judge alone, in camera)",
            "Gun Court — Circuit Court Division (jury, murder/treason)",
            "Gun Court — RM Division (preliminary exam)",
            "Parish Court (summary drug offence)",
            "Circuit Court (indictable drug offence)",
            "Supreme Court"]),
        13: ("BailStatus", [
            "Bail Granted", "Bail Denied", "Remanded in Custody",
            "Released — No Charge", "Released — Insufficient Evidence",
            "Stop Order Issued", "Electronic Monitoring"]),
        14: ("DPPStatus", [
            "File Being Prepared", "File Submitted to DPP",
            "Awaiting Forensic Certificate", "Awaiting Ballistic Certificate",
            "Crown Counsel Reviewing", "Ruling — Charge Approved",
            "Ruling — No Charge (Insufficient Evidence)",
            "Voluntary Bill of Indictment", "Preliminary Exam Ordered",
            "Returned for Further Investigation"]),
        15: ("ForensicStatus", [
            "Not Yet Submitted", "Submitted to IFSLM", "Analysis In Progress",
            "Certificate Issued", "Certificate Overdue", "Inconclusive",
            "Returned to FNID"]),
        16: ("Parish", [
            "Kingston", "St. Andrew", "St. Thomas", "Portland", "St. Mary",
            "St. Ann", "Trelawny", "St. James", "Hanover", "Westmoreland",
            "St. Elizabeth", "Manchester", "Clarendon", "St. Catherine"]),
        17: ("YesNo", ["Yes", "No"]),
        18: ("OpOutcome", [
            "Successful — Seizure + Arrest", "Successful — Seizure Only",
            "Successful — Arrest Only", "Partial — Intel Developed",
            "Negative — No Finds", "Aborted — Compromised",
            "Aborted — Safety Concern", "Ongoing"]),
        19: ("SeizureLocation", [
            "Port Bustamante", "Kingston Logistics Centre", "Newport West",
            "Norman Manley Airport", "Sangster International Airport",
            "Ian Fleming Airport", "Roadway/Vehicle", "Residential Premises",
            "Commercial Premises", "Beach/Coastal", "Open Land/Bushes",
            "Courier/Postal Facility", "ZOSO Checkpoint", "Other"]),
        20: ("IBIS_Status", [
            "Not Submitted", "Submitted — Pending", "No Match",
            "Hit — Confirmed Match", "Hit — Unconfirmed", "Serial Obliterated"]),
        21: ("eTrace_Status", [
            "Not Submitted", "Submitted — Pending", "Traced — US Origin",
            "Traced — Non-US Origin", "Untraceable — No Serial",
            "Untraceable — Obliterated", "Trace Complete"]),
        22: ("ExhibitDisposal", [
            "Held — Active Case", "Held — Court Order",
            "Destroyed — Authorized", "Returned to Owner",
            "Forfeited — POCA", "Pending Disposal Authorization"]),
        23: ("POCAStatus", [
            "Not Applicable", "Referred to FID", "Restraint Order Applied",
            "Restraint Order Granted", "Civil Recovery Filed",
            "Forfeiture Order — Post Conviction",
            "Consent Order", "Dismissed"]),
        24: ("CaseClassification", [
            "Firearms", "Narcotics", "Firearms + Narcotics",
            "Trafficking — Firearms", "Trafficking — Narcotics",
            "Trafficking — Multi-Commodity", "POCA / Financial",
            "Murder with Firearm", "Shooting with Intent"]),
        25: ("CaseStatus", [
            "Open — Active Investigation", "Open — Pending Forensics",
            "Open — Pending Witness Statements", "Open — Surveillance",
            "Open — Pending DPP Submission", "Referred to DPP",
            "Closed — Convicted", "Closed — Acquitted",
            "Closed — No Charge (Insufficient Evidence)",
            "Closed — Withdrawn", "Cold Case — Under Review",
            "Cold Case — Dormant", "Merged — See Linked Case"]),
        26: ("InvestigationType", [
            "Murder Investigation (Firearm)", "Illegal Possession of Firearm",
            "Trafficking — Firearms", "Trafficking — Narcotics",
            "Shooting with Intent", "Drug Dealing / Distribution",
            "Import/Export — Narcotics", "Import/Export — Firearms",
            "POCA / Financial Investigation", "Stockpiling (s.6)",
            "Manufacture / Assembly", "Courier / Parcel Intercept",
            "Guns-for-Drugs Pipeline", "Gang / Network Investigation"]),
    }

    for col, (name, values) in lists.items():
        ws.cell(row=1, column=col, value=name).font = HEADER_FONT
        ws.cell(row=1, column=col).fill = HEADER_FILL
        for i, v in enumerate(values, 2):
            ws.cell(row=i, column=col, value=v)
        ws.column_dimensions[get_column_letter(col)].width = max(20, len(name)+4)
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 1: INTELLIGENCE LOG
#  Where tips/intel come in — Crime Stop 311, NIB 811, DEA, informants, JCA
# ═══════════════════════════════════════════════════════════════
def build_intel_log(wb):
    ws = wb.create_sheet("tbl_IntelLog")
    ws.sheet_properties.tabColor = INTEL_GOLD
    headers = [
        "IntelID", "DateReceived", "TimeReceived", "Source",
        "SourceRef", "Priority", "SubjectMatter",
        "FirearmsRelated", "NarcoticsRelated", "TraffickingRelated",
        "TargetPerson", "TargetLocation", "Parish",
        "SubstanceOfIntel", "TriageDecision", "TriageBy", "TriageDate",
        "LinkedOpID", "LinkedCaseID",
        "Outcome", "OutcomeNotes",
        "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,10,22,14,10,20,10,10,10,18,22,14,36,28,16,12,12,12,22,30,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","INT-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_IntelLog", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "D2:D1000", "=Lookups!$A$2:$A$18")
    dv_list(ws, "F2:F1000", "=Lookups!$B$2:$B$5")
    dv_list(ws, "H2:H1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "I2:I1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "J2:J1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "M2:M1000", "=Lookups!$P$2:$P$15")
    dv_list(ws, "O2:O1000", "=Lookups!$C$2:$C$9")
    ws.conditional_formatting.add("F2:F1000", CellIsRule(operator="equal", formula=['"Critical"'], fill=RED_FILL))
    ws.conditional_formatting.add("F2:F1000", CellIsRule(operator="equal", formula=['"High"'], fill=AMBER_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 2: OPERATION REGISTER
#  Every planned operation — warrants, teams, outcomes
# ═══════════════════════════════════════════════════════════════
def build_operation_register(wb):
    ws = wb.create_sheet("tbl_Operations")
    ws.sheet_properties.tabColor = OPS_RED
    headers = [
        "OpsID", "OpsName", "OpsDate", "OpsType",
        "LinkedIntelID", "Parish", "Location",
        # Warrant & Authorization
        "WarrantObtained", "WarrantBasis", "WarrantNumber", "IssuingJP",
        "AuthorizingOfficer", "AuthorizingRank",
        # Team
        "TeamLeader", "TeamLeaderRank", "TeamSize",
        "InspectorPresent", "BodyCamActivated",
        # Execution
        "StartTime", "EndTime", "Outcome",
        # Seizure Summary
        "FirearmsSeized", "AmmunitionSeized", "NarcoticsSeized_kg",
        "CashSeized_JMD", "VehiclesSeized", "OtherExhibits",
        "ArrestsMade",
        # Post-Op
        "AfterActionCompleted", "AfterActionNotes",
        "INCOMPlianceFlags",
        "CreatedBy", "CreatedDate"
    ]
    widths = [12,22,12,26,12,14,24,10,30,14,18,20,14,18,14,10,10,10,10,10,22,10,10,12,14,10,14,10,10,30,28,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","OPS-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_Operations", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "D2:D1000", "=Lookups!$D$2:$D$18")
    dv_list(ws, "F2:F1000", "=Lookups!$P$2:$P$15")
    dv_list(ws, "H2:H1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "I2:I1000", "=Lookups!$E$2:$E$7")
    dv_list(ws, "Q2:Q1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "R2:R1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "U2:U1000", "=Lookups!$R$2:$R$9")
    dv_list(ws, "AC2:AC1000", "=Lookups!$Q$2:$Q$3")
    # INDECOM compliance: flag if no inspector present
    ws.conditional_formatting.add("Q2:Q1000", CellIsRule(operator="equal", formula=['"No"'], fill=RED_FILL))
    ws.conditional_formatting.add("R2:R1000", CellIsRule(operator="equal", formula=['"No"'], fill=AMBER_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 3: FIREARM SEIZURE LOG
#  Every firearm recovered — IBIS, eTrace, ballistic cert tracking
# ═══════════════════════════════════════════════════════════════
def build_firearm_seizures(wb):
    ws = wb.create_sheet("tbl_FirearmSeizures")
    ws.sheet_properties.tabColor = OPS_RED
    headers = [
        "SeizureID", "LinkedOpsID", "LinkedCaseID", "SeizureDate",
        "SeizureLocation", "Parish",
        # Firearm Details
        "FirearmType", "Make", "Model", "Calibre",
        "SerialNumber", "SerialStatus",
        "Condition", "Loaded", "RoundsChamber", "MagazineCapacity",
        # IBIS / Ballistics
        "IBIS_Submitted", "IBIS_SubmitDate", "IBIS_Status",
        "IBIS_MatchCaseID", "IBIS_MatchDetails",
        # ATF eTrace
        "eTrace_Submitted", "eTrace_Status", "eTrace_TraceNumber",
        "CountryOfOrigin", "LastKnownPurchaser", "US_State",
        # Ballistic Certificate
        "BallisticCertStatus", "BallisticCertDate", "BallisticExaminer",
        "BallisticResult",
        # Court Readiness
        "ForensicRecoveryBag", "EvidencePhotographed", "MarkedByCommissioner",
        "CourtReady",
        # Disposal
        "DisposalStatus",
        "SeizingOfficer", "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,12,12,22,14,20,16,16,12,16,14,12,8,8,10,
              10,12,18,14,24,10,18,16,16,20,14,16,12,18,24,10,10,10,10,16,16,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","FS-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_FirearmSeizures", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "E2:E1000", "=Lookups!$S$2:$S$15")
    dv_list(ws, "F2:F1000", "=Lookups!$P$2:$P$15")
    dv_list(ws, "G2:G1000", "=Lookups!$F$2:$F$17")
    dv_list(ws, "J2:J1000", "=Lookups!$G$2:$G$16")
    dv_list(ws, "S2:S1000", "=Lookups!$T$2:$T$7")
    dv_list(ws, "W2:W1000", "=Lookups!$U$2:$U$8")
    dv_list(ws, "AB2:AB1000", "=Lookups!$O$2:$O$8")
    dv_list(ws, "AF2:AF1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "AG2:AG1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "AH2:AH1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "AI2:AI1000", "=Lookups!$V$2:$V$7")
    # Highlight if ballistic cert overdue (submitted >90 days, no cert)
    ws.conditional_formatting.add("AB2:AB1000",
        CellIsRule(operator="equal", formula=['"Certificate Overdue"'], fill=RED_FILL))
    # Highlight IBIS hits
    ws.conditional_formatting.add("S2:S1000",
        FormulaRule(formula=['=LEFT(S2,3)="Hit"'], fill=GOLD_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 4: NARCOTICS SEIZURE LOG
#  Every drug seizure — field test, IFSLM submission, forensic cert
# ═══════════════════════════════════════════════════════════════
def build_narcotics_seizures(wb):
    ws = wb.create_sheet("tbl_NarcoticSeizures")
    ws.sheet_properties.tabColor = NARCO_TEAL
    headers = [
        "SeizureID", "LinkedOpsID", "LinkedCaseID", "SeizureDate",
        "SeizureLocation", "Parish",
        # Drug Details
        "DrugType", "Quantity", "Unit", "PackagingDesc",
        "EstStreetValue_JMD", "Concealment",
        # Field Test
        "FieldTestConducted", "FieldTestResult",
        # IFSLM Forensic
        "SubmittedToIFSLM", "IFSLM_SubmitDate", "IFSLM_LabNumber",
        "ForensicCertStatus", "ForensicCertDate", "ForensicResult",
        "AnalystName",
        # Guns-for-Drugs Nexus
        "LinkedToGunTrade", "EstEquivFirearms",
        # Disposal
        "DisposalAuthorized", "DisposalAuthorizedBy", "DisposalDate",
        "DisposalWitnesses",
        # Tracking
        "CourtReady",
        "SeizingOfficer", "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,12,12,22,14,24,10,8,20,14,22,10,14,10,12,14,16,12,20,16,10,10,10,18,12,18,10,16,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","NS-"&TEXT(ROW()-1,"0000"))')
        # Auto-calc equivalent firearms (50 lbs ganja = 1 handgun)
        ws.cell(row=r, column=23, value=f'=IF(AND(V{r}="Yes",I{r}="lbs"),ROUND(H{r}/50,0),"")')
    add_table(ws, "tbl_NarcoticSeizures", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "E2:E1000", "=Lookups!$S$2:$S$15")
    dv_list(ws, "F2:F1000", "=Lookups!$P$2:$P$15")
    dv_list(ws, "G2:G1000", "=Lookups!$H$2:$H$12")
    dv_list(ws, "I2:I1000", "=Lookups!$I$2:$I$9")
    dv_list(ws, "M2:M1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "O2:O1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "R2:R1000", "=Lookups!$O$2:$O$8")
    dv_list(ws, "V2:V1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "X2:X1000", "=Lookups!$Q$2:$Q$3")
    ws.conditional_formatting.add("R2:R1000",
        CellIsRule(operator="equal", formula=['"Certificate Overdue"'], fill=RED_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 5: AMMUNITION & OTHER SEIZURES
# ═══════════════════════════════════════════════════════════════
def build_other_seizures(wb):
    ws = wb.create_sheet("tbl_OtherSeizures")
    ws.sheet_properties.tabColor = OPS_RED
    headers = [
        "SeizureID", "LinkedOpsID", "LinkedCaseID", "SeizureDate",
        "ItemType", "Description", "Quantity",
        "Calibre", "SerialNumber",
        "SeizureLocation", "Parish",
        "ForensicSubmitted", "ForensicStatus",
        "CourtReady", "DisposalStatus",
        "EstValue_JMD", "POCARelevant",
        "SeizingOfficer", "Notes", "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,12,12,18,28,10,12,16,22,14,10,16,10,16,14,10,16,28,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","OS-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_OtherSeizures", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "F2:F1000", "=Lookups!$P$2:$P$15")
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 6: ARREST REGISTER
#  48hr rule, charge, caution, bail, Gun Court / Parish Court routing
# ═══════════════════════════════════════════════════════════════
def build_arrest_register(wb):
    ws = wb.create_sheet("tbl_Arrests")
    ws.sheet_properties.tabColor = OPS_RED
    headers = [
        "ArrestID", "LinkedOpsID", "LinkedCaseID", "ArrestDate", "ArrestTime",
        "ArrestLocation", "Parish",
        # Suspect
        "SuspectName", "DOB", "Gender", "Address",
        "NationalID_TRN", "PriorConvictions",
        # Caution & Charge
        "CautionAdministered", "CautionTime",
        "ChargeDate", "ChargeTime",
        "PrimaryOffence", "SecondaryOffence",
        "ActSection",
        # 48-Hour Rule
        "48hr_Deadline", "FirstCourtAppearance", "48hr_Met",
        # Court Routing
        "CourtType", "CourtLocation",
        # Bail
        "BailStatus", "BailAmount_JMD", "BailConditions",
        "RemandFacility", "StopOrderIssued",
        # POCA
        "POCAStatus", "POCARef",
        # Tracking
        "ArrestingOfficer", "Notes", "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,12,12,10,22,14,22,12,8,24,16,10,10,10,12,10,30,30,28,14,14,10,36,18,14,14,24,18,10,14,14,16,28,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","ARR-"&TEXT(ROW()-1,"0000"))')
        # Auto-calc 48hr deadline
        ws.cell(row=r, column=21, value=f'=IF(D{r}="","",D{r}+2)')
        ws.cell(row=r, column=21).number_format = 'DD/MM/YYYY HH:MM'
        # Auto-check 48hr met
        ws.cell(row=r, column=23, value=f'=IF(OR(D{r}="",V{r}=""),"",IF(V{r}<=U{r},"Yes","BREACH"))')
    add_table(ws, "tbl_Arrests", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "G2:G1000", "=Lookups!$P$2:$P$15")
    dv_list(ws, "N2:N1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "X2:X1000", "=Lookups!$L$2:$L$7")
    dv_list(ws, "Z2:Z1000", "=Lookups!$M$2:$M$8")
    dv_list(ws, "AE2:AE1000", "=Lookups!$W$2:$W$9")
    # RED if 48hr breached
    ws.conditional_formatting.add("W2:W1000",
        CellIsRule(operator="equal", formula=['"BREACH"'], fill=RED_FILL))
    ws.conditional_formatting.add("W2:W1000",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=GREEN_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 7: EXHIBIT CHAIN OF CUSTODY
# ═══════════════════════════════════════════════════════════════
def build_chain_of_custody(wb):
    ws = wb.create_sheet("tbl_ChainOfCustody")
    ws.sheet_properties.tabColor = EVIDENCE_GREEN
    headers = [
        "TransferID", "ExhibitRef", "ExhibitType", "ExhibitDescription",
        "LinkedCaseID",
        "TransferDate", "TransferTime",
        "ReleasedBy", "ReleasedByRank",
        "ReceivedBy", "ReceivedByRank",
        "Location_From", "Location_To",
        "Reason", "ForensicRecoveryBag",
        "ConditionOnTransfer",
        "SignedReceiptObtained",
        "Notes", "CreatedDate"
    ]
    widths = [12,14,16,28,12,12,10,18,14,18,14,22,22,22,10,20,10,28,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(F{r}="","","COC-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_ChainOfCustody", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "O2:O1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "Q2:Q1000", "=Lookups!$Q$2:$Q$3")
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 8: CASE FILE REGISTRY & INVESTIGATIONS
#  Master register of every FNID case — from opening through
#  full investigation lifecycle to resolution/closure
# ═══════════════════════════════════════════════════════════════
REGISTRY_BLUE = "2E75B6"

def build_case_registry(wb):
    ws = wb.create_sheet("tbl_CaseRegistry")
    ws.sheet_properties.tabColor = REGISTRY_BLUE
    headers = [
        # Registration
        "CaseRegID", "CaseFileNumber", "DateOpened", "CaseClassification",
        "CasePriority", "CaseStatus",
        # Investigation Type & Offence
        "InvestigationType", "PrimaryOffence", "ActSection",
        # Linked Records
        "LinkedIntelIDs", "LinkedOpsIDs", "LinkedArrestIDs",
        "LinkedFirearmSeizureIDs", "LinkedNarcoticSeizureIDs",
        "LinkedDPPCaseFileID",
        # Assignment
        "InvestigatingOfficer", "InvestigatingOfficerRank",
        "Supervisor", "DateAssigned",
        # Location / Target
        "Parish", "Division", "TargetPerson",
        "TargetOrganisation", "GangAffiliation",
        # Investigation Milestones (Yes/No checkboxes)
        "SceneExamComplete", "WitnessStatementsTaken",
        "ForensicEvidenceSubmitted", "ForensicResultsReceived",
        "SurveillanceConducted", "PhoneRecordsObtained",
        "FinancialRecordsObtained", "CCTVFootageObtained",
        "DigitalForensicsComplete",
        # Progress Tracking
        "InvestigationProgress", "LastActivityDate",
        "NextActionRequired", "NextActionDeadline",
        # Supervisor Review
        "LastReviewDate", "ReviewedBy", "ReviewNotes",
        # Escalation
        "Escalated", "EscalatedTo", "EscalationReason",
        # Cold Case
        "ColdCaseReviewDate", "ColdCaseReviewedBy",
        # Resolution
        "ResolutionDate", "ResolutionSummary",
        # Tracking
        "DaysOpen", "Notes", "CreatedBy", "CreatedDate"
    ]
    widths = [
        12, 16, 12, 22,       # Registration
        10, 28,                # Priority, Status
        28, 30, 28,            # Investigation Type & Offence
        14, 14, 14, 14, 14, 14,  # Linked Records
        18, 14, 18, 12,        # Assignment
        14, 16, 22, 22, 18,    # Location / Target
        10, 10, 10, 10, 10, 10, 10, 10, 10,  # Milestones
        12, 12, 28, 14,        # Progress Tracking
        12, 16, 28,            # Supervisor Review
        10, 18, 24,            # Escalation
        12, 16,                # Cold Case
        12, 30,                # Resolution
        10, 28, 14, 12         # Tracking
    ]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))

    for r in range(2, 7):
        # Auto-generate CaseRegID
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","CR-"&TEXT(ROW()-1,"0000"))')
        # Auto-calc DaysOpen
        ws.cell(row=r, column=headers.index("DaysOpen")+1,
                value=f'=IF(C{r}="","",IF(ISBLANK(AR{r}),TODAY()-C{r},AR{r}-C{r}))')
        # Auto-calc InvestigationProgress (9 milestone checkboxes)
        prog_col = headers.index("InvestigationProgress") + 1
        ws.cell(row=r, column=prog_col, value=(
            f'=IF(C{r}="","",('
            f'(Z{r}="Yes")+(AA{r}="Yes")+(AB{r}="Yes")+(AC{r}="Yes")+'
            f'(AD{r}="Yes")+(AE{r}="Yes")+(AF{r}="Yes")+(AG{r}="Yes")+'
            f'(AH{r}="Yes"))/9)'))
        ws.cell(row=r, column=prog_col).number_format = '0%'

    add_table(ws, "tbl_CaseRegistry", f"A1:{get_column_letter(len(headers))}6")

    # Data validations — dropdowns
    dv_list(ws, "D2:D1000", "=Lookups!$X$2:$X$10")    # CaseClassification
    dv_list(ws, "E2:E1000", "=Lookups!$B$2:$B$5")      # Priority (reuse IntelPriority)
    dv_list(ws, "F2:F1000", "=Lookups!$Y$2:$Y$14")     # CaseStatus
    dv_list(ws, "G2:G1000", "=Lookups!$Z$2:$Z$15")     # InvestigationType
    dv_list(ws, "T2:T1000", "=Lookups!$P$2:$P$15")     # Parish
    # Milestone Yes/No validations
    for col_letter in ["Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH"]:
        dv_list(ws, f"{col_letter}2:{col_letter}1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "AL2:AL1000", "=Lookups!$Q$2:$Q$3")    # Escalated

    # Conditional formatting — Case Status
    ws.conditional_formatting.add("F2:F1000",
        CellIsRule(operator="beginsWith", formula=['"Open"'], fill=AMBER_FILL))
    ws.conditional_formatting.add("F2:F1000",
        CellIsRule(operator="beginsWith", formula=['"Cold Case"'], fill=RED_FILL))
    ws.conditional_formatting.add("F2:F1000",
        CellIsRule(operator="beginsWith", formula=['"Closed — Convicted"'], fill=GREEN_FILL))
    # Priority — Critical/High highlights
    ws.conditional_formatting.add("E2:E1000",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=RED_FILL))
    ws.conditional_formatting.add("E2:E1000",
        CellIsRule(operator="equal", formula=['"High"'], fill=AMBER_FILL))
    # Investigation Progress — low progress warning
    prog_range = f"{get_column_letter(headers.index('InvestigationProgress')+1)}2:" \
                 f"{get_column_letter(headers.index('InvestigationProgress')+1)}1000"
    ws.conditional_formatting.add(prog_range,
        CellIsRule(operator="lessThan", formula=['0.33'], fill=RED_FILL))
    ws.conditional_formatting.add(prog_range,
        CellIsRule(operator="between", formula=['0.33', '0.66'], fill=AMBER_FILL))
    ws.conditional_formatting.add(prog_range,
        CellIsRule(operator="greaterThanOrEqual", formula=['0.66'], fill=GREEN_FILL))
    # Escalation flag
    ws.conditional_formatting.add("AL2:AL1000",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=RED_FILL))
    # Overdue next action
    ws.conditional_formatting.add("AK2:AK1000",
        FormulaRule(formula=['=AND(AK2<>"",AK2<TODAY())'], fill=RED_FILL))

    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 9: CASE FILE TRACKER (DPP Pipeline)
#  From investigation through DPP ruling to court listing
# ═══════════════════════════════════════════════════════════════
def build_case_file_tracker(wb):
    ws = wb.create_sheet("tbl_CaseFiles")
    ws.sheet_properties.tabColor = COURT_PURPLE
    headers = [
        "CaseFileID", "LinkedOpsID", "CaseType",
        "PrimaryOffence", "ActSection",
        "InvestigatingOfficer", "Supervisor",
        # Linked Records
        "ArrestIDs", "FirearmSeizureIDs", "NarcoticSeizureIDs", "OtherSeizureIDs",
        # File Completeness (DPP requirements)
        "StatementsComplete", "ForensicCertObtained", "BallisticCertObtained",
        "PostMortemObtained", "DisclosureComplete",
        "FileInTriplicate",
        # DPP Submission
        "DPPSubmitDate", "DPPStatus", "CrownCounselAssigned",
        "DPPRulingDate", "DPPRuling",
        "ChargesApproved",
        # Court
        "CourtType", "FirstHearingDate", "NextHearingDate",
        "CourtOutcome", "SentenceIfConvicted",
        # POCA
        "POCAStatus", "POCAAmount_JMD",
        # Tracking
        "DaysInPipeline", "FileCompletenessScore",
        "Notes", "CreatedBy", "CreatedDate"
    ]
    widths = [12,12,16,30,28,18,18,16,16,16,16,10,10,10,10,10,10,12,22,18,12,20,24,36,12,12,18,22,14,14,12,10,28,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(D{r}="","","CF-"&TEXT(ROW()-1,"0000"))')
        # Days in pipeline
        ws.cell(row=r, column=31, value=f'=IF(R{r}="","",TODAY()-R{r})')
        # File completeness score (6 checkboxes)
        ws.cell(row=r, column=32, value=(
            f'=IF(D{r}="","",('
            f'(L{r}="Yes")+(M{r}="Yes")+(N{r}="Yes")+'
            f'(O{r}="Yes")+(P{r}="Yes")+(Q{r}="Yes"))/6)'))
        ws.cell(row=r, column=32).number_format = '0%'
    add_table(ws, "tbl_CaseFiles", f"A1:{get_column_letter(len(headers))}6")
    dv_list(ws, "L2:Q1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "S2:S1000", "=Lookups!$N$2:$N$11")
    dv_list(ws, "X2:X1000", "=Lookups!$L$2:$L$7")
    dv_list(ws, "AC2:AC1000", "=Lookups!$W$2:$W$9")
    ws.conditional_formatting.add("AF2:AF1000",
        CellIsRule(operator="lessThan", formula=['0.5'], fill=RED_FILL))
    ws.conditional_formatting.add("AF2:AF1000",
        CellIsRule(operator="between", formula=['0.5', '0.8'], fill=AMBER_FILL))
    ws.conditional_formatting.add("AF2:AF1000",
        CellIsRule(operator="greaterThanOrEqual", formula=['0.8'], fill=GREEN_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 9: WITNESS / STATEMENT REGISTER
# ═══════════════════════════════════════════════════════════════
def build_witness_register(wb):
    ws = wb.create_sheet("tbl_Witnesses")
    ws.sheet_properties.tabColor = COURT_PURPLE
    headers = [
        "WitnessID", "LinkedCaseFileID", "WitnessName",
        "ContactNumber", "Address", "Parish",
        "WitnessType", "StatementTaken", "StatementDate",
        "CautionGiven", "WillingToTestify",
        "ProtectionRequired", "WitnessProtectionReferred",
        "SummonsIssued", "SummonsServed", "SummonsServedDate",
        "Vulnerable", "IntimidationReported",
        "CourtAppearanceDate",
        "Notes", "CreatedBy", "CreatedDate"
    ]
    widths = [12,14,22,16,24,14,14,10,12,10,10,10,10,10,10,12,10,10,12,28,14,12]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","","WIT-"&TEXT(ROW()-1,"0000"))')
    add_table(ws, "tbl_Witnesses", f"A1:{get_column_letter(len(headers))}6")
    for col in "H,J,K,L,M,N,O,Q,R".split(","):
        dv_list(ws, f"{col}2:{col}1000", "=Lookups!$Q$2:$Q$3")
    dv_list(ws, "F2:F1000", "=Lookups!$P$2:$P$15")
    ws.conditional_formatting.add("R2:R1000",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=RED_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 10: INFORMANT REGISTER (Restricted)
# ═══════════════════════════════════════════════════════════════
def build_informant_register(wb):
    ws = wb.create_sheet("tbl_Informants")
    ws.sheet_properties.tabColor = INTEL_GOLD
    headers = [
        "InformantCode", "Status", "RegisteredDate",
        "Handler", "HandlerRank",
        "Parish", "Territory",
        "ReliabilityRating", "TotalTips",
        "TipsToOperations", "TipsToSeizures", "TipsToArrests",
        "LastContactDate", "NextContactDue",
        "PolygraphDate", "PolygraphResult",
        "Notes", "CreatedBy"
    ]
    widths = [16,10,12,18,14,14,20,14,10,12,12,12,12,12,12,12,28,14]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    add_table(ws, "tbl_Informants", f"A1:{get_column_letter(len(headers))}6")
    ws.conditional_formatting.add("N2:N1000",
        FormulaRule(formula=['=AND(N2<>"",N2<TODAY())'], fill=RED_FILL))
    ws.protection = openpyxl.worksheet.protection.SheetProtection(
        sheet=False, password="fnid_restricted")
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 11: PERSONNEL REGISTER
# ═══════════════════════════════════════════════════════════════
def build_personnel(wb):
    ws = wb.create_sheet("tbl_Personnel")
    ws.sheet_properties.tabColor = MED_GRAY
    headers = [
        "PersonnelID", "RegNo", "Rank", "Name",
        "Unit", "Assignment", "Status",
        "PolygraphDate", "PolygraphResult",
        "FirearmsQualified", "TacticalTraining",
        "CourtTestimonyTrained",
        "JoinedFNID", "PriorPosting",
        "Phone", "Notes"
    ]
    widths = [12,12,14,22,16,20,12,12,12,10,10,10,12,18,14,24]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    add_table(ws, "tbl_Personnel", f"A1:{get_column_letter(len(headers))}6")
    return ws


# ═══════════════════════════════════════════════════════════════
#  TABLE 12: VEHICLE / FLEET LOG
# ═══════════════════════════════════════════════════════════════
def build_vehicles(wb):
    ws = wb.create_sheet("tbl_Vehicles")
    ws.sheet_properties.tabColor = MED_GRAY
    headers = [
        "VehicleID", "Registration", "Make", "Model", "Year",
        "Type", "Status", "AssignedTo",
        "FuelType", "LastServiceDate", "NextServiceDue",
        "InsuranceExpiry", "FitnessExpiry",
        "Source", "Notes"
    ]
    widths = [12,14,14,14,8,14,14,18,10,12,12,12,12,18,24]
    write_headers(ws, headers, widths)
    data_rows(ws, len(headers))
    add_table(ws, "tbl_Vehicles", f"A1:{get_column_letter(len(headers))}6")
    ws.conditional_formatting.add("K2:K1000",
        FormulaRule(formula=['=AND(K2<>"",K2<TODAY())'], fill=RED_FILL))
    ws.conditional_formatting.add("L2:L1000",
        FormulaRule(formula=['=AND(L2<>"",L2<TODAY())'], fill=RED_FILL))
    return ws


# ═══════════════════════════════════════════════════════════════
#  DASHBOARD: COMMAND DASHBOARD
#  What the Area Commander needs every morning
# ═══════════════════════════════════════════════════════════════
def build_command_dashboard(wb):
    ws = wb.create_sheet("DASH_Command")
    ws.sheet_properties.tabColor = NAVY

    # Banner
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="FNID AREA 3 — DAILY COMMAND DASHBOARD")
    c.font = Font(name="Calibri", bold=True, size=18, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 42
    ws.merge_cells("A2:L2")
    ws.cell(row=2, column=1,
        value="Intelligence → Operations → Seizures → Arrests → Forensics → DPP → Court").font = Font(
        name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws.cell(row=2, column=1).alignment = CENTER

    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    def section(r, title, color):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
        return r + 1

    def metrics(r, items, color=NAVY):
        for i, (label, _) in enumerate(items):
            c = ws.cell(row=r, column=i+1, value=label)
            c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = CENTER
            c.border = THIN_BORDER
        r += 1
        for i, (_, formula) in enumerate(items):
            c = ws.cell(row=r, column=i+1, value=formula)
            c.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
            c.alignment = CENTER
            c.border = THIN_BORDER
        return r + 2

    r = 4
    r = section(r, "INTELLIGENCE PIPELINE", INTEL_GOLD)
    r = metrics(r, [
        ("Total Intel (YTD)", '=COUNTA(tbl_IntelLog[IntelID])-COUNTBLANK(tbl_IntelLog[IntelID])'),
        ("Pending Triage", '=COUNTBLANK(tbl_IntelLog[TriageDecision])'),
        ("Critical/High", '=COUNTIF(tbl_IntelLog[Priority],"Critical")+COUNTIF(tbl_IntelLog[Priority],"High")'),
        ("Firearms-Related", '=COUNTIF(tbl_IntelLog[FirearmsRelated],"Yes")'),
        ("Narcotics-Related", '=COUNTIF(tbl_IntelLog[NarcoticsRelated],"Yes")'),
        ("Trafficking Intel", '=COUNTIF(tbl_IntelLog[TraffickingRelated],"Yes")'),
        ("→ Converted to Op", '=COUNTIF(tbl_IntelLog[TriageDecision],"Action*")'),
        ("Conversion Rate", '=IFERROR(COUNTIF(tbl_IntelLog[TriageDecision],"Action*")/(COUNTA(tbl_IntelLog[IntelID])-COUNTBLANK(tbl_IntelLog[IntelID])),"—")'),
    ], INTEL_GOLD)

    r = section(r, "OPERATIONS", OPS_RED)
    r = metrics(r, [
        ("Total Ops (YTD)", '=COUNTA(tbl_Operations[OpsID])-COUNTBLANK(tbl_Operations[OpsID])'),
        ("Search Warrants", '=COUNTIF(tbl_Operations[OpsType],"Search Warrant*")'),
        ("Port/Cargo", '=COUNTIF(tbl_Operations[OpsType],"Port*")+COUNTIF(tbl_Operations[OpsType],"Airport*")'),
        ("Vehicle Intercepts", '=COUNTIF(tbl_Operations[OpsType],"Vehicle*")'),
        ("Joint Ops", '=COUNTIF(tbl_Operations[OpsType],"Joint*")'),
        ("Success Rate", '=IFERROR(COUNTIF(tbl_Operations[Outcome],"Successful*")/(COUNTA(tbl_Operations[OpsID])-COUNTBLANK(tbl_Operations[OpsID])),"—")'),
        ("Inspector Present %", '=IFERROR(COUNTIF(tbl_Operations[InspectorPresent],"Yes")/COUNTA(tbl_Operations[InspectorPresent]),"—")'),
        ("Body Cam Active %", '=IFERROR(COUNTIF(tbl_Operations[BodyCamActivated],"Yes")/COUNTA(tbl_Operations[BodyCamActivated]),"—")'),
    ], OPS_RED)

    r = section(r, "FIREARMS SEIZURES", OPS_RED)
    r = metrics(r, [
        ("Firearms Seized", '=COUNTA(tbl_FirearmSeizures[SeizureID])-COUNTBLANK(tbl_FirearmSeizures[SeizureID])'),
        ("Pistols", '=COUNTIF(tbl_FirearmSeizures[FirearmType],"Pistol*")'),
        ("Rifles", '=COUNTIF(tbl_FirearmSeizures[FirearmType],"Rifle*")'),
        ("Shotguns", '=COUNTIF(tbl_FirearmSeizures[FirearmType],"Shotgun")'),
        ("IBIS Hits", '=COUNTIF(tbl_FirearmSeizures[IBIS_Status],"Hit*")'),
        ("eTrace Complete", '=COUNTIF(tbl_FirearmSeizures[eTrace_Status],"Trace*")'),
        ("US Origin", '=COUNTIF(tbl_FirearmSeizures[eTrace_Status],"Traced*US*")'),
        ("Ballistic Cert Done", '=COUNTIF(tbl_FirearmSeizures[BallisticCertStatus],"Certificate Issued")'),
    ], OPS_RED)

    r = section(r, "NARCOTICS SEIZURES", NARCO_TEAL)
    r = metrics(r, [
        ("Drug Seizures", '=COUNTA(tbl_NarcoticSeizures[SeizureID])-COUNTBLANK(tbl_NarcoticSeizures[SeizureID])'),
        ("Cannabis Items", '=COUNTIF(tbl_NarcoticSeizures[DrugType],"Cannabis*")'),
        ("Cocaine Items", '=COUNTIF(tbl_NarcoticSeizures[DrugType],"Cocaine*")'),
        ("Est Value (JMD)", '=SUM(tbl_NarcoticSeizures[EstStreetValue_JMD])'),
        ("Linked to Gun Trade", '=COUNTIF(tbl_NarcoticSeizures[LinkedToGunTrade],"Yes")'),
        ("Est Equiv Firearms", '=SUM(tbl_NarcoticSeizures[EstEquivFirearms])'),
        ("Forensic Cert Done", '=COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Certificate Issued")'),
        ("Cert Overdue", '=COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Certificate Overdue")'),
    ], NARCO_TEAL)

    r = section(r, "ARRESTS & 48-HOUR COMPLIANCE", OPS_RED)
    r = metrics(r, [
        ("Total Arrests", '=COUNTA(tbl_Arrests[ArrestID])-COUNTBLANK(tbl_Arrests[ArrestID])'),
        ("Firearms Charges", '=COUNTIF(tbl_Arrests[ActSection],"*Firearms*")'),
        ("Drug Charges", '=COUNTIF(tbl_Arrests[ActSection],"*DDA*")+COUNTIF(tbl_Arrests[ActSection],"*Dangerous*")'),
        ("48hr Met", '=COUNTIF(tbl_Arrests[48hr_Met],"Yes")'),
        ("48hr BREACH", '=COUNTIF(tbl_Arrests[48hr_Met],"BREACH")'),
        ("48hr Rate", '=IFERROR(COUNTIF(tbl_Arrests[48hr_Met],"Yes")/COUNTA(tbl_Arrests[48hr_Met]),"—")'),
        ("Bail Granted", '=COUNTIF(tbl_Arrests[BailStatus],"Bail Granted")'),
        ("Remanded", '=COUNTIF(tbl_Arrests[BailStatus],"Remanded*")'),
    ], OPS_RED)

    r = section(r, "CASE REGISTRY & INVESTIGATIONS", REGISTRY_BLUE)
    r = metrics(r, [
        ("Total Cases", '=COUNTA(tbl_CaseRegistry[CaseRegID])-COUNTBLANK(tbl_CaseRegistry[CaseRegID])'),
        ("Open — Active", '=COUNTIF(tbl_CaseRegistry[CaseStatus],"Open — Active*")'),
        ("Open — Pending", '=COUNTIF(tbl_CaseRegistry[CaseStatus],"Open — Pending*")'),
        ("Firearms Cases", '=COUNTIF(tbl_CaseRegistry[CaseClassification],"Firearms")+COUNTIF(tbl_CaseRegistry[CaseClassification],"Firearms*Narcotics")'),
        ("Narcotics Cases", '=COUNTIF(tbl_CaseRegistry[CaseClassification],"Narcotics")+COUNTIF(tbl_CaseRegistry[CaseClassification],"Firearms*Narcotics")'),
        ("Escalated", '=COUNTIF(tbl_CaseRegistry[Escalated],"Yes")'),
        ("Cold Cases", '=COUNTIF(tbl_CaseRegistry[CaseStatus],"Cold Case*")'),
        ("Avg Progress", '=IFERROR(AVERAGE(tbl_CaseRegistry[InvestigationProgress]),"—")'),
    ], REGISTRY_BLUE)

    r = section(r, "DPP CASE PIPELINE", COURT_PURPLE)
    r = metrics(r, [
        ("Total Case Files", '=COUNTA(tbl_CaseFiles[CaseFileID])-COUNTBLANK(tbl_CaseFiles[CaseFileID])'),
        ("Being Prepared", '=COUNTIF(tbl_CaseFiles[DPPStatus],"File Being*")'),
        ("Submitted to DPP", '=COUNTIF(tbl_CaseFiles[DPPStatus],"File Submitted*")'),
        ("Awaiting Forensic", '=COUNTIF(tbl_CaseFiles[DPPStatus],"Awaiting Forensic*")'),
        ("Awaiting Ballistic", '=COUNTIF(tbl_CaseFiles[DPPStatus],"Awaiting Ballistic*")'),
        ("Charge Approved", '=COUNTIF(tbl_CaseFiles[DPPStatus],"Ruling*Charge*")'),
        ("Avg File Score", '=IFERROR(AVERAGE(tbl_CaseFiles[FileCompletenessScore]),"—")'),
        ("Avg Days Pipeline", '=IFERROR(AVERAGE(tbl_CaseFiles[DaysInPipeline]),"—")'),
    ], COURT_PURPLE)

    r = section(r, "FORENSIC BOTTLENECK (IFSLM)", EVIDENCE_GREEN)
    r = metrics(r, [
        ("Ballistic Pending", '=COUNTIF(tbl_FirearmSeizures[BallisticCertStatus],"Submitted*")+COUNTIF(tbl_FirearmSeizures[BallisticCertStatus],"Analysis*")'),
        ("Ballistic Overdue", '=COUNTIF(tbl_FirearmSeizures[BallisticCertStatus],"Certificate Overdue")'),
        ("Ballistic Complete", '=COUNTIF(tbl_FirearmSeizures[BallisticCertStatus],"Certificate Issued")'),
        ("Drug Cert Pending", '=COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Submitted*")+COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Analysis*")'),
        ("Drug Cert Overdue", '=COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Certificate Overdue")'),
        ("Drug Cert Complete", '=COUNTIF(tbl_NarcoticSeizures[ForensicCertStatus],"Certificate Issued")'),
        ("POCA Referrals", '=COUNTIF(tbl_Arrests[POCAStatus],"Referred*")'),
        ("POCA Forfeited", '=COUNTIF(tbl_Arrests[POCAStatus],"Forfeiture*")'),
    ], EVIDENCE_GREEN)

    return ws


# ═══════════════════════════════════════════════════════════════
#  HOME SHEET — Navigation Hub
# ═══════════════════════════════════════════════════════════════
def build_home(wb):
    ws = wb.create_sheet("HOME")
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
        value="JAMAICA CONSTABULARY FORCE — FIREARM AND NARCOTICS INVESTIGATION DIVISION")
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1,
        value="AREA 3 (Manchester | St. Elizabeth | Clarendon) — OPERATIONAL WORKBOOK v2.0").font = Font(
        name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws.cell(row=2, column=1).alignment = CENTER

    ws.merge_cells("A3:H3")
    ws.cell(row=3, column=1,
        value="Director: SSP Patrae Rowe | Commissioner: Dr. Kevin Blake | FNID Direct: 876-923-6184").font = Font(
        name="Calibri", italic=True, size=10, color=MED_GRAY)
    ws.cell(row=3, column=1).alignment = CENTER

    ws.merge_cells("A4:H4")
    ws.cell(row=4, column=1,
        value="INTELLIGENCE → OPERATION → SEIZURE → ARREST → FORENSICS → DPP → COURT → CONVICTION").font = Font(
        name="Calibri", bold=True, size=11, color=OPS_RED)
    ws.cell(row=4, column=1).alignment = CENTER

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 18

    sections = [
        ("DAILY COMMAND", NAVY, [
            ("DASH_Command", "Daily Command Dashboard — all key metrics at a glance"),
        ]),
        ("STAGE 1: INTELLIGENCE", INTEL_GOLD, [
            ("tbl_IntelLog", "Intelligence Log — Crime Stop, NIB, DEA, JCA, informant tips"),
            ("tbl_Informants", "Informant Register — RESTRICTED ACCESS (code names only)"),
        ]),
        ("STAGE 2: OPERATIONS", OPS_RED, [
            ("tbl_Operations", "Operation Register — warrants, teams, execution, outcomes"),
        ]),
        ("STAGE 3: SEIZURES", OPS_RED, [
            ("tbl_FirearmSeizures", "Firearm Seizure Log — IBIS, eTrace, ballistic cert tracking"),
            ("tbl_NarcoticSeizures", "Narcotics Seizure Log — field test, IFSLM, forensic cert, disposal"),
            ("tbl_OtherSeizures", "Ammunition, Cash, Electronics, Vehicles — other exhibits"),
        ]),
        ("STAGE 4: ARRESTS", OPS_RED, [
            ("tbl_Arrests", "Arrest Register — 48hr rule, charge, caution, bail, court routing"),
        ]),
        ("STAGE 5: CASE REGISTRY & INVESTIGATIONS", REGISTRY_BLUE, [
            ("tbl_CaseRegistry", "Master Case File Registry — every case, investigation milestones, progress tracking"),
        ]),
        ("STAGE 6: EVIDENCE & FORENSICS", EVIDENCE_GREEN, [
            ("tbl_ChainOfCustody", "Exhibit Chain of Custody — every handover logged"),
        ]),
        ("STAGE 7: DPP & COURT", COURT_PURPLE, [
            ("tbl_CaseFiles", "Case File Tracker — DPP submission pipeline, completeness scoring"),
            ("tbl_Witnesses", "Witness & Statement Register — testimony, protection, summons"),
        ]),
        ("SUPPORT", MED_GRAY, [
            ("tbl_Personnel", "Personnel Register — staff, qualifications, polygraph"),
            ("tbl_Vehicles", "Vehicle/Fleet Log — status, service, fitness"),
            ("Lookups", "Reference Data — Jamaica-specific lookups"),
        ]),
    ]

    r = 6
    for title, color, items in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[r].height = 26
        r += 1
        for sheet_name, desc in items:
            ws.cell(row=r, column=1, value=f"  >> {sheet_name}").font = Font(
                name="Calibri", bold=True, size=10, color=DARK_BLUE)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.cell(row=r, column=2, value=desc).font = BODY_FONT
            ws.cell(row=r, column=2).alignment = LEFT_WRAP
            r += 1
        r += 1

    # Legal references footer
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
        value="LEGAL FRAMEWORK: Firearms (Prohibition, Restriction and Regulation) Act 2022 | "
              "Dangerous Drugs Act (as amended 2015) | Gun Court Act 1974 | "
              "Proceeds of Crime Act 2007 | Bail Act 2023 | "
              "Constabulary Force Act s.15 (48-Hour Rule)").font = Font(
        name="Calibri", italic=True, size=9, color=MED_GRAY)
    ws.cell(row=r, column=1).alignment = LEFT_WRAP
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1,
        value="CONFIDENTIAL — FOR OFFICIAL USE ONLY | FNID Area 3 | "
              "Search warrants under Section 91, Firearms Act 2022").font = Font(
        name="Calibri", bold=True, size=9, color=OPS_RED)
    return ws


# ═══════════════════════════════════════════════════════════════
#  MAIN BUILD
# ═══════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building FNID v2.1 — Purpose-Driven Operational Workbook...")
    print("=" * 60)

    print("[1/14] HOME — Navigation hub")
    build_home(wb)

    print("[2/14] DASH_Command — Daily command dashboard")
    build_command_dashboard(wb)

    print("[3/14] Lookups — Jamaica-specific reference data")
    build_lookups(wb)

    print("[4/14] tbl_IntelLog — Intelligence pipeline")
    build_intel_log(wb)

    print("[5/14] tbl_Operations — Operation register")
    build_operation_register(wb)

    print("[6/14] tbl_FirearmSeizures — Firearm seizure log")
    build_firearm_seizures(wb)

    print("[7/14] tbl_NarcoticSeizures — Narcotics seizure log")
    build_narcotics_seizures(wb)

    print("[8/14] tbl_OtherSeizures — Ammunition, cash, electronics")
    build_other_seizures(wb)

    print("[9/14] tbl_Arrests — Arrest register (48hr compliance)")
    build_arrest_register(wb)

    print("[10/14] tbl_CaseRegistry — Case file registry & investigations")
    build_case_registry(wb)

    print("[11/14] tbl_ChainOfCustody — Exhibit custody chain")
    build_chain_of_custody(wb)

    print("[12/14] tbl_CaseFiles — DPP case file pipeline")
    build_case_file_tracker(wb)

    print("[13/14] tbl_Witnesses — Witness & statement register")
    build_witness_register(wb)

    print("[14/14] Support tables — Informants, Personnel, Vehicles")
    build_informant_register(wb)
    build_personnel(wb)
    build_vehicles(wb)

    # Move HOME to first position
    wb.move_sheet("HOME", offset=-wb.sheetnames.index("HOME"))

    fname = "FNID_Area3_Operational_Workbook.xlsx"
    wb.save(fname)

    import os
    size = os.path.getsize(fname)
    print(f"\n{'=' * 60}")
    print(f"FNID v2.0 COMPLETE: {fname}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Size:   {size/1024:.0f} KB")
    print(f"\n  Operational Tables:")
    for name in wb.sheetnames:
        if name.startswith("tbl_"):
            print(f"    - {name}")
    print(f"\n  Dashboards:")
    for name in wb.sheetnames:
        if name.startswith("DASH_"):
            print(f"    - {name}")
    print(f"\n  Lifecycle: INTELLIGENCE → OPERATION → SEIZURE → ARREST → FORENSICS → DPP → COURT")


if __name__ == "__main__":
    main()
