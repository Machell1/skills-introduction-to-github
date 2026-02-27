#!/usr/bin/env python3
"""
FNID Area 3 - Siloed Unit Portal System v1.0

Generates per-unit Excel portal workbooks for FNID Area 3
(Manchester, St. Elizabeth, Clarendon).

Units:
  1. Intel Unit          - Intelligence pipeline
  2. Operations Unit     - Operation planning & execution
  3. Seizures Unit       - Firearms, narcotics, other seizures
  4. Arrests/Court Unit  - Arrests, bail, court tracking
  5. Forensics/Evidence  - Chain of custody, lab tracking
  6. Registry Unit       - Case management SOP compliance

Each portal includes:
  - Data entry sheet(s) with dropdowns and submission tracking
  - Dashboard with KPI summary cards and charts
  - Lookups reference sheet
  - Auto-fill support from Word/Excel imports

Applicable legislation:
  - Firearms (Prohibition, Restriction and Regulation) Act, 2022
  - Dangerous Drugs Act (as amended 2015)
  - Gun Court Act, 1974
  - Proceeds of Crime Act (POCA), 2007
  - Bail Act, 2023
  - Constabulary Force Act s.15 (48-Hour Rule)
  - Evidence Act / Evidence (Amendment) Act, 2021
  - DPP Prosecution Protocol (April 2012)
  - DPP Disclosure Protocol (September 2013)
"""

import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel

# =================================================================
#  STYLE CONSTANTS
# =================================================================
NAVY       = "1F3864"
DARK_BLUE  = "002060"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
MED_GRAY   = "808080"
LIGHT_GRAY = "F2F2F2"
OPS_RED    = "C00000"
INTEL_GOLD = "BF8F00"
EVIDENCE_GREEN = "538135"
COURT_PURPLE   = "7030A0"
NARCO_TEAL     = "00B0F0"
REGISTRY_BROWN = "8B4513"

UNIT_COLORS = {
    "Intel":      INTEL_GOLD,
    "Operations": OPS_RED,
    "Seizures":   NARCO_TEAL,
    "Arrests":    COURT_PURPLE,
    "Forensics":  EVIDENCE_GREEN,
    "Registry":   REGISTRY_BROWN,
}

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT   = Font(name="Calibri", size=10)
BODY_BOLD   = Font(name="Calibri", bold=True, size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, size=16, color=NAVY)
SUB_FONT    = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
KPI_NUM     = Font(name="Calibri", bold=True, size=24, color=NAVY)
KPI_LABEL   = Font(name="Calibri", size=9, color=MED_GRAY)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL  = PatternFill("solid", fgColor="FFE699")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
GOLD_FILL   = PatternFill("solid", fgColor="FFD700")
DRAFT_FILL  = PatternFill("solid", fgColor="FFF2CC")
SUBMITTED_FILL = PatternFill("solid", fgColor="D9EAD3")
EDITED_FILL = PatternFill("solid", fgColor="D0E0F0")

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portals")


# =================================================================
#  UTILITY FUNCTIONS
# =================================================================
def write_headers(ws, headers, widths=None):
    """Write styled header row. No sheet-level autoFilter (tables handle it)."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if widths and c <= len(widths):
            ws.column_dimensions[get_column_letter(c)].width = widths[c-1]
    ws.freeze_panes = "A2"


def add_table(ws, name, ref):
    """Add Excel Table object (handles its own autoFilter)."""
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TABLE_STYLE
    ws.add_table(t)


def dv_list(ws, cell_range, formula):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Select from dropdown"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def data_rows(ws, num_cols, start_row=2, count=5):
    """Pre-format empty data rows with borders."""
    for r in range(start_row, start_row + count):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT


def submission_columns():
    """Standard submission tracking columns appended to every data sheet."""
    return ["RecordStatus", "SubmittedBy", "SubmittedDate",
            "LastEditedBy", "LastEditedDate"]


def submission_widths():
    return [14, 16, 14, 16, 14]


def add_submission_tracking(ws, num_data_cols, data_row_count=5):
    """Add submission tracking columns and formatting to a data sheet."""
    sub_cols = submission_columns()
    sub_w = submission_widths()
    start_col = num_data_cols + 1
    for i, (h, w) in enumerate(zip(sub_cols, sub_w)):
        c = start_col + i
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    # Status dropdown
    status_col = get_column_letter(start_col)
    dv_list(ws, f"{status_col}2:{status_col}{1+data_row_count}",
            '"Draft,Submitted,Edited"')
    # Conditional formatting on status
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}1000",
        CellIsRule(operator="equal", formula=['"Draft"'], fill=DRAFT_FILL))
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}1000",
        CellIsRule(operator="equal", formula=['"Submitted"'], fill=SUBMITTED_FILL))
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}1000",
        CellIsRule(operator="equal", formula=['"Edited"'], fill=EDITED_FILL))
    return len(sub_cols)


def build_home(wb, unit_name, color, sheet_list):
    """Build unit-specific HOME navigation sheet."""
    ws = wb.active
    ws.title = "HOME"
    ws.sheet_properties.tabColor = color
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("B2:C2")
    cell = ws.cell(row=2, column=2, value=f"FNID AREA 3 - {unit_name.upper()} UNIT PORTAL")
    cell.font = Font(name="Calibri", bold=True, size=20, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:C3")
    ws.cell(row=3, column=2,
            value="Jamaica Constabulary Force | Manchester | St. Elizabeth | Clarendon"
            ).font = Font(name="Calibri", size=11, color=MED_GRAY, italic=True)
    ws["B3"].alignment = Alignment(horizontal="center")

    row = 5
    ws.cell(row=row, column=2, value="PORTAL NAVIGATION").font = SUB_FONT
    row += 1
    for sheet_name, desc in sheet_list:
        ws.cell(row=row, column=2, value=f"  {sheet_name}").font = BODY_BOLD
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="SUBMISSION WORKFLOW").font = SUB_FONT
    row += 1
    steps = [
        "1. Enter data in the data entry sheet(s)",
        "2. Set RecordStatus to 'Draft' while working",
        "3. Fill SubmittedBy with your name/badge when ready",
        "4. Change RecordStatus to 'Submitted' to finalize",
        "5. To edit a submitted record, change status to 'Edited'",
        "6. Use the Dashboard sheet to view KPIs and charts",
    ]
    for s in steps:
        ws.cell(row=row, column=2, value=s).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=2,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ).font = Font(name="Calibri", size=9, color=MED_GRAY)
    return ws


def build_lookups(wb, lookup_dict, tab_color=MED_GRAY):
    """Build Lookups reference sheet from a dict of {col: (name, [values])}."""
    ws = wb.create_sheet("Lookups")
    ws.sheet_properties.tabColor = tab_color
    for col, (name, values) in lookup_dict.items():
        ws.cell(row=1, column=col, value=name).font = HEADER_FONT
        ws.cell(row=1, column=col).fill = HEADER_FILL
        for i, v in enumerate(values, 2):
            ws.cell(row=i, column=col, value=v)
        ws.column_dimensions[get_column_letter(col)].width = max(20, len(name)+4)
    return ws


def kpi_card(ws, row, col, label, formula, width=2):
    """Place a KPI card on the dashboard."""
    if width > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+width-1)
        ws.merge_cells(start_row=row+1, start_column=col,
                       end_row=row+1, end_column=col+width-1)
    cell_val = ws.cell(row=row, column=col, value=formula)
    cell_val.font = KPI_NUM
    cell_val.alignment = CENTER
    cell_val.border = THIN_BORDER
    cell_lbl = ws.cell(row=row+1, column=col, value=label)
    cell_lbl.font = KPI_LABEL
    cell_lbl.alignment = CENTER
    for c in range(col, col+width):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row+1, column=c).border = THIN_BORDER


def add_pie_chart(ws, title, cats_ref, vals_ref, anchor, width=14, height=10):
    """Add a pie chart to the dashboard."""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    data = Reference(ws, **vals_ref)
    cats = Reference(ws, **cats_ref)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    ws.add_chart(chart, anchor)
    return chart


def add_bar_chart(ws, title, cats_ref, vals_ref, anchor,
                  width=14, height=10, horizontal=False):
    """Add a bar chart to the dashboard."""
    chart = BarChart()
    if horizontal:
        chart.type = "bar"
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    data = Reference(ws, **vals_ref)
    cats = Reference(ws, **cats_ref)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)
    return chart


def add_line_chart(ws, title, cats_ref, data_refs, anchor,
                   width=18, height=10, labels=None):
    """Add a line chart with one or more series."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    cats = Reference(ws, **cats_ref)
    chart.set_categories(cats)
    if isinstance(data_refs, list):
        for i, dr in enumerate(data_refs):
            data = Reference(ws, **dr)
            chart.add_data(data, titles_from_data=False)
            if labels and i < len(labels):
                chart.series[i].title = labels[i]
    else:
        data = Reference(ws, **data_refs)
        chart.add_data(data, titles_from_data=False)
    ws.add_chart(chart, anchor)
    return chart


# =================================================================
#  SHARED LOOKUP DATA
# =================================================================
COMMON_LOOKUPS = {
    1: ("IntelSource", [
        "Crime Stop (311)", "NIB (811)", "Informant", "DEA", "ATF",
        "HSI", "MOCA", "JCA Detection", "Divisional CIB", "Patrol Intercept",
        "C-TOC/SIB", "INTERPOL", "Walk-In Report", "FNID Direct (923-6184)",
        "UK NCA", "US Marshals", "Anonymous"]),
    2: ("Priority", ["Critical", "High", "Medium", "Low"]),
    3: ("Parish", [
        "Manchester", "St. Elizabeth", "Clarendon",
        "Kingston", "St. Andrew", "St. Thomas", "Portland", "St. Mary",
        "St. Ann", "Trelawny", "St. James", "Hanover", "Westmoreland",
        "St. Catherine"]),
    4: ("YesNo", ["Yes", "No"]),
    5: ("RecordStatus", ["Draft", "Submitted", "Edited"]),
}

INTEL_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("TriageDecision", [
        "Action - Mount Operation", "Action - Surveillance", "Action - Port Alert",
        "Refer to Divisional CIB", "Refer to MOCA", "Refer to SIB",
        "Intel Filed - Monitor", "Closed - Insufficient Info"]),
}

OPERATIONS_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("OperationType", [
        "Search Warrant Execution", "Snap Raid", "Port/Cargo Interdiction",
        "Airport Interdiction", "Vehicle Interception", "Coastal/Beach Operation",
        "Checkpoint", "Surveillance", "Joint Op - JCA", "Joint Op - JDF",
        "Joint Op - MOCA", "Joint Op - DEA", "Fugitive Apprehension",
        "Follow-Up Secondary Op", "ZOSO Operation", "SOE Operation",
        "Courier/Parcel Investigation"]),
    7: ("WarrantBasis", [
        "Section 91 - Firearms Act 2022", "Section 21 - Dangerous Drugs Act",
        "Section 14 - POCA 2007", "ZOSO - Warrantless (s.5 ZOSO Act)",
        "SOE - Warrantless", "Sergeant Written Directive (DDA s.21)"]),
    8: ("OpOutcome", [
        "Successful - Seizure + Arrest", "Successful - Seizure Only",
        "Successful - Arrest Only", "Partial - Intel Developed",
        "Negative - No Finds", "Aborted - Compromised",
        "Aborted - Safety Concern", "Ongoing"]),
}

SEIZURES_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("FirearmType", [
        "Pistol - Semi-Auto", "Pistol - Revolver", "Rifle - Semi-Auto",
        "Rifle - Bolt Action", "Rifle - Assault", "Shotgun", "Submachine Gun",
        "Machine Pistol", "Improvised/Homemade", "3D-Printed", "Imitation/Replica",
        "Component - Frame/Receiver", "Component - Barrel", "Component - Slide",
        "Magazine", "Silencer/Suppressor"]),
    7: ("Calibre", [
        "9mm", ".40 S&W", ".45 ACP", ".380 ACP", ".22 LR", ".25 ACP",
        ".38 Special", ".357 Magnum", "5.56mm/.223", "7.62x39mm",
        "7.62x51mm/.308", "12 Gauge", "20 Gauge", ".44 Magnum", "Other"]),
    8: ("DrugType", [
        "Cannabis/Ganja - Compressed", "Cannabis/Ganja - Loose",
        "Cannabis/Ganja - Oil/Edibles", "Cocaine - Powder",
        "Cocaine - Crack", "Heroin", "MDMA/Ecstasy", "Methamphetamine",
        "Synthetic Cannabinoids", "Prescription Opioids", "Other Controlled"]),
    9: ("DrugUnit", ["kg", "g", "lbs", "oz", "plants", "tablets", "ml", "litres"]),
    10: ("SeizureLocation", [
        "Port Bustamante", "Kingston Logistics Centre", "Newport West",
        "Norman Manley Airport", "Sangster International Airport",
        "Roadway/Vehicle", "Residential Premises", "Commercial Premises",
        "Beach/Coastal", "Open Land/Bushes", "Courier/Postal Facility",
        "ZOSO Checkpoint", "Other"]),
}

ARRESTS_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("FirearmsActOffence", [
        "s.5 - Possession of Prohibited Weapon (15-25yr)",
        "s.6 - Stockpiling (3+ weapons / 50+ rounds)",
        "s.7 - Trafficking in Prohibited Weapon (20yr min)",
        "s.8 - Possession with Intent to Traffic",
        "s.9 - Manufacture of Prohibited Weapon",
        "s.10 - Dealing in Prohibited Weapon",
        "s.12 - Diversion of Lawful Firearms",
        "s.29 - Unmarked Firearm",
        "Shooting with Intent", "Illegal Possession of Ammunition",
        "Wounding with Intent (firearm)"]),
    7: ("DDAOffence", [
        "Import/Export - Cannabis (DDA Part IIIA)",
        "Import/Export - Cocaine (DDA Part IV)",
        "Dealing/Trafficking - Cannabis",
        "Dealing/Trafficking - Cocaine",
        "Possession - Cannabis (>2oz)",
        "Possession - Cocaine",
        "Cultivation - Cannabis (>5 plants)",
        "Deemed Dealing - School Premises"]),
    8: ("BailStatus", [
        "Bail Granted", "Bail Denied", "Remanded in Custody",
        "Released - No Charge", "Released - Insufficient Evidence",
        "Stop Order Issued", "Electronic Monitoring"]),
    9: ("CourtType", [
        "Gun Court - High Court Division",
        "Gun Court - Circuit Court Division",
        "Gun Court - RM Division",
        "Parish Court", "Circuit Court", "Supreme Court"]),
}

FORENSICS_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("ExhibitType", [
        "Firearm", "Ammunition", "Magazine", "Drug Sample",
        "Cash/Currency", "Electronic Device", "Vehicle", "Document",
        "Clothing/Personal Item", "Biological Sample", "Other"]),
    7: ("IBIS_Status", [
        "Not Submitted", "Submitted - Pending", "No Match",
        "Hit - Confirmed Match", "Hit - Unconfirmed", "Serial Obliterated"]),
    8: ("eTrace_Status", [
        "Not Submitted", "Submitted - Pending", "Traced - US Origin",
        "Traced - Non-US Origin", "Untraceable - No Serial",
        "Untraceable - Obliterated", "Trace Complete"]),
    9: ("ForensicStatus", [
        "Not Yet Submitted", "Submitted to IFSLM", "Analysis In Progress",
        "Certificate Issued", "Certificate Overdue", "Inconclusive"]),
    10: ("ExhibitDisposal", [
        "Held - Active Case", "Held - Court Order",
        "Destroyed - Authorized", "Returned to Owner",
        "Forfeited - POCA", "Pending Disposal Authorization"]),
}

REGISTRY_LOOKUPS = {
    **COMMON_LOOKUPS,
    6: ("CaseClassification", [
        "Firearms", "Narcotics", "Firearms + Narcotics",
        "Trafficking - Firearms", "Trafficking - Narcotics",
        "Trafficking - Multi-Commodity", "POCA / Financial",
        "Murder with Firearm", "Shooting with Intent"]),
    7: ("CaseStatus", [
        "Open - Active Investigation", "Open - Pending Forensics",
        "Open - Pending Witness Statements", "Open - Surveillance",
        "Open - Pending DPP Submission", "Referred to DPP",
        "Closed - Convicted", "Closed - Acquitted",
        "Closed - No Charge", "Closed - Withdrawn",
        "Cold Case - Under Review", "Cold Case - Dormant"]),
    8: ("DPPStatus", [
        "File Being Prepared", "File Submitted to DPP",
        "Awaiting Forensic Certificate", "Awaiting Ballistic Certificate",
        "Crown Counsel Reviewing", "Ruling - Charge Approved",
        "Ruling - No Charge", "Voluntary Bill of Indictment",
        "Preliminary Exam Ordered", "Returned for Further Investigation"]),
    9: ("SOPCompliance", [
        "Compliant", "Non-Compliant - Missing Statements",
        "Non-Compliant - Missing Forensics",
        "Non-Compliant - Missing Exhibit Register",
        "Non-Compliant - Overdue DPP Submission",
        "Non-Compliant - 48hr Breach", "Under Review"]),
    10: ("DisclosureStatus", [
        "Not Required Yet", "Disclosure Package Prepared",
        "Served on Defence", "Partial Disclosure (PII applied)",
        "Supplementary Disclosure Pending", "Complete"]),
    11: ("FileCompleteness", [
        "Complete", "Substantially Complete (>80%)",
        "Incomplete - Missing Key Documents",
        "Incomplete - Missing Forensics",
        "Incomplete - Missing Statements", "Not Started"]),
}


# =================================================================
#  SAMPLE DATA (realistic FNID Area 3 entries for chart population)
# =================================================================
SAMPLE_INTEL = [
    ["INT-2025-001", "2025-01-15", "08:30", "Crime Stop (311)", "CS-4421", "High",
     "Firearm cache reported at abandoned property", "Yes", "No", "No",
     "Unknown", "Lot 14 Green Ave, Mandeville", "Manchester",
     "Caller reports seeing men burying items in back lot at night",
     "Action - Mount Operation", "Insp. Brown", "2025-01-15",
     "OP-2025-001", "CR-2025-001", "Successful - Seizure + Arrest",
     "2 firearms recovered", "DC Williams", "2025-01-15"],
    ["INT-2025-002", "2025-01-22", "14:15", "Informant", "INF-009", "Critical",
     "Cocaine shipment via fishing vessel", "No", "Yes", "Yes",
     "Marcus COLE", "Black River Bay", "St. Elizabeth",
     "50kg cocaine expected landfall Friday night",
     "Action - Mount Operation", "Supt. Davis", "2025-01-22",
     "OP-2025-002", "CR-2025-002", "Successful - Seizure Only",
     "32kg cocaine seized", "DC Thompson", "2025-01-22"],
    ["INT-2025-003", "2025-02-03", "10:00", "DEA", "DEA-JM-0087", "High",
     "Trafficking network moving guns from Florida", "Yes", "No", "Yes",
     "Ryan SMITH", "May Pen area", "Clarendon",
     "DEA identifies shipping container with concealed weapons",
     "Action - Port Alert", "Insp. Brown", "2025-02-03",
     "OP-2025-003", "", "Ongoing",
     "Container identified, surveillance active", "DC Williams", "2025-02-03"],
    ["INT-2025-004", "2025-02-10", "09:45", "Divisional CIB", "CIB-MC-112", "Medium",
     "Ganja cultivation site in hills", "No", "Yes", "No",
     "Unknown", "Newport, Manchester", "Manchester",
     "Large scale cultivation reported by patrol",
     "Action - Surveillance", "Insp. Morgan", "2025-02-10",
     "", "", "", "", "Cpl. James", "2025-02-10"],
    ["INT-2025-005", "2025-02-18", "16:30", "Walk-In Report", "WIR-0033", "Low",
     "Suspicious activity at warehouse", "No", "No", "No",
     "Unknown", "Spaldings Road", "Manchester",
     "Citizen reports trucks loading at night",
     "Intel Filed - Monitor", "Insp. Brown", "2025-02-18",
     "", "", "", "", "DC Thompson", "2025-02-18"],
    ["INT-2025-006", "2025-03-01", "07:00", "MOCA", "MOCA-2025-019", "Critical",
     "Gang leader with illegal arsenal", "Yes", "Yes", "Yes",
     "Devon GRANT", "Christiana", "Manchester",
     "MOCA identifies key target with multiple firearms and drug links",
     "Action - Mount Operation", "Supt. Davis", "2025-03-01",
     "OP-2025-004", "CR-2025-003", "Successful - Seizure + Arrest",
     "5 firearms, 2kg cocaine seized, 3 arrested", "DC Williams", "2025-03-01"],
    ["INT-2025-007", "2025-03-10", "11:20", "Anonymous", "ANON-071", "Medium",
     "Drug dealing at school zone", "No", "Yes", "No",
     "Unknown", "Mandeville High School area", "Manchester",
     "Ongoing drug sales near school perimeter",
     "Refer to Divisional CIB", "Insp. Morgan", "2025-03-10",
     "", "", "", "", "Cpl. James", "2025-03-10"],
    ["INT-2025-008", "2025-03-15", "13:45", "JCA Detection", "JCA-2025-055", "High",
     "Ammunition in shipping barrel", "Yes", "No", "Yes",
     "Shipping Co. ABC Ltd", "Kingston Logistics Centre", "St. Catherine",
     "JCA X-ray detects ammunition in household barrel to Manchester",
     "Action - Port Alert", "Insp. Brown", "2025-03-15",
     "OP-2025-005", "", "Successful - Seizure Only",
     "500 rounds 9mm seized", "DC Thompson", "2025-03-15"],
    ["INT-2025-009", "2025-03-22", "08:00", "Patrol Intercept", "PI-MC-044", "High",
     "Vehicle with concealed compartment", "Yes", "Yes", "Yes",
     "Andrew BLAKE", "Highway 2000 checkpoint", "Clarendon",
     "Patrol finds modified vehicle with hidden compartment",
     "Action - Mount Operation", "Insp. Brown", "2025-03-22",
     "OP-2025-006", "CR-2025-004", "Successful - Seizure + Arrest",
     "1 pistol, 200g cocaine, $500K cash", "DC Williams", "2025-03-22"],
    ["INT-2025-010", "2025-04-01", "15:00", "UK NCA", "NCA-JM-2025-003", "Critical",
     "International cocaine trafficking ring", "No", "Yes", "Yes",
     "Multiple targets", "South coast", "St. Elizabeth",
     "NCA requests assistance with Jamaica end of trafficking network",
     "Action - Surveillance", "Supt. Davis", "2025-04-01",
     "", "", "Ongoing",
     "Joint surveillance operation planned", "DC Thompson", "2025-04-01"],
]

MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
SAMPLE_MONTHLY_COUNTS = [3, 4, 5, 2, 3, 4, 5, 3, 4, 6, 3, 2]  # generic


# =================================================================
#  1. INTEL UNIT PORTAL
# =================================================================
def build_intel_portal():
    """Generate FNID_Intel_Portal.xlsx"""
    wb = Workbook()

    # -- HOME --
    sheets = [
        ("Intel_Log", "Intelligence pipeline - data entry & tracking"),
        ("Dashboard", "KPI summary cards and visual charts"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Intel", INTEL_GOLD, sheets)

    # -- INTEL LOG DATA SHEET --
    ws = wb.create_sheet("Intel_Log")
    ws.sheet_properties.tabColor = INTEL_GOLD
    headers = [
        "IntelID", "DateReceived", "TimeReceived", "Source", "SourceRef",
        "Priority", "SubjectMatter", "FirearmsRelated", "NarcoticsRelated",
        "TraffickingRelated", "TargetPerson", "TargetLocation", "Parish",
        "SubstanceOfIntel", "TriageDecision", "TriageBy", "TriageDate",
        "LinkedOpID", "LinkedCaseID", "Outcome", "OutcomeNotes",
        "CreatedBy", "CreatedDate",
    ]
    widths = [
        14, 12, 10, 18, 14, 10, 25, 10, 10, 10,
        18, 22, 14, 30, 22, 16, 12, 14, 14, 22, 25, 14, 12,
    ]
    write_headers(ws, headers, widths)

    # Populate sample data
    for i, row_data in enumerate(SAMPLE_INTEL):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    # More empty rows for data entry
    data_rows(ws, len(headers), start_row=len(SAMPLE_INTEL)+2, count=20)

    # Submission tracking
    n_sub = add_submission_tracking(ws, len(headers), len(SAMPLE_INTEL)+20)
    total_cols = len(headers) + n_sub
    last_col = get_column_letter(total_cols)

    # Mark sample rows as Submitted
    status_col_idx = len(headers) + 1
    for r in range(2, len(SAMPLE_INTEL)+2):
        ws.cell(row=r, column=status_col_idx, value="Submitted").font = BODY_FONT
        ws.cell(row=r, column=status_col_idx+1, value="System").font = BODY_FONT
        ws.cell(row=r, column=status_col_idx+2, value="2025-04-01").font = BODY_FONT

    # Dropdowns
    dv_list(ws, "D2:D500", "Lookups!$A$2:$A$20")
    dv_list(ws, "F2:F500", "Lookups!$B$2:$B$10")
    dv_list(ws, "H2:H500", "Lookups!$D$2:$D$5")
    dv_list(ws, "I2:I500", "Lookups!$D$2:$D$5")
    dv_list(ws, "J2:J500", "Lookups!$D$2:$D$5")
    dv_list(ws, "M2:M500", "Lookups!$C$2:$C$20")
    dv_list(ws, "O2:O500", "Lookups!$F$2:$F$12")

    # Table
    end_row = len(SAMPLE_INTEL) + 1 + 20
    add_table(ws, "tbl_IntelLog",
              f"A1:{get_column_letter(len(headers))}{end_row}")

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = INTEL_GOLD
    dash.sheet_view.showGridLines = False

    # Title
    dash.merge_cells("B2:K2")
    dash.cell(row=2, column=2,
              value="INTEL UNIT DASHBOARD - FNID Area 3").font = TITLE_FONT
    dash["B2"].alignment = CENTER

    # KPI Cards
    kpi_card(dash, 4, 2, "Total Intel Reports",
             f"=COUNTA(Intel_Log!A2:A{end_row})")
    kpi_card(dash, 4, 5, "Critical Priority",
             f'=COUNTIF(Intel_Log!F2:F{end_row},"Critical")')
    kpi_card(dash, 4, 8, "Actioned",
             f'=COUNTIF(Intel_Log!O2:O{end_row},"Action*")')
    kpi_card(dash, 4, 11, "Firearms Related",
             f'=COUNTIF(Intel_Log!H2:H{end_row},"Yes")')

    # Summary tables for charts
    # Source distribution
    sources = ["Crime Stop (311)", "Informant", "DEA", "MOCA", "Anonymous",
               "JCA Detection", "Patrol Intercept", "Walk-In Report",
               "UK NCA", "Other"]
    r = 8
    dash.cell(row=r, column=2, value="Source").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, src in enumerate(sources):
        dash.cell(row=r+1+i, column=2, value=src).font = BODY_FONT
        if src == "Other":
            dash.cell(row=r+1+i, column=3,
                      value=f'=COUNTA(Intel_Log!D2:D{end_row})-SUM(C{r+1}:C{r+i})'
                      ).font = BODY_FONT
        else:
            dash.cell(row=r+1+i, column=3,
                      value=f'=COUNTIF(Intel_Log!D2:D{end_row},B{r+1+i})'
                      ).font = BODY_FONT

    # Pie chart - Source distribution
    add_pie_chart(dash, "Intelligence by Source",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(sources)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(sources)},
                  anchor="E8")

    # Priority breakdown
    priorities = ["Critical", "High", "Medium", "Low"]
    pr = r + len(sources) + 3
    dash.cell(row=pr, column=2, value="Priority").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, p in enumerate(priorities):
        dash.cell(row=pr+1+i, column=2, value=p).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Intel_Log!F2:F{end_row},B{pr+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Intel by Priority",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+4},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+4},
                  anchor=f"E{pr}")

    # Monthly trend
    mr = pr + 7
    dash.cell(row=mr, column=2, value="Month").font = BODY_BOLD
    dash.cell(row=mr, column=3, value="Count").font = BODY_BOLD
    for i, m in enumerate(MONTHS_SHORT):
        dash.cell(row=mr+1+i, column=2, value=m).font = BODY_FONT
        month_num = str(i+1).zfill(2)
        dash.cell(row=mr+1+i, column=3,
                  value=f'=COUNTIFS(Intel_Log!B2:B{end_row},">="&"2025-{month_num}-01",Intel_Log!B2:B{end_row},"<"&"2025-{str(i+2).zfill(2) if i<11 else "2026-01"}-01")'
                  ).font = BODY_FONT

    add_line_chart(dash, "Monthly Intelligence Trend",
                   cats_ref={"min_col": 2, "min_row": mr+1, "max_row": mr+12},
                   data_refs={"min_col": 3, "min_row": mr+1, "max_row": mr+12},
                   anchor=f"E{mr}", width=20)

    # -- LOOKUPS --
    build_lookups(wb, INTEL_LOOKUPS, INTEL_GOLD)

    # Save
    path = os.path.join(OUTPUT_DIR, "FNID_Intel_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  2. OPERATIONS UNIT PORTAL
# =================================================================
def build_operations_portal():
    """Generate FNID_Operations_Portal.xlsx"""
    wb = Workbook()

    sheets = [
        ("Op_Register", "Operation planning & execution tracking"),
        ("Dashboard", "KPI summary and operational charts"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Operations", OPS_RED, sheets)

    # -- OPERATIONS REGISTER --
    ws = wb.create_sheet("Op_Register")
    ws.sheet_properties.tabColor = OPS_RED
    headers = [
        "OpID", "OpName", "OpDate", "OpType", "WarrantBasis",
        "WarrantNumber", "IssuingJP", "Parish", "TargetLocation",
        "TargetPerson", "TeamLead", "TeamSize", "JointAgency",
        "LinkedIntelID", "StartTime", "EndTime", "Duration_hrs",
        "FirearmsSeized", "NarcoticsSeized", "AmmoSeized", "CashSeized",
        "ArrestsMade", "Outcome", "OutcomeNotes", "RiskAssessment",
        "BodyCamFootage", "EvidenceTagged",
    ]
    widths = [
        14, 20, 12, 22, 24, 14, 16, 14, 22, 18,
        16, 10, 16, 14, 10, 10, 10, 10, 10, 10, 12,
        10, 22, 28, 14, 10, 10,
    ]
    write_headers(ws, headers, widths)

    # Sample operations data
    sample_ops = [
        ["OP-2025-001", "Op THUNDER", "2025-01-20", "Search Warrant Execution",
         "Section 91 - Firearms Act 2022", "WN-2025-011", "JP Sinclair",
         "Manchester", "Lot 14 Green Ave, Mandeville", "Unknown",
         "Insp. Brown", 8, "None", "INT-2025-001", "05:00", "08:30", 3.5,
         2, 0, 45, 0, 1, "Successful - Seizure + Arrest",
         "2 pistols and ammo recovered, 1 male arrested",
         "High", "Yes", "Yes"],
        ["OP-2025-002", "Op NEPTUNE", "2025-01-25", "Coastal/Beach Operation",
         "Section 21 - Dangerous Drugs Act", "", "",
         "St. Elizabeth", "Black River Bay", "Marcus COLE",
         "Supt. Davis", 15, "JDF Coast Guard", "INT-2025-002", "22:00", "03:00", 5,
         0, 1, 0, 0, 0, "Successful - Seizure Only",
         "32kg cocaine recovered from fishing vessel",
         "Critical", "Yes", "Yes"],
        ["OP-2025-003", "Op CONTAINER", "2025-02-10", "Port/Cargo Interdiction",
         "Section 91 - Firearms Act 2022", "", "",
         "St. Catherine", "Kingston Logistics Centre", "Ryan SMITH",
         "Insp. Brown", 6, "JCA", "INT-2025-003", "10:00", "16:00", 6,
         0, 0, 0, 0, 0, "Ongoing",
         "Container under surveillance",
         "High", "N/A", "N/A"],
        ["OP-2025-004", "Op VIPER", "2025-03-05", "Search Warrant Execution",
         "Section 91 - Firearms Act 2022", "WN-2025-018", "JP Clarke",
         "Manchester", "Christiana district", "Devon GRANT",
         "Supt. Davis", 20, "MOCA", "INT-2025-006", "04:30", "09:00", 4.5,
         5, 1, 200, 850000, 3, "Successful - Seizure + Arrest",
         "Major seizure: 5 firearms, 2kg cocaine, $850K cash, 3 arrested",
         "Critical", "Yes", "Yes"],
        ["OP-2025-005", "Op BARREL", "2025-03-18", "Port/Cargo Interdiction",
         "Section 91 - Firearms Act 2022", "", "",
         "St. Catherine", "Kingston Logistics Centre", "ABC Ltd",
         "Insp. Brown", 4, "JCA", "INT-2025-008", "09:00", "12:00", 3,
         0, 0, 500, 0, 0, "Successful - Seizure Only",
         "500 rounds 9mm in household barrel",
         "Medium", "Yes", "Yes"],
    ]

    for i, row_data in enumerate(sample_ops):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws, len(headers), start_row=len(sample_ops)+2, count=20)
    n_sub = add_submission_tracking(ws, len(headers), len(sample_ops)+20)
    end_row = len(sample_ops) + 1 + 20

    # Mark samples as submitted
    sc = len(headers) + 1
    for r in range(2, len(sample_ops)+2):
        ws.cell(row=r, column=sc, value="Submitted")
        ws.cell(row=r, column=sc+1, value="System")

    # Dropdowns
    dv_list(ws, "D2:D500", "Lookups!$F$2:$F$20")
    dv_list(ws, "E2:E500", "Lookups!$G$2:$G$10")
    dv_list(ws, "H2:H500", "Lookups!$C$2:$C$20")
    dv_list(ws, "W2:W500", "Lookups!$H$2:$H$12")

    add_table(ws, "tbl_Operations",
              f"A1:{get_column_letter(len(headers))}{end_row}")

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = OPS_RED
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:K2")
    dash.cell(row=2, column=2,
              value="OPERATIONS UNIT DASHBOARD - FNID Area 3").font = TITLE_FONT
    dash["B2"].alignment = CENTER

    kpi_card(dash, 4, 2, "Total Operations",
             f"=COUNTA(Op_Register!A2:A{end_row})")
    kpi_card(dash, 4, 5, "Successful Ops",
             f'=COUNTIF(Op_Register!W2:W{end_row},"Successful*")')
    kpi_card(dash, 4, 8, "Total Arrests",
             f"=SUM(Op_Register!V2:V{end_row})")
    kpi_card(dash, 4, 11, "Firearms Seized",
             f"=SUM(Op_Register!R2:R{end_row})")

    # Op type chart
    op_types = ["Search Warrant Execution", "Port/Cargo Interdiction",
                "Coastal/Beach Operation", "Vehicle Interception",
                "Checkpoint", "Joint Op - MOCA", "Joint Op - JCA"]
    r = 8
    dash.cell(row=r, column=2, value="Op Type").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, ot in enumerate(op_types):
        dash.cell(row=r+1+i, column=2, value=ot).font = BODY_FONT
        dash.cell(row=r+1+i, column=3,
                  value=f'=COUNTIF(Op_Register!D2:D{end_row},B{r+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Operations by Type",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(op_types)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(op_types)},
                  anchor="E8", horizontal=True)

    # Outcome pie
    outcomes = ["Successful - Seizure + Arrest", "Successful - Seizure Only",
                "Successful - Arrest Only", "Partial - Intel Developed",
                "Negative - No Finds", "Ongoing"]
    pr = r + len(op_types) + 3
    dash.cell(row=pr, column=2, value="Outcome").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, o in enumerate(outcomes):
        dash.cell(row=pr+1+i, column=2, value=o).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Op_Register!W2:W{end_row},B{pr+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "Operation Outcomes",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+len(outcomes)},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+len(outcomes)},
                  anchor=f"E{pr}")

    # Monthly trend
    mr = pr + len(outcomes) + 3
    dash.cell(row=mr, column=2, value="Month").font = BODY_BOLD
    dash.cell(row=mr, column=3, value="Ops Count").font = BODY_BOLD
    for i, m in enumerate(MONTHS_SHORT):
        dash.cell(row=mr+1+i, column=2, value=m).font = BODY_FONT
        mn = str(i+1).zfill(2)
        nx = str(i+2).zfill(2) if i < 11 else "01"
        yr = "2025" if i < 11 else "2026"
        dash.cell(row=mr+1+i, column=3,
                  value=f'=COUNTIFS(Op_Register!C2:C{end_row},">="&"2025-{mn}-01",Op_Register!C2:C{end_row},"<"&"{yr}-{nx}-01")'
                  ).font = BODY_FONT

    add_line_chart(dash, "Monthly Operations Trend",
                   cats_ref={"min_col": 2, "min_row": mr+1, "max_row": mr+12},
                   data_refs={"min_col": 3, "min_row": mr+1, "max_row": mr+12},
                   anchor=f"E{mr}", width=20)

    build_lookups(wb, OPERATIONS_LOOKUPS, OPS_RED)

    path = os.path.join(OUTPUT_DIR, "FNID_Operations_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  3. SEIZURES UNIT PORTAL
# =================================================================
def build_seizures_portal():
    """Generate FNID_Seizures_Portal.xlsx with Firearms, Narcotics, Other sheets."""
    wb = Workbook()

    sheets = [
        ("Firearm_Seizures", "Firearm and ammunition seizure log"),
        ("Narcotics_Seizures", "Narcotics seizure log"),
        ("Other_Seizures", "Cash, electronics, vehicle seizures"),
        ("Dashboard", "Seizure KPIs and charts"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Seizures", NARCO_TEAL, sheets)

    DATA_ROWS = 20

    # -- FIREARM SEIZURES --
    ws = wb.create_sheet("Firearm_Seizures")
    ws.sheet_properties.tabColor = NARCO_TEAL
    fa_headers = [
        "SeizureID", "SeizureDate", "LinkedOpID", "Parish", "Location",
        "FirearmType", "Make", "Model", "SerialNumber", "Calibre",
        "CountryOfOrigin", "AmmoCount", "MagazineCount",
        "IBIS_Status", "eTrace_Status", "ExhibitTag", "StorageLocation",
        "SeizedBy", "WitnessOfficer", "Notes",
    ]
    fa_widths = [
        14, 12, 14, 14, 22, 18, 14, 14, 16, 12,
        16, 10, 10, 16, 16, 14, 16, 16, 16, 28,
    ]
    write_headers(ws, fa_headers, fa_widths)

    sample_firearms = [
        ["FS-2025-001", "2025-01-20", "OP-2025-001", "Manchester",
         "Lot 14 Green Ave, Mandeville", "Pistol - Semi-Auto",
         "Glock", "19", "ABC12345", "9mm", "USA", 15, 1,
         "Submitted - Pending", "Submitted - Pending", "EXH-FA-001",
         "FNID Armoury", "DC Williams", "Cpl. James",
         "Found buried in backyard"],
        ["FS-2025-002", "2025-01-20", "OP-2025-001", "Manchester",
         "Lot 14 Green Ave, Mandeville", "Pistol - Revolver",
         "Smith & Wesson", "686", "XYZ98765", ".38 Special", "USA", 30, 0,
         "Submitted - Pending", "Traced - US Origin", "EXH-FA-002",
         "FNID Armoury", "DC Williams", "Cpl. James",
         "Loaded, found under mattress"],
        ["FS-2025-003", "2025-03-05", "OP-2025-004", "Manchester",
         "Christiana district", "Rifle - Assault",
         "Unknown", "AR-15 pattern", "OBLITERATED", "5.56mm/.223", "Unknown", 60, 2,
         "Submitted - Pending", "Untraceable - Obliterated", "EXH-FA-003",
         "FNID Armoury", "Supt. Davis", "Insp. Brown",
         "Serial number filed off"],
        ["FS-2025-004", "2025-03-05", "OP-2025-004", "Manchester",
         "Christiana district", "Shotgun",
         "Mossberg", "500", "MOB55443", "12 Gauge", "USA", 25, 0,
         "No Match", "Traced - US Origin", "EXH-FA-004",
         "FNID Armoury", "DC Thompson", "Cpl. James", ""],
        ["FS-2025-005", "2025-03-05", "OP-2025-004", "Manchester",
         "Christiana district", "Pistol - Semi-Auto",
         "Taurus", "G2C", "TAU99001", "9mm", "Brazil", 17, 1,
         "Hit - Confirmed Match", "Traced - Non-US Origin", "EXH-FA-005",
         "FNID Armoury", "DC Thompson", "Insp. Brown",
         "IBIS hit - linked to 2024 shooting in Kingston"],
    ]

    for i, row_data in enumerate(sample_firearms):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws, len(fa_headers), start_row=len(sample_firearms)+2, count=DATA_ROWS)
    add_submission_tracking(ws, len(fa_headers), len(sample_firearms)+DATA_ROWS)
    fa_end = len(sample_firearms) + 1 + DATA_ROWS
    add_table(ws, "tbl_FirearmSeizures",
              f"A1:{get_column_letter(len(fa_headers))}{fa_end}")

    dv_list(ws, "D2:D500", "Lookups!$C$2:$C$20")
    dv_list(ws, "F2:F500", "Lookups!$F$2:$F$20")
    dv_list(ws, "J2:J500", "Lookups!$G$2:$G$20")
    dv_list(ws, "N2:N500", "Lookups!$G$2:$G$10")  # IBIS reuse

    # -- NARCOTICS SEIZURES --
    ws2 = wb.create_sheet("Narcotics_Seizures")
    ws2.sheet_properties.tabColor = NARCO_TEAL
    na_headers = [
        "SeizureID", "SeizureDate", "LinkedOpID", "Parish", "Location",
        "DrugType", "Quantity", "Unit", "EstStreetValue_JMD",
        "PackagingMethod", "ConcealmentMethod", "FieldTestResult",
        "LabCertStatus", "ExhibitTag", "StorageLocation",
        "SeizedBy", "WitnessOfficer", "Notes",
    ]
    na_widths = [
        14, 12, 14, 14, 22, 22, 10, 8, 16,
        18, 20, 14, 16, 14, 16, 16, 16, 28,
    ]
    write_headers(ws2, na_headers, na_widths)

    sample_narcotics = [
        ["NS-2025-001", "2025-01-25", "OP-2025-002", "St. Elizabeth",
         "Black River Bay", "Cocaine - Powder", 32, "kg", 96000000,
         "Brick wrapped in tape", "Hidden in fish hold", "Positive",
         "Submitted to IFSLM", "EXH-NA-001", "FNID Evidence Room",
         "Supt. Davis", "Insp. Morgan", "32 bricks recovered from vessel"],
        ["NS-2025-002", "2025-03-05", "OP-2025-004", "Manchester",
         "Christiana district", "Cocaine - Powder", 2, "kg", 6000000,
         "Vacuum sealed bags", "False wall in house", "Positive",
         "Certificate Issued", "EXH-NA-002", "FNID Evidence Room",
         "DC Williams", "Cpl. James", "Found with firearms cache"],
        ["NS-2025-003", "2025-03-22", "OP-2025-006", "Clarendon",
         "Highway 2000 checkpoint", "Cocaine - Powder", 0.2, "kg", 600000,
         "Small bags", "Vehicle hidden compartment", "Positive",
         "Submitted to IFSLM", "EXH-NA-003", "FNID Evidence Room",
         "DC Williams", "Cpl. James", "Found with pistol and cash"],
    ]

    for i, row_data in enumerate(sample_narcotics):
        for j, val in enumerate(row_data):
            cell = ws2.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws2, len(na_headers), start_row=len(sample_narcotics)+2, count=DATA_ROWS)
    add_submission_tracking(ws2, len(na_headers), len(sample_narcotics)+DATA_ROWS)
    na_end = len(sample_narcotics) + 1 + DATA_ROWS
    add_table(ws2, "tbl_NarcoticSeizures",
              f"A1:{get_column_letter(len(na_headers))}{na_end}")

    dv_list(ws2, "D2:D500", "Lookups!$C$2:$C$20")
    dv_list(ws2, "F2:F500", "Lookups!$H$2:$H$15")
    dv_list(ws2, "H2:H500", "Lookups!$I$2:$I$12")

    # -- OTHER SEIZURES --
    ws3 = wb.create_sheet("Other_Seizures")
    ws3.sheet_properties.tabColor = NARCO_TEAL
    ot_headers = [
        "SeizureID", "SeizureDate", "LinkedOpID", "Parish", "Location",
        "ItemType", "Description", "Quantity", "EstValue_JMD",
        "ExhibitTag", "StorageLocation", "SeizedBy", "WitnessOfficer", "Notes",
    ]
    ot_widths = [14, 12, 14, 14, 22, 16, 28, 10, 14, 14, 16, 16, 16, 28]
    write_headers(ws3, ot_headers, ot_widths)

    sample_other = [
        ["OS-2025-001", "2025-03-05", "OP-2025-004", "Manchester",
         "Christiana district", "Cash/Currency", "JMD banknotes in bag",
         1, 850000, "EXH-OT-001", "FNID Evidence Room",
         "DC Thompson", "Insp. Brown", "Seized with firearms and drugs"],
        ["OS-2025-002", "2025-03-22", "OP-2025-006", "Clarendon",
         "Highway 2000", "Cash/Currency", "JMD and USD mixed",
         1, 500000, "EXH-OT-002", "FNID Evidence Room",
         "DC Williams", "Cpl. James", "Found in vehicle compartment"],
    ]
    for i, row_data in enumerate(sample_other):
        for j, val in enumerate(row_data):
            cell = ws3.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws3, len(ot_headers), start_row=len(sample_other)+2, count=DATA_ROWS)
    add_submission_tracking(ws3, len(ot_headers), len(sample_other)+DATA_ROWS)
    ot_end = len(sample_other) + 1 + DATA_ROWS
    add_table(ws3, "tbl_OtherSeizures",
              f"A1:{get_column_letter(len(ot_headers))}{ot_end}")

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = NARCO_TEAL
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:K2")
    dash.cell(row=2, column=2,
              value="SEIZURES UNIT DASHBOARD - FNID Area 3").font = TITLE_FONT
    dash["B2"].alignment = CENTER

    kpi_card(dash, 4, 2, "Firearms Seized",
             f"=COUNTA(Firearm_Seizures!A2:A{fa_end})")
    kpi_card(dash, 4, 5, "Narcotics (kg)",
             f'=SUMIF(Narcotics_Seizures!H2:H{na_end},"kg",Narcotics_Seizures!G2:G{na_end})')
    kpi_card(dash, 4, 8, "IBIS Hits",
             f'=COUNTIF(Firearm_Seizures!N2:N{fa_end},"Hit*")')
    kpi_card(dash, 4, 11, "Total Exhibits",
             f"=COUNTA(Firearm_Seizures!P2:P{fa_end})+COUNTA(Narcotics_Seizures!N2:N{na_end})+COUNTA(Other_Seizures!J2:J{ot_end})")

    # Firearm type breakdown
    fa_types = ["Pistol - Semi-Auto", "Pistol - Revolver", "Rifle - Assault",
                "Shotgun", "Submachine Gun", "Improvised/Homemade"]
    r = 8
    dash.cell(row=r, column=2, value="Firearm Type").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, ft in enumerate(fa_types):
        dash.cell(row=r+1+i, column=2, value=ft).font = BODY_FONT
        dash.cell(row=r+1+i, column=3,
                  value=f'=COUNTIF(Firearm_Seizures!F2:F{fa_end},B{r+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "Firearms by Type",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(fa_types)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(fa_types)},
                  anchor="E8")

    # Drug type breakdown
    drug_types = ["Cannabis/Ganja - Compressed", "Cannabis/Ganja - Loose",
                  "Cocaine - Powder", "Cocaine - Crack", "Heroin", "Other Controlled"]
    pr = r + len(fa_types) + 3
    dash.cell(row=pr, column=2, value="Drug Type").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, dt in enumerate(drug_types):
        dash.cell(row=pr+1+i, column=2, value=dt).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Narcotics_Seizures!F2:F{na_end},B{pr+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "Narcotics by Type",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+len(drug_types)},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+len(drug_types)},
                  anchor=f"E{pr}")

    # Parish distribution
    parishes = ["Manchester", "St. Elizabeth", "Clarendon"]
    qr = pr + len(drug_types) + 3
    dash.cell(row=qr, column=2, value="Parish").font = BODY_BOLD
    dash.cell(row=qr, column=3, value="Firearms").font = BODY_BOLD
    dash.cell(row=qr, column=4, value="Narcotics").font = BODY_BOLD
    for i, p in enumerate(parishes):
        dash.cell(row=qr+1+i, column=2, value=p).font = BODY_FONT
        dash.cell(row=qr+1+i, column=3,
                  value=f'=COUNTIF(Firearm_Seizures!D2:D{fa_end},B{qr+1+i})'
                  ).font = BODY_FONT
        dash.cell(row=qr+1+i, column=4,
                  value=f'=COUNTIF(Narcotics_Seizures!D2:D{na_end},B{qr+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Seizures by Parish",
                  cats_ref={"min_col": 2, "min_row": qr+1, "max_row": qr+3},
                  vals_ref={"min_col": 3, "min_row": qr+1, "max_row": qr+3},
                  anchor=f"E{qr}")

    build_lookups(wb, SEIZURES_LOOKUPS, NARCO_TEAL)

    path = os.path.join(OUTPUT_DIR, "FNID_Seizures_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  4. ARRESTS & COURT UNIT PORTAL
# =================================================================
def build_arrests_court_portal():
    """Generate FNID_ArrestsCourt_Portal.xlsx"""
    wb = Workbook()

    sheets = [
        ("Arrest_Register", "Arrest log with 48-hour compliance"),
        ("Bail_Court", "Bail status and court tracking"),
        ("Dashboard", "Arrest KPIs and court pipeline charts"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Arrests & Court", COURT_PURPLE, sheets)
    DATA_ROWS = 20

    # -- ARREST REGISTER --
    ws = wb.create_sheet("Arrest_Register")
    ws.sheet_properties.tabColor = COURT_PURPLE
    ar_headers = [
        "ArrestID", "ArrestDate", "ArrestTime", "LinkedOpID",
        "SuspectName", "SuspectDOB", "SuspectAddress", "SuspectOccupation",
        "Parish", "ArrestLocation", "ArrestingOfficer", "ArrestingOfficerRank",
        "Offence1", "LawAndSection1", "Offence2", "LawAndSection2",
        "48hr_Deadline", "ChargeDate", "ChargeWithin48hr",
        "BailStatus", "CourtType", "FirstCourtDate",
        "RemandLocation", "LegalRepresentation",
        "StatementsTaken", "WitnessCount", "Notes",
    ]
    ar_widths = [
        14, 12, 10, 14, 20, 12, 24, 16,
        14, 22, 18, 14,
        28, 28, 28, 28,
        14, 12, 10,
        18, 24, 14,
        18, 18,
        10, 10, 28,
    ]
    write_headers(ws, ar_headers, ar_widths)

    sample_arrests = [
        ["AR-2025-001", "2025-01-20", "06:45", "OP-2025-001",
         "John BROWN", "1990-05-12", "15 Main St, Mandeville", "Labourer",
         "Manchester", "Lot 14 Green Ave", "DC Williams", "Detective Corporal",
         "Illegal Possession of Firearm", "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "Illegal Possession of Ammunition", "Illegal Possession of Ammunition",
         "2025-01-22 06:45", "2025-01-20", "Yes",
         "Remanded in Custody", "Gun Court - High Court Division", "2025-02-03",
         "Tower Street Adult Correctional", "Legal Aid Council",
         "Yes", 3, "Found at premises during warrant execution"],
        ["AR-2025-002", "2025-03-05", "05:15", "OP-2025-004",
         "Devon GRANT", "1985-08-23", "Christiana, Manchester", "Unemployed",
         "Manchester", "Christiana district", "Supt. Davis", "Superintendent",
         "s.6 - Stockpiling (3+ weapons / 50+ rounds)",
         "s.6 - Stockpiling (3+ weapons / 50+ rounds)",
         "Dealing/Trafficking - Cocaine", "Dealing/Trafficking - Cocaine",
         "2025-03-07 05:15", "2025-03-05", "Yes",
         "Bail Denied", "Gun Court - High Court Division", "2025-03-19",
         "Tower Street Adult Correctional", "Private Attorney",
         "Yes", 5, "Major target - MOCA joint operation"],
        ["AR-2025-003", "2025-03-05", "05:30", "OP-2025-004",
         "Michael REID", "1992-11-07", "Christiana, Manchester", "Fisherman",
         "Manchester", "Christiana district", "DC Thompson", "Detective Corporal",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "", "",
         "2025-03-07 05:30", "2025-03-06", "Yes",
         "Remanded in Custody", "Gun Court - RM Division", "2025-03-20",
         "Mandeville Police Lock-Up", "None",
         "Yes", 2, "Associate of GRANT"],
        ["AR-2025-004", "2025-03-22", "14:00", "OP-2025-006",
         "Andrew BLAKE", "1988-02-15", "May Pen, Clarendon", "Driver",
         "Clarendon", "Highway 2000 checkpoint", "DC Williams", "Detective Corporal",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "Possession - Cocaine", "Possession - Cocaine",
         "2025-03-24 14:00", "2025-03-22", "Yes",
         "Bail Granted", "Gun Court - RM Division", "2025-04-05",
         "", "Private Attorney",
         "Yes", 2, "Vehicle checkpoint intercept"],
    ]

    for i, row_data in enumerate(sample_arrests):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws, len(ar_headers), start_row=len(sample_arrests)+2, count=DATA_ROWS)
    add_submission_tracking(ws, len(ar_headers), len(sample_arrests)+DATA_ROWS)
    ar_end = len(sample_arrests) + 1 + DATA_ROWS

    # 48-hour compliance conditional formatting
    ws.conditional_formatting.add(
        f"S2:S{ar_end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(
        f"S2:S{ar_end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=RED_FILL))

    add_table(ws, "tbl_Arrests",
              f"A1:{get_column_letter(len(ar_headers))}{ar_end}")

    dv_list(ws, "I2:I500", "Lookups!$C$2:$C$20")
    dv_list(ws, "M2:M500", "Lookups!$F$2:$F$15")
    dv_list(ws, "O2:O500", "Lookups!$G$2:$G$12")
    dv_list(ws, "S2:S500", "Lookups!$D$2:$D$5")
    dv_list(ws, "T2:T500", "Lookups!$H$2:$H$12")
    dv_list(ws, "U2:U500", "Lookups!$I$2:$I$10")

    # -- BAIL & COURT --
    ws2 = wb.create_sheet("Bail_Court")
    ws2.sheet_properties.tabColor = COURT_PURPLE
    bc_headers = [
        "ArrestID", "SuspectName", "CourtType", "CourtDate",
        "BailStatus", "BailAmount_JMD", "BailConditions",
        "NextHearingDate", "CrownCounsel", "Defence",
        "Plea", "Verdict", "Sentence", "AppealFiled", "Notes",
    ]
    bc_widths = [14, 20, 24, 12, 18, 14, 24, 14, 18, 18, 10, 14, 20, 10, 28]
    write_headers(ws2, bc_headers, bc_widths)
    data_rows(ws2, len(bc_headers), count=DATA_ROWS)
    add_submission_tracking(ws2, len(bc_headers), DATA_ROWS)
    bc_end = 1 + DATA_ROWS
    add_table(ws2, "tbl_BailCourt",
              f"A1:{get_column_letter(len(bc_headers))}{bc_end}")

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = COURT_PURPLE
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:K2")
    dash.cell(row=2, column=2,
              value="ARRESTS & COURT DASHBOARD - FNID Area 3").font = TITLE_FONT
    dash["B2"].alignment = CENTER

    kpi_card(dash, 4, 2, "Total Arrests",
             f"=COUNTA(Arrest_Register!A2:A{ar_end})")
    kpi_card(dash, 4, 5, "48hr Compliant",
             f'=COUNTIF(Arrest_Register!S2:S{ar_end},"Yes")')
    kpi_card(dash, 4, 8, "Bail Denied",
             f'=COUNTIF(Arrest_Register!T2:T{ar_end},"Bail Denied")')
    kpi_card(dash, 4, 11, "Remanded",
             f'=COUNTIF(Arrest_Register!T2:T{ar_end},"Remanded*")')

    # Bail status breakdown
    bail_statuses = ["Bail Granted", "Bail Denied", "Remanded in Custody",
                     "Released - No Charge", "Stop Order Issued"]
    r = 8
    dash.cell(row=r, column=2, value="Bail Status").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, bs in enumerate(bail_statuses):
        dash.cell(row=r+1+i, column=2, value=bs).font = BODY_FONT
        dash.cell(row=r+1+i, column=3,
                  value=f'=COUNTIF(Arrest_Register!T2:T{ar_end},B{r+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "Bail Status Distribution",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(bail_statuses)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(bail_statuses)},
                  anchor="E8")

    # Court type breakdown
    courts = ["Gun Court - High Court Division", "Gun Court - RM Division",
              "Parish Court", "Circuit Court"]
    pr = r + len(bail_statuses) + 3
    dash.cell(row=pr, column=2, value="Court Type").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, ct in enumerate(courts):
        dash.cell(row=pr+1+i, column=2, value=ct).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Arrest_Register!U2:U{ar_end},B{pr+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Cases by Court Type",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+len(courts)},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+len(courts)},
                  anchor=f"E{pr}")

    # Monthly arrest trend
    mr = pr + len(courts) + 3
    dash.cell(row=mr, column=2, value="Month").font = BODY_BOLD
    dash.cell(row=mr, column=3, value="Arrests").font = BODY_BOLD
    for i, m in enumerate(MONTHS_SHORT):
        dash.cell(row=mr+1+i, column=2, value=m).font = BODY_FONT
        mn = str(i+1).zfill(2)
        nx = str(i+2).zfill(2) if i < 11 else "01"
        yr = "2025" if i < 11 else "2026"
        dash.cell(row=mr+1+i, column=3,
                  value=f'=COUNTIFS(Arrest_Register!B2:B{ar_end},">="&"2025-{mn}-01",Arrest_Register!B2:B{ar_end},"<"&"{yr}-{nx}-01")'
                  ).font = BODY_FONT

    add_line_chart(dash, "Monthly Arrests Trend",
                   cats_ref={"min_col": 2, "min_row": mr+1, "max_row": mr+12},
                   data_refs={"min_col": 3, "min_row": mr+1, "max_row": mr+12},
                   anchor=f"E{mr}", width=20)

    build_lookups(wb, ARRESTS_LOOKUPS, COURT_PURPLE)

    path = os.path.join(OUTPUT_DIR, "FNID_ArrestsCourt_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  5. FORENSICS & EVIDENCE UNIT PORTAL
# =================================================================
def build_forensics_portal():
    """Generate FNID_Forensics_Portal.xlsx"""
    wb = Workbook()

    sheets = [
        ("Chain_of_Custody", "Exhibit chain of custody log"),
        ("Lab_Tracking", "Forensic lab submission and certificate tracking"),
        ("Dashboard", "Forensic pipeline KPIs and charts"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Forensics & Evidence", EVIDENCE_GREEN, sheets)
    DATA_ROWS = 20

    # -- CHAIN OF CUSTODY --
    ws = wb.create_sheet("Chain_of_Custody")
    ws.sheet_properties.tabColor = EVIDENCE_GREEN
    coc_headers = [
        "ExhibitTag", "ExhibitType", "Description", "LinkedCaseID",
        "LinkedSeizureID", "SeizedDate", "SeizedBy", "SeizedLocation",
        "CurrentCustodian", "StorageLocation", "TransferDate",
        "TransferFrom", "TransferTo", "TransferReason",
        "Condition", "PhotosTaken", "SealIntact", "DisposalStatus", "Notes",
    ]
    coc_widths = [
        14, 16, 28, 14, 14, 12, 16, 22,
        16, 18, 12, 16, 16, 20,
        14, 10, 10, 22, 28,
    ]
    write_headers(ws, coc_headers, coc_widths)

    sample_coc = [
        ["EXH-FA-001", "Firearm", "Glock 19 9mm pistol", "CR-2025-001",
         "FS-2025-001", "2025-01-20", "DC Williams", "Mandeville",
         "FNID Armoury NCO", "FNID Armoury Cage A", "2025-01-21",
         "DC Williams", "FNID Armoury NCO", "Evidence booking",
         "Good", "Yes", "Yes", "Held - Active Case", ""],
        ["EXH-FA-005", "Firearm", "Taurus G2C 9mm - IBIS Hit", "CR-2025-003",
         "FS-2025-005", "2025-03-05", "DC Thompson", "Christiana",
         "IFSLM Ballistics", "IFSLM Lab", "2025-03-08",
         "FNID Armoury NCO", "IFSLM Ballistics", "IBIS confirmation testing",
         "Good", "Yes", "Yes", "Held - Active Case",
         "IBIS hit linked to Kingston shooting"],
        ["EXH-NA-001", "Drug Sample", "32kg cocaine from vessel", "CR-2025-002",
         "NS-2025-001", "2025-01-25", "Supt. Davis", "Black River Bay",
         "IFSLM Chemistry", "IFSLM Lab", "2025-01-28",
         "FNID Evidence NCO", "IFSLM Chemistry", "Purity analysis",
         "Sealed", "Yes", "Yes", "Held - Active Case",
         "Bulk stored at FNID, sample sent to IFSLM"],
    ]

    for i, row_data in enumerate(sample_coc):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws, len(coc_headers), start_row=len(sample_coc)+2, count=DATA_ROWS)
    add_submission_tracking(ws, len(coc_headers), len(sample_coc)+DATA_ROWS)
    coc_end = len(sample_coc) + 1 + DATA_ROWS
    add_table(ws, "tbl_ChainOfCustody",
              f"A1:{get_column_letter(len(coc_headers))}{coc_end}")

    dv_list(ws, "B2:B500", "Lookups!$F$2:$F$15")
    dv_list(ws, "R2:R500", "Lookups!$J$2:$J$10")

    # -- LAB TRACKING --
    ws2 = wb.create_sheet("Lab_Tracking")
    ws2.sheet_properties.tabColor = EVIDENCE_GREEN
    lab_headers = [
        "LabRef", "ExhibitTag", "LinkedCaseID", "SubmissionDate",
        "LabType", "ExamType", "Analyst", "ExpectedDate",
        "ActualCompletionDate", "CertificateNumber", "CertificateStatus",
        "Result", "CollectedBy", "CollectionDate",
        "IBIS_Status", "eTrace_Status", "Notes",
    ]
    lab_widths = [
        14, 14, 14, 12, 14, 18, 16, 12,
        14, 16, 16, 20, 16, 12,
        16, 16, 28,
    ]
    write_headers(ws2, lab_headers, lab_widths)

    sample_lab = [
        ["LAB-2025-001", "EXH-FA-001", "CR-2025-001", "2025-01-22",
         "Ballistics", "Firearm identification & test fire", "Dr. Patel", "2025-03-22",
         "", "", "Submitted to IFSLM", "",
         "", "", "Submitted - Pending", "Submitted - Pending",
         "Standard turnaround 8 weeks"],
        ["LAB-2025-002", "EXH-FA-005", "CR-2025-003", "2025-03-08",
         "Ballistics", "IBIS confirmation + test fire", "Dr. Patel", "2025-05-08",
         "", "", "Analysis In Progress", "",
         "", "", "Hit - Confirmed Match", "Traced - Non-US Origin",
         "Priority - linked to Kingston shooting"],
        ["LAB-2025-003", "EXH-NA-001", "CR-2025-002", "2025-01-28",
         "Chemistry", "Drug purity analysis", "Ms. Wong", "2025-03-28",
         "2025-03-15", "CERT-CHEM-2025-044", "Certificate Issued",
         "Cocaine HCl, 87% purity",
         "DC Thompson", "2025-03-16", "", "",
         "Certificate available for DPP file"],
    ]

    for i, row_data in enumerate(sample_lab):
        for j, val in enumerate(row_data):
            cell = ws2.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws2, len(lab_headers), start_row=len(sample_lab)+2, count=DATA_ROWS)
    add_submission_tracking(ws2, len(lab_headers), len(sample_lab)+DATA_ROWS)
    lab_end = len(sample_lab) + 1 + DATA_ROWS
    add_table(ws2, "tbl_LabTracking",
              f"A1:{get_column_letter(len(lab_headers))}{lab_end}")

    dv_list(ws2, "K2:K500", "Lookups!$I$2:$I$10")
    dv_list(ws2, "O2:O500", "Lookups!$G$2:$G$10")
    dv_list(ws2, "P2:P500", "Lookups!$H$2:$H$10")

    # Overdue highlight
    ws2.conditional_formatting.add(
        f"K2:K{lab_end}",
        CellIsRule(operator="equal", formula=['"Certificate Overdue"'], fill=RED_FILL))

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = EVIDENCE_GREEN
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:K2")
    dash.cell(row=2, column=2,
              value="FORENSICS & EVIDENCE DASHBOARD - FNID Area 3").font = TITLE_FONT
    dash["B2"].alignment = CENTER

    kpi_card(dash, 4, 2, "Total Exhibits",
             f"=COUNTA(Chain_of_Custody!A2:A{coc_end})")
    kpi_card(dash, 4, 5, "Certs Issued",
             f'=COUNTIF(Lab_Tracking!K2:K{lab_end},"Certificate Issued")')
    kpi_card(dash, 4, 8, "Pending at Lab",
             f'=COUNTIF(Lab_Tracking!K2:K{lab_end},"*IFSLM*")+COUNTIF(Lab_Tracking!K2:K{lab_end},"Analysis*")')
    kpi_card(dash, 4, 11, "Overdue",
             f'=COUNTIF(Lab_Tracking!K2:K{lab_end},"*Overdue*")')

    # Exhibit type breakdown
    ex_types = ["Firearm", "Ammunition", "Drug Sample", "Cash/Currency",
                "Electronic Device", "Other"]
    r = 8
    dash.cell(row=r, column=2, value="Exhibit Type").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, et in enumerate(ex_types):
        dash.cell(row=r+1+i, column=2, value=et).font = BODY_FONT
        dash.cell(row=r+1+i, column=3,
                  value=f'=COUNTIF(Chain_of_Custody!B2:B{coc_end},B{r+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "Exhibits by Type",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(ex_types)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(ex_types)},
                  anchor="E8")

    # Lab status pipeline
    lab_stats = ["Not Yet Submitted", "Submitted to IFSLM", "Analysis In Progress",
                 "Certificate Issued", "Certificate Overdue"]
    pr = r + len(ex_types) + 3
    dash.cell(row=pr, column=2, value="Lab Status").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, ls in enumerate(lab_stats):
        dash.cell(row=pr+1+i, column=2, value=ls).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Lab_Tracking!K2:K{lab_end},B{pr+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Forensic Certificate Pipeline",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+len(lab_stats)},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+len(lab_stats)},
                  anchor=f"E{pr}", horizontal=True)

    # IBIS/eTrace status
    ibis_stats = ["Not Submitted", "Submitted - Pending", "No Match",
                  "Hit - Confirmed Match", "Serial Obliterated"]
    qr = pr + len(lab_stats) + 3
    dash.cell(row=qr, column=2, value="IBIS Status").font = BODY_BOLD
    dash.cell(row=qr, column=3, value="Count").font = BODY_BOLD
    for i, ib in enumerate(ibis_stats):
        dash.cell(row=qr+1+i, column=2, value=ib).font = BODY_FONT
        dash.cell(row=qr+1+i, column=3,
                  value=f'=COUNTIF(Lab_Tracking!O2:O{lab_end},B{qr+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "IBIS Status Distribution",
                  cats_ref={"min_col": 2, "min_row": qr+1, "max_row": qr+len(ibis_stats)},
                  vals_ref={"min_col": 3, "min_row": qr+1, "max_row": qr+len(ibis_stats)},
                  anchor=f"E{qr}")

    build_lookups(wb, FORENSICS_LOOKUPS, EVIDENCE_GREEN)

    path = os.path.join(OUTPUT_DIR, "FNID_Forensics_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  6. REGISTRY UNIT PORTAL - Case Management & SOP Compliance
#
#  Based on JCF case management requirements and DPP protocols:
#    - Case file completeness tracking per DPP evidential sufficiency test
#    - SOP compliance checklist (statements, forensics, exhibits, disclosure)
#    - 48-hour rule compliance tracking
#    - DPP submission pipeline with ruling tracking
#    - Disclosure protocol compliance (DPP September 2013)
#    - Witness and statement management
#    - Court scheduling and outcome tracking
# =================================================================
def build_registry_portal():
    """Generate FNID_Registry_Portal.xlsx - Case Management SOP Compliance"""
    wb = Workbook()

    sheets = [
        ("Case_Registry", "Master case file registry and SOP compliance"),
        ("DPP_Pipeline", "DPP submission tracking and rulings"),
        ("SOP_Checklist", "Per-case SOP compliance checklist"),
        ("Witness_Statements", "Witness and statement management"),
        ("Disclosure_Log", "Disclosure protocol compliance"),
        ("Dashboard", "Case management KPIs, pipeline charts, SOP compliance"),
        ("Lookups", "Dropdown reference data"),
    ]
    build_home(wb, "Registry", REGISTRY_BROWN, sheets)
    DATA_ROWS = 20

    # -- CASE REGISTRY (Master File) --
    ws = wb.create_sheet("Case_Registry")
    ws.sheet_properties.tabColor = REGISTRY_BROWN
    cr_headers = [
        "CaseID", "RegistrationDate", "Classification", "OIC_Name",
        "OIC_Rank", "OIC_BadgeNo", "Parish", "Division",
        "OffenceDescription", "LawAndSection",
        "SuspectName", "SuspectDOB", "SuspectAddress", "SuspectOccupation",
        "VictimName", "VictimAddress",
        "LinkedIntelID", "LinkedOpID", "LinkedArrestID",
        "CaseStatus", "FileCompleteness", "SOPCompliance",
        "DPPSubmissionDate", "DPPStatus", "DPPRuling",
        "CourtType", "NextCourtDate", "TrialDate",
        "Verdict", "Sentence",
        "POCAReferred", "POCAStatus",
        "CreatedBy", "CreatedDate", "LastUpdated", "Notes",
    ]
    cr_widths = [
        14, 14, 22, 18, 14, 12, 14, 14,
        30, 30,
        20, 12, 24, 16, 20, 24,
        14, 14, 14,
        22, 22, 22,
        14, 22, 22,
        24, 14, 12,
        14, 20,
        10, 18,
        14, 12, 12, 30,
    ]
    write_headers(ws, cr_headers, cr_widths)

    sample_cases = [
        ["CR-2025-001", "2025-01-20", "Firearms", "DC Williams",
         "Detective Corporal", "JCF-4455", "Manchester", "Area 3 HQ",
         "Illegal possession of 2 firearms and ammunition",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "John BROWN", "1990-05-12", "15 Main St, Mandeville", "Labourer",
         "", "",
         "INT-2025-001", "OP-2025-001", "AR-2025-001",
         "Open - Pending Forensics", "Incomplete - Missing Forensics",
         "Non-Compliant - Missing Forensics",
         "", "Awaiting Ballistic Certificate", "",
         "Gun Court - High Court Division", "2025-04-15", "",
         "", "",
         "No", "Not Applicable",
         "DC Williams", "2025-01-20", "2025-03-01",
         "Awaiting ballistic cert from IFSLM before DPP submission"],
        ["CR-2025-002", "2025-01-25", "Narcotics", "Supt. Davis",
         "Superintendent", "JCF-2201", "St. Elizabeth", "Area 3 HQ",
         "Import of 32kg cocaine via fishing vessel",
         "Import/Export - Cocaine (DDA Part IV)",
         "Unknown - vessel crew fled", "", "", "",
         "", "",
         "INT-2025-002", "OP-2025-002", "",
         "Open - Active Investigation", "Incomplete - Missing Statements",
         "Non-Compliant - Missing Statements",
         "", "File Being Prepared", "",
         "", "", "",
         "", "",
         "Yes", "Referred to FID",
         "Supt. Davis", "2025-01-25", "2025-02-15",
         "Vessel crew not yet identified. Coast Guard assisting."],
        ["CR-2025-003", "2025-03-05", "Firearms + Narcotics", "Supt. Davis",
         "Superintendent", "JCF-2201", "Manchester", "Area 3 HQ",
         "Stockpiling 5 firearms, trafficking 2kg cocaine",
         "s.6 - Stockpiling (3+ weapons / 50+ rounds)",
         "Devon GRANT", "1985-08-23", "Christiana", "Unemployed",
         "", "",
         "INT-2025-006", "OP-2025-004", "AR-2025-002",
         "Referred to DPP", "Substantially Complete (>80%)",
         "Compliant",
         "2025-03-25", "Crown Counsel Reviewing", "",
         "Gun Court - High Court Division", "2025-04-20", "",
         "", "",
         "Yes", "Restraint Order Applied",
         "Supt. Davis", "2025-03-05", "2025-03-25",
         "Major case - MOCA joint op. DPP file submitted."],
        ["CR-2025-004", "2025-03-22", "Firearms + Narcotics", "DC Williams",
         "Detective Corporal", "JCF-4455", "Clarendon", "Area 3 HQ",
         "Possession of firearm and cocaine during vehicle stop",
         "s.5 - Possession of Prohibited Weapon (15-25yr)",
         "Andrew BLAKE", "1988-02-15", "May Pen", "Driver",
         "", "",
         "INT-2025-009", "OP-2025-006", "AR-2025-004",
         "Open - Pending DPP Submission", "Substantially Complete (>80%)",
         "Compliant",
         "", "File Being Prepared", "",
         "Gun Court - RM Division", "2025-04-05", "",
         "", "",
         "No", "Not Applicable",
         "DC Williams", "2025-03-22", "2025-03-30",
         "Straightforward case. Lab cert for cocaine pending."],
    ]

    for i, row_data in enumerate(sample_cases):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i+2, column=j+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    data_rows(ws, len(cr_headers), start_row=len(sample_cases)+2, count=DATA_ROWS)
    add_submission_tracking(ws, len(cr_headers), len(sample_cases)+DATA_ROWS)
    cr_end = len(sample_cases) + 1 + DATA_ROWS
    add_table(ws, "tbl_CaseRegistry",
              f"A1:{get_column_letter(len(cr_headers))}{cr_end}")

    # Conditional formatting for compliance
    ws.conditional_formatting.add(
        f"V2:V{cr_end}",
        CellIsRule(operator="equal", formula=['"Compliant"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(
        f"V2:V{cr_end}",
        CellIsRule(operator="beginsWith", formula=['"Non-Compliant"'], fill=RED_FILL))

    # File completeness formatting
    ws.conditional_formatting.add(
        f"U2:U{cr_end}",
        CellIsRule(operator="equal", formula=['"Complete"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(
        f"U2:U{cr_end}",
        CellIsRule(operator="beginsWith", formula=['"Incomplete"'], fill=RED_FILL))

    dv_list(ws, "C2:C500", "Lookups!$F$2:$F$12")
    dv_list(ws, "G2:G500", "Lookups!$C$2:$C$20")
    dv_list(ws, "T2:T500", "Lookups!$G$2:$G$15")
    dv_list(ws, "U2:U500", "Lookups!$K$2:$K$10")
    dv_list(ws, "V2:V500", "Lookups!$I$2:$I$10")
    dv_list(ws, "X2:X500", "Lookups!$H$2:$H$15")

    # -- DPP PIPELINE --
    ws2 = wb.create_sheet("DPP_Pipeline")
    ws2.sheet_properties.tabColor = REGISTRY_BROWN
    dpp_headers = [
        "CaseID", "Classification", "OIC_Name", "SuspectName",
        "OffenceSummary", "DPPFileDate", "CrownCounselAssigned",
        "DPPStatus", "EvidentialSufficiency", "PublicInterestMet",
        "RulingDate", "RulingOutcome", "RulingNotes",
        "VoluntaryBillRequired", "PrelimExamRequired",
        "ReturnedForInvestigation", "ReturnReason",
        "FileResubmissionDate", "Notes",
    ]
    dpp_widths = [
        14, 22, 18, 20, 30, 12, 20,
        22, 14, 14,
        12, 22, 30,
        14, 14,
        14, 28,
        14, 30,
    ]
    write_headers(ws2, dpp_headers, dpp_widths)
    data_rows(ws2, len(dpp_headers), count=DATA_ROWS)
    add_submission_tracking(ws2, len(dpp_headers), DATA_ROWS)
    dpp_end = 1 + DATA_ROWS
    add_table(ws2, "tbl_DPPPipeline",
              f"A1:{get_column_letter(len(dpp_headers))}{dpp_end}")

    dv_list(ws2, "H2:H500", "Lookups!$H$2:$H$15")
    dv_list(ws2, "I2:I500", "Lookups!$D$2:$D$5")
    dv_list(ws2, "J2:J500", "Lookups!$D$2:$D$5")
    dv_list(ws2, "N2:N500", "Lookups!$D$2:$D$5")
    dv_list(ws2, "O2:O500", "Lookups!$D$2:$D$5")
    dv_list(ws2, "P2:P500", "Lookups!$D$2:$D$5")

    # -- SOP CHECKLIST --
    ws3 = wb.create_sheet("SOP_Checklist")
    ws3.sheet_properties.tabColor = REGISTRY_BROWN
    sop_headers = [
        "CaseID", "OIC_Name", "ChecklistDate",
        # Station diary & initial records
        "StationDiaryEntry", "CrimeReportFiled", "OffenceRegisterUpdated",
        # Suspect processing
        "SuspectCautioned", "RightsAdvised", "DetaineeBookEntry",
        "PropertyBookEntry", "48hrComplianceMet",
        # Evidence & exhibits
        "ExhibitRegisterUpdated", "ExhibitsPhotographed", "ExhibitsSealed",
        "ChainOfCustodyStarted", "ForensicSubmissionsMade",
        # Statements
        "VictimStatementTaken", "WitnessStatementsTaken",
        "SuspectStatementCautioned", "OfficerStatements",
        # Investigation
        "ScenePhotographed", "SceneSketchDrawn", "IDParadeRequired",
        "IDParadeConducted",
        # DPP readiness
        "ForensicCertsReceived", "BallisticCertReceived",
        "PostMortemReceived", "DPPFileComplete",
        "DisclosurePackagePrepared",
        # Overall
        "OverallCompliance", "ComplianceNotes",
    ]
    sop_widths = [14, 18, 12] + [12]*26 + [22, 30]
    write_headers(ws3, sop_headers, sop_widths)
    data_rows(ws3, len(sop_headers), count=DATA_ROWS)
    add_submission_tracking(ws3, len(sop_headers), DATA_ROWS)
    sop_end = 1 + DATA_ROWS
    add_table(ws3, "tbl_SOPChecklist",
              f"A1:{get_column_letter(len(sop_headers))}{sop_end}")

    # Yes/No dropdowns for all checklist columns
    for col_idx in range(4, len(sop_headers)-1):
        col_letter = get_column_letter(col_idx)
        dv_list(ws3, f"{col_letter}2:{col_letter}500", "Lookups!$D$2:$D$5")
    dv_list(ws3, f"{get_column_letter(len(sop_headers)-1)}2:{get_column_letter(len(sop_headers)-1)}500",
            "Lookups!$I$2:$I$10")

    # Conditional formatting - No = red, Yes = green
    for col_idx in range(4, len(sop_headers)-1):
        col_letter = get_column_letter(col_idx)
        ws3.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{sop_end}",
            CellIsRule(operator="equal", formula=['"No"'], fill=RED_FILL))
        ws3.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{sop_end}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=GREEN_FILL))

    # -- WITNESS STATEMENTS --
    ws4 = wb.create_sheet("Witness_Statements")
    ws4.sheet_properties.tabColor = REGISTRY_BROWN
    wit_headers = [
        "StatementID", "CaseID", "WitnessName", "WitnessType",
        "WitnessAddress", "WitnessPhone", "RelationToCase",
        "StatementDate", "StatementTakenBy", "StatementPages",
        "StatementSigned", "WitnessWilling", "SpecialMeasuresNeeded",
        "SpecialMeasuresType", "AvailableForCourt", "Notes",
    ]
    wit_widths = [14, 14, 20, 16, 24, 14, 20, 12, 18, 10, 10, 10, 12, 20, 12, 28]
    write_headers(ws4, wit_headers, wit_widths)
    data_rows(ws4, len(wit_headers), count=DATA_ROWS)
    add_submission_tracking(ws4, len(wit_headers), DATA_ROWS)
    wit_end = 1 + DATA_ROWS
    add_table(ws4, "tbl_WitnessStatements",
              f"A1:{get_column_letter(len(wit_headers))}{wit_end}")

    # -- DISCLOSURE LOG --
    ws5 = wb.create_sheet("Disclosure_Log")
    ws5.sheet_properties.tabColor = REGISTRY_BROWN
    disc_headers = [
        "DisclosureID", "CaseID", "DisclosureDate", "DisclosureType",
        "MaterialDisclosed", "ServedOnDefence", "DefenceSolicitor",
        "ServiceMethod", "ServiceDate", "AcknowledgementReceived",
        "PIIApplicationMade", "PIIOutcome",
        "SupplementaryDisclosureNeeded", "SupplementaryDate",
        "DisclosureStatus", "PreparedBy", "Notes",
    ]
    disc_widths = [14, 14, 12, 18, 30, 10, 20, 14, 12, 12, 10, 16, 12, 12, 18, 16, 28]
    write_headers(ws5, disc_headers, disc_widths)
    data_rows(ws5, len(disc_headers), count=DATA_ROWS)
    add_submission_tracking(ws5, len(disc_headers), DATA_ROWS)
    disc_end = 1 + DATA_ROWS
    add_table(ws5, "tbl_DisclosureLog",
              f"A1:{get_column_letter(len(disc_headers))}{disc_end}")

    dv_list(ws5, "O2:O500", "Lookups!$J$2:$J$10")

    # -- DASHBOARD --
    dash = wb.create_sheet("Dashboard")
    dash.sheet_properties.tabColor = REGISTRY_BROWN
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:L2")
    dash.cell(row=2, column=2,
              value="REGISTRY UNIT DASHBOARD - Case Management & SOP Compliance"
              ).font = TITLE_FONT
    dash["B2"].alignment = CENTER

    dash.merge_cells("B3:L3")
    dash.cell(row=3, column=2,
              value="Per DPP Prosecution Protocol (2012) & Disclosure Protocol (2013)"
              ).font = Font(name="Calibri", size=9, color=MED_GRAY, italic=True)
    dash["B3"].alignment = CENTER

    # KPIs
    kpi_card(dash, 5, 2, "Total Cases",
             f"=COUNTA(Case_Registry!A2:A{cr_end})")
    kpi_card(dash, 5, 5, "SOP Compliant",
             f'=COUNTIF(Case_Registry!V2:V{cr_end},"Compliant")')
    kpi_card(dash, 5, 8, "Files Complete",
             f'=COUNTIF(Case_Registry!U2:U{cr_end},"Complete")')
    kpi_card(dash, 5, 11, "At DPP",
             f'=COUNTIF(Case_Registry!T2:T{cr_end},"Referred to DPP")')

    # Case status breakdown
    case_stats = ["Open - Active Investigation", "Open - Pending Forensics",
                  "Open - Pending DPP Submission", "Referred to DPP",
                  "Closed - Convicted", "Closed - Acquitted", "Closed - No Charge"]
    r = 9
    dash.cell(row=r, column=2, value="Case Status").font = BODY_BOLD
    dash.cell(row=r, column=3, value="Count").font = BODY_BOLD
    for i, cs in enumerate(case_stats):
        dash.cell(row=r+1+i, column=2, value=cs).font = BODY_FONT
        dash.cell(row=r+1+i, column=3,
                  value=f'=COUNTIF(Case_Registry!T2:T{cr_end},B{r+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Cases by Status",
                  cats_ref={"min_col": 2, "min_row": r+1, "max_row": r+len(case_stats)},
                  vals_ref={"min_col": 3, "min_row": r+1, "max_row": r+len(case_stats)},
                  anchor="E9", horizontal=True)

    # DPP pipeline
    dpp_stats = ["File Being Prepared", "File Submitted to DPP",
                 "Crown Counsel Reviewing", "Ruling - Charge Approved",
                 "Ruling - No Charge", "Returned for Further Investigation",
                 "Awaiting Forensic Certificate", "Awaiting Ballistic Certificate"]
    pr = r + len(case_stats) + 3
    dash.cell(row=pr, column=2, value="DPP Status").font = BODY_BOLD
    dash.cell(row=pr, column=3, value="Count").font = BODY_BOLD
    for i, ds in enumerate(dpp_stats):
        dash.cell(row=pr+1+i, column=2, value=ds).font = BODY_FONT
        dash.cell(row=pr+1+i, column=3,
                  value=f'=COUNTIF(Case_Registry!X2:X{cr_end},B{pr+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "DPP Submission Pipeline",
                  cats_ref={"min_col": 2, "min_row": pr+1, "max_row": pr+len(dpp_stats)},
                  vals_ref={"min_col": 3, "min_row": pr+1, "max_row": pr+len(dpp_stats)},
                  anchor=f"E{pr}")

    # SOP compliance chart
    sop_stats = ["Compliant", "Non-Compliant - Missing Statements",
                 "Non-Compliant - Missing Forensics",
                 "Non-Compliant - Missing Exhibit Register",
                 "Non-Compliant - Overdue DPP Submission", "Under Review"]
    qr = pr + len(dpp_stats) + 3
    dash.cell(row=qr, column=2, value="SOP Compliance").font = BODY_BOLD
    dash.cell(row=qr, column=3, value="Count").font = BODY_BOLD
    for i, ss in enumerate(sop_stats):
        dash.cell(row=qr+1+i, column=2, value=ss).font = BODY_FONT
        dash.cell(row=qr+1+i, column=3,
                  value=f'=COUNTIF(Case_Registry!V2:V{cr_end},B{qr+1+i})'
                  ).font = BODY_FONT

    add_pie_chart(dash, "SOP Compliance Status",
                  cats_ref={"min_col": 2, "min_row": qr+1, "max_row": qr+len(sop_stats)},
                  vals_ref={"min_col": 3, "min_row": qr+1, "max_row": qr+len(sop_stats)},
                  anchor=f"E{qr}")

    # Classification breakdown
    classifs = ["Firearms", "Narcotics", "Firearms + Narcotics",
                "Trafficking - Firearms", "Trafficking - Narcotics",
                "POCA / Financial"]
    xr = qr + len(sop_stats) + 3
    dash.cell(row=xr, column=2, value="Classification").font = BODY_BOLD
    dash.cell(row=xr, column=3, value="Count").font = BODY_BOLD
    for i, cl in enumerate(classifs):
        dash.cell(row=xr+1+i, column=2, value=cl).font = BODY_FONT
        dash.cell(row=xr+1+i, column=3,
                  value=f'=COUNTIF(Case_Registry!C2:C{cr_end},B{xr+1+i})'
                  ).font = BODY_FONT

    add_bar_chart(dash, "Cases by Classification",
                  cats_ref={"min_col": 2, "min_row": xr+1, "max_row": xr+len(classifs)},
                  vals_ref={"min_col": 3, "min_row": xr+1, "max_row": xr+len(classifs)},
                  anchor=f"E{xr}")

    build_lookups(wb, REGISTRY_LOOKUPS, REGISTRY_BROWN)

    path = os.path.join(OUTPUT_DIR, "FNID_Registry_Portal.xlsx")
    wb.save(path)
    return path


# =================================================================
#  7. COMMAND DASHBOARD - Aggregated cross-unit view
# =================================================================
def build_command_dashboard():
    """Generate FNID_Command_Dashboard.xlsx - Aggregated cross-unit KPIs.

    Since this workbook cannot directly reference other workbooks with
    formulas (Excel external references require absolute file paths),
    this dashboard uses a Python-populated summary approach:
    it reads the sample data counts and creates static summary + charts.
    Run the refresh script to update from live portal data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Command_Dashboard"
    ws.sheet_properties.tabColor = NAVY
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:M2")
    ws.cell(row=2, column=2,
            value="FNID AREA 3 - COMMAND DASHBOARD").font = Font(
                name="Calibri", bold=True, size=22, color=NAVY)
    ws["B2"].alignment = CENTER

    ws.merge_cells("B3:M3")
    ws.cell(row=3, column=2,
            value="Manchester | St. Elizabeth | Clarendon | Aggregated Operational View"
            ).font = Font(name="Calibri", size=11, color=MED_GRAY, italic=True)
    ws["B3"].alignment = CENTER

    ws.merge_cells("B4:M4")
    ws.cell(row=4, column=2,
            value=f"Last Refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ).font = Font(name="Calibri", size=9, color=MED_GRAY)
    ws["B4"].alignment = CENTER

    # UNIT SUMMARY TABLE
    r = 6
    summary_headers = ["Unit", "Total Records", "Submitted", "Pending",
                       "Compliance Rate"]
    for i, h in enumerate(summary_headers):
        cell = ws.cell(row=r, column=2+i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Pre-computed from sample data
    unit_data = [
        ("Intel", 10, 10, 0, "100%"),
        ("Operations", 5, 5, 0, "100%"),
        ("Seizures - Firearms", 5, 5, 0, "100%"),
        ("Seizures - Narcotics", 3, 3, 0, "100%"),
        ("Seizures - Other", 2, 2, 0, "100%"),
        ("Arrests", 4, 4, 0, "100%"),
        ("Forensics - Exhibits", 3, 3, 0, "100%"),
        ("Forensics - Lab", 3, 3, 0, "100%"),
        ("Registry - Cases", 4, 4, 0, "50%"),
    ]
    for i, (unit, total, sub, pend, rate) in enumerate(unit_data):
        row = r + 1 + i
        ws.cell(row=row, column=2, value=unit).font = BODY_BOLD
        ws.cell(row=row, column=3, value=total).font = BODY_FONT
        ws.cell(row=row, column=4, value=sub).font = BODY_FONT
        ws.cell(row=row, column=5, value=pend).font = BODY_FONT
        ws.cell(row=row, column=6, value=rate).font = BODY_FONT
        for c in range(2, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = CENTER

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    for c in range(3, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # OPERATIONAL HIGHLIGHTS
    sr = r + len(unit_data) + 2
    ws.merge_cells(f"B{sr}:F{sr}")
    ws.cell(row=sr, column=2, value="OPERATIONAL HIGHLIGHTS").font = SUB_FONT

    highlights = [
        "Total firearms seized: 7 (including 1 IBIS hit linked to Kingston shooting)",
        "Total narcotics seized: 34.2 kg cocaine (est. street value $102.6M JMD)",
        "Total arrests: 4 (100% within 48-hour rule)",
        "Active cases at DPP: 1 (CR-2025-003 - Devon GRANT stockpiling)",
        "Pending forensic certificates: 2 (ballistics for CR-2025-001, chemistry for CR-2025-004)",
        "POCA referrals: 2 cases referred to FID",
        "SOP compliance rate: 50% (2 of 4 cases fully compliant)",
    ]
    for i, h in enumerate(highlights):
        ws.cell(row=sr+1+i, column=2, value=f"  {h}").font = BODY_FONT

    # UNIT RECORDS BAR CHART
    chart_data_start = r + 1
    chart_data_end = r + len(unit_data)

    chart = BarChart()
    chart.title = "Records by Unit"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    data = Reference(ws, min_col=3, min_row=chart_data_start,
                     max_row=chart_data_end)
    cats = Reference(ws, min_col=2, min_row=chart_data_start,
                     max_row=chart_data_end)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"H{r}")

    # SEIZURES PIE CHART
    seizure_r = sr + len(highlights) + 2
    ws.cell(row=seizure_r, column=2, value="Seizure Type").font = BODY_BOLD
    ws.cell(row=seizure_r, column=3, value="Count").font = BODY_BOLD
    seizure_data = [("Firearms", 7), ("Narcotics (entries)", 3),
                    ("Other", 2), ("Ammunition (rounds)", 500)]
    for i, (st, cnt) in enumerate(seizure_data):
        ws.cell(row=seizure_r+1+i, column=2, value=st).font = BODY_FONT
        ws.cell(row=seizure_r+1+i, column=3, value=cnt).font = BODY_FONT

    pie = PieChart()
    pie.title = "Seizure Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    pie_data = Reference(ws, min_col=3, min_row=seizure_r+1,
                         max_row=seizure_r+len(seizure_data))
    pie_cats = Reference(ws, min_col=2, min_row=seizure_r+1,
                         max_row=seizure_r+len(seizure_data))
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, f"E{seizure_r}")

    # MONTHLY TREND (cross-unit)
    trend_r = seizure_r + len(seizure_data) + 8
    ws.cell(row=trend_r, column=2, value="Month").font = BODY_BOLD
    ws.cell(row=trend_r, column=3, value="Intel").font = BODY_BOLD
    ws.cell(row=trend_r, column=4, value="Operations").font = BODY_BOLD
    ws.cell(row=trend_r, column=5, value="Seizures").font = BODY_BOLD
    ws.cell(row=trend_r, column=6, value="Arrests").font = BODY_BOLD

    # Estimated monthly data from samples
    monthly_data = [
        ("Jan", 2, 2, 3, 1), ("Feb", 2, 1, 0, 0), ("Mar", 4, 2, 5, 3),
        ("Apr", 2, 0, 0, 0), ("May", 0, 0, 0, 0), ("Jun", 0, 0, 0, 0),
        ("Jul", 0, 0, 0, 0), ("Aug", 0, 0, 0, 0), ("Sep", 0, 0, 0, 0),
        ("Oct", 0, 0, 0, 0), ("Nov", 0, 0, 0, 0), ("Dec", 0, 0, 0, 0),
    ]
    for i, (m, intel, ops, seiz, arr) in enumerate(monthly_data):
        row = trend_r + 1 + i
        ws.cell(row=row, column=2, value=m).font = BODY_FONT
        ws.cell(row=row, column=3, value=intel).font = BODY_FONT
        ws.cell(row=row, column=4, value=ops).font = BODY_FONT
        ws.cell(row=row, column=5, value=seiz).font = BODY_FONT
        ws.cell(row=row, column=6, value=arr).font = BODY_FONT

    line = LineChart()
    line.title = "Monthly Operational Tempo"
    line.style = 10
    line.width = 22
    line.height = 12
    for col_idx in range(3, 7):
        data = Reference(ws, min_col=col_idx, min_row=trend_r+1,
                         max_row=trend_r+12)
        line.add_data(data, titles_from_data=False)
    cats = Reference(ws, min_col=2, min_row=trend_r+1, max_row=trend_r+12)
    line.set_categories(cats)
    line.series[0].tx = SeriesLabel(v="Intel")
    line.series[1].tx = SeriesLabel(v="Operations")
    line.series[2].tx = SeriesLabel(v="Seizures")
    line.series[3].tx = SeriesLabel(v="Arrests")
    ws.add_chart(line, f"H{trend_r-2}")

    # PORTAL INDEX
    idx_r = trend_r + 15
    ws.merge_cells(f"B{idx_r}:F{idx_r}")
    ws.cell(row=idx_r, column=2, value="UNIT PORTAL FILES").font = SUB_FONT
    portals = [
        ("FNID_Intel_Portal.xlsx", "Intelligence pipeline"),
        ("FNID_Operations_Portal.xlsx", "Operation planning & execution"),
        ("FNID_Seizures_Portal.xlsx", "Firearms, narcotics, other seizures"),
        ("FNID_ArrestsCourt_Portal.xlsx", "Arrests, bail, court tracking"),
        ("FNID_Forensics_Portal.xlsx", "Chain of custody, lab tracking"),
        ("FNID_Registry_Portal.xlsx", "Case management & SOP compliance"),
    ]
    for i, (fname, desc) in enumerate(portals):
        ws.cell(row=idx_r+1+i, column=2, value=fname).font = BODY_BOLD
        ws.cell(row=idx_r+1+i, column=4, value=desc).font = BODY_FONT

    path = os.path.join(OUTPUT_DIR, "FNID_Command_Dashboard.xlsx")
    wb.save(path)
    return path


# =================================================================
#  MAIN - Generate all portals
# =================================================================
def generate_all():
    """Generate all 6 unit portals + command dashboard."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 60)
    print("FNID Area 3 - Siloed Unit Portal System v1.0")
    print("=" * 60)

    builders = [
        ("Intel Unit", build_intel_portal),
        ("Operations Unit", build_operations_portal),
        ("Seizures Unit", build_seizures_portal),
        ("Arrests & Court Unit", build_arrests_court_portal),
        ("Forensics & Evidence Unit", build_forensics_portal),
        ("Registry Unit (SOP Compliance)", build_registry_portal),
        ("Command Dashboard", build_command_dashboard),
    ]

    for i, (name, builder) in enumerate(builders, 1):
        print(f"  [{i}/{len(builders)}] {name}...", end=" ", flush=True)
        path = builder()
        size_kb = os.path.getsize(path) / 1024
        print(f"OK ({size_kb:.0f} KB)")

    print("=" * 60)
    print(f"ALL PORTALS GENERATED in: {OUTPUT_DIR}/")
    print()
    print("Portal files:")
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.endswith(".xlsx"):
            fpath = os.path.join(OUTPUT_DIR, f)
            print(f"  {f} ({os.path.getsize(fpath)/1024:.0f} KB)")
    print()
    print("Features per portal:")
    print("  - Data entry sheets with dropdown validations")
    print("  - Submission tracking (Draft/Submitted/Edited)")
    print("  - Dashboard with KPI cards and charts")
    print("  - Sample data pre-populated for demonstration")
    print("  - Conditional formatting for compliance/status")
    print()
    print("Next steps:")
    print("  1. Open each portal in Excel to verify")
    print("  2. Use fnid_data_import.py to import from Word/Excel documents")
    print("  3. Officers enter data and set RecordStatus to Submit/Edit")
    print("=" * 60)


if __name__ == "__main__":
    generate_all()
