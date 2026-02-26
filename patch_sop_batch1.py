#!/usr/bin/env python3
"""
SOP Compliance Patch — Batch 1: New Back-End Tables
Adds: Witnesses, Statements, CaseAssignments, VictimContact,
      InvestigationPlans, CaseFileChecklist, Disclosure, CommandConference
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from generate_fnid_workbook import (
    HEADER_FONT, HEADER_FILL, BODY_FONT, BODY_FONT_BOLD,
    CENTER, LEFT_WRAP, THIN_BORDER,
    GREEN_FILL, AMBER_FILL, RED_FILL, LIGHT_GRAY_FILL,
    TABLE_STYLE, NAVY, DARK_BLUE, LIGHT_BLUE, WHITE,
    write_headers, add_table, add_data_validation_list,
    SECTION_FONT, SUBHEADER_FONT, SUBHEADER_FILL, GOLD_FILL,
    MED_GRAY
)


# ══════════════════════════════════════════════════════════════════════
#  TABLE 1 — WITNESS MANAGEMENT
# ══════════════════════════════════════════════════════════════════════

def build_witnesses_table(wb):
    ws = wb.create_sheet("tbl_Witnesses")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "WitnessID","LinkedCaseID","FullName","Alias","DOB","Gender",
        "Address","Parish","Phone","Email",
        "WitnessType","RelationToCase",
        "StatementTaken","StatementDate","StatementID",
        "WillingToTestify","AvailableForCourt","CourtDate",
        "SummonsIssued","SummonsDate","SummonsServed","SummonsServedDate",
        "ProtectionRequired","ProtectionType","ProtectionApproved",
        "VulnerableWitness","SpecialMeasures",
        "CredibilityAssessment","InterviewCount",
        "LastContactDate","NextContactDue",
        "Status","Notes",
        "CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,22,16,14,10,
              24,16,16,20,
              16,20,
              12,14,14,
              14,14,14,
              12,14,12,14,
              14,18,14,
              14,20,
              18,12,
              14,14,
              12,24,
              14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","WIT-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Witnesses", ref)
    add_data_validation_list(ws, "H2:H1000", "=Lookups!$D$2:$D$16")
    # Yes/No validations
    for col in ["M","P","Q","S","U","W","Y","Z"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    # Contact overdue
    ws.conditional_formatting.add(
        "AE2:AE1000",
        FormulaRule(formula=['=AND(AE2<>"",AE2<TODAY())'], fill=RED_FILL)
    )
    # Summons not served
    ws.conditional_formatting.add(
        "U2:U1000",
        FormulaRule(formula=['=AND(S2="Yes",U2="No")'], fill=AMBER_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 2 — FORMAL STATEMENT REGISTER
# ══════════════════════════════════════════════════════════════════════

def build_statements_table(wb):
    ws = wb.create_sheet("tbl_Statements")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "StatementID","LinkedCaseID","LinkedWitnessID","LinkedPersonID",
        "StatementDate","StatementTime","Location",
        "StatementType","DeponentName","DeponentRole",
        "TakenBy","TakenByRank","TakenByBadge",
        "CautionGiven","CautionTime","RightsExplained",
        "LegalRepPresent","LegalRepName",
        "InterpreterUsed","InterpreterName","Language",
        "StatementPages","ExhibitsReferenced",
        "VideoRecorded","AudioRecorded",
        "DeponentSignature","WitnessToSignature",
        "DisclosedToProsecution","DisclosureDate",
        "Retracted","RetractionDate","RetractionReason",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,14,
              14,12,20,
              16,22,16,
              18,14,14,
              12,12,12,
              12,20,
              12,18,14,
              10,20,
              12,12,
              12,18,
              14,14,
              12,14,24,
              24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","STMT-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Statements", ref)
    # Yes/No validations
    for col in ["N","P","Q","S","X","Y","Z","AA","AC"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 3 — CASE ASSIGNMENT / HANDOVER LOG
# ══════════════════════════════════════════════════════════════════════

def build_case_assignments(wb):
    ws = wb.create_sheet("tbl_CaseAssignments")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "AssignmentID","LinkedCaseID","AssignmentDate","AssignmentTime",
        "AssignedFrom","AssignedFromRank","AssignedFromBadge",
        "AssignedTo","AssignedToRank","AssignedToBadge","AssignedToUnit",
        "AssignmentType","Reason",
        "AuthorizedBy","AuthorizedByRank",
        "HandoverBriefingDone","HandoverDate","HandoverNotes",
        "CaseFileTransferred","FileCondition",
        "ExhibitsAccountedFor","ExhibitDiscrepancies",
        "PendingActions","OutstandingLeads",
        "AcknowledgedByReceiver","AcknowledgementDate",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,14,12,
              18,14,14,
              18,14,14,14,
              16,24,
              18,14,
              14,14,30,
              14,18,
              14,24,
              24,20,
              14,14,
              24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","ASGN-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_CaseAssignments", ref)
    for col in ["P","S","U","Y"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 4 — VICTIM / COMPLAINANT CONTACT LOG
# ══════════════════════════════════════════════════════════════════════

def build_victim_contact(wb):
    ws = wb.create_sheet("tbl_VictimContact")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "ContactID","LinkedCaseID","VictimName","ContactDate","ContactTime",
        "ContactMethod","ContactedBy","ContactedByRank",
        "Purpose","UpdateProvided","VictimSatisfied",
        "ConcernsRaised","ActionRequired","ActionDueDate",
        "ReferralMade","ReferralAgency","ReferralDate",
        "VictimImpactAssessment","SafetyConcerns","SafetyPlanInPlace",
        "NextScheduledContact","ContactFrequency",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,22,14,12,
              16,18,14,
              20,24,12,
              24,24,14,
              12,20,14,
              20,18,14,
              14,14,
              24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","VCON-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_VictimContact", ref)
    for col in ["K","O","T"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    # Overdue next contact
    ws.conditional_formatting.add(
        "U2:U1000",
        FormulaRule(formula=['=AND(U2<>"",U2<TODAY())'], fill=RED_FILL)
    )
    # Overdue action
    ws.conditional_formatting.add(
        "N2:N1000",
        FormulaRule(formula=['=AND(N2<>"",N2<TODAY())'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 5 — INVESTIGATION PLANS
# ══════════════════════════════════════════════════════════════════════

def build_investigation_plans(wb):
    ws = wb.create_sheet("tbl_InvestigationPlans")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "PlanID","LinkedCaseID","PlanVersion","PlanDate",
        "PreparedBy","PreparedByRank","ApprovedBy","ApprovalDate",
        "Hypothesis","ObjectivesOfInvestigation",
        "LinesOfEnquiry","KeyEvidenceRequired",
        "WitnessesToInterview","SuspectsToInterview",
        "ForensicRequirements","SurveillanceRequired",
        "SearchesPlanned","WarrantsRequired",
        "InterAgencySupport","ResourcesRequired",
        "TimelineTarget","MilestonesPlanned",
        "RiskAssessment","OfficerSafetyConsiderations",
        "LegalConsiderations","DisclosureObligations",
        "ReviewSchedule","NextReviewDate",
        "PlanStatus","LastRevisedDate",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,12,14,
              18,14,18,14,
              36,36,
              36,30,
              24,24,
              24,14,
              14,14,
              20,24,
              14,30,
              24,24,
              24,20,
              16,14,
              14,14,
              24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","IPLAN-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_InvestigationPlans", ref)
    for col in ["P","R"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    # Overdue review
    ws.conditional_formatting.add(
        "AB2:AB1000",
        FormulaRule(formula=['=AND(AB2<>"",AB2<TODAY())'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 6 — CASE FILE COMPLETENESS CHECKLIST
# ══════════════════════════════════════════════════════════════════════

def build_case_file_checklist(wb):
    ws = wb.create_sheet("tbl_CaseFileChecklist")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "ChecklistID","LinkedCaseID","ReviewDate","ReviewedBy",
        # Core documents
        "CR1_Present","CR2_Present","CR5_Present",
        "StationDiaryExtract","InvestigationPlan",
        "AllStatementsPresent","StatementCount",
        "CautionStatementPresent","ChargeSheet",
        # Evidence & Forensics
        "ExhibitRegisterComplete","ChainOfCustodyComplete",
        "ForensicReportsReceived","ForensicReportsCount",
        "ScenePhotos","SceneSketch","SceneReport",
        # Court readiness
        "WitnessList","WitnessSummonses","WitnessAvailabilityConfirmed",
        "ProsecutionBrief","DisclosureComplete",
        # Suspect documentation
        "SuspectCautionRecord","SuspectInterviewRecord",
        "IdentificationEvidence","AlibiChecked",
        # Administrative
        "SupervisoryEndorsements","CommandConferenceNotes",
        "CorrespondenceFileFiled","VictimContactLog",
        # Scoring
        "TotalItemsChecked","TotalItemsPresent","CompletenessScore",
        "FileStatus",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,14,18,
              12,12,12,
              14,14,
              14,12,
              14,12,
              14,14,
              14,14,
              12,12,14,
              14,14,14,
              14,14,
              14,14,
              14,12,
              14,14,
              14,14,
              14,14,14,
              14,
              24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","FCHK-"&TEXT(ROW()-1,"0000"))')
        # Count checked items (cols E through AF = cols 5 to 32, but only Yes/No ones)
        # 28 checkable items (E through AF)
        ws.cell(row=r, column=33, value=28)  # TotalItemsChecked = fixed 28
        # TotalItemsPresent = count of "Yes" in the check columns
        ws.cell(row=r, column=34,
                value=f'=COUNTIF(E{r}:AF{r},"Yes")')
        # CompletenessScore = Present/Total
        ws.cell(row=r, column=35,
                value=f'=IF(AG{r}=0,"",AH{r}/AG{r})')
        ws.cell(row=r, column=35).number_format = '0%'
        # FileStatus auto
        ws.cell(row=r, column=36,
                value=f'=IF(AI{r}="","",IF(AI{r}>=0.95,"Complete",IF(AI{r}>=0.75,"Substantially Complete",IF(AI{r}>=0.5,"Incomplete","Critically Incomplete"))))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_CaseFileChecklist", ref)
    # Yes/No for all check columns (E through AF = columns 5 through 32)
    for col_num in range(5, 33):
        cl = get_column_letter(col_num)
        if cl not in ['K','Q']:  # StatementCount and ForensicReportsCount are numbers
            add_data_validation_list(ws, f"{cl}2:{cl}1000", "=Lookups!$I$2:$I$3")
    # Completeness color coding
    ws.conditional_formatting.add(
        "AI2:AI1000",
        FormulaRule(formula=['=AND(AI2<>"",AI2>=0.95)'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        "AI2:AI1000",
        FormulaRule(formula=['=AND(AI2<>"",AI2>=0.5,AI2<0.95)'], fill=AMBER_FILL)
    )
    ws.conditional_formatting.add(
        "AI2:AI1000",
        FormulaRule(formula=['=AND(AI2<>"",AI2<0.5)'], fill=RED_FILL)
    )
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 7 — PROSECUTION DISCLOSURE LOG
# ══════════════════════════════════════════════════════════════════════

def build_disclosure_log(wb):
    ws = wb.create_sheet("tbl_Disclosure")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "DisclosureID","LinkedCaseID","DisclosureDate",
        "MaterialType","MaterialDescription","ExhibitRef","StatementRef",
        "DisclosedTo","DisclosedToRole","DisclosedBy","DisclosedByRank",
        "Method","ReceiptAcknowledged","AcknowledgementDate",
        "Sensitive","PIIMaterial","RedactionRequired","RedactionDone",
        "ThirdPartyMaterial","ThirdPartyConsent",
        "ScheduleRef","ContinuingDuty","LastReviewDate",
        "Notes","CreatedBy","CreatedDate"
    ]
    widths = [14,14,14,
              16,28,14,14,
              20,16,18,14,
              14,14,14,
              12,12,12,12,
              12,12,
              14,12,14,
              24,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","DISC-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_Disclosure", ref)
    for col in ["M","O","P","Q","R","S","T","V"]:
        add_data_validation_list(ws, f"{col}2:{col}1000", "=Lookups!$I$2:$I$3")
    return ws


# ══════════════════════════════════════════════════════════════════════
#  TABLE 8 — COMMAND / CASE CONFERENCE NOTES
# ══════════════════════════════════════════════════════════════════════

def build_command_conferences(wb):
    ws = wb.create_sheet("tbl_CommandConference")
    ws.sheet_properties.tabColor = "1F3864"
    headers = [
        "ConferenceID","LinkedCaseID","ConferenceDate","ConferenceTime",
        "ConferenceType","Chairperson","ChairpersonRank",
        "Attendees","AttendeeCount",
        "CaseSynopsis","InvestigationProgress",
        "EvidenceStatusSummary","ForensicStatusSummary",
        "KeyIssuesDiscussed","DirectivesIssued",
        "DirectiveAssignedTo","DirectiveDueDate","DirectiveStatus",
        "ResourceApprovals","EscalationRequired","EscalatedTo",
        "NextConferenceDate","MinutesTakenBy",
        "Notes","CreatedBy","CreatedDate","LastUpdatedBy","LastUpdatedDate"
    ]
    widths = [14,14,14,12,
              16,20,14,
              30,10,
              36,36,
              24,24,
              36,36,
              20,14,14,
              24,12,18,
              14,18,
              24,14,14,14,14]
    write_headers(ws, headers, widths=widths)
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","CONF-"&TEXT(ROW()-1,"0000"))')
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    ref = f"A1:{get_column_letter(len(headers))}6"
    add_table(ws, "tbl_CommandConference", ref)
    add_data_validation_list(ws, "T2:T1000", "=Lookups!$I$2:$I$3")
    # Overdue directives
    ws.conditional_formatting.add(
        "Q2:Q1000",
        FormulaRule(formula=['=AND(Q2<>"",Q2<TODAY(),R2<>"Completed")'], fill=RED_FILL)
    )
    # Next conference overdue
    ws.conditional_formatting.add(
        "V2:V1000",
        FormulaRule(formula=['=AND(V2<>"",V2<TODAY())'], fill=AMBER_FILL)
    )
    return ws


def main_batch1():
    fname = "FNID_Area3_Operational_Workbook.xlsx"
    print(f"SOP Batch 1: Loading {fname}...")
    wb = openpyxl.load_workbook(fname)

    print("[1/8] Adding tbl_Witnesses...")
    build_witnesses_table(wb)
    print("[2/8] Adding tbl_Statements...")
    build_statements_table(wb)
    print("[3/8] Adding tbl_CaseAssignments...")
    build_case_assignments(wb)
    print("[4/8] Adding tbl_VictimContact...")
    build_victim_contact(wb)
    print("[5/8] Adding tbl_InvestigationPlans...")
    build_investigation_plans(wb)
    print("[6/8] Adding tbl_CaseFileChecklist...")
    build_case_file_checklist(wb)
    print("[7/8] Adding tbl_Disclosure...")
    build_disclosure_log(wb)
    print("[8/8] Adding tbl_CommandConference...")
    build_command_conferences(wb)

    wb.save(fname)
    print(f"Batch 1 saved: {fname} — {len(wb.sheetnames)} sheets")
    return fname

if __name__ == "__main__":
    main_batch1()
